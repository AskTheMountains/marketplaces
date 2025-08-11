import pandas as pd
import numpy as np
import shutil
from loguru import logger
import openpyxl
from datetime import date,datetime,timedelta
from itertools import product
from openpyxl import Workbook, formatting
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.formatting.rule import ColorScale, ColorScaleRule, FormatObject
from openpyxl.utils import get_column_letter
# from options import client_name

# from options import settings, headers, client_number
# # Файл с некоторыми константами
# from constants import clusters_mapping, clusters_mapping_jewelry, catalog_supply_columns, svod_total_columns, svod_clusters_columns
# # Таблица объединения кластеров в один (какие в какой объединяем)
# clusters_mapping_df = pd.DataFrame(clusters_mapping)
# clusters_mapping_jewelry_df = pd.DataFrame(clusters_mapping_jewelry)

# date_report_created = '2024-12-17'
# svod_excel = pd.read_excel(f"Clients/{client_name}/SupplySvod/{date_report_created}_Расчет_Поставок_{client_name}_Ozon.xlsx", sheet_name='Всего')
# svod_excel_clusters = pd.read_excel(f"Clients/{client_name}/SupplySvod/{date_report_created}_Расчет_Поставок_{client_name}_Ozon.xlsx", sheet_name='По кластерам')

# Создание копии оригинального файла, в котором будет производиться форматирование
def copy_supply_svod_file(path_supply_svod, client_name, date_report_created):
    src_file = f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_Ozon.xlsx"
    dst_file = f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_Ozon_formatted.xlsx"
    shutil.copy(src_file, dst_file)

# Форматирование листа "Всего"
def format_sheet_total(
        path_supply_svod,
        client_name,
        date_report_created,
        svod_excel,
        svod_excel_clusters,
        clusters_mapping_df,
        sheet_name='Всего'
    ):

    logger.info(f"Formatting sheet \"{sheet_name}\"")

    # Открываем файл с расчетами Excel
    wb = openpyxl.load_workbook(f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_Ozon_formatted.xlsx")
    ws = wb[sheet_name]

    # Кол-во строк в df
    svod_len = svod_excel.shape[0]

    # Номер строки, откуда начинается запись
    row_start = 2

    # df с соответствием заголовков и названий столбцов
    excel_columns = pd.DataFrame({"column": svod_excel.columns,
                                  "column_number": np.arange(1, len(svod_excel.columns) + 1)})
    excel_columns['excel_column'] = excel_columns['column_number'].apply(lambda x: get_column_letter(x))

    # Выборка различных групп колонок по их названиям
    # Мин. и макс. колонка
    min_col = excel_columns.loc[excel_columns['column_number'].idxmin(), 'excel_column']
    max_col = excel_columns.loc[excel_columns['column_number'].idxmax(), 'excel_column']
    # Заголовки
    header_cells = ws[f"{min_col}{row_start - 1}:{max_col}{row_start - 1}"]
    # Все колонки, кроме заголовков
    all_cells = ws[f"{min_col}{row_start}:{max_col}{svod_len + 1}"]
    # Артикул
    article = excel_columns.loc[excel_columns['column'] == 'Артикул', 'excel_column'].values[0]
    # Штрихкод
    barcode = excel_columns.loc[excel_columns['column'] == 'Штрихкод', 'excel_column'].values[0]
    # Оборачиваемость
    turnover = excel_columns.loc[excel_columns['column'] == 'Оборачиваемость', 'excel_column'].values[0]
    # Заказы
    orders = excel_columns.loc[excel_columns['column'].str.contains('ЗАКАЗЫ'), 'excel_column'].values[0]
    # Продажи
    sales = excel_columns.loc[excel_columns['column'].str.contains('ПРОДАЖИ'), 'excel_column'].values[0]
    # Остатки
    reminders = excel_columns.loc[excel_columns['column'].str.contains('ОСТАТОК') & ~(excel_columns['column'].str.contains('FBS')), 'excel_column'].values[0]
    # Остатки FBS
    reminders_fbs = excel_columns.loc[excel_columns['column'].str.contains('ОСТАТОК') & excel_columns['column'].str.contains('FBS'), 'excel_column'].values[0]
    # Товары_в_пути
    # products_delivering = excel_columns.loc[excel_columns['column'] == 'Товары_в_пути', 'excel_column'].values[0]
    # ОЖИДАЕТСЯ_ПОСТУПЛЕНИЕ
    supplying = excel_columns.loc[excel_columns['column'] == 'ОЖИДАЕТСЯ_ПОСТУПЛЕНИЕ', 'excel_column'].values[0]
    # Потребность на 40 дней
    demand_40_days = excel_columns.loc[excel_columns['column'] == 'Потребность на 40 дней', 'excel_column'].values[0]
    # Потребность на 60 дней
    demand_60_days = excel_columns.loc[excel_columns['column'] == 'Потребность на 60 дней', 'excel_column'].values[0]
    # Дефицит\Избыток 40 дней
    deficit_40_days = excel_columns.loc[excel_columns['column'] == 'Дефицит/Избыток 40 дней', 'excel_column'].values[0]
    # Дефицит\Избыток 40 дней с учетом FBS
    deficit_40_days_fbs = excel_columns.loc[excel_columns['column'] == 'Дефицит/Избыток 40 дней с учетом FBS', 'excel_column'].values[0]
    # Дефицит\Избыток 60 дней
    deficit_60_days = excel_columns.loc[excel_columns['column'] == 'Дефицит/Избыток 60 дней', 'excel_column'].values[0]
    # Потребность на 40 дней (округл.)
    demand_40_days_rounded = excel_columns.loc[excel_columns['column'] == 'Потребность на 40 дней (округл.)', 'excel_column'].values[0]
    # Потребность на 40 дней с учетом FBS (округл.)
    demand_40_days_rounded_fbs = excel_columns.loc[excel_columns['column'] == 'Потребность на 40 дней с учетом FBS (округл.)', 'excel_column'].values[0]
    # Потребность на 60 дней (округл.)
    demand_60_days_rounded = excel_columns.loc[excel_columns['column'] == 'Потребность на 60 дней (округл.)', 'excel_column'].values[0]

    # Нужные колонки с листа По кластерам
    # Кол-во строк в df
    svod_len_clusters = svod_excel_clusters.shape[0]

    # Номер строки, откуда начинается запись
    row_start_clusters = 2

    # df с соответствием заголовков и названий столбцов
    excel_columns_clusters = pd.DataFrame({"column": svod_excel_clusters.columns,
                                  "column_number": np.arange(1, len(svod_excel_clusters.columns) + 1)})
    excel_columns_clusters['excel_column'] = excel_columns_clusters['column_number'].apply(lambda x: get_column_letter(x))

    # Артикул (Кластеры)
    article_from_clusters = excel_columns_clusters.loc[excel_columns_clusters['column'] == 'Артикул', 'excel_column'].values[0]
    # Кластер (Кластеры)
    cluster_from_clusters = excel_columns_clusters.loc[excel_columns_clusters['column'] == 'Кластер', 'excel_column'].values[0]
    # Корректировка с учетом нулевых остатков (Кластеры)
    correction_null_reminders_from_clusters = excel_columns_clusters.loc[excel_columns_clusters['column'] == 'Корректировка с учетом нулевых остатков', 'excel_column'].values[0]
    # Остаток (Кластеры)
    reminders_from_clusters = excel_columns_clusters.loc[excel_columns_clusters['column'].str.contains('ОСТАТОК') & ~(excel_columns_clusters['column'].str.contains('FBS')), 'excel_column'].values[0]

    # Форматирование
    # Стили границ
    thin = Side(border_style="thin", color="000000")

    # Форматирование заголовков таблицы
    for row in header_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11,
                                       bold=True)
            cell.alignment = Alignment(horizontal='center',
                                                 wrap_text=True)
            cell.border = Border(top = thin, bottom = thin,
                                                        right = thin, left = thin)
    # Значок фильтра на столбцы
    ws.auto_filter.ref = ws.dimensions

    # Границы (сетка)
    for row in all_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11)
            cell.border = Border(top = thin, bottom = thin,
                                                        right = thin, left = thin)
    # Автоподбор ширины столбца
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Формулы
    for i in range(row_start, svod_len + 1):
        # Оборачиваемость
        ws[f'{turnover}{i}'] = f'= ({orders}{i} + {sales}{i}) / 2 / 30'
        # Потребность на 40 дней
        ws[f'{demand_40_days}{i}'] = f'= {turnover}{i} * 40'
        # Потребность на 60 дней
        ws[f'{demand_60_days}{i}'] = f'= {turnover}{i} * 60'
        # Дефицит/избыток 40 дней
        ws[f'{deficit_40_days}{i}'] = f'= {reminders}{i} + {supplying}{i} - {demand_40_days}{i}'
        # Дефицит/избыток 40 дней с учетом FBS
        ws[f'{deficit_40_days_fbs}{i}'] = f'= {reminders}{i} + {reminders_fbs}{i} + {supplying}{i} - {demand_40_days}{i}'
        # Дефицит/избыток 60 дней
        ws[f'{deficit_60_days}{i}'] = f'= {reminders}{i} + {supplying}{i} - {demand_60_days}{i}'
        # Потребность на 40 дней (округл.)
        ws[f'{demand_40_days_rounded}{i}'] = f'= ROUNDUP(IF({deficit_40_days}{i} > 0, 0, -{deficit_40_days}{i}),0)'
        # Потребность на 40 дней с учетом FBS (округл.)
        ws[f'{demand_40_days_rounded_fbs}{i}'] = f'= ROUNDUP(IF({deficit_40_days_fbs}{i} > 0, 0, -{deficit_40_days_fbs}{i}),0)'
        # Потребность на 60 дней (округл.)
        ws[f'{demand_60_days_rounded}{i}'] = f'= ROUNDUP(IF({deficit_60_days}{i} > 0, 0, -{deficit_60_days}{i}),0)'

    # Суммирование данных с листа по кластерам
    clusters_mapping_df_ = clusters_mapping_df.copy()
    # Объединяем кластер и его субкластеры в один лист
    # clusters_mapping_df_['cluster_and_subcluster'] = clusters_mapping_df_['cluster'].map(list) + clusters_mapping_df_['subcluster']
    for i in range(clusters_mapping_df_.shape[0]):
        # Выбираем список кластеров из таблицы соответствия
        cluster = clusters_mapping_df_['cluster'][i]
        subcluster_list = clusters_mapping_df_['subcluster'][i].copy()
        # Добавляем кластер в список субкластеров
        subcluster_list.insert(0, cluster)
        # Удаляем дубликаты из субкластеров
        subcluster_list = list(dict.fromkeys(subcluster_list))
        # Выбираем колонку с потребностью и остатками для кластера на листе всего
        demand_cluster_column = excel_columns.loc[excel_columns['column'] == f"Потребность {cluster}", 'excel_column'].values[0]
        reminders_cluster_column = excel_columns.loc[excel_columns['column'] == f"Остаток {cluster}", 'excel_column'].values[0]
        for k in range(row_start, svod_len + 1):
            # Строка, куда будет помещаться итоговая формула
            subcluster_formula_demand = ''
            subcluster_formula_reminders = ''
            for subcluster in subcluster_list:
                # Прописываем формулу, включающую в себя сумму по субкластерам
                subcluster_formula_demand_tmp = (
                     f"SUMIFS('По кластерам'!${correction_null_reminders_from_clusters}${row_start_clusters}:${correction_null_reminders_from_clusters}${svod_len_clusters + 1}, "
                     f"'По кластерам'!${article_from_clusters}${row_start_clusters}:${article_from_clusters}${svod_len_clusters + 1}, ${article}{k}, "
                     f"'По кластерам'!${cluster_from_clusters}${row_start_clusters}:${cluster_from_clusters}${svod_len_clusters + 1}, \"{subcluster}\""
                     f")\n"
                )
                subcluster_formula_reminders_tmp = (
                     f"SUMIFS('По кластерам'!${reminders_from_clusters}${row_start_clusters}:${reminders_from_clusters}${svod_len_clusters + 1}, "
                     f"'По кластерам'!${article_from_clusters}${row_start_clusters}:${article_from_clusters}${svod_len_clusters + 1}, ${article}{k}, "
                     f"'По кластерам'!${cluster_from_clusters}${row_start_clusters}:${cluster_from_clusters}${svod_len_clusters + 1}, \"{subcluster}\""
                     f")\n"
                )
                # Суммируем с предыдущим кластером
                subcluster_formula_demand = f"{subcluster_formula_demand_tmp} + {subcluster_formula_demand}"
                subcluster_formula_reminders = f"{subcluster_formula_reminders_tmp} + {subcluster_formula_reminders}"
                # Удаляем лишние символы в конце строки, если это последний субкластер
                subcluster_formula_demand = subcluster_formula_demand.rstrip(" + ''")
                subcluster_formula_reminders = subcluster_formula_reminders.rstrip(" + ''")
            # Добавляем = вначале формулы, чтобы эксель воспринимал это как формулу
            subcluster_formula_demand = f"= {subcluster_formula_demand}"
            subcluster_formula_reminders = f"= {subcluster_formula_reminders}"
            # Вставляем формулу в нужную ячейку
            ws[f'{demand_cluster_column}{k}'] = subcluster_formula_demand
            ws[f'{reminders_cluster_column}{k}'] = subcluster_formula_reminders

    # Условное форматирование для колонок Дефицит\Избыток на 40 и 60 дней
    red_font = Font(color='9C0006')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_font = Font(color='006100')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    deficit_cols = [deficit_40_days, deficit_40_days_fbs, deficit_60_days]
    for i in range(len(deficit_cols)):
        deficit_cols[i] = f"{deficit_cols[i]}{row_start}:{deficit_cols[i]}{svod_len + 1}"
    for cell_range in deficit_cols:
        ws.conditional_formatting.add(cell_range,
                                formatting.rule.CellIsRule(
                                    operator='lessThan',
                                    formula=['0'],
                                    fill=red_fill, font=red_font))
        ws.conditional_formatting.add(cell_range,
                                formatting.rule.CellIsRule(
                                    operator='greaterThan',
                                    formula=['0'],
                                    fill=green_fill, font=green_font))
    # 2 знака после запятой у чисел
    svod_2num_cells = [turnover, demand_40_days, demand_60_days]
    for i in range(len(svod_2num_cells)):
        svod_2num_cells[i] = f"{svod_2num_cells[i]}{row_start}:{svod_2num_cells[i]}{svod_len + 1}"
    for cell_range in svod_2num_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.number_format = "0.00"

    # 1 знак после запятой у чисел
    svod_1num_cells = [deficit_40_days, deficit_40_days_fbs, deficit_60_days]
    for i in range(len(svod_1num_cells)):
        svod_1num_cells[i] = f"{svod_1num_cells[i]}{row_start}:{svod_1num_cells[i]}{svod_len + 1}"
    for cell_range in svod_1num_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.number_format = "0.0"

    # 0 знаков после запятой у чисел
    svod_0num_cells = [barcode]
    for i in range(len(svod_0num_cells)):
        svod_0num_cells[i] = f"{svod_0num_cells[i]}{row_start}:{svod_0num_cells[i]}{svod_len + 1}"
    for cell_range in svod_0num_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.number_format = "0"

    # Выравнивание по центру
    svod_middle_cells = excel_columns.loc[~excel_columns['column'].str.contains('Сезон|Статус|Основной артикул|Артикул|Штрихкод|Наименование|Категория|Цвет'), 'excel_column'].to_list()
    for i in range(len(svod_middle_cells)):
        svod_middle_cells[i] = f"{svod_middle_cells[i]}{row_start}:{svod_middle_cells[i]}{svod_len + 1}"
    for cell_range in svod_middle_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.alignment = Alignment(horizontal = "center",
                                                            vertical = "center")
    # Выравнивание слева
    svod_left_cells = excel_columns.loc[excel_columns['column'].str.contains('Сезон|Статус|Основной артикул|Артикул|Штрихкод|Наименование|Категория|Цвет'), 'excel_column'].to_list()
    for i in range(len(svod_left_cells)):
        svod_left_cells[i] = f"{svod_left_cells[i]}{row_start}:{svod_left_cells[i]}{svod_len + 1}"
    for cell_range in svod_left_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.alignment = Alignment(horizontal = "left",
                                                            vertical = "center")

    #return wb
    # Сохранение файла Excel
    wb.save(f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_Ozon_formatted.xlsx")
    logger.info("Finished formatting sheet \"Всего\"'")


# Форматирование листа "Сводная по кластерам"
def format_sheet_pivot_clusters(date_report_created, svod_excel):
    print('jopa')

# Форматирование листа "По кластерам"
def format_sheet_clusters(
        path_supply_svod,
        client_name,
        date_report_created,
        svod_excel
    ):
    # Номер строки, откуда начинается запись
    logger.info('Formatting sheet \"По кластерам\"')
    # Открываем файл с расчетами Excel
    wb = openpyxl.load_workbook(f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_Ozon_formatted.xlsx")
    ws = wb['По кластерам']

    # Кол-во строк в df
    svod_len = svod_excel.shape[0]

    # Номер строки, откуда начинается запись
    row_start = 2

    # df с соответствием заголовков и названий столбцов
    excel_columns = pd.DataFrame({"column": svod_excel.columns,
                                  "column_number": np.arange(1, len(svod_excel.columns) + 1)})
    excel_columns['excel_column'] = excel_columns['column_number'].apply(lambda x: get_column_letter(x))

    # Выборка различных групп колонок по их названиям
    # Мин. и макс. колонка
    min_col = excel_columns.loc[excel_columns['column_number'].idxmin(), 'excel_column']
    max_col = excel_columns.loc[excel_columns['column_number'].idxmax(), 'excel_column']
    # Заголовки
    header_cells = ws[f"{min_col}{row_start - 1}:{max_col}{row_start - 1}"]
    # Все колонки, кроме заголовков
    all_cells = ws[f"{min_col}{row_start}:{max_col}{svod_len + 1}"]
    # Штрихкод
    barcode = excel_columns.loc[excel_columns['column'] == 'Штрихкод', 'excel_column'].values[0]
    # Оборачиваемость
    turnover = excel_columns.loc[excel_columns['column'] == 'Оборачиваемость', 'excel_column'].values[0]
    # Заказы
    orders = excel_columns.loc[excel_columns['column'].str.contains('ЗАКАЗЫ'), 'excel_column'].values[0]
    # Продажи
    sales = excel_columns.loc[excel_columns['column'].str.contains('ПРОДАЖИ'), 'excel_column'].values[0]
    # Остатки
    reminders = excel_columns.loc[excel_columns['column'].str.contains('ОСТАТОК'), 'excel_column'].values[0]
    # Товары_в_пути
    # products_delivering = excel_columns.loc[excel_columns['column'] == 'Товары_в_пути', 'excel_column'].values[0]
    # ОЖИДАЕТСЯ_ПОСТУПЛЕНИЕ
    supplying = excel_columns.loc[excel_columns['column'] == 'ОЖИДАЕТСЯ_ПОСТУПЛЕНИЕ', 'excel_column'].values[0]
    # Потребность на 40 дней
    demand_40_days = excel_columns.loc[excel_columns['column'] == 'Потребность на 40 дней', 'excel_column'].values[0]
    # Потребность на 60 дней
    demand_60_days = excel_columns.loc[excel_columns['column'] == 'Потребность на 60 дней', 'excel_column'].values[0]
    # Дефицит\Избыток 40 дней
    deficit_40_days = excel_columns.loc[excel_columns['column'] == 'Дефицит/Избыток 40 дней', 'excel_column'].values[0]
    # Дефицит\Избыток 60 дней
    deficit_60_days = excel_columns.loc[excel_columns['column'] == 'Дефицит/Избыток 60 дней', 'excel_column'].values[0]
    # Потребность на 40 дней (округл.)
    demand_40_days_rounded = excel_columns.loc[excel_columns['column'] == 'Потребность на 40 дней (округл.)', 'excel_column'].values[0]
    # Потребность на 60 дней (округл.)
    demand_60_days_rounded = excel_columns.loc[excel_columns['column'] == 'Потребность на 60 дней (округл.)', 'excel_column'].values[0]
    # Корректировка с учетом нулевых остатков
    correction_null_reminders = excel_columns.loc[excel_columns['column'] == 'Корректировка с учетом нулевых остатков', 'excel_column'].values[0]

    # Форматирование
    # Стили границ
    thin = Side(border_style="thin", color="000000")

    # Форматирование заголовков таблицы
    for row in header_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11,
                                       bold=True)
            cell.alignment = Alignment(horizontal='center',
                                                 wrap_text=True)
            cell.border = Border(top = thin, bottom = thin,
                                                        right = thin, left = thin)
    # Значок фильтра на столбцы
    ws.auto_filter.ref = ws.dimensions

    # Границы (сетка)
    for row in all_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11)
            cell.border = Border(top = thin, bottom = thin,
                                                        right = thin, left = thin)
    # Автоподбор ширины столбца
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Формулы
    for i in range(row_start, svod_len + 1):
        # Оборачиваемость
        ws[f'{turnover}{i}'] = f'= ({orders}{i} + {sales}{i}) / 2 / 30'
        # Потребность на 40 дней
        ws[f'{demand_40_days}{i}'] = f'= {turnover}{i} * 40'
        # Потребность на 60 дней
        ws[f'{demand_60_days}{i}'] = f'= {turnover}{i} * 60'
        # Дефицит/избыток 40 дней
        ws[f'{deficit_40_days}{i}'] = f'= {reminders}{i} + {supplying}{i} - {demand_40_days}{i}'
        # Дефицит/избыток 60 дней
        ws[f'{deficit_60_days}{i}'] = f'= {reminders}{i} + {supplying}{i} - {demand_60_days}{i}'
        # Потребность на 40 дней (округл.)
        ws[f'{demand_40_days_rounded}{i}'] = f'= ROUNDUP(IF({deficit_40_days}{i} > 0, 0, -{deficit_40_days}{i}),0)'
        # Потребность на 60 дней (округл.)
        ws[f'{demand_60_days_rounded}{i}'] = f'= ROUNDUP(IF({deficit_60_days}{i} > 0, 0, -{deficit_60_days}{i}),0)'
        # Корректировка с учетом нулевых остатков
        ws[f'{correction_null_reminders}{i}'] = f'= {demand_60_days_rounded}{i}'

    # Условное форматирование для колонок Дефицит\Избыток на 40 и 60 дней
    red_font = Font(color='9C0006')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_font = Font(color='006100')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    deficit_cols = [deficit_40_days, deficit_60_days]
    for i in range(len(deficit_cols)):
        deficit_cols[i] = f"{deficit_cols[i]}{row_start}:{deficit_cols[i]}{svod_len + 1}"
    for cell_range in deficit_cols:
        ws.conditional_formatting.add(cell_range,
                                formatting.rule.CellIsRule(
                                    operator='lessThan',
                                    formula=['0'],
                                    fill=red_fill, font=red_font))
        ws.conditional_formatting.add(cell_range,
                                formatting.rule.CellIsRule(
                                    operator='greaterThan',
                                    formula=['0'],
                                    fill=green_fill, font=green_font))
    # 2 знака после запятой у чисел
    svod_2num_cells = [turnover, demand_40_days, demand_60_days]
    for i in range(len(svod_2num_cells)):
        svod_2num_cells[i] = f"{svod_2num_cells[i]}{row_start}:{svod_2num_cells[i]}{svod_len + 1}"
    for cell_range in svod_2num_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.number_format = "0.00"

    # 1 знак после запятой у чисел
    svod_1num_cells = [deficit_40_days, deficit_60_days]
    for i in range(len(svod_1num_cells)):
        svod_1num_cells[i] = f"{svod_1num_cells[i]}{row_start}:{svod_1num_cells[i]}{svod_len + 1}"
    for cell_range in svod_1num_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.number_format = "0.0"

    # 0 знаков после запятой у чисел
    svod_0num_cells = [barcode]
    for i in range(len(svod_0num_cells)):
        svod_0num_cells[i] = f"{svod_0num_cells[i]}{row_start}:{svod_0num_cells[i]}{svod_len + 1}"
    for cell_range in svod_0num_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.number_format = "0"

    # Выравнивание по центру
    svod_middle_cells = excel_columns.loc[~excel_columns['column'].str.contains('Сезон|Статус|Основной артикул|Артикул|Штрихкод|Наименование|Категория|Кластер|Цвет'), 'excel_column'].to_list()
    for i in range(len(svod_middle_cells)):
        svod_middle_cells[i] = f"{svod_middle_cells[i]}{row_start}:{svod_middle_cells[i]}{svod_len + 1}"
    for cell_range in svod_middle_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.alignment = Alignment(horizontal = "center",
                                                            vertical = "center")
    # Выравнивание слева
    svod_left_cells = excel_columns.loc[excel_columns['column'].str.contains('Сезон|Статус|Основной артикул|Артикул|Штрихкод|Наименование|Категория|Кластер|Цвет'), 'excel_column'].to_list()
    for i in range(len(svod_left_cells)):
        svod_left_cells[i] = f"{svod_left_cells[i]}{row_start}:{svod_left_cells[i]}{svod_len + 1}"
    for cell_range in svod_left_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.alignment = Alignment(horizontal = "left",
                                                            vertical = "center")
    #return wb
    # Сохранение файла Excel
    wb.save(f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_Ozon_formatted.xlsx")
    logger.info("Finished formatting sheet \"По кластерам\"'")
