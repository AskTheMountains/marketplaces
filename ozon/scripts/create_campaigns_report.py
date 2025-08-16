# %% Определение функций
import requests
import json
import time
from datetime import date,datetime,timedelta
import pandas as pd
import shutil
import os
import glob
import send2trash
import csv
import zipfile
from zipfile import ZipFile
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_00
import re
from loguru import logger
import getopt
import sys
from pathlib import Path

# Папка, где лежит текущий скрипт
BASE_DIR = Path(__file__).parent.parent


# Файл с настройками и номером клиента
# from options import settings, client_number, headers
# Некоторые константы
from ozon.scripts.constants import (
    headers,
    client_name,
    marketplace_dir_name,
    client_id_performance,
    client_secret_performance,
    ozon_performance_api_url,
    promotion_companies
)
# Функции выгрузки данных по API Seller
from ozon.scripts.uploadDataFromOzon import(
    get_ozon_product,
    getOrders,
    getStockReminders_fbo_v2,
    getStockReminders_fbs
)
# Функции выгрузки данных по API Performane
from ozon.scripts.uploadDataFromOzonPerformance import (
    getAuthorizationToken,
    getCompanyList,
    getCompanyStatistics,
)
# Некоторые вспомогательные функции
from generic_functions import move_columns


# Создание нужных директорий
# def create_dirs():
#     # Папка с клиентом
#     client_dir = f"Clients/{client_name}/"
#     # Список директорий для создания
#     dir_names = [
#         'CampaignsReport',
#         f"CampaignsReport/{str(date.today())}/{str(date.today())}_Кампании_API"
#     ]
#     for dir_name in dir_names:
#         dir_path = os.path.join(client_dir, dir_name)
#         if not os.path.exists(dir_path):
#             logger.info(f"Creating folder {dir_path} for client {client_name}")
#             os.makedirs(dir_path)


# Создание директории для выгрузки статистики РК за конкретные даты
def create_upload_statistic_dir(date_campaign_report, client_name):
    # Задаем путь к директории
    # upload_path = (
    #     f"Clients/{client_name}/CampaignsReport/{str(date.today())}/"
    #     f"{dt_start_dir}-{dt_end_dir}_API/"
    # )
    upload_path = (
        f"{BASE_DIR}/"
        f"Clients/{client_name}/CampaignsReport/"
        f"{date_campaign_report}/{date_campaign_report}_Кампании_API"
    )
    if not os.path.exists(upload_path):
        os.makedirs(upload_path, exist_ok=True)
        logger.info(f"Creating Companies Upload directory:{upload_path}")

    return upload_path

# Функция создания директории, куда будет помещаться статистика кампаний
def create_upload_statistic_dir(client_name, date_report=str(date.today())):
    # Задаем путь к директории
    # upload_path = (
    #     f"{BASE_DIR}/"
    #     f"Clients/{client_name}/ClientSvod/"
    #     f"{date_report}/{date_report}_Кампании_API"
    # )
    upload_path = (
        f"{BASE_DIR}/"
        f"Clients/{client_name}/CampaignsReport/"
        f"{date_report}/{date_report}_Кампании_API"
    )
    if not os.path.exists(upload_path):
        os.makedirs(upload_path, exist_ok=True)
        logger.info(f"Creating Companies Upload directory:{upload_path}")

    return upload_path


# Функция создания диапазона дат
# GPT START ----
def generate_dates(
        reference_date: str = None,
        start_date: str = None,
        end_date: str = None
    ) -> pd.DataFrame:
    # Проверка на несовместимость параметров
    if reference_date and (start_date or end_date):
        raise ValueError("Нельзя одновременно задавать reference_date и start_date/end_date.")

    # Если указан диапазон — используем его
    if start_date and end_date:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    elif start_date or end_date:
        raise ValueError("Нужно задать одновременно start_date и end_date.")
    else:
        # Старая логика по reference_date
        if reference_date is None:
            today = datetime.now()
        else:
            today = datetime.strptime(reference_date, "%Y-%m-%d")

        # Логика в зависимости от дня недели
        if today.weekday() == 0:
            # Если сегодня понедельник (weekday() == 0)
            # Понедельник предыдущей недели:
            start_dt = today - timedelta(days=7)
            start_dt = start_dt - timedelta(days=start_dt.weekday())
            # Воскресенье предыдущей недели:
            end_dt = start_dt + timedelta(days=6)
        else:
            # Если не понедельник:
            # Понедельник текущей недели:
            start_dt = today - timedelta(days=today.weekday())
            # Вчерашний день:
            end_dt = today - timedelta(days=1)

    start_date_str = start_dt.strftime('%Y-%m-%d')
    end_date_str = end_dt.strftime('%Y-%m-%d')
    start_date_iso = start_dt.strftime('%Y-%m-%dT00:00:00Z')
    end_date_iso = end_dt.strftime('%Y-%m-%dT23:59:59Z')

    df_dates = pd.DataFrame([{
        'date_start': start_date_str,
        'date_end': end_date_str,
        'datetime_start': start_date_iso,
        'datetime_end': end_date_iso
    }])

    return df_dates
# GPT END ----

# Функция получения начальной и конечной даты в формате строки
def get_start_end_date(df_date_range, type_dates='companies'):
    # Если задан тип дат - даты для заказов, то берем даты со временм
    if type_dates == 'orders':
        date_start, date_end = df_date_range.loc[0, ['datetime_start', 'datetime_end']]
    # В остальных случаях берем даты без времени
    else:
        date_start, date_end = df_date_range.loc[0, ['date_start', 'date_end']]

    return date_start, date_end


# Функция создания списка РК с датами
def create_input_companies(promotion_companies, df_date_range):
    # Создаем копию списка кампаний из констант
    input_companies = promotion_companies
    # Получаем даты для отчета РК
    date_start_performance, date_end_performance = get_start_end_date(df_date_range, type_dates='companies')
    # Добавляем даты к списку РК
    for company in input_companies:
        company.extend([date_start_performance, date_end_performance])

    return input_companies

# Функция проверки того, что нужна статистика по всем кампаниям
def check_all_companies_stats(input_companies):
    # Если в словаре по кампаниям не указан id кампании,
    # то выгружаем статистику по всем кампаниям
    if '' in input_companies:
        # Флаг проверки
        all_companies_stats_check = True
    else:
        all_companies_stats_check = False

    return all_companies_stats_check


# Функция создания датафрейма со списком кампаний и датами их выгрузки
def create_companies_upload_df(
        input_companies,
        all_companies_stats_check,
        upload_path,
):

    # Создаем df из словаря
    df_input_companies = (
        pd.DataFrame(
            input_companies,
            columns=['Номер кампании', 'date_start', 'date_end']
        )
        # .reset_index()
        # .rename(columns={
        #     'index': 'Номер кампании'
        # })
    )

    # Если нужна статистика по всем кампаниям, удаляем остальные строки из df
    if all_companies_stats_check:
        df_input_companies = (
            df_input_companies
            .loc[df_input_companies['Номер кампании'].isin(['']), :]
        )
    # Если нужна статистика только по определенным кампаниям,
    # то оставляем только строки с указанными номерами кампаний
    # Удаляем дубликаты
    else:
        df_input_companies = (
            df_input_companies
            .loc[~df_input_companies['Номер кампании'].isin(['']), :]
            .drop_duplicates(subset=['Номер кампании', 'date_start', 'date_end'])
        )

    # Сбрасываем index после loc
    df_input_companies = df_input_companies.reset_index(drop=True)

    # Переводим даты в нужный формат
    df_input_companies['dt_start_file'] = pd.to_datetime(df_input_companies['date_start']).dt.strftime('%d.%m.%Y')
    df_input_companies['dt_end_file'] = pd.to_datetime(df_input_companies['date_end']).dt.strftime('%d.%m.%Y')
    # Генерируем имя файла кампании
    df_input_companies['file_name_company'] = (
        df_input_companies['Номер кампании'] + '_' +
        df_input_companies['dt_start_file'] + '-' +
        df_input_companies['dt_end_file'] + '.csv'
    )
    # Формируем полный путь к файлу
    df_input_companies['file_path_company'] = df_input_companies['file_name_company'].apply(
        lambda x: os.path.join(upload_path, x)
    )
    # Выводим информацию о выгружаемых кампаниях в консоль
    df_input_companies_print = (
        df_input_companies
        .loc[:, ['Номер кампании', 'dt_start_file', 'dt_end_file']]
        .to_string(index=False)
    )
    logger.info(
        f"Creating report for companies:\n"
        f"{df_input_companies_print}"
    )

    return df_input_companies


# Функция выгрузка списка кампаний АПИ
def get_company_list(
        all_companies_stats_check,
        df_input_companies,
        auth_token
):
    # Если нужна статистика по всем кампаниям, выгружаем список всех кампаний
    if all_companies_stats_check:
        # Оставляем пустой лист
        companies_list = []
        # Получаем список всех кампаний по АПИ
        df_companies = getCompanyList(auth_token, companies_list)

        # Возвращаем мин и макс дату
        min_date_start = min(df_input_companies['date_start'])
        max_date_end = max(df_input_companies['date_end'])

        # Добавляем мин. и макс. даты в столбцы
        df_input_companies['date_start'] = min_date_start
        df_input_companies['date_end'] = max_date_end
        # Переводим даты в нужный формат
        df_input_companies['dt_start_file'] = pd.to_datetime(df_input_companies['date_start']).dt.strftime('%d.%m.%Y')
        df_input_companies['dt_end_file'] = pd.to_datetime(df_input_companies['date_end']).dt.strftime('%d.%m.%Y')
        # Генерируем имя файла кампании
        df_input_companies['file_name_company'] = (
            df_input_companies['Номер кампании'] + '_' +
            df_input_companies['dt_start_file'] + '-' +
            df_input_companies['dt_end_file'] + '.csv'
        )
        # Формируем полный путь к файлу
        df_input_companies['file_path_company'] = df_input_companies['file_name_company'].apply(
            lambda x: os.path.join(upload_path, x)
        )

        # Добавляем вспомогательные столбцы
        # df_input_companies_parsed = (
        #     df_companies
        #     # Переводим даты в формат, в котором они приходят из АПИ Performance
        #     .assign(
        #         date_start=min_date_start,
        #         date_end=max_date_end,
        #         dt_start_file=lambda x: pd.to_datetime(x['date_start']).dt.strftime('%d.%m.%Y'),
        #         dt_end_file=lambda x: pd.to_datetime(x['date_end']).dt.strftime('%d.%m.%Y')
        #     )
        #     # Получаем ожидаемое имя файла кампании
        #     .assign(
        #         file_name_company=lambda x: x['Номер кампании'] + '_' + x['dt_start_file'] + '-' + x['dt_end_file'] + '.csv',
        #         file_path_company=lambda x: BASE_DIR / x['file_name_company'],
        #     )
        # )

    # Если нужна статистика только определенных кампаний, формируем список этих кампаний
    else:
        # Формируем список только тех кампаний, которые нужны
        companies_list = df_input_companies['Номер кампании'].to_list()
        # Получаем список кампаний по апи
        df_companies = getCompanyList(auth_token, companies_list)
        # Объединяем с df с номерами кампаний
        df_input_companies_parsed = df_companies.merge(
            df_input_companies,
            how='left',
            on='Номер кампании'
        )

    # Убираем дубликаты по одинаковым кампаниям и датам для них
    df_input_companies_parsed = (
        df_input_companies_parsed
        .groupby(['Номер кампании', 'date_start', 'date_end'], as_index=False)
        .last()
    )
    # Создаем столбец с ID кампании по дате
    df_input_companies_parsed['unique_company_id'] = (
        df_input_companies_parsed
        .groupby(['Номер кампании', 'date_start', 'date_end'], as_index=False)
        .ngroup() + 1
    )
    # Переименовываем колонку с типом кампании
    df_input_companies_parsed = df_input_companies_parsed.rename(columns={
        'advObjectType': 'Тип кампании'
    })
    # Заменяем названия типов кампаний
    companies_ru_names = {
        'SKU': 'Трафареты',
        'BANNER': 'Баннер',
        'SEARCH_PROMO': 'Продвижение в поиске'
    }
    df_input_companies_parsed['Тип кампании'] = df_input_companies_parsed['Тип кампании'].replace(companies_ru_names)
    # Выбираем нужные колонки
    df_input_companies_parsed = df_input_companies_parsed.loc[:,[
        'unique_company_id',
        'Номер кампании',
        'Тип кампании',
        'file_name_company',
        'file_path_company',
        'date_start',
        'date_end',
        'dt_start_file',
        'dt_end_file',
    ]]

    return df_input_companies_parsed


# Функция получения списка кампаний, по которым статистика ещё не загружена
def get_missing_companies(
        # date_start,
        # date_end,
        df_input_companies_parsed,
        upload_path,
        # df_companies
):
    # # Получаем список кампаний
    # companies_list = df_companies['Номер кампании'].to_list()
    # # Проверяем,есть ли выгруженная статистика кампаний за конкретные даты
    # # Переводим даты в формат, в котором они находятся в именах файлах по РК
    # dt_start_file = pd.to_datetime(date_start).strftime('%d.%m.%Y')
    # dt_end_file = pd.to_datetime(date_end).strftime('%d.%m.%Y')
    # date_part = f'{dt_start_file}-{dt_end_file}'    # Формируем часть с датами в названии файлов
    # extension = '.csv'                     # расширение файлов

    # GPT START----
    # Формируем имя шаблона файла для каждого идентификатора
    # expected_files = [f"{company_id}_{date_part}{extension}" for company_id in companies_list]
    expected_files = df_input_companies_parsed['file_name_company'].to_list()
    # Получаем список всех файлов в папке
    actual_files = set(os.listdir(upload_path))

    # Проверяем наличие файлов для нужных кампаний и составляем итоговый список
    result_files = []
    missing = []
    for fname in expected_files:
        if fname in actual_files:
            result_files.append(os.path.join(upload_path, fname))
        else:
            missing.append(fname)

    # GPT END----
    # Если есть пропущенные кампании, то помещаем их в df
    if missing:
        # Создаем df с пропущенными кампаниями
        # df_missing_companies = pd.DataFrame({
        #     'file_name': missing
        # })
        # Получаем номера пропущенных кампаний
        missing_companies_ids = [x.split('_')[0] for x in missing]
        # Выбираем из df со списком кампаний пропущенные кампании
        df_missing_companies = (
            df_input_companies_parsed
            .loc[df_input_companies_parsed['Номер кампании'].isin(missing_companies_ids), :]
            .reset_index(drop=True)
        )
        # df для вывода в консоль, потому что я так хочу!
        df_missing_companies_print = df_missing_companies.loc[:, ['Номер кампании', 'dt_start_file', 'dt_end_file']].to_string(index=False)
        logger.info(
            f"\nБудет загружена статистика для кампаний:\n"
            f"{df_missing_companies_print}"
        )
        # Номер кампании
        # df_missing_companies['Номер кампании'] = df_missing_companies['file_name'].str.split('_').str[0]

    # Если все кампании присутствуют в директории, то возвращаем пустой df
    else:
        df_missing_companies = pd.DataFrame()

    return df_missing_companies


# Функция получения списка выгруженных файлов по РК
def get_companies_files_list(
        date_start,
        date_end,
        upload_path,
        # companies_list=[]
):
    # Переводим даты в формат, в котором они находятся в именах файлах по РК
    dt_start_file = pd.to_datetime(date_start).strftime('%d.%m.%Y')
    dt_end_file = pd.to_datetime(date_end).strftime('%d.%m.%Y')
    date_part = f'{dt_start_file}-{dt_end_file}'    # Формируем часть с датами в названии файлов
    extension = '.csv'                     # расширение файлов

    # Список файлов с наименованиями кампаний (будет заполнен далее)
    filenames_companies_files = []

    # # Формируем список файлов по заданным кампаниям, если они заданы
    # if companies_list:
    #     filenames_companies_files = [f"{upload_path}/{id_}_{date_part}{extension}" for id_ in companies_list]
    # # Если список кампаний не задан, считываем все файлы по кампаниям
    # else:
    # Маска имени файлов статистики РК
    path_companies = f"{upload_path}/*_{date_part}.{extension}"
    # Считывание файла с путем до него
    for file in glob.glob(path_companies):
        filenames_companies_files.append(file)

    # Получаем список csv файлов по кампаниям
    df_filenames_companies = pd.DataFrame({
        "path": filenames_companies_files,
    })
    # Считывание только имени файла
    df_filenames_companies['file_name'] = df_filenames_companies['path'].apply(lambda x: os.path.basename(x))
    # [os.path.basename(x) for x in glob.glob(filenames_companies['path'])]
    # Размер файла
    df_filenames_companies['file_size'] = df_filenames_companies['path'].apply(lambda x: os.path.getsize(x))
    # filenames_companies['file_size'] = [os.path.getsize(x) for x in glob.glob(filenames_companies_files)]

    # Номер кампании
    df_filenames_companies['Номер кампании'] = df_filenames_companies['file_name'].str.split('_').str[0]
    # Сортировка по размеру
    df_filenames_companies = df_filenames_companies.sort_values(by='file_size', ascending=False, ignore_index=True)

    return df_filenames_companies


# Функция выгрузки статистики по рекламным кампаниям
def get_company_statistics(
        df_input_companies_parsed,
        upload_path,
        auth_token,
):
    # Проверяем, есть ли кампании, по которым не загружена статистика
    df_missing_companies = get_missing_companies(
        df_input_companies_parsed,
        upload_path,
    )
    # Если файлов по каким-то кампаниям за указанные даты нет, то загружаем их
    if not df_missing_companies.empty:
        # Разделяем кампании по датам
        groups = dict(tuple(df_missing_companies.groupby(['date_start', 'date_end'])))
        # Перебор всех групп:
        for key, group_df in groups.items():
            # print(f'Группа: {key}')
            # print(group_df)
            # Получаем даты кампаний
            date_from = key[0]
            date_to = key[1]
            # Выгружаем статистику по нужным кампаниям за указанные даты
            getCompanyStatistics(
                date_start='',
                date_end='',
                auth_token=auth_token,
                df_companies=group_df,
                companies_stats_dir=upload_path,
                date_from=date_from,
                date_to=date_to,
                task_id=None,
            )
    else:
        logger.info("Файлы по всем кампаниями уже загружены")


# Функция парсинга файла по рекламной кампании
# def parse_company_file(df_company):
#     # Определяем тип отчета (по первой колонке в отчете)
#     # TODO: придумать более надежный способ
#     if df_company.columns[0] == 'Дата':
#         # Убираем строку Всего и Корректировка
#         df_company = df_company.loc[~df_company['Дата'].isin(['Всего', 'Корректировка']), :]
#         # Ozon ID = SKU в данном отчете
#         df_company = df_company.rename(columns={'Ozon ID': 'sku'})
#         # Переименовываем колонку с расходами для удобства
#         df_company = df_company.rename(columns={'Расход, ₽': 'Расходы'})
#         # Переименовываем колонку с названием товара
#         df_company = df_company.rename(columns={'Наименование': 'Наименование товара'})
#         # Добавляем колонку с номером кампании
#         df_company['Номер кампании'] = df_filenames_companies['Номер кампании'][i]
#         # Выбираем нужные колонки
#         # df_company = df_company.loc[:, ['Номер кампании','Дата', 'sku', 'Наименование товара', 'Расходы']]
#         # df_company.to_csv(f"{parsed_dir_path}/{df_filenames_companies['Номер кампании'][i]}.csv", sep=';')
#     else:
#         # Убираем строку Всего и Корректировка
#         df_company = df_company.loc[~df_company['sku'].isin(['Всего', 'Корректировка']), :]
#         # Переименовываем колонку с расходами для удобства
#         df_company = df_company.rename(columns={'Расход, ₽, с НДС': 'Расходы'})
#         # Переименовываем колонку с названием товара
#         df_company = df_company.rename(columns={'Название товара': 'Наименование товара'})
#         # Добавляем колонку с номером кампании
#         df_company['Номер кампании'] = df_filenames_companies['Номер кампании'][i]
#         # Выбираем нужные колонки
#         # df_company = df_company.loc[:, ['Номер кампании','sku', 'Наименование товара', 'Расходы']]
#         # df_company.to_csv(f"{parsed_dir_path}/{filenames_companies['Номер кампании'][i]}.csv", sep=';', index=False)

#     return df_company



# Функция парсинга данных по РК
def parse_companies_files(
        date_campaign_report,
        df_input_companies_parsed,

):

    # Директория, куда будем помещать обработанные файлы
    # parsed_dir_path = (
    #     f"{BASE_DIR}/"
    #     f"Clients/{client_name}/CampaignsReport/"
    #     f"{date_campaign_report}/{date_campaign_report}_Кампании_Parsed"
    # )
    # if not os.path.exists(parsed_dir_path):
    #     os.makedirs(parsed_dir_path, exist_ok=True)
    #     logger.info(f"Creating Companies Parsed directory:{parsed_dir_path}")

    # Получаем df со списком кампаний
    # df_filenames_companies = get_companies_files_list(
    #     date_start,
    #     date_end,
    #     upload_path,
    #     companies_list
    # )
    # df, куда будем помещать результаты обработки кампаний
    df_companies_parsed = pd.DataFrame()

    # Обработка csv по статистике кампаний в цикле
    for idx, row in df_input_companies_parsed.iterrows():
        # Путь до файла
        path_statistics_file = row['file_path_company']
        # Уникальный ID Кампании
        unique_company_id = row['unique_company_id']
        # Номер кампании
        company_number = row['Номер кампании']
        # Тип кампании
        company_type = row['Тип кампании']
        # Даты кампании
        date_start_company = row['date_start']
        date_end_company = row['date_end']
        # Даты кампании в названии файла
        dt_start_file = row['dt_start_file']
        dt_end_file = row['dt_end_file']
        # Если файла по данной кампании нет, выводим предупреждение
        if not os.path.exists(path_statistics_file):
            logger.warning(f"No statistics file for company {company_number}_{dt_start_file}_{dt_end_file}.csv")
        # Если файл есть, то начинаем его обратоку
        else:
            # Считываем csv
            df_company = pd.read_csv(path_statistics_file, sep=';', decimal=',', skiprows=1)
            # Если по кампании есть расходы, начинаем их обрабатывать
            if df_company.shape[0] >= 2:
                # Определяем тип отчета (по первой колонке в отчете)
                # TODO: придумать более надежный способ
                if df_company.columns[0] == 'Дата':
                    # Убираем строку Всего и Корректировка
                    df_company = df_company.loc[~df_company['Дата'].isin(['Всего', 'Корректировка']), :]
                    # Ozon ID = SKU в данном отчете
                    df_company = df_company.rename(columns={
                        'Наименование': 'Наименование товара',
                        'sku': 'SKU',
                        'Цена товара, ₽': 'Цена товара, руб',
                        'Ср. цена клика, ₽': 'Ср. цена клика, руб',
                        'Расход, ₽, с НДС': 'Расход с НДС, руб',
                        'Заказы': 'Заказы шт (из отчета РК)',
                        'Заказы модели': 'Заказы модели шт (из отчета РК)',
                        'Продажи, ₽': 'Заказы руб (из отчета РК)',
                        'Продажи с заказов модели, ₽': 'Заказы модели руб (из отчета РК)',
                    })
                    # Добавляем колонки с информацией о кампании
                    df_company = df_company.assign(**{
                        'unique_company_id': unique_company_id,
                        'Номер кампании': company_number,
                        'Тип кампании': company_type,
                        'date_start_company': date_start_company,
                        'date_end_company': date_end_company,
                    })
                    # Выбираем нужные колонки
                    # df_company = df_company.loc[:, ['Номер кампании','Дата', 'sku', 'Наименование товара', 'Расходы']]
                    # df_company.to_csv(f"{parsed_dir_path}/{df_filenames_companies['Номер кампании'][i]}.csv", sep=';')
                else:
                    # Убираем строку Всего и Корректировка
                    df_company = df_company.loc[~df_company['sku'].isin(['Всего', 'Корректировка']), :]
                    # Ozon ID = SKU в данном отчете
                    df_company = df_company.rename(columns={
                        'Наименование': 'Наименование товара',
                        'sku': 'SKU',
                        'Цена товара, ₽': 'Цена товара, руб',
                        'Ср. цена клика, ₽': 'Ср. цена клика, руб',
                        'Расход, ₽, с НДС': 'Расход с НДС, руб',
                        'Заказы': 'Заказы шт (из отчета РК)',
                        'Заказы модели': 'Заказы модели шт (из отчета РК)',
                        'Продажи, ₽': 'Заказы руб (из отчета РК)',
                        'Продажи с заказов модели, ₽': 'Заказы модели руб (из отчета РК)',
                    })
                    # Добавляем колонку с номером кампании
                    df_company = df_company.assign(**{
                        'unique_company_id': unique_company_id,
                        'Номер кампании': company_number,
                        'Тип кампании': company_type,
                        'date_start_company': date_start_company,
                        'date_end_company': date_end_company,
                    })
                    # Выбираем нужные колонки
                    # df_company = df_company.loc[:, ['Номер кампании','sku', 'Наименование товара', 'Расходы']]
                    # df_company.to_csv(f"{parsed_dir_path}/{filenames_companies['Номер кампании'][i]}.csv", sep=';', index=False)

            # Объединяем с предыдущим проходом цикла
            df_companies_parsed = pd.concat([df_companies_parsed, df_company], ignore_index=True)

    # Переименовываем колонки с датами РК
    df_companies_parsed = df_companies_parsed.rename(columns={
        'date_start_company': 'Дата начала статистики РК',
        'date_end_company': 'Дата окончания статистики РК',
    })

    # Переводим колонки с датами кампании в timestamp
    timestamp_cols = ['Дата начала статистики РК', 'Дата окончания статистики РК']
    # for col in timestamp_cols:
    #     df_companies_parsed[col] = pd.to_datetime(df_companies_parsed[col]).dt.tz_localize(None)
    df_companies_parsed[timestamp_cols] = (
        df_companies_parsed[timestamp_cols]
        .apply(
            lambda col: pd.to_datetime(col).dt.tz_localize(None)
        )
    )

    # Добавляем время к дате окончания РК
    time_to_add = timedelta(
        hours=23,
        minutes=59,
        seconds=59
    )
    df_companies_parsed['Дата окончания статистики РК'] = (
        df_companies_parsed['Дата окончания статистики РК'] + time_to_add
    )

    # Считаем суммы по кампаниям
    # df_companies_stats = (
    #     df_companies_parsed
    #     .groupby(['Номер кампании', 'sku', 'Наименование товара'])
    #     .agg({
    #         'Цена товара, ₽': 'sum'
    #     })
    #     .reset_index()
    # )
    return df_companies_parsed


# Функция обработка списка товаров АПИ
def process_product_list(df_products):
    # Создаем копию для избежания изменений в оригинальном df
    df_products_processed = df_products.copy()
    # Переводим SKU в строку чтобы не было ошибок при merge
    df_products_processed['SKU'] = df_products_processed['SKU'].astype(str)
    # Переводим Артикул в строку
    df_products_processed['Артикул'] = df_products_processed['Артикул'].astype(str)
    # Создаем колонку основного артикула для определения модельного ряда
    df_products_processed['Артикул размера'] = df_products_processed['Артикул']

    if client_name in ['SENS_IP']:
        # GPT START----
        # Разделяем с конца через rsplit, берем суффикс
        last = df_products_processed['Артикул'].str.rsplit('_', n=1, expand=True)
        # (если нет "_", во второй части будет NaN)
        if last.shape[1] < 2:
            last = last.reindex(columns=[0,1], fill_value=None)
        # Явно приводим второй столбец к строке
        last[1] = last[1].astype(str).fillna('')
        # Проверка: есть ли буквы (русские/латинские) во второй части
        mask = last[1].str.contains(r'[A-Za-zА-Яа-я]', na=False)
        # Если есть буквы — берем оригинал, иначе — часть до последнего '_'
        df_products_processed['Артикул размера'] = (
            df_products_processed['Артикул']
            .where(mask, last[0])
        )
    elif client_name in ['SENS']:
        pattern = (
            r'^(?:'
            r'([^/\s]+?)(?=[/\s]\d)'             # 1) короткий артикул типа УБ24 перед / или пробелом+число
            r'|'
            r'(.+?)(?:(?<=\S)(?=[_\-\s]\d))'      # 2) длинный артикул до разделителя перед размером
            r')'
        )

        df_products_processed['Артикул размера'] = (
            df_products_processed['Артикул']
            .str.extract(pattern, expand=False)   # даст DF из двух колонок
            .bfill(axis=1)                        # заполняет NaN справа на основе непустой
            .iloc[:, 0]                           # выбираем полученный артикул
            .fillna(df_products_processed['Артикул'])  # если не совпало — возвращаем оригинал

        )

    # Делаем сортировку
    df_products_processed = df_products_processed.sort_values(by='Артикул', ignore_index=True)
    # Присваиваем id каждому артикулу размера
    df_products_processed['ID Размерного ряда'] = (
        df_products_processed
        .groupby(['Артикул размера'])
        .ngroup() + 1
    )
    # Получаем количество размеров в каждой группе
    df_products_processed['Количество размеров в артикуле'] = (
        df_products_processed
        .groupby(['ID Размерного ряда'])
        .transform('size')
    )
    # GPT END----

    # Перемещаем колонки в начало df
    df_products_processed = move_columns(
        df_products_processed,
        [
            'Артикул размера',
            'ID Размерного ряда',
            'Количество размеров в артикуле',
            'model_id',
            'model_count'
        ],
        position='Артикул'
    )
    # Переименовываем некоторые
    # df_products_processed = df_products_processed.rename(columns={
    #     'model_id': 'ID Модели товара'
    # })

    return df_products_processed


# Функция добавления данных по товарам к отчетам РК
def add_products_data_to_companies_reports(df_companies_parsed, df_products_processed):

    # Создаем копию для избежания изменений в оригинальном df
    df_products_processed_ = df_products_processed.copy()
    # Переводим SKU в строку чтобы не было ошибок при merge
    df_products_processed_['SKU'] = df_products_processed_['SKU'].astype(str)

    # Объединяем датафреймы с информацией о кампаниях со списком товаров
    df_companies_products_data = df_companies_parsed.merge(
        df_products_processed_[[
            'SKU',
            'Артикул',
            'ID Размерного ряда',
            'Количество размеров в артикуле',
            'model_id'
        ]],
        on='SKU',
        how='left',
        indicator=True
    )
    # Перемещаем колонку с артикулом в начало
    df_companies_products_data = move_columns(
        df_companies_products_data,
        ['Артикул', 'ID Размерного ряда', 'Количество размеров в артикуле', 'model_id'],
        position='SKU'
    )
    # Перемещаем колонку с номером кампании в начало
    df_companies_products_data = move_columns(
        df_companies_products_data,
        ['unique_company_id', 'Номер кампании', 'Тип кампании', 'Дата начала статистики РК', 'Дата окончания статистики РК'],
        position=0
    )
    # Переименовываем некоторые колонки
    # df_companies_products_data = df_companies_products_data.rename(columns={
    #     # 'model_id': 'ID Модели товара',
    #     'date_start_company': 'Дата начала статистики РК',
    #     'date_end_company': 'Дата окончания статистики РК',
    # })
    # Заполняем пропуски, если есть
    df_companies_products_data = df_companies_products_data.fillna({
        # 'ID Модели товара': 'Неизвестная модель',
        # 'model_id': 'Неизвестная модель',
        'model_id': 0,
        'Артикул': 'Неизвестный артикул',
    })

    # Проверяем, что подтянулась информация по всем товарам
    df_missing_products = (
        df_companies_products_data
        .loc[df_companies_products_data['_merge'] == 'left_only',
            ['Номер кампании', 'Артикул']
        ]
    )
    if not df_missing_products.empty:
        logger.warning(f"No info about products:\n{df_missing_products.to_string(index=False)}")

    # Удаляем кампании, по которым не выгрузилась статистика
    df_companies_products_data = (
        df_companies_products_data
        .dropna(subset=['SKU'], ignore_index=True)
    )
    # Удаляем колонку с индикатором мерджа
    df_companies_products_data = (
        df_companies_products_data
        .loc[:, ~df_companies_products_data.columns.isin(['_merge'])]
    )

    return df_companies_products_data


# Функция определения минимальной и максимальной даты в кампаниях
def get_min_max_dates(df_input_companies_parsed):
    # Создаем копию для избежания изменений в оригинальном df
    df_input_companies_parsed_ = df_input_companies_parsed.copy()
    # Переводим колонки в datetime
    df_input_companies_parsed_['datetime_start'] = pd.to_datetime(df_input_companies_parsed_['date_start'])
    df_input_companies_parsed_['datetime_end'] = pd.to_datetime(df_input_companies_parsed_['date_end'])
    # Находим мин. и макс. дату
    # GPT START----
    # Создаем столбец с минимумом по строке
    df_input_companies_parsed_['row_min'] = df_input_companies_parsed_[['datetime_start', 'datetime_end']].min(axis=1)
    df_input_companies_parsed_['row_max'] = df_input_companies_parsed_[['datetime_start', 'datetime_end']].max(axis=1)

    # Индекс общего минимума среди этих двух столбцов (по всем строкам)
    idx_min = df_input_companies_parsed_['row_min'].idxmin()
    idx_max = df_input_companies_parsed_['row_max'].idxmax()
    # Получаем label для строки с общим минимумом
    min_date = df_input_companies_parsed_.loc[idx_min, 'date_start']
    max_date = df_input_companies_parsed_.loc[idx_max, 'date_end']
    # Переводим в нужный формат
    min_date = min_date + 'T00:00:00Z'
    max_date = max_date + 'T23:59:59Z'
    # GPT END----

    return min_date, max_date


# Функция выгрузки отчета о заказах
def get_orders(headers, date_start, date_end):
    # Выгружаем отчет об отправлениях fbo и fbs
    df_orders_fbo = getOrders(headers, date_start, date_end, delivery_schema='fbo', to_save=False)
    df_orders_fbs = getOrders(headers, date_start, date_end, delivery_schema='fbs', to_save=False)
    # Убираем лишние колонки из отчета fbo
    df_orders_fbo = df_orders_fbo.loc[:, ~df_orders_fbo.columns.isin(['Объемный вес товаров, кг'])]

    # Объединяем два отчета в один
    common_columns = ['Артикул', 'Наименование товара', 'OZON id', 'Принят в обработку', 'Сумма отправления']

    # GPT START----
    dfs = []
    for df in [df_orders_fbo, df_orders_fbs]:
        if df is not None and not df.empty:
            # Берем только те common_columns, которые реально есть
            cols = [col for col in common_columns if col in df.columns]
            dfs.append(df.loc[:, cols])

    if dfs:
        df_orders_all = pd.concat(dfs, ignore_index=True)
    else:
        # Если оба датафрейма пустые — создаём пустой датафрейм с нужными колонками
        df_orders_all = pd.DataFrame(columns=common_columns)
    # GPT END----

    # Переводим колонку с датой заказа в datetime
    df_orders_all['datetime_orders'] = pd.to_datetime(df_orders_all['Принят в обработку'])
    # Указываем, какие колонки считаем за заказы в штуках и рублях
    df_orders_all['Заказы шт'] = 1
    df_orders_all['Заказы руб'] = df_orders_all['Сумма отправления']
    # Переименовываем колонку с Ozon ID
    df_orders_all = df_orders_all.rename(columns={
        'OZON id': 'Ozon Product ID'
    })
    # Переводим столбец с артикулом в строку
    df_orders_all['Артикул'] = df_orders_all['Артикул'].astype(str)

    return df_orders_all


# Функция добавления информации о товарах в заказы
def add_products_data_to_orders(df_orders_all, df_products_processed):

    # Создаем копию для избежания изменений в оригинальном df
    # df_products_processed_ = df_products_processed.copy()
    # Переводим SKU в строку чтобы не было ошибок при merge
    # df_products_processed_['SKU'] = df_products_processed_['SKU'].astype(str)

    # Объединяем датафреймы с информацией о кампаниях со списком товаров
    df_orders_products_data = df_orders_all.merge(
        df_products_processed[[
            'SKU',
            'Артикул',
            'ID Размерного ряда',
            'Количество размеров в артикуле',
            'model_id'
        ]],
        on='Артикул',
        how='left',
        indicator=True
    )
    # Перемещаем колонку с артикулом в начало
    df_orders_products_data = move_columns(
        df_orders_products_data,
        [
            'Артикул',
            'ID Размерного ряда',
            'Количество размеров в артикуле',
            'model_id'
        ],
        position='Наименование товара'
    )

    # Переименовываем некоторые колонки
    # df_orders_products_data = df_orders_products_data.rename(columns={
    #     'model_id': 'ID Модели товара',
    # })
    # Заполняем пропуски, если есть
    df_orders_products_data = df_orders_products_data.fillna({
        # 'ID Модели товара': 'Неизвестная модель',
        # 'model_id': 'Неизвестная модель',
        'model_id': 0,
        'Артикул': 'Неизвестный артикул',
    })

    # Проверяем, что подтянулась информация по всем товарам
    df_missing_products = (
        df_orders_products_data
        .loc[df_orders_products_data['_merge'] == 'left_only',
            ['Артикул']
        ]
    )

    if not df_missing_products.empty:
        logger.warning(f"No info about products: {df_missing_products.to_string(index=False)}")

    # Удаляем колонку с индикатором мерджа
    df_orders_products_data = (
        df_orders_products_data
        .loc[:, ~df_orders_products_data.columns.isin(['_merge'])]
    )

    return df_orders_products_data

# Функция выгрузки остатков
def get_reminders(df_products, headers):
    # Выгружаем остатки FBO и FBS
    df_reminders_fbo = getStockReminders_fbo_v2(headers, df_products, to_save=False)
    df_reminders_fbs = getStockReminders_fbs(headers, to_save=False)
    # Переименовываем колонки для удобства
    df_reminders_fbo = df_reminders_fbo.rename(columns={
        'Доступный к продаже товар': 'Остатки FBO',
    })
    df_reminders_fbs = df_reminders_fbs.rename(columns={
        'Доступно на моем складе, шт': 'Остатки FBS',
    })
    # Объединяем остатки FBO и FBS в один df
    df_reminders_all = pd.concat([
        df_reminders_fbo.loc[:, ['Артикул', 'Остатки FBO']],
        df_reminders_fbs.loc[:, ['Артикул', 'Остатки FBS']],
    ])

    # Переводим колонку с артикулом в строку
    df_reminders_all['Артикул'] = df_reminders_all['Артикул'].astype(str)

    # Считаем остатки FBO + FBS
    df_reminders_total = (
        df_reminders_all
        .groupby(['Артикул'], as_index=False)
        .sum()
        .assign(**{
            'Остаток': lambda df:df[['Остатки FBO', 'Остатки FBS']].sum(axis=1)
        })
        # .reset_index()
    )

    return df_reminders_total


# Функция расчета заказов по моделям и размерам
def calc_orders_by_models(
        date_start_company,
        date_end_company,
        unique_company_id,
        company_number,
        df_orders_products_data,
        df_products_processed,
):
    # Создаем копию для избежания изменений в оригинальном df
    df_orders_products_data__ = df_orders_products_data.copy()
    # Фильтруем заказы по датам кампании
    df_orders_date_filtered = (
        df_orders_products_data__
        .loc[
            df_orders_products_data__['datetime_orders'].between(date_start_company, date_end_company),
              :
            ]
            .copy()
    )
    # Задаем правила агрегирования
    agg_rules = {
        'sum': [  # Функция, которую применяем
            {
                'agg_col': 'Заказы шт', # Колонка, по которой считаем значения
                'group_col': 'Артикул', # Колонка с группой
                'new_col': 'Заказы шт (из отчета по Заказам)' # Название новой колонки
            },
            {
                'agg_col': 'Заказы руб',
                'group_col': 'Артикул',
                'new_col': 'Заказы руб (из отчета по Заказам)'
            },
            {
                'agg_col': 'Заказы шт',
                'group_col': 'model_id',
                'new_col': 'Заказы модели шт (из отчета по Заказам)'
            },
            {
                'agg_col': 'Заказы руб',
                'group_col': 'model_id',
                'new_col': 'Заказы модели руб (из отчета по Заказам)'
            },
            {
                'agg_col': 'Заказы шт',
                'group_col': 'ID Размерного ряда',
                'new_col': 'Заказы размерного ряда шт (из отчета по Заказам)'
            },
            {
                'agg_col': 'Заказы руб',
                'group_col': 'ID Размерного ряда',
                'new_col': 'Заказы размерного ряда руб (из отчета по Заказам)'
            },
        ],
    }

    # Лист, в который будут помещены итоговые колонки
    new_cols_list = []
    # Цикл по функциям агрегаций
    for func, params_list in agg_rules.items():
        # Цикл по колонкам агрегаций
        for params in params_list:
            group_col = params['group_col']
            agg_col   = params['agg_col']
            new_col   = params['new_col']
            df_orders_date_filtered[new_col] = (
                df_orders_date_filtered
                .groupby(group_col)
                [agg_col]
                .transform(func)
            )
            # Добавляем колонки в лист колонок
            new_cols_list.append(new_col)

    # Выбираем нужные колонки
    tmp_df_orders_for_company = (
        df_orders_date_filtered
        .copy()
        .pipe(
            lambda df: df.loc[
                :,
                df.columns.isin(
                    [
                        'Артикул',
                        # 'ID Размерного ряда',
                        # 'Количество размеров в артикуле',
                        # 'model_id',
                    ]
                    + new_cols_list
                )
            ]
        )
        # Удаляем дубликаты у артикула, т.к. нужные метрики мы посчитали отдельными столбцами
        .drop_duplicates(subset=['Артикул'], ignore_index=True)
    )

    # Добавляем данные по кампаниям
    tmp_df_orders_for_company = tmp_df_orders_for_company.assign(**{
        'unique_company_id': unique_company_id,
        'Номер кампании': company_number,
    })

    # Объединяем данные заказов со списком товаров,
    # чтобы увидеть, какие артикулы были в сшивке
    # tmp_df_products_with_orders = (
    #     df_products_processed
    #     .loc[:, ['Артикул', 'model_id', 'ID Размерного ряда', 'Количество размеров в артикуле']]
    #     .merge(
    #         tmp_df_orders,
    #         how='right',
    #         on='Артикул'
    #     )
    # )

    return tmp_df_orders_for_company

# Функция добавления сшитых товаров
def add_model_products(
        tmp_df_company,
        tmp_df_orders_for_company,
        df_products_processed,
        company_sku_list,
):
    # Создаем копию df с заказами по кампании
    tmp_df_orders_for_company_ = tmp_df_orders_for_company.copy()
    # Убираем лишние колонки для избежания дубликатов
    tmp_df_orders_for_company_ = (
        tmp_df_orders_for_company_
        .loc[:, ~tmp_df_orders_for_company_.columns.isin([
            'unique_company_id',
            'Номер кампании',
        ])]
    )

    # Мерджим заказы со списком товаров
    df_products_with_orders = (
        df_products_processed
        # .assign(**{
        #     'Артикул для моделей': (lambda df: df['Артикул'])
        # })
        .loc[:, ['Артикул', 'model_id', 'ID Размерного ряда', 'Количество размеров в артикуле']]
        .merge(
            tmp_df_orders_for_company,
            how='left',
            on='Артикул'
        )
        .fillna(0)
    )

    # Далее будет разбивка merge на два датафрейма (тупая хуйня):
    #   один - объединение по артикулу (товары одного артикула, но разных размеров)
    #   второй - объединение по идентификатору модели товара (для товаров в сшивке Озон)

    # Определяем артикулы, по которым нужны заказы размеров
    df_company_products_with_sizes = (
        tmp_df_company
        .loc[tmp_df_company['Количество размеров в артикуле'] > 1, :]
    )
    # Добавляем колонку с типом товара
    df_company_products_with_sizes['Тип товара'] = 'Товар из РК'
    # Определяем колонки для мерджа по артикулу
    columns_for_size_merge = (
        df_products_with_orders
        .columns[
            df_products_with_orders.columns.str.contains('Артикул|Заказы')
        ]
    )
    # Объединяем со списком товаров
    df_company_products_with_sizes_orders = df_company_products_with_sizes.merge(
        df_products_with_orders[columns_for_size_merge],
        on='Артикул',
        how='left',
    )

    # Определяем артикулы, по которым нужно добавить сшивку
    df_company_products_with_models = (
        tmp_df_company
        .loc[tmp_df_company['Количество размеров в артикуле'] <= 1, :]
    )
    # Добавляем колонку с типом товара
    df_company_products_with_models['Тип товара'] = 'Товар из РК'

    # Удаляем артикулы, которые уже есть в кампании, чтобы не было дубликатов
    # df_company_products_with_models = (
    #     df_company_products_with_models
    #     .loc[~df_company_products_with_models['Артикул'].isin(company_sku_list), :]
    # )
    # Определяем колонки для мерджа по id модели товара
    columns_for_model_merge = (
        df_products_with_orders
        .columns[
            df_products_with_orders.columns.str.contains('Артикул|model_id|Заказы')
        ]
    )
    # Определяем список моделей в РК
    company_model_list = df_company_products_with_models['model_id'].unique().tolist()
    # Из списка товаров выбираем товары данной модели, за исключением товаров в РК
    df_models_products = (
        df_products_processed
        .loc[
            df_products_processed['model_id'].isin(company_model_list),
            [
                'Артикул',
                'Название товара',
                'SKU',
                'ID Размерного ряда',
                'Количество размеров в артикуле',
                'model_id',
            ]
        ]
        .pipe(lambda df:
             df
             # Убираем артикулы кампании
             .loc[~df['Артикул'].isin(company_sku_list), :]
             # Добавляем столбец с типом товара
             .assign(**{
                 'Тип товара': 'Товар из сшивки'
             })
             # Добавляем информацию о товаре из списка товаров
            #  .merge(
            #      df_products_processed[[
            #          'Артикул',
            #          'Название товара'
            #      ]],
            #      on='Артикул',
            #      how='left'
            #  )
        )
    )
    # Добавляем товары в df с РК
    df_company_products_with_models_orders = (
        pd.concat([
            df_company_products_with_models,
            df_models_products
        ], ignore_index=True)
    )
    # Заполняем NA в столбцах с инфой по кампании
    columns_to_fill_models = [
        'unique_company_id',
        'Номер кампании',
        'Дата начала статистики РК',
       'Дата окончания статистики РК',
    ]
    df_company_products_with_models_orders[columns_to_fill_models] = (
        df_company_products_with_models_orders[columns_to_fill_models].ffill()
    )
    # Добавляем информацию по заказам для товаров из сшивки
    df_company_products_with_models_orders = (
        df_company_products_with_models_orders
        .merge(
            df_products_with_orders[columns_for_size_merge],
            on='Артикул',
            how='left'
        )
    )
    # Выполняем сортировку, чтобы первым шел товар РК, а под ним - товары его модели
    df_company_products_with_models_orders = (
        df_company_products_with_models_orders
        .sort_values(
            by=['model_id', 'Тип товара'],
            ascending=[True, True],
            ignore_index=True
        )
    )

    # Способ № 1
    # Объединяем со списком товаров по id модели
    # df_company_products_with_models_orders = (
    #     df_company_products_with_models
    #     .merge(
    #         df_products_with_orders[columns_for_model_merge],
    #         on='model_id',
    #         how='left',
    #     )
    # )
    # # У нас появилось две колонки с артикулами:
    # # одна - из отчета по кампании;
    # # вторая - из отчета по товарам из моделей
    # # Переименовываем колонку, которая пришла из моделей, удаляем первую
    # df_company_products_with_models_orders = (
    #     df_company_products_with_models_orders
    #     .rename(columns={
    #         'Артикул_y': 'Артикул'
    #     })
    #     .drop(columns=['Артикул_x'])
    # )
    # # Заполняем пропуски в артикуле
    # df_company_products_with_models_orders['Артикул'] = (
    #     df_company_products_with_models_orders['Артикул']
    #     .fillna('Неизвестная модель')
    # )
    # # ---- GPT START ----
    # # Для каждой группы model_id выделим строки, которые уже были в РК
    # # и которые пришли из сшивки моделей
    # df_company_products_with_models_orders['Признак товара из сшивки'] = (
    #     df_company_products_with_models_orders
    #     .groupby('model_id')
    #     .cumcount() != 0
    # )

    # # Сортируем: сначала по model_id,
    # # а затем по строкам, пришедшим из списка товаров ("сшитые модели")
    # df_company_products_with_models_orders = (
    #     df_company_products_with_models_orders
    #     .sort_values(['model_id', 'Признак товара из сшивки'], ascending=[True, True])
    #     .reset_index(drop=True)
    #     # .drop(columns=['Признак товара из сшивки'])
    # )
    # # ---- GPT END ----
    # # Способ № 1
    # # У товаров из сшивки делаем столбец Тип товара из сшивки
    # df_company_products_with_models_orders['Тип товара'] = np.where(
    #     df_company_products_with_models_orders['Признак товара из сшивки'],
    #     'Товар из сшивки',
    #     'Товар из РК'
    # )
    # # Удаляем артикулы, которые уже были в РК, но также попали в сшивку
    # mask_models_products = (
    #     (df_company_products_with_models_orders['Тип товара'] == 'Товар из сшивки') &
    #     (df_company_products_with_models_orders['Артикул'].isin(company_sku_list))
    # )
    # df_company_products_with_models_orders = (
    #     df_company_products_with_models_orders[~mask_models_products]
    #     .reset_index(drop=True)
    # )


    # Удаляем дубликаты после мерджа по id модели,
    # чтобы исключить товары, которые уже были в кампании
    # df_company_products_with_models_orders = (
    #     df_company_products_with_models_orders
    #     .drop_duplicates(subset=['Артикул'], keep='first', ignore_index=True)
    # )


    # Объединяем результат в один df
    df_company_result_orders = pd.concat([
        df_company_products_with_sizes_orders,
        df_company_products_with_models_orders
    ], ignore_index=True)


    # Способ №2
    # Добавляем столбец с типом товара, чтобы понимать, каике товары пришли из сшивки
    # df_company_result_orders['Тип товара'] = np.where(
    #     df_company_result_orders['Артикул'].isin(company_sku_list),
    #     'Товар из РК',
    #     'Товар из сшивки'
    # )
    # # У товаров из сшивки делаем пустыми столбцы по РК
    # company_stat_cols = (
    #     df_company_result_orders
    #     .loc[:,'Цена товара, руб':'Дата добавления']
    #     .columns
    # )
    # # Выбираем товары только РК
    # mask_model_products = df_company_result_orders['Тип товара'] == 'Товар из сшивки'
    # # Делаем столбцы пустыми
    # df_company_result_orders.loc[mask_model_products, company_stat_cols] = np.nan


    return df_company_result_orders


# Функция расчета заказов по моделям и размерам (v2)
def calc_orders_by_models_v2(
        date_start_company,
        date_end_company,
        unique_company_id,
        company_number,
        df_orders_products_data,
        df_products_processed,
):
    # Создаем копию для избежания изменений в оригинальном df
    df_orders_products_data__ = df_orders_products_data.copy()
    # Фильтруем заказы по датам кампании
    df_orders_date_filtered = (
        df_orders_products_data__
        .loc[
            df_orders_products_data__['datetime_orders'].between(date_start_company, date_end_company),
              :
            ]
            .copy()
    )
    # Задаем правила агрегирования
    agg_rules = {
        'sum': [  # Функция, которую применяем
            {
                'agg_col': 'Заказы шт', # Колонка, по которой считаем значения
                'group_col': 'Артикул', # Колонка с группой
                'new_col': 'Заказы шт (из отчета по Заказам)' # Название новой колонки
            },
            {
                'agg_col': 'Заказы руб',
                'group_col': 'Артикул',
                'new_col': 'Заказы руб (из отчета по Заказам)'
            },
            {
                'agg_col': 'Заказы шт',
                'group_col': 'model_id',
                'new_col': 'Заказы модели шт (из отчета по Заказам)'
            },
            {
                'agg_col': 'Заказы руб',
                'group_col': 'model_id',
                'new_col': 'Заказы модели руб (из отчета по Заказам)'
            },
            {
                'agg_col': 'Заказы шт',
                'group_col': 'ID Размерного ряда',
                'new_col': 'Заказы размерного ряда шт (из отчета по Заказам)'
            },
            {
                'agg_col': 'Заказы руб',
                'group_col': 'ID Размерного ряда',
                'new_col': 'Заказы размерного ряда руб (из отчета по Заказам)'
            },
        ],
    }

    # Словарь с итоговыми датафреймами
    grouped_dfs_company = {}

    # Соберем все уникальные group_col
    all_group_cols = set(params['group_col'] for rule in agg_rules.values() for params in rule)

    for group_col in all_group_cols:
        # Найдем все параметры агрегации для этой group_col
        relevant_aggs = [
            (func, params)
            for func, params_list in agg_rules.items()
            for params in params_list
            if params['group_col'] == group_col
        ]
        # Строим словарь для передачи в .agg()
        agg_dict = {}
        for func, params in relevant_aggs:
            agg_col = params['agg_col']
            new_col = params['new_col']
            # Pandas поддерживает агрегаты в формате output_column=(input_column, func)
            # Соберём в список, если несколько агрегаций по одной колонке
            agg_dict[new_col] = (agg_col, func)
        # Делаем группировку
        df_grouped_company = (
            df_orders_date_filtered
            .groupby(group_col)
            .agg(**agg_dict)
            .reset_index()
            # Добавляем номер кампании
            # .assign(**{
            #     'unique_company_id': unique_company_id,
            #     'Номер кампании': company_number,
            # })
        )
        grouped_dfs_company[group_col] = df_grouped_company


    return grouped_dfs_company

# Функция расчета заказов и остатков для товаров в РК
def calc_orders_for_companies(
        df_companies_products_data,
        df_orders_products_data,
        df_products_processed,
):
    # Создаем копии для избежания изменений в оригинальных df
    # df_orders_products_data_ = df_orders_products_data.copy()
    df_companies_with_orders = df_companies_products_data.copy()

    # df, куда будем помещать итоговый результат
    df_companies_with_orders_and_models = pd.DataFrame()

    # Убираем лишние столбцы
    df_companies_with_orders = (
        df_companies_with_orders
        .loc[:, df_companies_with_orders.columns.isin([
                'unique_company_id',
                'Номер кампании',
                'Тип кампании',
                'Дата начала статистики РК',
                'Дата окончания статистики РК',
                'SKU',
                'Артикул',
                'ID Размерного ряда',
                'Количество размеров в артикуле',
                'model_id',
                'Наименование товара',
                'Цена товара, ₽',
                'Показы',
                'Клики',
                'CTR (%)',
                'В корзину',
                'Ср. цена клика, ₽',
                'Расход с НДС, руб',
                'Дата добавления',
        ])]
    )
    # Делаем цикл по каждой рекламной кампании
    for unique_company_id in df_companies_products_data['unique_company_id'].unique():
        # Временный df с данными текущей кампании
        tmp_df_company = (
            df_companies_products_data
            .loc[df_companies_products_data['unique_company_id'] == unique_company_id, :]
            # .drop_duplicates(subset=['Номер кампании'])
        )
        # Получаем номер кампании
        company_number = tmp_df_company['Номер кампании'].iloc[0]
        # Получаем даты начала и окончания кампании
        date_start_company = tmp_df_company['Дата начала статистики РК'].iloc[0]
        date_end_company = tmp_df_company['Дата окончания статистики РК'].iloc[0]
        # Получаем список исходных SKU кампании
        company_sku_list = tmp_df_company['Артикул'].to_list()
        # Считаем заказы с фильтром по данным датам
        grouped_dfs_company = calc_orders_by_models_v2(
            date_start_company,
            date_end_company,
            unique_company_id,
            company_number,
            df_orders_products_data,
            df_products_processed
        )
        # Добавляем товары по сшивке
        tmp_df_company_with_models = add_model_products_v2(
            tmp_df_company,
            grouped_dfs_company,
            df_products_processed,
            unique_company_id,
            company_number,
            company_sku_list
        )
        # Объединяем с данными предыдущей кампании
        df_companies_with_orders_and_models = pd.concat([
            df_companies_with_orders_and_models,
            tmp_df_company_with_models
        ], ignore_index=True)

    return df_companies_with_orders_and_models


# Функция добавления сшитых товаров (v2)
def add_model_products_v2(
        tmp_df_company,
        grouped_dfs_company,
        df_products_processed,
        unique_company_id,
        company_number,
        company_sku_list,

):
    # Создаем копию df списка товаров, куда будем помещать заказы
    df_products_with_orders = df_products_processed.copy()
    # Добавляем заказы к списку товаров
    for join_key, df_company_orders in grouped_dfs_company.items():
        # Мерджим заказы со списком товаров
        df_products_with_orders = (
            df_products_with_orders
            # .assign(**{
            #     'Артикул для моделей': (lambda df: df['Артикул'])
            # })
            # .loc[:, ['Артикул', 'model_id', 'ID Размерного ряда', 'Количество размеров в артикуле']]
            .merge(
                grouped_dfs_company[join_key],
                how='left',
                on=join_key,
            )
            .fillna(0)
        )
    # Выбираем нужные колонки
    orders_columns = (
        df_products_with_orders
        .columns[
            df_products_with_orders.columns.str.contains('Заказы')
        ]
        .tolist()
    )
    df_products_with_orders = (
        df_products_with_orders
        .loc[
            :,
            df_products_with_orders.columns.isin([
                    'Артикул',
                    'SKU',
                    'Название товара',
                    'model_id',
                    'ID Размерного ряда',
                    'Количество размеров в артикуле'
                ] + orders_columns
            )
        ]
    )

    # Добавляем заказы к df кампании
    tmp_df_company_with_orders = (
        tmp_df_company
        # Создаем копию, чтобы не изменять исходный df
        .copy()
        # Добавляем тип товара - товар из РК
        .assign(**{
            'Тип товара': 'Товар из РК'
        })
        # К товарам из РК добавляем различные типы заказов
        .merge(
            df_products_with_orders[['Артикул'] + orders_columns],
            on='Артикул',
            how='left'
        )
        .fillna(0)
    )

    # Далее будет разбивка merge на два датафрейма (тупая хуйня):
    #   один - объединение по артикулу (товары одного артикула, но разных размеров)
    #   второй - объединение по идентификатору модели товара (для товаров в сшивке Озон)

    # Определяем артикулы, по которым нужны заказы размеров
    # df_company_products_with_sizes = (
    #     tmp_df_company
    #     .loc[tmp_df_company['Количество размеров в артикуле'] > 1, :]
    # )
    # # Добавляем колонку с типом товара
    # df_company_products_with_sizes['Тип товара'] = 'Товар из РК'
    # # Определяем колонки для мерджа по артикулу
    # columns_for_size_merge = (
    #     df_products_with_orders
    #     .columns[
    #         df_products_with_orders.columns.str.contains('Артикул|Заказы')
    #     ]
    # )
    # # Объединяем с заказами
    # df_company_products_with_sizes_orders = df_company_products_with_sizes.merge(
    #     df_products_with_orders[columns_for_size_merge],
    #     on='Артикул',
    #     how='left',
    # )

    # Определяем артикулы, по которым нужно добавить сшивку
    df_company_products_with_models = (
        tmp_df_company_with_orders
        .loc[tmp_df_company_with_orders['Количество размеров в артикуле'] <= 1, :]
    )
    # Добавляем колонку с типом товара
    # df_company_products_with_models['Тип товара'] = 'Товар из РК'


    # Удаляем артикулы, которые уже есть в кампании, чтобы не было дубликатов
    # df_company_products_with_models = (
    #     df_company_products_with_models
    #     .loc[~df_company_products_with_models['Артикул'].isin(company_sku_list), :]
    # )
    # Определяем колонки для мерджа по id модели товара
    columns_for_model_merge = (
        df_products_with_orders
        .columns[
            df_products_with_orders.columns.str.contains('Артикул|model_id|Заказы')
        ]
    )
    # Определяем список моделей в РК
    company_model_list = df_company_products_with_models['model_id'].unique().tolist()
    # Из списка товаров выбираем товары данной модели, за исключением товаров в РК
    df_models_products = (
        df_products_with_orders
        .loc[
            df_products_with_orders['model_id'].isin(company_model_list),
            [
                'Артикул',
                'Название товара',
                'SKU',
                'ID Размерного ряда',
                'Количество размеров в артикуле',
                'model_id',
                # 'Заказы модели шт (из отчета по Заказам)',
                # 'Заказы модели руб (из отчета по Заказам)',
            ] + orders_columns
        ]
        .pipe(
            lambda df:
                df
                # Убираем артикулы кампании
                .loc[~df['Артикул'].isin(company_sku_list), :]
                # Добавляем столбец с типом товара
                .assign(**{
                    'Тип товара': 'Товар из сшивки'
                })
                # Добавляем информацию о товаре из списка товаров
                #  .merge(
                #      df_products_processed[[
                #          'Артикул',
                #          'Название товара'
                #      ]],
                #      on='Артикул',
                #      how='left'
                #  )
        )
    )
    # Добавляем товары в df с РК
    df_company_products_with_models_orders = (
        pd.concat([
            tmp_df_company_with_orders,
            df_models_products
        ], ignore_index=True)
    )
    # Заполняем NA в столбцах с инфой по кампании
    columns_to_fill_models = [
        'unique_company_id',
        'Номер кампании',
        'Тип кампании',
        'Дата начала статистики РК',
        'Дата окончания статистики РК',
    ]
    df_company_products_with_models_orders[columns_to_fill_models] = (
        df_company_products_with_models_orders[columns_to_fill_models].ffill()
    )
    # Добавляем информацию по заказам для товаров из сшивки
    # df_company_products_with_models_orders = (
    #     df_company_products_with_models_orders
    #     .merge(
    #         df_products_with_orders[columns_for_size_merge],
    #         on='Артикул',
    #         how='left'
    #     )
    # )
    # Выполняем сортировку, чтобы первым шел товар РК, а под ним - товары его модели
    df_company_products_with_models_orders = (
        df_company_products_with_models_orders
        .sort_values(
            by=['model_id', 'Тип товара'],
            ascending=[True, True],
            ignore_index=True
        )
    )

    # Способ № 1
    # Объединяем со списком товаров по id модели
    # df_company_products_with_models_orders = (
    #     df_company_products_with_models
    #     .merge(
    #         df_products_with_orders[columns_for_model_merge],
    #         on='model_id',
    #         how='left',
    #     )
    # )
    # # У нас появилось две колонки с артикулами:
    # # одна - из отчета по кампании;
    # # вторая - из отчета по товарам из моделей
    # # Переименовываем колонку, которая пришла из моделей, удаляем первую
    # df_company_products_with_models_orders = (
    #     df_company_products_with_models_orders
    #     .rename(columns={
    #         'Артикул_y': 'Артикул'
    #     })
    #     .drop(columns=['Артикул_x'])
    # )
    # # Заполняем пропуски в артикуле
    # df_company_products_with_models_orders['Артикул'] = (
    #     df_company_products_with_models_orders['Артикул']
    #     .fillna('Неизвестная модель')
    # )
    # # ---- GPT START ----
    # # Для каждой группы model_id выделим строки, которые уже были в РК
    # # и которые пришли из сшивки моделей
    # df_company_products_with_models_orders['Признак товара из сшивки'] = (
    #     df_company_products_with_models_orders
    #     .groupby('model_id')
    #     .cumcount() != 0
    # )

    # # Сортируем: сначала по model_id,
    # # а затем по строкам, пришедшим из списка товаров ("сшитые модели")
    # df_company_products_with_models_orders = (
    #     df_company_products_with_models_orders
    #     .sort_values(['model_id', 'Признак товара из сшивки'], ascending=[True, True])
    #     .reset_index(drop=True)
    #     # .drop(columns=['Признак товара из сшивки'])
    # )
    # # ---- GPT END ----
    # # Способ № 1
    # # У товаров из сшивки делаем столбец Тип товара из сшивки
    # df_company_products_with_models_orders['Тип товара'] = np.where(
    #     df_company_products_with_models_orders['Признак товара из сшивки'],
    #     'Товар из сшивки',
    #     'Товар из РК'
    # )
    # # Удаляем артикулы, которые уже были в РК, но также попали в сшивку
    # mask_models_products = (
    #     (df_company_products_with_models_orders['Тип товара'] == 'Товар из сшивки') &
    #     (df_company_products_with_models_orders['Артикул'].isin(company_sku_list))
    # )
    # df_company_products_with_models_orders = (
    #     df_company_products_with_models_orders[~mask_models_products]
    #     .reset_index(drop=True)
    # )


    # Удаляем дубликаты после мерджа по id модели,
    # чтобы исключить товары, которые уже были в кампании
    # df_company_products_with_models_orders = (
    #     df_company_products_with_models_orders
    #     .drop_duplicates(subset=['Артикул'], keep='first', ignore_index=True)
    # )


    # Объединяем результат в один df
    # df_company_result_orders = pd.concat([
    #     df_company_products_with_sizes_orders,
    #     df_company_products_with_models_orders
    # ], ignore_index=True)
    df_company_result_orders = df_company_products_with_models_orders.copy()


    # Способ №2
    # Добавляем столбец с типом товара, чтобы понимать, каике товары пришли из сшивки
    # df_company_result_orders['Тип товара'] = np.where(
    #     df_company_result_orders['Артикул'].isin(company_sku_list),
    #     'Товар из РК',
    #     'Товар из сшивки'
    # )
    # # У товаров из сшивки делаем пустыми столбцы по РК
    # company_stat_cols = (
    #     df_company_result_orders
    #     .loc[:,'Цена товара, руб':'Дата добавления']
    #     .columns
    # )
    # # Выбираем товары только РК
    # mask_model_products = df_company_result_orders['Тип товара'] == 'Товар из сшивки'
    # # Делаем столбцы пустыми
    # df_company_result_orders.loc[mask_model_products, company_stat_cols] = np.nan


    return df_company_result_orders



# Функция добавления остатков
def add_reminders(df_companies_with_orders_and_models, df_reminders_all):
    # Добавляем ранее выгруженные остатки
    df_companies_with_orders_and_reminders = df_companies_with_orders_and_models.merge(
        df_reminders_all[['Артикул', 'Остаток']],
        how='left',
        on='Артикул'
    )

    return df_companies_with_orders_and_reminders


# Функция расчета итоговых заказов
def calc_final_orders(df_companies_with_orders_and_reminders):
    # Создаем копию для избежания изменений в оригинальном df
    df_final_orders = df_companies_with_orders_and_reminders.copy()
    # Заполняем NA чтобы при сумме не появлялся NA
    df_final_orders = df_final_orders.infer_objects(copy=False)
    # df_final_orders = df_final_orders.fillna(0)
    # Считаем итоговую колонку заказов в рублях и штуках
    df_final_orders['Заказы всего шт'] = np.where(
        df_final_orders['Количество размеров в артикуле'] > 1,
        # У артикулов с размерным рядом - берем сумму всех заказов размерного ряда артикула
        df_final_orders['Заказы размерного ряда шт (из отчета по Заказам)'],
        # У артикулов без размерного ряда - суммируем заказы модели из отчета РК и заказы только этого артикула
        df_final_orders['Заказы модели шт (из отчета РК)'] + df_final_orders['Заказы шт (из отчета по Заказам)']
    )
    df_final_orders['Заказы всего руб'] = np.where(
        df_final_orders['Количество размеров в артикуле'] > 1,
        # У артикулов с размерным рядом - берем сумму всех заказов размерного ряда артикула
        df_final_orders['Заказы размерного ряда руб (из отчета по Заказам)'],
        # У артикулов без размерного ряда - суммируем заказы модели из отчета РК и заказы только этого артикула
        df_final_orders['Заказы модели руб (из отчета РК)'] + df_final_orders['Заказы руб (из отчета по Заказам)']
    )
    # Если заказы модели из отчета по заказам нулевые, а из отчета РК - нет, то берем заказы модели из отчета РК
    # epsilon = 1e-5
    # df_final_orders['Заказы всего шт'] = np.where(
    #     (df_final_orders['Заказы модели шт (из отчета РК)'] > 0) &
    #     (np.isclose(df_final_orders['Заказы модели шт (из отчета по Заказам)'], 0, atol=epsilon)),

    #     df_final_orders['Заказы модели шт (из отчета РК)'],
    #     df_final_orders['Заказы модели шт (из отчета по Заказам)']
    # )
    # df_final_orders['Заказы всего руб'] = np.where(
    #     (df_final_orders['Заказы модели руб (из отчета РК)'] > 0) &
    #     (np.isclose(df_final_orders['Заказы модели руб (из отчета по Заказам)'], 0, atol=epsilon)),

    #     df_final_orders['Заказы модели руб (из отчета РК)'],
    #     df_final_orders['Заказы модели руб (из отчета по Заказам)']
    # )
    # Считаем ДРР
    df_final_orders = df_final_orders.loc[:, ~df_final_orders.columns.isin(['ДРР, %',])]
    df_final_orders['ДРР, %'] = np.where(
        df_final_orders['Заказы всего руб'] == 0,
        np.nan,  # Заменить на nan при делении на 0
        np.round(
            df_final_orders['Расход с НДС, руб'] / df_final_orders['Заказы всего руб'] * 100,
            2
        )
    )
    # Переводим целые числа в тип int
    for col in df_final_orders.select_dtypes(include=['float']):
        if np.all(df_final_orders[col].dropna() % 1 == 0):  # Убираем NaN, проверяем дробную часть
            df_final_orders[col] = df_final_orders[col].astype('Int64')  # Int64 поддерживает пропуски (NA). Если нет NA, можно int

    return df_final_orders


# Функция создания датафрейма для записи в excel
def create_excel_report(df_final_orders):
    # Создаем копию для избежания изменений в оригинальном df
    df_excel_companies_report = df_final_orders.copy()

    # Убираем лишние колонки
    df_excel_companies_report = df_excel_companies_report.loc[:, ~df_excel_companies_report.columns.isin([
        'unique_company_id',
        'ID Размерного ряда',
        'model_id',
    ])]

    # Делаем "удобный" порядок колонок
    columns_order = [ # 'Тип кампании',
        'Номер кампании', 'Дата начала статистики РК', 'Дата окончания статистики РК',
        'SKU', 'Артикул', 'Название товара', 'Тип товара',
        'Количество размеров в артикуле', 'Цена товара, руб',
        'Показы', 'Клики', 'CTR (%)', 'В корзину', 'Ср. цена клика, руб',
        'Заказы шт (из отчета РК)', 'Заказы руб (из отчета РК)',
        'Заказы модели шт (из отчета РК)', 'Заказы модели руб (из отчета РК)',
        'Заказы шт (из отчета по Заказам)',
        'Заказы руб (из отчета по Заказам)',
        'Заказы модели шт (из отчета по Заказам)',
        'Заказы модели руб (из отчета по Заказам)',
        'Заказы размерного ряда шт (из отчета по Заказам)',
        'Заказы размерного ряда руб (из отчета по Заказам)',
        'Заказы всего шт', 'Заказы всего руб',
        'Расход с НДС, руб', 'ДРР, %',
        'Остаток',
        'Дата добавления',
    ]
    # Если какой-то колонки нет, добавляем её
    for col in columns_order:
        if col not in df_excel_companies_report.columns:
            df_excel_companies_report[col] = np.nan

    # Делаем нужный порядок колонок
    df_excel_companies_report =df_excel_companies_report.loc[:, columns_order]

    return df_excel_companies_report


# Функция сохранения и форматирования отчета excel
def save_and_format_excel(
        df_excel_companies_report,
        date_campaign_report
):
    # Имя файла для сохранения
    file_name_report = (
        f"{BASE_DIR}/Clients/{client_name}/CampaignsReport/{date_campaign_report}/"
        f"{date_campaign_report}_Отчет_РК_{client_name}_Ozon.xlsx"
    )
    # Сохраняем отчет в файл excel
    with pd.ExcelWriter(file_name_report) as w:
        df_excel_companies_report.to_excel(w, index=False)

    logger.info('Formatting companies report')

    # Форматируем файл с отчетом

    # Открываем книгу excel
    wb = load_workbook(file_name_report)
    # Выбираем нужный лист
    ws = wb['Sheet1']

    # Определяем стиль границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Создаем границы
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

    # Применение числового форматирования с разделением разрядов (через пробелы)
    # Получаем типы данных из исходного датафрейма
    columns = list(df_excel_companies_report.columns)
    dtypes = df_excel_companies_report.dtypes

    for idx, col in enumerate(columns, start=1):
        if pd.api.types.is_integer_dtype(dtypes[col]):
            num_format = '# ##0'  # Excel формат - разделение пробелом между тысячами
        elif pd.api.types.is_float_dtype(dtypes[col]):
            num_format = '# ##0.00'  # Разделение и 2 знака после запятой
        else:
            continue  # Пропускаем нечисловые

        for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):  # min_row=2 чтобы пропустить шапку
            for c in cell:
                c.number_format = num_format

    # Автоподбор ширины столбцов + фильтр на столбцы
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for i, cell in enumerate(col):
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        # Если это заголовок (первая строка), учитываем значок фильтра
        head_len = len(str(col[0].value)) if col[0].value else 0
        filter_icon_length = 3 if head_len == max_len else 2  # 3 для filter icon
        ws.column_dimensions[col_letter].width = max_len + filter_icon_length

    # Значок фильтра на столбцы
    ws.auto_filter.ref = ws.dimensions

    # Сохранение обновленного файла
    wb.save(file_name_report)


# Вызов всех функций
def create_campaigns_report(
        client_name,
        client_id_performance,
        client_secret_performance,
        headers,
        input_companies,
        upload_path,
        date_campaign_report = str(date.today()),
        to_save=True,
        delete_files=False,
):
    logger.info(f"Creating companies report for client {client_name} for date {date_campaign_report}")

    # Проверяем, нужна ли статистика по всем кампаниям
    all_companies_stats_check = check_all_companies_stats(input_companies)
    # Создаем датафрейм с ID кампаний, их датами и ожидаемым списком файлов
    df_input_companies = create_companies_upload_df(
        input_companies,
        all_companies_stats_check,
        upload_path
    )

    # Получаем токен авторизации
    auth_token = getAuthorizationToken(
        client_id_performance,
        client_secret_performance,
    )

    # Получаем список кампаний АПИ
    df_input_companies_parsed = get_company_list(
        all_companies_stats_check,
        df_input_companies,
        auth_token
    )
    # Получаем статистику кампаний
    get_company_statistics(
        df_input_companies_parsed,
        upload_path,
        auth_token,
    )
    # Обрабатываем загруженные файлы по рекламным кампаниям
    df_companies_parsed = parse_companies_files(
        date_campaign_report,
        df_input_companies_parsed
    )
    # Выгружаем список товаров
    df_products = get_ozon_product(headers, to_save=False)
    # Обрабатываем список товаров
    df_products_processed = process_product_list(df_products)
    # Добавляем данные по товарам в отчет по РК
    df_companies_products_data = add_products_data_to_companies_reports(
        df_companies_parsed,
        df_products_processed
    )
    # Получаем мин. и макс. дату для заказов
    date_start_orders, date_end_orders = get_min_max_dates(df_input_companies_parsed)
    # Выгружаем заказы
    df_orders_all = get_orders(headers, date_start_orders, date_end_orders)
    # Выгружаем остатки
    df_reminders_all = get_reminders(df_products_processed, headers)

    # Добавляем данные по товарам в заказы
    df_orders_products_data = add_products_data_to_orders(df_orders_all, df_products_processed)

    # Считаем различные виды заказов
    df_companies_with_orders_and_models = calc_orders_for_companies(
        df_companies_products_data,
        df_orders_products_data,
        df_products_processed
    )
    # Добавляем остатки
    df_companies_with_orders_and_reminders = add_reminders(
        df_companies_with_orders_and_models,
        df_reminders_all,
    )
    # Считаем итоговые заказы
    df_final_orders = calc_final_orders(df_companies_with_orders_and_reminders)
    # Создаем df для записи в excel
    df_excel_companies_report = create_excel_report(df_final_orders)
    # Если стоит флаг сохранения, то сохраняем отчет
    if to_save:
        # Сохраняем и форматируем отчет в excel
        save_and_format_excel(
            df_excel_companies_report,
            date_campaign_report
        )
    # Если стоит флаг удаления файлов после выполнения скрипта
    if delete_files:
        # Нормализуем путь (прямые\обратные слеши)
        upload_path_norm = os.path.normpath(upload_path)
        # Проверка, что директория существует
        if os.path.isdir(upload_path):
            try:
                send2trash.send2trash(upload_path_norm)
                print(f"Directory '{upload_path_norm}' and its contents deleted successfully.")
            except OSError as e:
                print(f"Error: {upload_path_norm} : \n{e.strerror}")
        else:
            print(f"Directory '{upload_path_norm}' does not exist.")

    logger.info(f"Done creating campaigns report for client {client_name}")

    return df_excel_companies_report




# %% Вызов функций
if __name__ == '__main__':

    # Генерируем диапазон дат
    df_date_range = generate_dates(
        start_date='2025-07-01',
        end_date='2025-07-31'
    )
    # Формируем список кампаний, по которым нужно сформировать отчет
    input_companies = create_input_companies(promotion_companies, df_date_range)
    # Дата, за которую формируется (или формировался) отчет
    date_campaign_report = str(date.today())
    # date_campaign_report = '2025-05-16'

    logger.info(f"Creating companies report for client {client_name} for date {date_campaign_report}")

   # Создаем директорию, куда будет помещена статистика кампаний из АПИ
    upload_path = create_upload_statistic_dir(
        date_campaign_report,
        client_name
    )

    # Проверяем, нужна ли статистика по всем кампаниям
    all_companies_stats_check = check_all_companies_stats(input_companies)
    # Создаем датафрейм с ID кампаний, их датами и ожидаемым списком файлов
    df_input_companies = create_companies_upload_df(
        input_companies,
        all_companies_stats_check,
        upload_path
    )

    # Получаем токен авторизации
    auth_token = getAuthorizationToken(
        client_id_performance,
        client_secret_performance,
    )

    # Получаем список кампаний АПИ
    df_input_companies_parsed = get_company_list(
        all_companies_stats_check,
        df_input_companies,
        auth_token
    )
    # Получаем статистику кампаний
    get_company_statistics(
        df_input_companies_parsed,
        upload_path,
        auth_token,
    )
    # Обрабатываем загруженные файлы по рекламным кампаниям
    df_companies_parsed = parse_companies_files(
        date_campaign_report,
        df_input_companies_parsed
    )
    # Выгружаем список товаров
    df_products = get_ozon_product(headers, to_save=False)
    # Обрабатываем список товаров
    df_products_processed = process_product_list(df_products)
    # Добавляем данные по товарам в отчет по РК
    df_companies_products_data = add_products_data_to_companies_reports(
        df_companies_parsed,
        df_products_processed
    )
    # Получаем мин. и макс. дату для заказов
    date_start_orders, date_end_orders = get_min_max_dates(df_input_companies_parsed)
    # Выгружаем заказы
    df_orders_all = get_orders(headers, date_start_orders, date_end_orders)
    # Выгружаем остатки
    df_reminders_all = get_reminders(df_products_processed, headers)

    # Добавляем данные по товарам в заказы
    df_orders_products_data = add_products_data_to_orders(df_orders_all, df_products_processed)

    # Считаем различные виды заказов
    df_companies_with_orders_and_models = calc_orders_for_companies(
        df_companies_products_data,
        df_orders_products_data,
        df_products_processed
    )
    # Добавляем остатки
    df_companies_with_orders_and_reminders = add_reminders(
        df_companies_with_orders_and_models,
        df_reminders_all,
    )
    # Считаем итоговые заказы
    df_final_orders = calc_final_orders(df_companies_with_orders_and_reminders)
    # Создаем df для записи в excel
    df_excel_companies_report = create_excel_report(df_final_orders)
    # Сохраняем и форматируем отчет в excel
    save_and_format_excel(
        df_excel_companies_report,
        date_campaign_report
    )

    logger.info(f"Done creating campaigns report for client {client_name}")

# %%
