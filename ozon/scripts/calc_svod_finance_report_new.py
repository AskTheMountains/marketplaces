
# %% Определение всех функций
import requests
import json
import time
from datetime import date,datetime,timedelta
import pandas as pd
import shutil
import os
import glob
import csv
import zipfile
import openpyxl
from zipfile import ZipFile
import numpy as np
import re
from loguru import logger
import getopt
import sys
pd.options.mode.chained_assignment = None  # default='warn'
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color


# Файл с настройками и номером клиента
# from options import headers, settings, client_number

# Некоторые константы
from ozon.scripts.constants import (
    headers,
    client_name,
    marketplace_dir_name,
    client_id_performance,
    client_secret_performance,
    catalog_finace_svod_columns
)

# Некоторые доп. функции
from generic_functions import move_columns

# Функция выгрузки товаров в поставках
from ozon.scripts.uploadDataFromOzon import get_supply_orders

# Функция выгрузки отчета о релизации товаров
from ozon.scripts.uploadFinanceReports import get_products_realization_report

# Функция выгрузки данных о рекламных кампаниях
from ozon.scripts.uploadDataFromOzonPerformance import upload_data_performane



# Функция чтения файла с отчетом
def read_finance_report_file(finance_reports_dir, date_start, date_end):
    # Получения формата дат, которые содержатся в имени файла.
    date_start_file = datetime.strptime(date_start, '%Y-%m-%d').strftime('%d_%m_%Y')
    date_end_file = datetime.strptime(date_end, '%Y-%m-%d').strftime('%d_%m_%Y')
    # Чтение excel с исходными данными
    df_finance_report = pd.read_excel(f"{finance_reports_dir}/{client_name}_Отчет_по_начислениям_{date_start_file}_{date_end_file}.xlsx",
                                      sheet_name='Начисления',
                                      skiprows=1)
    # Переименование некоторых колонок для удобства
    # df_finance_report = df_finance_report.rename(columns={
    #     'За продажу или возврат до вычета комиссий и услуг': 'Выручка',
    #     'Обработка отправления (Drop-off/Pick-up) (разбивается по товарам пропорционально количеству в отправлении)': 'Обработка отправления',
    #     'Последняя миля (разбивается по товарам пропорционально доле цены товара в сумме отправления)': 'Последняя миля',
    #     'Обработка отмененного или невостребованного товара (разбивается по товарам в отправлении в одинаковой пропорции)': 'Обработка отмененного товара',
    #     })
    df_finance_report = df_finance_report.rename(columns={
        'Ozon SKU': 'SKU',
        'Сумма итого, руб.': 'Итого руб'
    })
    # Добавляем уникальный идентификатор строки
    df_finance_report['id'] = np.arange(1, df_finance_report.shape[0] + 1)

    return df_finance_report


# Функция выгрузки Продаж, Возвратов и Баллов Озон
def calc_sales_returns(date_start, date_end):
    # Создание диапазона дат выгрузки отчета о реализации
    dt_start = pd.to_datetime(date_start, format='%Y-%m-%d')
    dt_end = pd.to_datetime(date_end, format='%Y-%m-%d')
    date_range = (pd
                  .date_range(start=dt_start, end=dt_end, freq='MS')
                  .to_frame()
                  .rename(columns={0: 'date'})
                  .reset_index(drop=True)
                #   .assign(
                #       date_end=date_range['date_start'].shift(-1)
                #       )
    )
    df_realization_report_total = pd.DataFrame()
    for i in range(date_range.shape[0]):
        # Определение месяца и года, за который будет выгружен отчет о реализации
        month = date_range['date'][i].month
        year = date_range['date'][i].year

        logger.info(f"Uploading realization report year: {year} month: {month}")

        # Выгрузка отчета о релизации товаров
        df_realization_report = get_products_realization_report(headers, year, month, to_save=False)
        # Колонка с месяцем и годом для информации
        df_realization_report = df_realization_report.assign(
            month=month,
            year=year
        )
        # Расчет итоговых столбцов по продажам и возвратам
        df_realization_report['Продажи руб'] = df_realization_report[['Сумма_продажи',
                                                                    'Выплаты_по_механикам_лояльности_партнёров_звёзды_продажи',
                                                                    'Выплаты_по_механикам_лояльности_партнёров_зелёные_цены_продажи',
                                                                    'Выплаты_по_механикам_лояльности_партнёров_апвз_продажи']] \
                                                                        .sum(axis=1)
        df_realization_report['Продажи шт'] = df_realization_report['Количество_продажи']
        df_realization_report['Возвраты руб'] = -abs(df_realization_report[['Сумма_возвраты',
                                                                    'Выплаты_по_механикам_лояльности_партнёров_звёзды_возвраты',
                                                                    'Выплаты_по_механикам_лояльности_партнёров_зелёные_цены_возвраты',
                                                                    'Выплаты_по_механикам_лояльности_партнёров_апвз_возвраты']] \
                                                                        .sum(axis=1))
        df_realization_report['Возвраты шт'] = df_realization_report['Количество_возвраты']
        df_realization_report['Продажи - Возвраты (руб)'] = df_realization_report['Продажи руб'] - abs(df_realization_report['Возвраты руб'])
        df_realization_report['Продажи - Возвраты (шт)'] = df_realization_report['Продажи шт'] - df_realization_report['Возвраты шт']

        # Расчет итогового столбца по баллам озон
        df_realization_report['Баллы Озон Итог'] = df_realization_report['Баллы_за_скидки_продажи'] - df_realization_report['Баллы_за_скидки_возвраты']

        # Выбор нужных колонок
        df_realization_report = df_realization_report.loc[:, ['SKU', 'Артикул_продавца', 'Наименование_товара', 'Продажи шт', 'Возвраты шт', 'Продажи - Возвраты (шт)', 'Продажи руб', 'Сумма_продажи',  'Возвраты руб', 'Сумма_возвраты',  'Продажи - Возвраты (руб)', 'Баллы Озон Итог']]
        # Объединение с предыдущим проходом цикла
        df_realization_report_total = pd.concat([df_realization_report, df_realization_report_total])

    # Расчет итоговых сумм за все периоды
    df_realization_report_stats = (
        df_realization_report_total
        .groupby(['SKU', 'Артикул_продавца', 'Наименование_товара'])
        .sum()
        .reset_index()
        .rename(columns={
            'Артикул_продавца': 'Артикул продавца',
            'Наименование_товара': 'Наименование товара'
        })
    )

    # Расчет общего количества продаж, возвратов и баллов озон
    df_sales_returns_by_operation = (
        df_realization_report_stats
        .loc[:, ~df_realization_report_stats.columns.isin(['SKU', 'Артикул продавца', 'Наименование товара'])]
        .sum()
        .reset_index()
        .rename(columns={
            'index': 'Тип начисления',
            0: 'Итого'
        })
    )

    # Помещаем результат в словарь
    result_sales_returns = {
        'df_realization_report_stats': df_realization_report_stats,
        'df_sales_returns_by_operation': df_sales_returns_by_operation
    }

    return result_sales_returns


# Функция расчета расходов, привязанных к SKU
def calc_sku_costs(df_finance_report, groupby_cols = ['SKU', 'Артикул', 'Название товара']):
    # Выбираем затраты, которые идут с привязкой к SKU
    df_finance_report_sku = df_finance_report.loc[~df_finance_report['SKU'].isna(), :]
    # Список номеров строк с расходами
    sku_costs_ids = df_finance_report_sku['id'].to_list()
    # Колонки, по которым делаем группировку
    groupby_cols = groupby_cols

    # Типы начислений, относящиеся к продажам и возвратам
    sales_and_returns_ids = df_finance_report_sku.loc[df_finance_report_sku['Тип начисления'].isin(['Выручка', 'Возврат выручки']), 'id'].to_list()
    # Делаем выборку только по типам начислений с продажами и возвратами
    df_sales_returns= df_finance_report_sku.loc[df_finance_report_sku['id'].isin(sales_and_returns_ids), :]
    # Заменяем значения на "Продажа" и "Возврат"
    df_sales_returns['Тип начисления'] = (
        df_sales_returns['Тип начисления']
        .replace({
            'Выручка': 'Продажи',
            'Возврат выручки': 'Возвраты'
        })
    )
    # Считаем продажи и возвраты в штуках
    df_sku_costs_sales_returns_stats = (
        pd.pivot_table(
            df_sales_returns,
            values=['Количество', 'Итого руб'],
            index=groupby_cols,
            columns='Тип начисления',
            aggfunc='sum',
            fill_value=0
        )
    )
    # Избавление от MultiIndex и добавление префиксов
    df_sku_costs_sales_returns_stats.columns = [
        f"{col[1]} шт" if col[0] == 'Количество' else f"{col[1]} руб"
        for col in df_sku_costs_sales_returns_stats.columns
    ]
    # Получаем столбцы из Index
    # df_sku_costs_sales_returns_stats = df_sku_costs_sales_returns_stats.reset_index()

    # # Расчет возвратов и продаж в штуках и рублях
    # df_products_grouped = df_finance_report_sku.groupby(by=groupby_cols, dropna=False)
    # df_with_sku_stats = pd.DataFrame()
    # df_with_sku_stats['tmp_col'] = df_products_grouped.size()
    # # Продажи и возвраты считаем по столбцу "Выручка"
    # df_with_sku_stats['Продажи шт'] = df_finance_report_sku.loc[df_finance_report_sku['Выручка'] > 0, :] \
    #     .groupby(by=groupby_cols, dropna=False) \
    #         .size()
    # df_with_sku_stats['Продажи рубли'] = df_finance_report_sku.loc[df_finance_report_sku['Выручка'] > 0, :] \
    #     .groupby(by=groupby_cols, dropna=False)['Выручка'] \
    #         .sum()
    # df_with_sku_stats['Возвраты шт'] = df_finance_report_sku.loc[df_finance_report_sku['Выручка'] < 0, :] \
    #     .groupby(by=groupby_cols, dropna=False) \
    #         .size()
    # df_with_sku_stats['Возвраты рубли'] = df_finance_report_sku.loc[df_finance_report_sku['Выручка'] < 0, :] \
    #     .groupby(by=groupby_cols, dropna=False)['Выручка'] \
    #         .sum()

    # Делаем выборку только по типам начислений с продажами и возвратами
    df_sku_other_costs = df_finance_report_sku.loc[~df_finance_report_sku['id'].isin(sales_and_returns_ids), :]
    # Считаем прочие расходы в рублях
    df_sku_other_costs_stats = (
        pd.pivot_table(
            df_sku_other_costs,
            values='Итого руб',
            index=groupby_cols,
            columns='Тип начисления',
            aggfunc='sum',
            fill_value=0
        )
    )
    # Объединяем все расходы в один df
    df_sku_costs_all = (
        pd.concat([
            df_sku_costs_sales_returns_stats,
            df_sku_other_costs_stats
        ])
        #.reset_index()
        .groupby(level=[0, 1, 2]) # 0 - SKU, 1 - Артикул, 2 - Наименование
        .sum()
        .reset_index()
        .rename(columns={
            'Артикул': 'Артикул продавца',
            'Название товара': 'Наименование товара'
        })
    )

    # Если не было продаж или возвратов, искуственно создаем эти колонки
    for col in ['Продажи шт', 'Продажи руб', 'Возвраты шт', 'Возвраты руб']:
        if col not in df_sku_costs_all.columns:
            df_sku_costs_all[col] = 0

    # Продажи минус возвраты
    df_sku_costs_all['Продажи - Возвраты (шт)'] = df_sku_costs_all['Продажи шт'] - df_sku_costs_all['Возвраты шт']
    df_sku_costs_all['Продажи - Возвраты (руб)'] = df_sku_costs_all['Продажи руб'] - abs(df_sku_costs_all['Возвраты руб'])

    # Перемещаем колонки в начало df
    df_sku_costs_all = move_columns(
        df_sku_costs_all,
        columns_to_move=[
            'Продажи шт', 'Возвраты шт',
            'Продажи руб', 'Возвраты руб',
            'Продажи - Возвраты (шт)',
            'Продажи - Возвраты (руб)'
        ],
        position='Наименование товара',
        insert_type='after'
    )

    # Считаем итоговые суммы на каждый вид расходов
    df_sku_costs_by_operation = (
        df_sku_costs_all
        .loc[:, ~df_sku_costs_all.columns.isin(['SKU', 'Артикул продавца', 'Наименование товара'])]
        .sum()
        .reset_index()
        .rename(columns={
            'index': 'Тип начисления',
            0: 'Итого'
        })
    )

    # Продажи и возвраты берем из отчета о реализации
    # df_realization_report = result_sales_returns['df_realization_report_stats'].copy()
    # df_with_sku_stats = pd.concat([df_with_sku_stats, df_realization_report])
    # Считаем суммы после concat
    # df_with_sku_stats = df_with_sku_stats.groupby(groupby_cols).sum().reset_index()
    # Удаляем ненужную колонку
    # df_with_sku_stats = df_with_sku_stats.drop(columns=['tmp_col'])
    # Итоговый словарь с результатами
    result_sku_costs = {
        'df_sku_costs_all': df_sku_costs_all,
        'df_sku_costs_by_operation': df_sku_costs_by_operation,
        'sku_costs_ids': sku_costs_ids
        }

    return result_sku_costs


# Функция проверки наличия расходов по рекламным кампаниям
def check_companies_costs(df_finance_report):
    # Флаг наличия рекламных расходов
    companies_costs_flag = False
    # Выбираем затраты по рекламе, где есть кампании
    df_companies_costs = df_finance_report.loc[
        (df_finance_report['Название товара'].str.contains('Заказ №', na=False) )
        | (df_finance_report['ID начисления'].str.match('^\d{8}$', na=False)),
        ['id', 'Группа услуг', 'ID начисления', 'Тип начисления', 'Название товара', 'Итого руб']
    ]

    if not df_companies_costs.empty:
        companies_costs_flag = True

    return companies_costs_flag


# Функция нахождения идентификаторов рекламных кампаний
def get_companies_ids(df_finance_report):

    # Выбираем затраты по рекламе, где есть кампании
    df_companies_costs = df_finance_report.loc[
        (df_finance_report['Название товара'].str.contains('Заказ №', na=False) )
        | (df_finance_report['ID начисления'].str.match('^\d{8}$', na=False)),
        ['Группа услуг', 'ID начисления', 'Тип начисления', 'Название товара', 'Итого руб']
    ]

    if df_companies_costs.shape[0] > 0:
        # Создаем столбец с номером кампании
        conditions = [
            # Трафареты
            df_companies_costs['Название товара'].str.contains('Заказ №', na=False),
            # Выводы в топ
            df_companies_costs['ID начисления'].str.match('^\d{8}$', na=False)
        ]
        choices = [
            # Номер кампании для трафаретов
            (df_companies_costs['Название товара']
            .str.split('Заказ №').str[1]
            .str.split('.').str[0]),
            # Номер кампании для выводов в топ
            df_companies_costs['ID начисления']
        ]
        df_companies_costs['Номер кампании'] = np.select(conditions, choices, default=np.nan)

        # Считаем расходы на каждую кампанию
        df_companies_costs_total = (
            df_companies_costs
            .groupby(['Номер кампании', 'Тип начисления'])
            .agg(
                Итого_из_отчета=('Итого руб', 'sum')
            )
            .reset_index()
        )

    return df_companies_costs_total


# Функция парсинга файлов со статистикой по кампаниям
def parse_companies_files(finance_reports_dir, date_report):
    # Директория, где лежат данные по рекламным кампаниям
    companies_stats_dir = f"{finance_reports_dir}/{date_report}_Кампании"
    # Директория, куда будем складывать обработанные файлы
    companies_parsed_dir = f"{finance_reports_dir}/{date_report}_Кампании_parsed"
    # Копируем выгруженные файлы по рекламных расходам из отдельной папки
    # companies_upload_dir = f"Clients/{client_name}/UploadFilesPerformance/UploadFiles_{date_upload_performance}/{date_upload_performance}_Кампании"
    # companies_upload_dir = companies_stats_dir

    # Получаем список файлов в директории
    # companies_stats_files = os.listdir(companies_stats_dir)
    # Если директория с выгрузкой по рекламе существует за указанную дату
    # if os.path.exists(companies_upload_dir):
    #     companies_upload_files = os.listdir(companies_upload_dir)
    #     # Если есть выгруженные файлы по рекламе и в папке с отчетом нет скопированных ранее файлов,
    #     # то копируем файлы по рекламным кампаниям
    #     if (len(companies_upload_files) > 0) & (len(companies_stats_files) == 0):
    #         for upload_file in companies_upload_files:
    #             src_path = os.path.join(companies_upload_dir, upload_file)
    #             dst_path = os.path.join(companies_stats_dir, upload_file)
    #             shutil.copy(src_path, dst_path)

    # Получаем список csv файлов по кампаниям
    filenames_companies = {"path": []}
    path_companies = f"{companies_stats_dir}/*.csv"
    # Считывание файла с путем до него
    for file in glob.glob(path_companies):
        filenames_companies['path'].append(file)
    # Считывание только имени файла
    filenames_companies['file_name'] = [os.path.basename(x) for x in glob.glob(path_companies)]
    # Размер файла
    filenames_companies['file_size'] = [os.path.getsize(x) for x in glob.glob(path_companies)]
    # Перевод в df
    filenames_companies = pd.DataFrame(filenames_companies)
    # Номер заказа
    filenames_companies['order_number'] = filenames_companies['file_name'].str.split('_').str[0]
    # Сортировка по размеру
    filenames_companies = filenames_companies.sort_values(by='file_size', ascending=False, ignore_index=True)

    # Обработка csv по статистике кампаний в цикле
    for i in range(len(filenames_companies)):
        # Считываем csv
        df_company = pd.read_csv(filenames_companies['path'][i], sep=';', decimal=',', skiprows=1)
        # Если по кампании есть расходы, начинаем их обрабатывать
        if df_company.shape[0] >= 2:
            # Определяем тип отчета (по первой колонке в отчете)
            # TODO: придумать более надежный способ
            if df_company.columns[0] == 'Дата':
                # Убираем строку Всего и Корректировка
                df_company = df_company.loc[~df_company['Дата'].isin(['Всего', 'Корректировка']), :]
                # Ozon ID = SKU в данном отчете
                df_company = df_company.rename(columns={'Ozon ID': 'sku'})
                # Переименовываем колонку с расходами для удобства
                df_company = df_company.rename(columns={'Расход, ₽': 'Расходы'})
                # Переименовываем колонку с названием товара
                df_company = df_company.rename(columns={'Наименование': 'Наименование товара'})
                # Выбираем нужные колонки
                df_company = df_company.loc[:, ['Дата', 'sku', 'Наименование товара', 'Расходы']]
                df_company.to_csv(f"{companies_parsed_dir}/{filenames_companies['order_number'][i]}.csv", sep=';')
            else:
                # Убираем строку Всего и Корректировка
                df_company = df_company.loc[~df_company['sku'].isin(['Всего', 'Корректировка']), :]
                # Переименовываем колонку с расходами для удобства
                df_company = df_company.rename(columns={'Расход, ₽, с НДС': 'Расходы'})
                # Переименовываем колонку с названием товара
                df_company = df_company.rename(columns={'Название товара': 'Наименование товара'})
                # Выбираем нужные колонки
                df_company = df_company.loc[:, ['sku', 'Наименование товара', 'Расходы']]
                df_company.to_csv(f"{companies_parsed_dir}/{filenames_companies['order_number'][i]}.csv", sep=';', index=False)


# Функция расчета расходов по рекламным кампаниям
def calc_companies_cost_by_sku(date_report, df_finance_report):

    # df, куда будем помещать итоговый результат (разбивку расходов по SKU)
    df_result_companies_stats = pd.DataFrame()
    # df с рекламными расходами из отчета ЛК
    df_companies_stats_from_report = pd.DataFrame()
    # df с сравнением расходов по рекламе из АПИ и отчета ЛК
    df_companies_api_report_stats = pd.DataFrame()
    # id строк отчета, которые относятся к рекламным расходам
    companies_costs_ids = []

    # Выбираем затраты по рекламе, где есть кампании
    df_companies_costs = df_finance_report.loc[
        (df_finance_report['Название товара'].str.contains('Заказ №', na=False) )
        | (df_finance_report['ID начисления'].str.match('^\d{8}$', na=False)),
        ['id', 'Группа услуг', 'ID начисления', 'Тип начисления', 'Название товара', 'Итого руб']
    ]

    if df_companies_costs.shape[0] > 0:
        # Создаем столбец с номером кампании
        conditions = [
            # Трафареты
            df_companies_costs['Название товара'].str.contains('Заказ №', na=False),
            # Выводы в топ
            df_companies_costs['ID начисления'].str.match('^\d{8}$', na=False)
        ]
        choices = [
            # Номер кампании для трафаретов
            (df_companies_costs['Название товара']
            .str.split('Заказ №').str[1]
            .str.split('.').str[0]),
            # Номер кампании для выводов в топ
            df_companies_costs['ID начисления']
        ]
        df_companies_costs['Номер кампании'] = np.select(conditions, choices, default=np.nan)
        companies_costs_ids = df_companies_costs['id'].to_list()

        # Считаем расходы на каждую кампанию
        df_companies_costs_total = (
            df_companies_costs
            .groupby(['Номер кампании', 'Тип начисления'])
            .agg(
                Итого_из_отчета=('Итого руб', 'sum')
            )
            .reset_index()
        )
        # Помещаем расходы в df из начала функции
        df_companies_stats_from_report = df_companies_costs_total.copy()

        # Сопоставление номера заказа с расходами SKU из апи
        df_companies_api = pd.DataFrame()
        for i in range(len(df_companies_costs_total)):
            # Файл с данными по кампании
            company_file_path = f"{finance_reports_dir}/{date_report}_Кампании_parsed/{df_companies_costs_total['Номер кампании'][i]}.csv"
            # Если данный файл есть, начинаем считать расходы по данной кампании
            if os.path.exists(company_file_path):
                # Считываем данные по данной кампании
                df_company = pd.read_csv(company_file_path, sep=';')
            # Если нет, считаем, что расходы по данной кампании были равны 0
            else:
                df_company = pd.DataFrame(columns=['sku', 'Наименование товара', 'Расходы'])
                # Выводим предупреждение, что файла с расходами по данной кампании нет
                logger.warning(f"Company file {df_companies_costs_total['Номер кампании'][i]} has not been found")
            # Удаляем лишнюю колонку, если есть
            if 'Unnamed: 0' in df_company.columns:
                df_company = df_company.drop(columns='Unnamed: 0')
            # Считаем итоговые расходы по каждому SKU
            tmp_df = df_company.groupby(['sku', 'Наименование товара']).agg(
                Итого=('Расходы', 'sum')
            ).reset_index()
            # Добавляем столбцы с информацией о названии заказа
            tmp_df['Номер кампании'] = df_companies_costs_total['Номер кампании'][i]
            tmp_df['Тип начисления'] = df_companies_costs_total['Тип начисления'][i]

            # Объединяем с предыдущей кампанией
            df_companies_api = pd.concat([df_companies_api, tmp_df])

        # Считаем затраты на каждый тип начисления на каждый SKU
        df_companies_api_stats = (
            df_companies_api
            .groupby(['sku', 'Наименование товара', 'Тип начисления'])
            .agg(
            Итого=('Итого', 'sum')
            )
        .reset_index()
        )
        # Заменяем знак расходов с плюса на минус
        df_companies_api_stats['Итого'] = np.where(
            df_companies_api_stats['Итого'] > 0,
            -abs(df_companies_api_stats['Итого']),
            df_companies_api_stats['Итого']
        )
        # Переводим в длинный формат
        df_companies_api_stats = (
            pd.pivot_table(
                df_companies_api_stats,
                index=['sku', 'Наименование товара'],
                columns='Тип начисления',
                values='Итого',
                aggfunc='sum',
                fill_value=0
            )
            .reset_index()
        )
        # Переименовываем колонку с SKU
        df_companies_api_stats = df_companies_api_stats.rename(columns={'sku': 'SKU'})
        # Убираем имя у index
        df_companies_api_stats.index.name = None
        # Помещаем итоговый результат в переменную из начала функции
        df_result_companies_stats = df_companies_api_stats.copy()

        # Считаем расходы по типам начисления из апи
        df_companies_api_stats_by_operation = df_companies_api.groupby(['Номер кампании', 'Тип начисления']).agg(
            Итого_из_апи=('Итого', 'sum')
        ).reset_index()
        # Заменяем знак расходов с плюса на минус
        df_companies_api_stats_by_operation['Итого_из_апи'] = np.where(
            df_companies_api_stats_by_operation['Итого_из_апи'] > 0,
            -abs(df_companies_api_stats_by_operation['Итого_из_апи']),
            df_companies_api_stats_by_operation['Итого_из_апи']
        )
        # Мерджим с расходами из отчета
        df_companies_api_report_stats = df_companies_costs_total.merge(
            df_companies_api_stats_by_operation,
            on=['Номер кампании', 'Тип начисления'],
            how='left'
        )

    # Словарь с результатами
    result_companies_costs = {
        'df_companies_api_stats': df_result_companies_stats,
        'df_companies_stats_from_report': df_companies_stats_from_report,
        'df_companies_api_report_stats': df_companies_api_report_stats,
        'companies_costs_ids': companies_costs_ids
    }

    return result_companies_costs


# Функция добавления артикулов и наименований товаров к расходам по рекламным кампаниям
def add_articles_to_companies_costs(result_companies_costs, df_sku_list_all):
    # Получаем df с результатами расходов SKU по рекламным кампаниям
    df_result_companies_stats = result_companies_costs['df_companies_api_stats'].copy()
    # Мерджим со списком артикулов
    df_companies_stats_with_articles = (
        df_result_companies_stats
        .merge(
            df_sku_list_all[['SKU', 'Артикул продавца']],
            on='SKU',
            how='left'
        )
    )
    # Перемещаем колонку с артикулом в начало df
    df_companies_stats_with_articles = move_columns(
        df=df_companies_stats_with_articles,
        columns_to_move=['Артикул продавца'],
        position='SKU',
        insert_type='after'
    )
    # Заполняем пропуски
    df_companies_stats_with_articles = df_companies_stats_with_articles.fillna({
        'Артикул продавца': 'Неизвестный артикул'
    })
    # Изменяем df в словаре рекламных расходов на тот, что с артикулами
    result_companies_costs['df_companies_api_stats'] = df_companies_stats_with_articles.copy()

    return result_companies_costs


# Функция расчета данных по рекламным кампаниям
def calc_companies_costs(
        date_start,
        date_end,
        date_report,
        df_finance_report,
        df_sku_list_all,
        client_id_performance,
        client_secret_performance,
        task_id=None
    ):
    # Директория, где будут лежать данные по рекламным кампаниям
    companies_stats_dir = f"{finance_reports_dir}/{date_report}_Кампании"
    if not os.path.exists(companies_stats_dir):
        os.mkdir(companies_stats_dir)  # Создаем, если директория не существует
    # Директория, куда будем складывать обработанные файлы
    companies_parsed_dir = f"{finance_reports_dir}/{date_report}_Кампании_parsed"
    if not os.path.exists(companies_parsed_dir):
        os.mkdir(companies_parsed_dir)

    # Проверяем, что в отчете были расходы по рекламе
    companies_costs_flag = check_companies_costs(df_finance_report)
    if companies_costs_flag:
        logger.info('Found companies costs')
        # Получаем идентификаторы рекламных кампаний
        df_companies_ids = get_companies_ids(df_finance_report)
        # Получаем список файлов в директории с кампаниями
        companies_stats_files = os.listdir(companies_stats_dir)

        # Если нет выгруженных файлов по кампаниям, то делаем выгрузку статистики кампаний
        if not companies_stats_files:
            logger.info('No companies files found, uploading companies stats')
            # Формируем даты для выгрузки по API Performance
            date_start_performance = date_start + 'T00:00:00Z'
            date_end_performance = date_end + 'T23:59:59Z'
            # Выгружаем данные по рекламным кампаниям
            upload_data_performane(
                date_start_performance,
                date_end_performance,
                client_id_performance,
                client_secret_performance,
                companies_stats_dir,
                df_companies_ids,
                task_id
            )
        else:
            logger.info('Companies stats already downloaded')

        # Обрабатываем выгруженные данные
        parse_companies_files(finance_reports_dir, date_report)

    # Считаем расходы по рекламным кампаниям
    result_companies_costs_no_article = calc_companies_cost_by_sku(date_report, df_finance_report)
    # Добавляем артикул к SKU с рекламными расходами
    result_companies_costs = add_articles_to_companies_costs(result_companies_costs_no_article, df_sku_list_all)

    return result_companies_costs


# Функция выгрузки товаров в заявках на поставку
# def upload_supplies_items(supply_order_list):
#     # Если нет выгруженного файла с поставками, то выгружаем его
#     if not os.path.isfile(f"{finance_reports_dir}/{client_name}_Товары_в_поставках.xlsx"):
#         df_supplies_items = get_supply_orders(headers=headers, states=[], to_save=False)
#         df_supplies_items.to_excel(f"{finance_reports_dir}/{client_name}_Товары_в_поставках.xlsx")
#     # Если такой файл есть, считываем его
#     else:
#         df_supplies_items = pd.read_excel(f"{finance_reports_dir}/{client_name}_Товары_в_поставках.xlsx")

#     return df_supplies_items


# Функция расчета расходов по поставкам
def calc_supplies_costs(df_finance_report):

    # Итоговый df, который будет возвращен функцией
    supplies_costs_matched = pd.DataFrame()
    # Расходы по поставкам из отчета
    supplies_costs_from_report = pd.DataFrame()
    # id строк отчета с поставками, которые нашлись в апи
    supplies_costs_ids_matched = []

    # Выбираем расходы, которые относятся к поставкам
    df_supplies_from_report = df_finance_report.loc[
        # Все заявки на поставку начинаются с цифры 2
        (df_finance_report['ID начисления'].str.startswith('2', na=False))
        # Не содержат -
        & ~(df_finance_report['ID начисления'].str.contains('-', na=False))
        # Все заявки на поставку содержат ровно 13 цифр
        & (df_finance_report['ID начисления'].str.len() == 13),
        ['id', 'Группа услуг', 'ID начисления', 'Тип начисления', 'Название товара', 'Итого руб']
    ]


    # Если есть расходы на поставки в отчете
    if df_supplies_from_report.shape[0] > 0:
        # Список номеров строк с расходами на поставки, расходы по которым можно разбить на sku
        supplies_costs_ids_matched = []
        # Считаем расходы по конкретной заявке и по типу начисления
        df_supplies_from_report = df_supplies_from_report.rename(columns={'ID начисления': 'Номер заявки на поставку'})
        # Конвертация в Int64 для избежания ошибок при join
        df_supplies_from_report['Номер заявки на поставку'] = df_supplies_from_report['Номер заявки на поставку'].astype('Int64')
        df_supplies_from_report_total = df_supplies_from_report.groupby(['Номер заявки на поставку', 'Тип начисления']).agg(
            Расходы_поставки_из_отчета_ЛК=('Итого руб', 'sum')
        ).reset_index()
        # Помещаем расходы на поставки из ЛК в отдельный df
        supplies_costs_from_report = df_supplies_from_report_total.copy()
        # df_supplies_from_report_total['Номер заявки на поставку'] = df_supplies_from_report_total['Номер заявки на поставку'].astype('Int64')

        # Выгрузка товаров в заявке на поставку
        df_supplies_items = get_supply_orders(headers, states=[], to_save=False)
        # df_supplies_items.to_excel(f"{finance_reports_dir}/{client_name}_Товары_в_поставках.xlsx")

        # Переименовываем колонки для удобства
        df_supplies_items = df_supplies_items.rename(columns={
            'sku': 'SKU',
            'Артикул': 'Артикул продавца',
        })
        # Считаем итоговое количество sku в заявке на поставку
        df_supplies_items_total = (
            df_supplies_items
            .groupby(['SKU', 'Артикул продавца', 'Наименование товара', 'supply_id'])
            .agg(
                quantity=('quantity', 'sum')
            )
            .reset_index()
        )
        # Количество товаров в заявке
        df_supplies_items_total['Количество товаров в поставке'] = df_supplies_items_total.groupby('supply_id')['quantity'].transform('sum')

        # Конвертация в Int64 для избежания ошибок при join
        df_supplies_items_total['supply_id'] = df_supplies_items_total['supply_id'].astype('Int64')

        # Джойним таблицу с расходами из отчета для получения типа начисления
        # BUG: После джойна одна из колонок конвертируется во float
        # Из-за этого пришлось для соединяемых колонок делать конвертацию в Int64
        df_supplies_all = df_supplies_items_total.merge(
            df_supplies_from_report_total[['Номер заявки на поставку', 'Тип начисления', 'Расходы_поставки_из_отчета_ЛК']],
            left_on='supply_id',
            right_on='Номер заявки на поставку',
            how='outer',
            indicator=True
        )

        # Номера заявок на поставку, которые совпали с апи
        df_supplies_found = df_supplies_all.loc[df_supplies_all['_merge'] == 'both', :]
        # Номера заявок на поставку, которые не совпали с апи
        # df_supplies_mismatched = df_supplies_all.loc[df_supplies_all['_merge'] == 'right_only', :]

        # Если есть совпадающие заявки, считаем по ним расходы
        if df_supplies_found.shape[0] > 0:
            # Получаем id строк отчета из ЛК с поставками, расходы по которым мы разбили на sku
            supplies_costs_ids_matched = df_supplies_from_report.loc[df_supplies_from_report['Номер заявки на поставку'].isin(df_supplies_found['Номер заявки на поставку'].dropna().unique()), 'id'].to_list()
            # Считаем число товаров в каждой поставке
            # df_supplies_found['Количество товаров в поставке'] = df_supplies_found.groupby('supply_id')['quantity'].transform('sum')
            # Считаем расходы на отдельный sku (распределение от количества sku в поставке)
            df_supplies_found['Расходы по поставкам'] = df_supplies_found['quantity'] / df_supplies_found['Количество товаров в поставке'] * df_supplies_found['Расходы_поставки_из_отчета_ЛК']
            # Переводим в широкий формат
            df_supplies_found_total = (
                pd.pivot_table(
                    df_supplies_found,
                    index=['SKU', 'Артикул продавца', 'Наименование товара'],
                    columns='Тип начисления',
                    values='Расходы по поставкам',
                    aggfunc='sum',
                    fill_value=0
                )
                .reset_index()
            )
            # Переименовываем колонку с SKU
            # df_supplies_found_total = df_supplies_found_total.rename(columns={'sku': 'SKU'})
            # Переименовываем index
            df_supplies_found_total.index.name = None
            # Сохраняем результат в отдельную переменную, которую будет возвращать функция
            supplies_costs_matched = df_supplies_found_total.copy()

        # Если есть поставки, которых нет в апи, их будем считать отдельно в другой функции
        # if df_supplies_mismatched.shape[0] > 0:
        #     supplies_costs_ids_matched = df_supplies_from_report.loc[df_supplies_from_report['Номер заявки на поставку'].isin(df_supplies_mismatched['Номер заявки на поставку'].dropna().unique()), 'id'].to_list()

    # Формируем итоговый словарь с датафреймом расходов по поставкам
    # и id поставок, которых нет в апи
    result_supplies_costs = {
        'supplies_costs_matched': supplies_costs_matched,
        'supplies_costs_from_report': supplies_costs_from_report,
        'supplies_costs_ids_matched': supplies_costs_ids_matched
    }

    return result_supplies_costs


# Функция чтения справочной таблицы
def read_catalog(client_name):
    # Считывание справочной таблицы
    catalog = pd.read_excel(f"{marketplace_dir_name}/Clients/{client_name}/catalog/Справочная_таблица_{client_name}_Ozon.xlsx")
    # Из справочной таблицы убираем товары, у которых нет SKU или Артикула
    # catalog_ = catalog.dropna(subset=['SKU', 'Артикул продавца'])
    # В справочной таблице переводим себестоимость в число, т.к. там иногда бывают строки
    catalog['Себестоимость'] = catalog['Себестоимость'].apply(lambda x: pd.to_numeric(x, errors='coerce'))
    # Переименовываем колонку с наименованием товара
    catalog = catalog.rename(columns={
        'Название': 'Наименование товара'
    })

    return catalog


# Функция создания списка артикулов из расходов, которые уже разбиты по артикулам
def create_all_sku_df(result_sku_costs, result_supplies_costs, catalog):
    # Получаем df с расходами, которые уже разбиты по SKU и Артикулу
    df_sku_costs = result_sku_costs['df_sku_costs_all'].copy()
    df_supplies_costs = result_supplies_costs['supplies_costs_matched'].copy()
    catalog_ = catalog.copy()
    # Из справочной таблицы убираем товары, у которых нет SKU или Артикула
    catalog_ = catalog_.dropna(subset=['SKU', 'Артикул продавца'])
    # Формируем лист с датафреймами
    df_list = [df_sku_costs, df_supplies_costs, catalog_]
    # Выбираем нужные колонки
    sku_columns = ['SKU', 'Артикул продавца', 'Наименование товара']
    # Используем reindex для добавления отсутствующих колонок
    df_list = [df.reindex(columns=sku_columns) for df in df_list]
    # Объединяем в один df и убираем дубликаты
    df_sku_list_all = (
        pd.concat(df_list)
        .drop_duplicates(subset=['SKU', 'Артикул продавца'])
        .dropna()
    )

    return df_sku_list_all


# Функция объединения затрат, которые удалось распределить sku
def merge_sku_costs(result_sku_costs, result_companies_costs, result_supplies_costs):
    # Получаем df посчитанных затрат
    df_sales_returns = result_sku_costs['df_sku_costs_all']
    df_companies_stats = result_companies_costs['df_companies_api_stats']
    df_supplies_costs = result_supplies_costs['supplies_costs_matched']

    # Объединяем всё в один df
    df_sku_costs_all = pd.concat([df_sales_returns, df_supplies_costs, df_companies_stats])
    # Считаем сумму после union
    df_sku_costs_all_stats = (
        df_sku_costs_all
        .groupby(['SKU', 'Артикул продавца', 'Наименование товара'])
        .sum()
        .reset_index()
    )

    return df_sku_costs_all_stats


# Функция расчета расходов, которые распределяются в зависимости от количества продаж sku
def calc_other_costs(df_finance_report, df_sku_costs_all_stats, result_sku_costs, result_companies_stats, result_supplies_costs):

    # Переменные, в которые будем сохранять итоговый результат
    df_result_all_costs = pd.DataFrame()
    result_operations_other_costs = pd.DataFrame()

    # Копируем df с расходами по SKU для избежания изменений в оригинальном df
    df_sku_costs_all_stats_ = df_sku_costs_all_stats.copy()
    # Получаем id строк, по которым уже посчитали затраты, разбив их на sku
    sales_returns_ids = result_sku_costs['sku_costs_ids']
    companies_stats_ids = result_companies_stats['companies_costs_ids']
    supplies_costs_ids = result_supplies_costs['supplies_costs_ids_matched']
    calculated_costs_ids = sales_returns_ids + companies_stats_ids + supplies_costs_ids

    # Получаем из фин. отчета оставшиеся строки
    df_other_costs = df_finance_report.loc[~df_finance_report['id'].isin(calculated_costs_ids), :]

    if df_other_costs.shape[0] > 0:
        # Считаем суммарные расходы по каждому из типов начисления в прочих расходах
        df_other_costs_stats = df_other_costs.groupby('Тип начисления').agg(
            Итого_прочие_расходы=('Итого руб', 'sum')
        ).reset_index()

        # Для каждого из типов начислений разбиваем расходы на каждый SKU
        # Получаем список типов начислений
        operation_list_other = df_other_costs_stats['Тип начисления'].unique().tolist()
        # Всего продаж по всем SKU
        total_sales = df_sku_costs_all_stats_['Продажи шт'].sum()
        for operation_type in operation_list_other:
            # Расходы по данному типу операции
            operation_costs = df_other_costs_stats.loc[df_other_costs_stats['Тип начисления'] == operation_type, 'Итого_прочие_расходы'].values[0]
            df_sku_costs_all_stats_[operation_type] = df_sku_costs_all_stats_['Продажи шт'] / total_sales * operation_costs
            # Добавляем 0.0, чтобы избавиться от -0.0
            df_sku_costs_all_stats_[operation_type] = df_sku_costs_all_stats_[operation_type] + 0.0

        # Помещаем результаты расчетов в переменные из начала функции
        df_result_all_costs = df_sku_costs_all_stats_
        result_operations_other_costs = df_other_costs_stats

    # Формируем словарь с результатами
    result_all_costs = {'df_result_all_costs': df_result_all_costs,
                        'result_operations_other_costs': result_operations_other_costs}

    return result_all_costs


# Функция добавления данных из справочной таблицы
def add_data_from_catalog(result_all_costs):
    # Считывание справочной таблицы
    catalog_reference = pd.read_excel(f"{marketplace_dir_name}/Clients/{client_name}/catalog/Справочная_таблица_{client_name}_Ozon.xlsx")
    # Из справочной таблицы убираем товары, у которых нет SKU
    catalog_reference_ = catalog_reference.loc[~catalog_reference['SKU'].isna(), :]
    # В справочной таблице переводим себестоимость в число, т.к. там иногда бывают строки
    catalog_reference_['Себестоимость'] = catalog_reference_['Себестоимость'].apply(lambda x: pd.to_numeric(x, errors='coerce'))
    # Переименовываем колонку с наименованием товара
    catalog_reference_ = catalog_reference_.rename(columns={'Название': 'Наименование товара'})
    # Переводим SKU в float64, чтобы нормально проходило объединение
    catalog_reference_['SKU'] = catalog_reference_['SKU'].astype('float64')
    # Достаем df с результатами расчетов
    df_all_costs = result_all_costs['df_result_all_costs'].copy()
    # Колонки, которые берем из справочной таблицы
    catalog_columns = catalog_finace_svod_columns
    # Объединяем справочную таблицу с результатами расчетов
    df_all_costs_with_catalog = df_all_costs.merge(
        catalog_reference_[catalog_columns],
        on='SKU',
        how='left'
    )
    # # Перемещаем справочные колонки в начало df
    # df_all_costs_with_catalog = df_all_costs_with_catalog[ catalog_columns + [col for col in df_all_costs_with_catalog.columns if col not in catalog_columns]]
    # # Перемещаем колонку с себестоимостью в конец df
    # df_all_costs_with_catalog.insert(len(df_all_costs_with_catalog.columns) - 1, 'Себестоимость', df_all_costs_with_catalog.pop('Себестоимость'))
    columns_to_move=[col for col in catalog_columns if col!='SKU']
    df_all_costs_with_catalog = move_columns(
        df=df_all_costs_with_catalog,
        columns_to_move=columns_to_move,
        position='Наименование товара',
        insert_type='after'
    )

    return df_all_costs_with_catalog


# Функция расчета колонки итоговых расходов
def calc_final_costs(df_all_costs_with_catalog):
    # Достаем df с результатами расчетов
    df_final_costs = df_all_costs_with_catalog.copy()

    # Определяем колонки, которые будут исключены из затрат
    columns_to_exclude = ([
        'SKU', 'Наименование товара', 'Артикул продавца',
        'Продажи шт', 'Возвраты шт', 'Продажи руб',
        'Продажи - Возвраты (шт)', 'Продажи - Возвраты (руб)'
        ]
        + catalog_finace_svod_columns
    )
    # Считаем себестоимость
    df_final_costs['Себестоимость * Факт продаж'] = (-1) * abs(df_final_costs['Себестоимость'] * df_final_costs['Продажи - Возвраты (шт)'])
    # Выбираем итоговые колонки для расчета маржинальности
    final_costs_columns = df_final_costs.columns[~df_final_costs.columns.isin(columns_to_exclude)]

    # Считаем сумму по всем затратам
    df_final_costs['Все затраты (руб)'] = df_final_costs[final_costs_columns].sum(axis=1)

    # Вычитаем из продаж в рублях все затраты
    df_final_costs['Маржинальность руб'] = df_final_costs['Продажи руб'] + df_final_costs['Все затраты (руб)']
    # Там, где нет себестоимости, маржинальность не считаем
    df_final_costs.loc[df_final_costs['Себестоимость'].isna(), 'Маржинальность руб'] = np.nan
    df_final_costs['Маржинальность %'] = df_final_costs['Маржинальность руб'] / df_final_costs['Продажи руб']
    # Меняем inf на nan
    # df_final_costs['Маржинальность %'] = df_final_costs['Маржинальность %'].replace([-np.inf, np.inf], np.nan)

    return df_final_costs


# Функция расчета сводной таблицы по всем расходам
def create_svod_by_operations(df_final_costs):
    # Список колонок, по котором мы не считаем суммы
    columns_to_exclude = [
        'SKU',
        'Артикул продавца',
        'Основной артикул',
        'Размер',
        'Цвет',
        'Сезон',
        'Статус',
        'Наименование товара',
        'Себестоимость',
        'Категория',
        'РРЦ',
        # 'Себестоимость * Факт продаж',
        # 'Все затраты (руб)',
        # 'Маржинальность руб',
        'Маржинальность %',
    ]

    # Считаем суммы по остальным столбцам
    df_final_costs_by_operation = (
        df_final_costs
        # Убираем колонки, которые не должны быть включены в сумму
        .loc[
            :,
            ~df_final_costs.columns.isin(columns_to_exclude)
        ]
        # Считаем суммы по остальным
        .sum()
        # Достаем столбцы из index
        .reset_index()
        # Переименовываем колонки
        .rename(columns={
            'index': 'Тип операции',
            0: 'Итого руб'
        })
    )
    # Считаем сумму, от которой будем считать процент
    sum_columns = ['Продажи - Возвраты (руб)', 'Баллы за скидки',]
    # Считается сумма по тем колонкам, которые есть в df_final_costs
    sum_for_percent = (
        df_final_costs
        .loc[:, df_final_costs.columns.isin(sum_columns)]
        .sum().sum()
    )
    # Считаем процент
    df_final_costs_by_operation['Итого, %'] = (
        df_final_costs_by_operation['Итого руб'] / sum_for_percent * 100
    )
    # Делаем процент пропуском в некоторых расходах
    no_percent_operations = [
        'Продажи шт',
        'Возвраты шт',
    ]
    # Маска для loc
    mask_no_operations = df_final_costs_by_operation['Тип операции'].isin(no_percent_operations)
    # Создание пропусков
    df_final_costs_by_operation.loc[mask_no_operations, 'Итого, %'] = np.nan

    return df_final_costs_by_operation

# Функция сохранения расчетов в excel
def save_finance_svod_to_excel(
        date_report,
        df_finance_report,
        df_final_costs_by_operation,
        result_sku_costs,
        df_final_costs,
        result_all_costs,
        result_companies_stats,
        result_supplies_costs,
        date_start,
        date_end
    ):
    # Получаем нужные df из расчетов
    df_all_costs = df_final_costs.copy()
    df_sku_costs_by_operation = result_sku_costs['df_sku_costs_by_operation']
    result_operations_other_costs = result_all_costs['result_operations_other_costs']
    df_companies_api_report_stats = result_companies_stats['df_companies_api_report_stats']
    df_supplies_costs_from_report = result_supplies_costs['supplies_costs_from_report']

    # Формируем даты отчета
    df_report_dates = pd.DataFrame({'Дата начала периода': date_start, 'Дата окончания периода': date_end}, index=[0])

    # Путь для сохранения файла (с именем)
    filepath_finance_svod = (
        f"{finance_reports_dir}/{date_report}_Свод_из_фин_отчета_Озон_{client_name}.xlsx"
    )
    # Сохраняем отчет
    with pd.ExcelWriter(filepath_finance_svod) as w:
        # Даты формирования исходного отчета
        df_report_dates.to_excel(w, sheet_name='Даты формирования отчета', index=False)
        # Исходный отчет
        df_finance_report.to_excel(w, sheet_name='Исходный отчет', index=False)
        # Объединенные расходы с распределением по sku
        df_all_costs.to_excel(w, sheet_name='Итоговая таблица', index=False)
        # Суммы по всем начислениям
        df_final_costs_by_operation.to_excel(w, sheet_name='Сводная по начислениям', index=False)
        # Расходы, которые уже были распределены по SKU в исходном отчете
        df_sku_costs_by_operation.to_excel(w, sheet_name='Расходы по SKU', index=False)
        # Расходы по типам начисления, которые распределялись в зависимости от продаж SKU
        result_operations_other_costs.to_excel(w, sheet_name='Прочие расходы', index=False)
        # Расходы на поставки из отчета ЛК
        df_supplies_costs_from_report.to_excel(w, sheet_name='Расходы на поставки из ЛК', index=False)
        # Сопоставление расходов по рекламе из АПИ и из отчета
        df_companies_api_report_stats.to_excel(w, sheet_name='Расходы по рекламе', index=False)


# Функция форматирования excel
def format_excel(date_report, df_final_costs):
    logger.info(f"Formatting FinanceReport for client {client_name}")
    # Путь файла отчета с именем
    filepath_finance_svod = (
        f"{finance_reports_dir}/{date_report}_Свод_из_фин_отчета_Озон_{client_name}.xlsx"
    )
    # Открываем файл с расчетами Excel
    wb = openpyxl.load_workbook(filepath_finance_svod)

    # Цикл по всем листам
    for worksheet in wb.sheetnames:
        ws = wb[worksheet]
        # Автоподбор ширины столбца
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Определяем индексы столбцов, которым НЕ нужно форматировать числа
        exclude_format_idx = [
            idx + 1  # Openpyxl колонки считаются с 1
            for idx, header in enumerate(headers)
            if header and ('Артикул продавца' in str(header) or 'SKU' in str(header))
        ]

        # Формат чисел (пробелы как разделители)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for col_idx, cell in enumerate(row, start=1):
                # Не форматируем числа в исключённых столбцах
                if col_idx in exclude_format_idx:
                    continue
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '# ##0'

        # Значок фильтра на столбцы
        ws.auto_filter.ref = ws.dimensions

    # Отдельное форматирование для листа итоговых расходов
    final_costs_worksheet = wb['Итоговая таблица']

    # df с соответствием заголовков и названий столбцов
    excel_columns = pd.DataFrame({"column": df_final_costs.columns,
                                  "column_number": np.arange(1, len(df_final_costs.columns) + 1)})
    excel_columns['excel_column'] = excel_columns['column_number'].apply(lambda x: get_column_letter(x))
    # Кол-во строк в df
    svod_len = df_final_costs.shape[0]
    # Номер строки, откуда начинается запись
    row_start = 2

    # Мин. и макс. колонка
    min_col = excel_columns.loc[excel_columns['column_number'].idxmin(), 'excel_column']
    max_col = excel_columns.loc[excel_columns['column_number'].idxmax(), 'excel_column']
    # Заголовки
    header_cells = final_costs_worksheet[f"{min_col}{row_start - 1}:{max_col}{row_start - 1}"]
    # Все колонки, кроме заголовков
    all_cells = final_costs_worksheet[f"{min_col}{row_start}:{max_col}{svod_len + 1}"]

    # Формат заголовков
    thin_border = Side(border_style="thin", color="000000")
    for row in header_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11, bold=True)
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center',
                                       wrap_text=True)
            cell.border = Border(top = thin_border, bottom = thin_border,
                                 right = thin_border, left = thin_border)
    # Границы (сетка)
    for row in all_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11)
            cell.border = Border(top = thin_border, bottom = thin_border,
                                 right = thin_border, left = thin_border)


    # Колонка с маржинальностью
    marginality_percent = excel_columns.loc[excel_columns['column'].str.contains('Маржинальность %'), 'excel_column'].values[0]

    # Формат числа 0% у маржинальности
    percent_0_digit_cells = [marginality_percent]
    for i in range(len(percent_0_digit_cells)):
        percent_0_digit_cells[i] = f"{percent_0_digit_cells[i]}{row_start}:{percent_0_digit_cells[i]}{svod_len + 1}"

    for cell_range in percent_0_digit_cells:
        cell_range = final_costs_worksheet[cell_range]
        for row in cell_range:
            for cell in row:
                cell.number_format = '0%'

    wb.save(filepath_finance_svod)


# %% Вызов всех функций
if __name__ == "__main__":
    # Дата, с которой начинается имя папки с исходными данными
    date_report = '2025_07'
    # Даты, за которые был сформирован отчет (содержатся в имени файла с отчетом)
    date_start = '2025-07-01'
    date_end = '2025-07-31'
    # Дата, когда были выгружены данные по рекламе
    date_upload_performance = str(date.today())
    # Директория для сохранения результатов
    finance_reports_dir = f"{marketplace_dir_name}/Clients/{client_name}/FinanceReports/{date_report}"
    if not os.path.exists(finance_reports_dir):
        os.makedirs(finance_reports_dir)

    # Считывание файла с фин. отчетом
    df_finance_report = read_finance_report_file(finance_reports_dir, date_start, date_end)
    # Выгрузка отчета о реализации продаж, возвратов и баллов Озон
    # df_realization_report = calc_sales_returns(date_start, date_end)
    # Расчет расходов с привязкой к SKU
    result_sku_costs = calc_sku_costs(df_finance_report)
    # Расчет расходов по поставкам с разбивкой на SKU
    result_supplies_costs = calc_supplies_costs(df_finance_report)
    # Считываем справочную таблицу
    catalog = read_catalog(client_name)
    # Формируем список SKU с артикулами из расходов, которые уже разбиты по SKU
    df_sku_list_all = create_all_sku_df(
        result_sku_costs,
        result_supplies_costs,
        catalog
    )
    # Расчет расходов по рекламным кампаниям с привязкой к SKU
    # task_id = 'bb511716-e864-443d-94c7-66b326c042eb'
    result_companies_costs = calc_companies_costs(
        date_start,
        date_end,
        date_report,
        df_finance_report,
        df_sku_list_all,
        client_id_performance,
        client_secret_performance,
        # task_id=task_id
    )
    # Объединение расходов по SKU
    df_sku_costs_all_stats = merge_sku_costs(
        result_sku_costs,
        result_companies_costs,
        result_supplies_costs
    )
    # Расчет расходов по типам начислений, которые не удалось привязать к SKU
    result_all_costs = calc_other_costs(
        df_finance_report,
        df_sku_costs_all_stats,
        result_sku_costs,
        result_companies_costs,
        result_supplies_costs
    )

    # Добавляем данные из справочной таблицы
    df_all_costs_with_catalog = add_data_from_catalog(result_all_costs)
    # Расчет колонки с итоговыми расходами
    df_final_costs = calc_final_costs(df_all_costs_with_catalog)
    # Расчет сводной по всем операциям
    df_final_costs_by_operation = create_svod_by_operations(df_final_costs)
    # Сохранение в excel
    save_finance_svod_to_excel(
        date_report,
        df_finance_report,
        df_final_costs_by_operation,
        result_sku_costs,
        df_final_costs,
        result_all_costs,
        result_companies_costs,
        result_supplies_costs,
        date_start,
        date_end
    )
    # Функция форматирования Excel (на данный момент только ширина столбцов)
    format_excel(date_report, df_final_costs)
    print(f"\033[32m\033[47mDone calculating Finance Report for Client {client_name}")

# %%
