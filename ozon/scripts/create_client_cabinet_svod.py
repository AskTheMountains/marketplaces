
# %% Определение всех функций
import requests
import json
import time
from datetime import date,datetime,timedelta
import pandas as pd
import shutil
import os
import glob
from pathlib import Path
import csv
import zipfile
from zipfile import ZipFile
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from openpyxl.formatting.rule import ColorScaleRule
import re
from loguru import logger
import getopt
import sys
pd.set_option('future.no_silent_downcasting', True)


# Папка, где лежит текущий скрипт
BASE_DIR = Path(__file__).parent.parent


# Файл с настройками и номером клиента
# from options import settings, client_number, headers
# Некоторые константы
from ozon.scripts.constants import (
    headers,
    client_name,
    marketplace_dir_name,
    client_id_performance,
    client_secret_performance,
    ozon_performance_api_url,
    promotion_companies
)
# Функции выгрузки данных по API Seller
from ozon.scripts.uploadDataFromOzon import(
    get_ozon_product,
    getOrders,
    getStockReminders_fbo_v2,
    getStockReminders_fbs
)
# Функции выгрузки данных по API Seller (доп.)
from ozon.scripts.ozon_generic_functions import (
    get_orders,
    get_reminders_by_product
)
# Функции выгрузки данных по API Performane
from ozon.scripts.uploadDataFromOzonPerformance import (
    getAuthorizationToken,
    getCompanyList,
    getCompanyStatistics,
)
# Функция формирования статистики РК
from ozon.scripts.create_campaigns_report import(
    create_campaigns_report as upload_campaigns_report,
)
# Некоторые вспомогательные функции
from generic_functions import move_columns

# Функция создания директории, куда будет помещен итоговый отчет
def create_client_cabinet_svod_dir(client_name, date_report=str(date.today())):
    client_cabinet_path = (
        f"{BASE_DIR}/"
        f"Clients/{client_name}/ClientCabinetSvod/"
    )
    if not os.path.exists(client_cabinet_path):
        os.makedirs(client_cabinet_path, exist_ok=True)
        logger.info(f"Creating Companies Upload directory:{client_cabinet_path}")

    return client_cabinet_path


# Функция создания директории, куда будет помещаться статистика кампаний
def create_upload_statistic_dir(client_name, date_report=str(date.today())):
    # Задаем путь к директории
    # upload_path_statistic = (
    #     f"{BASE_DIR}/"
    #     f"Clients/{client_name}/ClientSvod/"
    #     f"{date_report}/{date_report}_Кампании_API"
    # )
    upload_path_statistic = (
        f"{BASE_DIR}/"
        f"Clients/{client_name}/CampaignsReport/"
        f"{date_report}/{date_report}_Кампании_API"
    )
    if not os.path.exists(upload_path_statistic):
        os.makedirs(upload_path_statistic, exist_ok=True)
        logger.info(f"Creating Companies Upload directory:{upload_path_statistic}")

    return upload_path_statistic


# Функция создания началной и конечной даты
# GPT START ----
def generate_dates(
        reference_date: str = None,
        start_date: str = None,
        end_date: str = None,
    ) -> pd.DataFrame:
    # Проверка на несовместимость параметров
    if reference_date and (start_date or end_date):
        raise ValueError("Нельзя одновременно задавать reference_date и start_date/end_date.")
    # Если задан диапазон — используем его
    if start_date and end_date:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    elif start_date or end_date:
        raise ValueError("Нужно задать одновременно start_date и end_date.")
    else:
        # Старая логика по reference_date
        if reference_date is None:
            today = datetime.now()
        else:
            today = datetime.strptime(reference_date, "%Y-%m-%d")

        # если день недели < 4 (до пятницы)
        if today.weekday() < 4:
            # понедельник предыдущей недели
            start_dt = today - timedelta(days=today.weekday() + 7)
            # воскресенье предыдущей недели
            end_dt = start_dt + timedelta(days=6)
        else:
            # понедельник текущей недели
            start_dt = today - timedelta(days=today.weekday())
            # вчерашний день
            end_dt = today - timedelta(days=1)

    start_date_str = start_dt.strftime('%Y-%m-%d')
    end_date_str = end_dt.strftime('%Y-%m-%d')
    start_date_iso = start_dt.strftime('%Y-%m-%dT00:00:00Z')
    end_date_iso = end_dt.strftime('%Y-%m-%dT23:59:59Z')

    df_dates = pd.DataFrame([{
        'date_start': start_date_str,
        'date_end': end_date_str,
        'datetime_start': start_date_iso,
        'datetime_end': end_date_iso
    }])

    return df_dates
# GPT END ----

# Функция получения начальной и конечной даты в формате строки
def get_start_end_date(df_dates, type_dates='companies'):
    # Если задан тип дат - даты для заказов, то берем даты со временем
    if type_dates == 'orders':
        date_start, date_end = df_dates.loc[0, ['datetime_start', 'datetime_end']]
    # В остальных случаях берем даты без времени
    else:
        date_start, date_end = df_dates.loc[0, ['date_start', 'date_end']]

    return date_start, date_end


# Функция создания диапазона дат с частотой одного дня
def generate_date_range(df_dates):
    # Получаем начальную и конечную дату
    date_start_range, date_end_range = get_start_end_date(df_dates)
    # Генерируем диапазон дат от начальной до конечной даты с интервалом в один день
    df_date_range = pd.DataFrame({'datetime': pd.date_range(date_start_range, date_end_range, freq='d')})
    # Переводим в формат даты, чтобы убрать время
    df_date_range['date'] = df_date_range['datetime'].dt.date

    return df_date_range


# Функция обработка списка товаров АПИ
def process_product_list(df_products):
    # Создаем копию для избежания изменений в оригинальном df
    df_products_processed = df_products.copy()
    # Убираем ненужные колонки
    # df_products_processed = df_products_processed.loc[:, ~df_products_processed.columns.isin([
    #     'FBS OZON SKU ID', 'Контент-рейтинг', 'Рейтинг', 'Причины скрытия', 'Бренд', 'Размер', 'Цвет',
    #     'Статус товара', 'Видимость FBO', 'Причины скрытия FBO (при наличии)',
    #     'Видимость FBS', 'Причины скрытия FBS (при наличии)', 'Дата создания',
    #     'Категория комиссии', 'Объем товара, л', 'Объемный вес, кг',
    #     'Доступно к продаже по схеме FBO, шт.',
    #     'Вывезти и нанести КИЗ (кроме Твери), шт', 'Зарезервировано, шт',
    #     'Доступно к продаже по схеме FBS, шт.',
    #     'Доступно к продаже по схеме realFBS, шт.',
    #     'Зарезервировано на моих складах, шт',
    #     'Актуальная ссылка на рыночную цену', 'Размер НДС, %'
    #    ])]
    df_products_processed = df_products_processed.loc[:, df_products_processed.columns.isin([
        'Артикул',
        'Название товара',
        'SKU',
        # 'Отзывы',
        # 'Текущая цена с учетом скидки, ₽',
    ])]
    # Удаляем лишние символы из артикула
    df_products_processed['Артикул'] = df_products_processed['Артикул'].str.replace("'", "", regex=False)
    # Переводим баркод в число
    # df_products_processed['Barcode'] = pd.to_numeric(df_products_processed['Barcode'], errors='coerce')
    # df_products_processed['Barcode'] = df_products_processed['Barcode'].apply(lambda x: format(x, 'f') if pd.notnull(x) else x)
    # df_products_processed['Barcode'] = df_products_processed['Barcode'].apply(lambda x: str(x).split('.')[0] if pd.notnull(x) else x).astype('Int64')
    # Переименовываем колонки с ценами, которые приходят из отчета по товарам из апи
    # чтобы потом не было путаницы с ценами, которые мы получаем из отдельного метода
    df_products_processed = df_products_processed.rename(columns={
        'Текущая цена с учетом скидки, ₽': 'Текущая цена с учетом скидки (из отчета по товарам)',
        # 'Цена до скидки (перечеркнутая цена), ₽': 'Цена до скидки (перечеркнутая цена) (из отчета по товарам)',
        # 'Рыночная цена, ₽': 'Рыночная цена (из отчета по товарам)',
        # 'Цена Premium, ₽': 'Цена Premium (из отчета по товарам)',
    })
    # Переименовываем колонки с нужными ценами
    df_products_processed = df_products_processed.rename(columns={
        # 'Цена до скидки (перечеркнутая цена) (из отчета по товарам)': 'Цена до скидки',
        'Текущая цена с учетом скидки (из отчета по товарам)': 'Текущая цена с учетом скидки',
        # 'Минимальная цена после применения всех скидок': 'Мин. цена Ozon',
        # 'Цена до учета скидок (зачеркнутая)': 'Цена до скидки',
        # 'Цена с учетом скидок (на карточке товара)': 'Цена после скидки',
        # 'Минимальная цена после применения всех скидок': 'Мин. цена Ozon',
        # 'Цена с учетом акций продавца': 'Цена по акции',
        # 'Цена с учетом всех акций': 'Цена с баллами Озон',
    })
    # Переводим SKU в строку чтобы не было ошибок при merge
    df_products_processed['SKU'] = df_products_processed['SKU'].astype(str)
    # Переводим Артикул в строку
    df_products_processed['Артикул'] = df_products_processed['Артикул'].astype(str)
    # Выбираем нужные колонки

    return df_products_processed

# Функция создания списка РК с датами
def create_input_companies(promotion_companies, df_dates):
    # Создаем копию списка кампаний из констант
    input_companies = promotion_companies
    # Получаем даты для отчета РК
    date_start_performance, date_end_performance = get_start_end_date(df_dates, type_dates='companies')
    # Добавляем даты к списку РК
    for company in input_companies:
        company.extend([date_start_performance, date_end_performance])

    return input_companies


# Функция выгрузки отчета по РК
def create_campaigns_report(
        df_dates,
        upload_path_statistic,
        client_name,
        headers,
        client_id_performance,
        client_secret_performance,
        to_save=True,
        delete_files=False,
    ):

    # Формируем список РК для отдельных клиентов
    input_companies = create_input_companies(
        promotion_companies,
        df_dates
    )
    # Формируем отчет РК
    df_campaigns_report = upload_campaigns_report(
        client_name,
        client_id_performance,
        client_secret_performance,
        headers,
        input_companies,
        upload_path_statistic,
        to_save=to_save,
        delete_files=delete_files
    )
    # Убираем сшитые товары
    df_campaigns_report = (
        df_campaigns_report
        .loc[df_campaigns_report['Тип товара'].isin(['Товар из РК']), :]
        .reset_index(drop=True)
    )
    # Переименовываем некоторые колонки
    df_campaigns_report = df_campaigns_report.rename(columns={
        'Расход с НДС, руб': 'РК',
        'ДРР, %': 'ДРР % по размеру'
    })

    return df_campaigns_report


# Функция выгрузки заказов
def get_orders_for_svod(df_dates, headers):
    # Получаем даты для отчета по заказам
    date_start_orders, date_end_orders = get_start_end_date(df_dates, type_dates='orders')
    # Выгружаем заказы
    df_orders_all = get_orders(
        headers,
        date_start_orders,
        date_end_orders
    )

    return df_orders_all

# Функция расчета заказов
def calc_orders(df_orders):
    # Убираем отмененные отправления
    df_orders_filtered = (
        df_orders
        .loc[~df_orders['Статус'].isin(['Отменён', 'Отменен']), :]
        .copy()
    )
    # Считаем заказы по артикулам
    df_orders_by_products = (
        df_orders_filtered
        .groupby(['Артикул'], as_index=False)
        .agg(**{
            'Сумма заказов': ('Заказы руб', 'sum'),
            'Кол-во заказов': ('Заказы шт', 'sum')
        })
    )

    return df_orders_by_products

# Функция добавления данных к списку товаров
def add_data_to_product_list(
        df_products_processed,
        df_campaigns_report,
        df_reminders,
        df_orders_by_products,
    ):
    # Создаем список датафреймов, которые объединяем
    merge_params = [
        {"df": df_reminders, "columns": ["Остаток",], "on": "Артикул"},
        {"df": df_campaigns_report, "columns": ["РК",], "on": "Артикул"},
        {"df": df_orders_by_products, "columns": ["Сумма заказов", "Кол-во заказов"], "on": "Артикул"}
    ]

    # Начинаем с основного датафрейма
    df_products_svod = df_products_processed.copy()

    for params in merge_params:
        # Ключ, по которому мерджим
        key = params["on"]
        # Берем указанные колонки из текущего df + колонку, по которой мерджим
        df_to_merge = params["df"][params["columns"] + [key]].copy()
        df_products_svod = df_products_svod.merge(df_to_merge, how='left', on=key, suffixes=('', f'_{key}_right'))

    return df_products_svod

# Функция расчета доп колонок
def calc_additional_columns(df_products_svod, date_report):
    # Создаем копию для избежания изменений в оригинальном df
    df_client_svod = df_products_svod.copy()
    # Заполняем пропуски
    columns_to_fillna = [
        'Остаток',
        'Сумма заказов',
        'Кол-во заказов'
    ]
    df_client_svod[columns_to_fillna] = (
        df_client_svod[columns_to_fillna]
        .fillna(0)
        .infer_objects(copy=False)
    )
    # Переводим некоторые колонки в integer
    columns_to_int = ['Остаток', 'Кол-во заказов']
    df_client_svod[columns_to_int] = (
        df_client_svod[columns_to_int]
        .astype(int)
    )
    # Средняя цена продажи
    df_client_svod['Средняя цена продажи'] = np.where(
        df_client_svod['Кол-во заказов'] > 0,
        df_client_svod['Сумма заказов'] / df_client_svod['Кол-во заказов'],
        np.nan
    )
    # Заменяем 0 на пустые ячейки в некоторых колонках
    zero_to_nan_columns = ['Средняя цена продажи', 'Кол-во заказов']
    for col in zero_to_nan_columns:
        # Для float точнее использовать np.isclose вместо ==, а для int можно просто == 0
        if np.issubdtype(df_client_svod[col].dtype, np.floating):
            mask = np.isclose(df_client_svod[col], 0)
        else:
            mask = df_client_svod[col] == 0
        # Меняем inf и -inf на nan
        df_client_svod[col] = df_client_svod[col].replace([np.inf, -np.inf], np.nan)
        # Заменяем 0 на nan
        df_client_svod.loc[mask, col] = np.nan

    # Дата добавления артикула
    df_client_svod['Дата добавления товара'] = str(date.today())
    df_client_svod['Дата добавления товара'] = (
        pd.to_datetime(df_client_svod['Дата добавления товара'])
        .dt.date
    )
    # ДРР %
    # df_client_svod['ДРР, %'] = df_client_svod['РК'] / df_client_svod['Сумма заказов'] * 100
    # df_client_svod['ДРР, %'] = df_client_svod['ДРР, %'].replace([-np.inf, np.inf], np.nan)

    return df_client_svod


# Функция создания df для записи в excel
def create_client_svod_excel(df_client_cabinet_svod, date_report, df_dates):
    # Создаем копию для избежания изменений в оригинальном df
    df_current_report = df_client_cabinet_svod.copy()
    # Делаем порядок колонок
    column_order = [
        'Артикул', 'SKU', 'Название товара',
        'Дата добавления товара',
        'Остаток',
        # 'Тип кампании',
        'РК',
        # 'Сумма заказов',
        'Кол-во заказов',
        'Средняя цена продажи',
        # 'ДРР, %',
        # 'Отзывы',
        # 'Текущая цена с учетом скидки (из отчета по товарам)',
        # 'ДРР % по размеру',

    ]
    df_current_report = df_current_report[column_order]

    # Создаем префикс колонок
    date_start_col = (
        pd.to_datetime(df_dates['date_start'].iloc[0])
        .strftime('%d.%m')
    )
    date_end_col = (
        pd.to_datetime(df_dates['date_end'].iloc[0])
        .strftime('%d.%m')
    )
    column_date_prefix = f"{date_start_col}-{date_end_col}"
    # Добавляем дату в колонку заказов
    date_report_formatted = pd.to_datetime(date_report).strftime('%d.%m')
    df_current_report = df_current_report.rename(columns={
        'Остаток': f'Остаток Озон {date_report_formatted}',
        'РК': f'РК {column_date_prefix}',
        'Кол-во заказов': f'Кол-во заказов {column_date_prefix}',
        'Средняя цена продажи': f'Ср. цена {column_date_prefix}',
    })

    # Сортировка по артикулу
    df_current_report = df_current_report.sort_values(by=['Артикул'], ignore_index=True)

    return df_current_report


# Функция чтения списка файлов в директории с отчетом
def read_report_files(
        client_name,
        client_cabinet_svod_dir,
        # date_current_report=str(date.today())
    ):
    # Маска поиска файлов
    cabinet_reports_file_mask = f'*Сводная_по_кабинету_{client_name}_Ozon.xlsx'
    search_pattern = os.path.join(client_cabinet_svod_dir, cabinet_reports_file_mask)

    # Фильтрация временных файлов excel
    files = [
        file_path
        for file_path in glob.glob(search_pattern)
        if not os.path.basename(file_path).startswith('~$')
    ]

    # Сбор информации о файлах
    data = []
    for file_path in files:
        file_name = os.path.basename(file_path)
        modified_time = os.path.getmtime(file_path)
        modified_datetime = datetime.fromtimestamp(modified_time)
        data.append({
            'Имя файла': file_name,
            'Дата изменения': modified_datetime.strftime('%Y-%m-%d %H:%M:%S'),
            'Полный путь': file_path
        })

    # Создание DataFrame
    columns = ['Имя файла', 'Дата изменения', 'Полный путь']
    df_cabinet_report_files = pd.DataFrame(data, columns=columns)

    # Сортировка по дате изменения
    df_cabinet_report_files = df_cabinet_report_files.sort_values(
        by=['Дата изменения'],
        ascending=False
    )

    return df_cabinet_report_files

# Функция чтения последнего отчета
def read_last_report_file(last_report_filepath):
    # Считываем файл с последним отчетом
    df_previous_report = pd.read_excel(last_report_filepath, sheet_name=0)
    # Переводим все имена колонок с датой в формат %d.%m.%Y
    new_columns = []
    for col in df_previous_report.columns:
        col_str = str(col)
        dt = pd.to_datetime(col_str, errors='coerce', format='mixed')
        if not pd.isnull(dt):
            new_columns.append(dt.strftime('%d.%m.%Y'))  # формат: день.месяц.две последние цифры года
        else:
            new_columns.append(col)
    df_previous_report.columns = new_columns
    # Переводим колонки с характеристиками товара в строку
    sku_columns = ['Артикул', 'Название товара', 'SKU']
    for col in sku_columns:
        df_previous_report[col] = df_previous_report[col].astype(str)

    return df_previous_report


# Функция добавления новых товаров в предыдущий отчет
def add_new_products(
        df_cabinet_report_files,
        df_current_report,
    ):

    # Создаем копию для избежания изменений в оригинальном df
    df_current_report_ = df_current_report.copy()
    # Берем файл с датой последнего изменения
    last_report_filepath = df_cabinet_report_files['Полный путь'].iloc[0]
    # Считываем файл с последним отчетом
    df_last_cabinet_report = read_last_report_file(last_report_filepath)
    # Разделяем объединение на два этапа:
    #  товары, которые отсутствовали в предыдущем отчете, мерджим через concat
    #  остальные товары через join

    # Создаем df с новыми артикулами
    df_new_products = (
        df_current_report_
        # .loc[:, ['Артикул', 'SKU', 'Название товара', 'Дата добавления товара']]
        .merge(
            df_last_cabinet_report[['Артикул']],
            on='Артикул',
            how='left',
            indicator='Наличие артикула'
        )
        .pipe(lambda df:
                df.loc[df['Наличие артикула'] == 'left_only', :]
        )
        .drop(columns=['Наличие артикула'])
    )
    if not df_new_products.empty:
        logger.info(f"Found {df_new_products.shape[0]} new products")

    # Формируем список с новыми артикулами
    new_products_list = df_new_products['Артикул'].to_list()

    # Создаем df с теми артикулами, которые уже были в предыдущих отчетах
    df_current_report_previous_products = (
        df_current_report_
        .loc[
            ~df_current_report['Артикул'].isin(new_products_list),
            ~df_current_report.columns.isin([
                'SKU',
                'Название товара',
                'Дата добавления товара'
            ])
        ]
        .copy()
    )

    # Старые товары объединяем через join
    df_current_and_previous_report = (
        df_last_cabinet_report
        .merge(
            df_current_report_previous_products,
            on='Артикул',
            how='left'
        )
    )
    # Новые товары объединяем через concat
    df_current_and_previous_report = pd.concat([
        df_current_and_previous_report,
        df_new_products
    ])

    # Сбрасываем index и делаем сортировку
    df_current_and_previous_report = (
        df_current_and_previous_report
        .reset_index(drop=True)
        .sort_values(by=['Артикул'], ignore_index=True)
    )
    # Заполняем пропуски в выбранных колонках
    columns_to_fill = (
        df_current_and_previous_report
        .columns[
            df_current_and_previous_report
            .columns.str.contains(
                # 'Остаток|Кол-во заказов|Ср. цена',
                'Остаток',
                regex=True
            )
        ]
        .to_list()
    )
    df_current_and_previous_report[columns_to_fill] = df_current_and_previous_report[columns_to_fill].fillna(0)

    # Переводим колонку с датой добавления в datetime для удобства
    df_current_and_previous_report['Дата добавления товара'] = (
        pd.to_datetime(df_current_and_previous_report['Дата добавления товара'])
        .dt.date
    )

    return df_current_and_previous_report


# Функция объединения текущего отчета с предыдущим
def add_data_from_previous_report(
        client_name,
        client_cabinet_svod_dir,
        df_current_report,
):
    # Считываем список файлов с отчетами по текущему клиенту
    df_cabinet_report_files = read_report_files(client_name, client_cabinet_svod_dir)
    # Если список файлов пустой, то считаем,
    # что отчет создается в первый раз и возвращаем копию исходного df
    if df_cabinet_report_files.empty:
        logger.info('Previous report has not been found, generating first report')
        df_current_and_previous_report = df_current_report.copy()
        return df_current_and_previous_report
    # Если список файлов не пустой, начинаем процесс объединения
    else:
        logger.info(f"Checking new products")
        df_current_and_previous_report = add_new_products(
            df_cabinet_report_files,
            df_current_report
        )

    return df_current_and_previous_report

# Функция сохранения и форматирования отчета excel
def save_and_format_excel(
        df_current_and_previous_report,
        client_cabinet_svod_dir,
        date_report,
        df_dates
    ):
    # Формируем директорию для сохранения
    path_report = (
        f"{client_cabinet_svod_dir}/"
    )
    name_report = (
        f"{date_report}_Сводная_по_кабинету_{client_name}_Ozon.xlsx"
    )
    file_name_report = f"{path_report}/{name_report}"
    if not os.path.exists(path_report):
        os.makedirs(path_report, exist_ok=True)
    # Формируем даты для названия листа
    # date_start_excel = (
    #     pd.to_datetime(df_dates['date_start'])
    #     .iloc[0]
    #     .strftime('%d.%m')
    # )
    # date_end_excel = (
    #     pd.to_datetime(df_dates['date_end'])
    #     .iloc[0]
    #     .strftime('%d.%m')
    # )
    # Задаем колонки с процентным форматом
    # percent_cols = ['ДРР, %', 'ДРР % по размеру']
    # Делим на 100, чтобы отображался процент
    # df_client_excel_svod[percent_cols] = df_client_excel_svod[percent_cols] / 100

    # Формируем название листа
    sheet_name_client_cabinet_svod = f"{client_name}"
    # Сохраняем книгу
    with pd.ExcelWriter(file_name_report) as w:
        df_current_and_previous_report.to_excel(w, index=False, sheet_name=sheet_name_client_cabinet_svod)

    # ---- Форматируем файл с отчетом ----

    # Открываем книгу excel
    wb = load_workbook(file_name_report)

    # Общее форматирование для всех листов
    for work_sheet in wb.sheetnames:
        # Список колонок
        svod_columns = list(df_current_and_previous_report.columns)
        # Выбираем нужный лист
        ws = wb[work_sheet]

        # Определяем стиль границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Создаем границы
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

        # Форматирование числовых ячеек
        exclude_number_format_substrings = ['SKU']
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Определяем номер колонки с заголовком
                col_idx = cell.column - 1  # openpyxl: колонка начинается с 1
                col_name = svod_columns[col_idx]

                # Пропускаем форматирование некоторых столбцов
                if any(substr in str(col_name) for substr in exclude_number_format_substrings):
                    continue

                if isinstance(cell.value, (int, float)):
                    if float(cell.value).is_integer():
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'

        # --- Перенос текста во всех заголовках ---
        for idx in range(1, ws.max_column + 1):
            ws.cell(row=1, column=idx).alignment = Alignment(
                wrap_text=True,
                #horizontal=ws.cell(row=1, column=idx).alignment.horizontal or 'center',
                horizontal='center',
                vertical='center'
            )

        # --- Автоподбор ширины столбцов + фильтр на столбцы ---
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for i, cell in enumerate(col):
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Если это заголовок (первая строка), учитываем значок фильтра
            head_len = len(str(col[0].value)) if col[0].value else 0
            filter_icon_length = 3 if head_len == max_len else 2  # 3 для filter icon
            ws.column_dimensions[col_letter].width = max_len + filter_icon_length

        # --- Значок фильтра на столбцы ---
        ws.auto_filter.ref = ws.dimensions

        # --- Форматирование по подстрокам в заголовках ---
        columns_fix_center = ['Остаток Озон', 'РК', 'Кол-во заказов']
        columns_fix_left = ['Артикул', 'SKU', 'Дата добавления товара', 'Ср. цена']

        for idx, col_name in enumerate(svod_columns):
            col_letter = get_column_letter(idx + 1)

            # Проверяем на подстроки с выравниванием по центру
            if any(sub in str(col_name) for sub in columns_fix_center):
                ws.column_dimensions[col_letter].width = 18  # фикс ширина
                align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for row in range(2, ws.max_row + 1):  # пропустить заголовки
                    ws.cell(row=row, column=idx+1).alignment = align
                ws.cell(row=1, column=idx+1).alignment = align  # заголовок центрируем

            # Проверяем на подстроки слева
            if any(sub in str(col_name) for sub in columns_fix_left):
                align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=idx+1).alignment = align
                ws.cell(row=1, column=idx+1).alignment = align  # заголовок тоже по левому краю

        # ---  Форматирование процентных колонок ---
        percent_substrings = ['%', 'ДРР']  # ищем колонки по наличию этих символов

        # Получаем номера подходящих колонок по наличию подстроки
        percent_col_indices = [
            idx + 1  # Excel columns start from 1
            for idx, name in enumerate(svod_columns)
            if any(substr in name for substr in percent_substrings)
        ]

        # Применяем форматирование к найденным колонкам
        for col_idx in percent_col_indices:
            for row in range(2, ws.max_row + 1):  # пропустить заголовки
                cell = ws.cell(row=row, column=col_idx)
                if cell.value:
                    cell.value = cell.value * 0.01  # для процентов
                cell.number_format = '0.00%'  # формат: проценты с 2 знаками

        # ---  Условное форматирование градиентом для выбранных колонок ---
        # gradient_cols = ['ДРР, %', 'ДРР % по размеру']
        # for col_name in gradient_cols:
        #     col_idx = list(df_current_and_previous_report.columns).index(col_name) + 1
        #     col_letter = ws.cell(row=1, column=col_idx).column_letter
        #     # Диапазон от второй строки до конца в данном столбце (без заголовков)
        #     cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"

        #     # Устанавливаем цвет: min (зеленый), max (оранжевый), середина (по желанию, желтый)
        #     rule = ColorScaleRule(
        #         start_type='min', start_color='63BE7B',   # Зеленый
        #         mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Желтый
        #         end_type='max', end_color='E06967',       # Красный
        #     )
        #     ws.conditional_formatting.add(cell_range, rule)

    # Сохранение обновленного файла
    wb.save(file_name_report)




# %% Вызов всех функций
if __name__ == '__main__':
    # Дата, за который считался (или считается) отчет
    date_report=str(date.today())
    # Директория, куда будет сохраняться итоговый отчет
    client_cabinet_svod_dir = create_client_cabinet_svod_dir(client_name, date_report)
    # Директория, куда будет загружена статистика кампаний
    upload_path_statistic = create_upload_statistic_dir(client_name, date_report)
    # Генерируем даты начала и окончания периода выгрузки
    df_dates = generate_dates()
    # Выбираем начальную и конечную дату
    date_start_print = df_dates['date_start'].iloc[0]
    date_end_print = df_dates['date_end'].iloc[0]
    logger.info(
        f"\nCreating client cabinet svod for client {client_name}\n"
        f"dates: {date_start_print} - {date_end_print}"
    )
    # Генерируем диапазон дат с частотой одного дня
    df_date_range = generate_date_range(df_dates)

    # Получаем список товаров АПИ
    df_products = get_ozon_product(headers, to_save=False)
    # Обрабатываем список товаров
    df_products_processed = process_product_list(df_products)
    # Получаем отчет РК
    df_campaigns_report = create_campaigns_report(
        df_dates,
        upload_path_statistic,
        client_name,
        headers,
        client_id_performance,
        client_secret_performance,
        to_save=True,
        delete_files=False
    )
    # Получаем заказы
    df_orders = get_orders_for_svod(df_dates, headers)
    # Получаем остатки
    df_reminders = get_reminders_by_product(df_products, headers)

    # Считаем заказы по товарам
    df_orders_by_products = calc_orders(df_orders)

    # Добавляем данные к списку товаров
    df_products_svod = add_data_to_product_list(
        df_products_processed,
        df_campaigns_report,
        df_reminders,
        df_orders_by_products,
    )
    # Расчитываем доп. колонки
    df_client_cabinet_svod = calc_additional_columns(
        df_products_svod,
        date_report,
    )
    # Создаем df для записи в excel
    df_current_report = create_client_svod_excel(
        df_client_cabinet_svod,
        date_report,
        df_dates
    )
    # Объединяем с предыдущим отчетом
    df_current_and_previous_report = add_data_from_previous_report(
        client_name,
        client_cabinet_svod_dir,
        df_current_report,
    )
    # Сохраняем отчет в excel
    save_and_format_excel(
        df_current_and_previous_report,
        client_cabinet_svod_dir,
        date_report,
        df_dates
    )

# %%
