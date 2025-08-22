
# %% Определение всех функций
import pandas as pd
import numpy as np
import requests
import time
import os
import openpyxl
import json
from loguru import logger
from datetime import date,datetime,timedelta
from itertools import product
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_00
import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
# pd.set_option('future.no_silent_downcasting', True)

# Файл с некоторыми константами
from ozon.scripts.constants import (
    headers,
    client_name,
    marketplace_dir_name,
    ozon_seller_api_url,
)

# Функции выгрузки данных по АПИ, которые уже готовы
from ozon.scripts.uploadDataFromOzon import getOrders
# Доп. функции
from generic_functions import move_columns

# Создание нужных директорий
def create_dirs():
    # Папка с клиентом
    client_dir = f"{marketplace_dir_name}/Clients/{client_name}/"
    # Список директорий для создания
    dir_names = ['SalesOrdersSvod']
    for dir_name in dir_names:
        dir_path = os.path.join(client_dir, dir_name)
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)


# Функция создания диапазона дат
# GPT START ----
def generate_date_range(
        reference_date: str = None,
        start_date: str = None,
        end_date: str = None
    ):
    # Проверка на несовместимость параметров:
    # либо используем reference_date, либо вручную заданный диапазон
    if reference_date and (start_date or end_date):
        raise ValueError("Нельзя одновременно задавать reference_date и start_date/end_date.")

    # Если оба manual-диапазона заданы
    if start_date and end_date:
        # Парсим строки дат в объекты datetime
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    elif start_date or end_date:
        # Один из диапазонов не указан — ошибка
        raise ValueError("Нужно задать одновременно start_date и end_date.")
    else:
        # Ни диапазон, ни reference_date не заданы явно
        #  — значит работаем с reference_date (по умолчанию — сегодня)
        if reference_date is None:
            ref_dt = datetime.now()
        else:
            # Парсим reference_date из строки
            ref_dt = datetime.strptime(reference_date, "%Y-%m-%d")

        # Если дата — первое число месяца
        if ref_dt.day == 1:
            # Берём последний день прошлого месяца
            prev_month_last_day = ref_dt - timedelta(days=1)
            # Начало диапазона — первое число прошлого месяца
            start_dt = prev_month_last_day.replace(day=1)
            # Конец диапазона — последний день прошлого месяца
            end_dt = prev_month_last_day
        else:
            # Если не первое число месяца:
            # Начало диапазона — первое число текущего месяца
            start_dt = ref_dt.replace(day=1)
            # Конец диапазона — предыдущие сутки относительно reference_date
            end_dt = ref_dt - timedelta(days=1)

    # Форматы для строк и для ISO-даты по времени суток
    start_date_str = start_dt.strftime('%Y-%m-%d')
    end_date_str = end_dt.strftime('%Y-%m-%d')
    start_date_iso = start_dt.strftime('%Y-%m-%dT00:00:00Z')
    end_date_iso = end_dt.strftime('%Y-%m-%dT23:59:59Z')

    # Собираем результат в DataFrame
    df_dates = pd.DataFrame([{
        'date_start': start_date_str,       # начало диапазона (дд.мм.гггг)
        'date_end': end_date_str,           # конец диапазона (дд.мм.гггг)
        'datetime_start': start_date_iso,   # начало ISO-датой с 00:00:00
        'datetime_end': end_date_iso        # конец ISO-датой с 23:59:59
    }])

    return start_date_iso, end_date_iso
# GPT END ----


# Функция создания датафрейма с диапазоном дат
def generate_date_range_df(date_start, date_end):
    # Генерируем диапазон дат от начальной до конечной даты с интервалом в один день
    date_range_df = pd.DataFrame({'date': pd.date_range(date_start, date_end, freq='d')})
    # Переводим в формат даты, чтобы убрать время
    date_range_df['date'] = date_range_df['date'].dt.date

    return date_range_df

# Функция выгрузки отчета списка транзакций
def getTransactionReport(headers, date_start, date_end):
    # Разбиваем диапазон дат на периоды по 1 месяцу каждый

    # GPT START ----
    # Преобразование строковых дат в datetime
    dt_start = pd.to_datetime(date_start).tz_localize(None)
    dt_end = pd.to_datetime(date_end).tz_localize(None)

    # Создание списка интервалов
    intervals = []

    # Первый интервал от начальной даты до конца месяца
    first_month_end = (dt_start + pd.DateOffset(months=1)).replace(day=1) - pd.Timedelta(seconds=1)
    if first_month_end > dt_end:
        first_month_end = dt_end

    intervals.append({
        'date_start': dt_start.strftime('%Y-%m-%dT%H:%M:%S.000Z'),
        'date_end': first_month_end.strftime('%Y-%m-%dT%H:%M:%S.000Z'),
        'dt_start': dt_start,
        'dt_end': first_month_end
    })

    # Следующий интервал от начала следующего месяца до конца месяца
    current_start = first_month_end + pd.Timedelta(seconds=1)
    while current_start <= dt_end:
        monthly_start = current_start.replace(day=1)
        monthly_end = (monthly_start + pd.DateOffset(months=1)).replace(day=1) - pd.Timedelta(seconds=1)

        if monthly_end > dt_end:
            monthly_end = dt_end

        intervals.append({
            'date_start': monthly_start.strftime('%Y-%m-%dT%H:%M:%S.000Z'),
            'date_end': monthly_end.strftime('%Y-%m-%dT%H:%M:%S.000Z'),
            'dt_start': monthly_start,
            'dt_end': monthly_end
        })

        # Переход к следующему месяцу
        current_start = monthly_start + pd.DateOffset(months=1)

    # Создание датафрейма
    date_range_df = pd.DataFrame(intervals)
    # GPT END ----

    # df, в который будем помещать результаты
    df_transaction_list = pd.DataFrame()

    for i in range(date_range_df.shape[0]):
        # Начальные значения для цикла
        page = 1
        page_count = 2
        while page_count > 0:
            # Выгружаем отдельно каждую страницу отчета
            params = json.dumps({
                "filter": {
                    "date": {
                        "from": date_range_df['date_start'][i],
                        "to": date_range_df['date_end'][i]
                    },
                    "operation_type": [],
                    "posting_number": "",
                    "transaction_type": "all"
                },
                "page": page,
                "page_size": 1000
            })

            resp_data_transaction_list = requests.post(f"{ozon_seller_api_url}/v3/finance/transaction/list", headers=headers, data=params).json()
            # Сколько нужно выгрузить страниц
            page_count =  resp_data_transaction_list['result']['page_count']
            # Увеличиваем страницу 1 для выгрузки следующей страницы
            page = page + 1
            # print(resp_data_transaction_list)
            # Промежуточный df, в который помещаем результаты текущей страницы
            tmp_df = pd.DataFrame(resp_data_transaction_list['result']['operations'])
            # Добавляем даты, за который был выгружен отчет по транзакциям
            tmp_df = tmp_df.assign(
                dt_start=date_range_df['dt_start'][i],
                dt_end=date_range_df['dt_end'][i]
            )
            # Объединяем с предыдущей страницей
            df_transaction_list = pd.concat([df_transaction_list, tmp_df])

    # Убираем дубликаты из index
    df_transaction_list = df_transaction_list.reset_index(drop=True)
    # Количество товаров в одной операции
    df_transaction_list['items_amount'] = df_transaction_list['items'].apply(lambda x: len(x))
    df_transaction_list['services_amount'] = df_transaction_list['services'].apply(lambda x: len(x))

    return df_transaction_list

# Функция выгрузки отчета о заказах
def get_orders(headers, date_start, date_end):
    # Выгружаем отчет об отправлениях fbo и fbs
    df_orders_fbo = getOrders(headers, date_start, date_end, delivery_schema='fbo', to_save=False)
    df_orders_fbs = getOrders(headers, date_start, date_end, delivery_schema='fbs', to_save=False)
    # Убираем лишние колонки из отчета fbo
    df_orders_fbo = df_orders_fbo.loc[:, ~df_orders_fbo.columns.isin(['Объемный вес товаров, кг'])]
    # Объединяем два отчета в один
    df_orders = pd.concat([df_orders_fbo, df_orders_fbs])

    return df_orders

# Функция объединения заказов fbo и fbs
def union_fbo_and_fbs_orders(df_postings_fbo, df_postings_fbs):
    df_postings_all = pd.concat([df_postings_fbo, df_postings_fbs])
    return df_postings_all

# Функция расчета продаж
def calc_sales(df_transactions_list, date_range_df):
    # Достаем информацию об отправлениях из списка транзакций
    postings = pd.json_normalize(df_transactions_list['posting'])
    # Добавляем тип (заказ\возврат), стоимость и количество товаров в отправлении
    postings = pd.concat([postings, df_transactions_list[['type', 'operation_date', 'operation_type', 'operation_type_name', 'amount', 'accruals_for_sale', 'items_amount']]], axis=1)
    # Там, где тип начисления возврат, делаем отрицательное число штук товара,
    # чтобы потом просто посчитать сумму
    postings['Количество шт'] = np.where(
        postings['operation_type_name'] == 'Получение возврата, отмены, невыкупа от покупателя',
        postings['items_amount'] * (-1),
        postings['items_amount']
    )
    # Колонка, по котрой берем дату для расчета продаж
    date_column = 'operation_date'
    # Колонка, по которой считаем итоговую сумму
    total_column = 'accruals_for_sale'
    # Переводим колонку с датой в timestamp
    postings[date_column] = pd.to_datetime(postings[date_column])
    # Оставляем только дату
    postings[date_column] = postings[date_column].dt.date
    # Переименовываем колонку для мерджа
    postings = postings.rename(columns={
        date_column: 'date',
        total_column: 'Продажи руб'
        })

    # Делаем выборку по нужным типам начислений и считаем итоговую цену для каждого отправления
    sales = (
        postings
        .loc[(postings['operation_type_name'] \
              .isin(
                  ['Доставка покупателю', 'Получение возврата, отмены, невыкупа от покупателя']
                  )
                  ),
            :]
        .groupby(['date'])
        .agg(**{
            'Продажи руб': ('Продажи руб', 'sum'),
            'Продажи шт': ('Количество шт', 'sum')
            })
        .reset_index()
    )
    # Мерджим с диапазоном дат, т.к. бывает, что не все дни присутствуют в заказах
    sales = sales.merge(
        date_range_df,
        on='date',
        how='outer'
    )
    # Заполняем пропуски после мерджа
    for col in ['Продажи руб', 'Продажи шт']:
        sales[col] = sales[col].fillna(0)

    return sales

    # returns = postings_total.loc[postings_total['type'] == 'returns', :]
    # orders = postings_total.loc[postings_total['type'] == 'orders', :]
    # a = orders.merge(returns, on='posting_number', how='outer', indicator=True)


# Функция расчета заказов
def calc_orders(df_orders_all, date_range_df):
    # Создаем копию для избежания изменений в оригинальном df
    df_orders_all_ = df_orders_all.copy()
    # Колонка, по котрой берем дату для расчета продаж
    date_column = 'Принят в обработку'
    # Колонка, по которой считаем итоговую сумму
    total_column = 'Сумма отправления'
    # Переводим колонку с датой в timestamp
    df_orders_all_[date_column] = pd.to_datetime(df_orders_all_[date_column])
    # Оставляем только дату
    df_orders_all_[date_column] = df_orders_all_[date_column].dt.date
    # Переименовываем колонку для мерджа
    df_orders_all_ = df_orders_all_.rename(columns={
        date_column: 'date',
        total_column: 'Заказы руб'
        })
    # Считаем заказы в штуках и рублях
    orders = (df_orders_all_
              .groupby(['date'])
              .agg(**{
                  'Заказы руб': ('Заказы руб', 'sum'),
                  'Заказы шт': ('Количество', 'sum')
              })
              .reset_index()
    )

    # Мерджим с диапазоном дат, т.к. бывает, что не все дни присутствуют в заказах
    orders = orders.merge(
        date_range_df,
        on='date',
        how='outer'
    )
    # Заполняем пропуски после мерджа
    for col in ['Заказы руб', 'Заказы шт']:
        orders[col] = orders[col].fillna(0)

    return orders

# Объединяем заказы и продажи в одну таблицу
def union_orders_and_sales(orders, sales):
    # Объединяем в один df
    df_orders_and_sales = pd.concat([
        orders.set_index(orders['date']),
        sales.set_index(sales['date'])
    ], axis=1)
    # Удаляем ненужную колонку
    df_orders_and_sales = df_orders_and_sales.drop(columns=['date'])
    # Достаем дату из индекса
    df_orders_and_sales = df_orders_and_sales.reset_index()
    # Форматируем дату в нужный формат
    df_orders_and_sales['Дата'] = df_orders_and_sales['date'].apply(lambda x: x.strftime('%d.%m.%Y'))
    # Разворачиваем в широкий вид
    df_orders_and_sales_wide = pd.pivot_table(
        df_orders_and_sales,
        values=['Заказы руб', 'Заказы шт', 'Продажи руб', 'Продажи шт'],
        # index=['Заказы руб', 'Заказы шт', 'Продажи руб', 'Продажи шт'],
        columns=['date'],
        aggfunc='sum'
    )
    # Делаем нужный формат даты у колонок
    df_orders_and_sales_wide.columns = pd.to_datetime(df_orders_and_sales_wide.columns).strftime('%d.%m.%Y')

    return df_orders_and_sales_wide


# Функция добавления плана и факта продаж
def add_plan(df_orders_and_sales_wide,
             client_name,
             plan_orders_rub = 0,
             plan_orders_amount = 0,
             plan_sales_rub = 0,
             plan_sales_amount = 0):
    # Создаем копию для избежания изменений в оригинальном df
    df_plan = df_orders_and_sales_wide.copy()
    # Считаем итоговые суммы
    df_plan['Факт'] = df_plan.sum(axis=1)
    # Добавляем План
    df_plan['План'] = np.nan
    df_plan.loc['Заказы руб', 'План'] = plan_orders_rub
    df_plan.loc['Заказы шт', 'План'] = plan_orders_amount
    df_plan.loc['Продажи руб', 'План'] = plan_sales_rub
    df_plan.loc['Продажи шт', 'План'] = plan_sales_amount
    # Считаем отклонение плана от факта
    df_plan['Факт/План, %'] = df_plan['Факт'].divide(df_plan['План'].replace({0: np.nan})) * 100
    # Заменяем inf, если получилось деление на ноль
    df_plan['Факт/План, %'] = df_plan['Факт/План, %'].replace([-np.inf, np.inf], np.nan)
    # Добавляем имя магазина
    # df_plan['Магазин'] = client_name
    # # Добавление пустой строки в начало DataFrame
    # empty_row = pd.DataFrame([[np.nan] * df_plan.shape[1]], columns=df_plan.columns, index=['Магазин'])
    # df = pd.concat([empty_row, df_plan])
    # # Перемещаем колонки в начало df
    df_plan = move_columns(df_plan, ['План', 'Факт', 'Факт/План, %'], 0)
    # Убираем имя у индекса
    df_plan.index.name = None

    return df_plan


# Функция сохранения отчета в excel
def save_excel(df_plan, date_start, date_end, date_report=str(date.today())):
    # Создаем df с именем клиента
    shop_name = pd.DataFrame({'Магазин': [client_name]})
    # Создаем df с диапазоном формирования отчета
    df_dates = pd.DataFrame({
        'Начальная дата': [date_start],
        'Конечная дата': [date_end]
    })
    # Даты для имени файла
    date_start_file = pd.to_datetime(date_start).strftime('%d-%m-%Y')
    date_end_file = pd.to_datetime(date_end).strftime('%d-%m-%Y')

    # Имя файла для сохранения отчета
    report_file_name = (f"{marketplace_dir_name}/Clients/{client_name}/SalesOrdersSvod/"
                        f"{date_start_file}_{date_end_file}"
                        f"_Свод_продажи_Ozon_{client_name}.xlsx"
    )

    # GPT START ----
    # Параметры для записи датафреймов
    dataframes = [
        (shop_name, 2, 2, False),  # Начало в строке 1, колонке 1, не записывать индекс датафрейма
        (df_plan, 6, 2, True),
        (df_dates, 12, 2, False),
    ]

    # Запись датафреймов на один лист Excel
    with pd.ExcelWriter(report_file_name, engine='openpyxl') as writer:
        ranges = []  # Здесь будем сохранять диапазоны (начало и конец) для границ
        for df, start_row, start_col, save_index in dataframes:
            # Если сохраняется индекс, добавляется 1 столбец
            end_row = start_row + len(df)  # Вычисляем конечную строку
            end_col = start_col + len(df.columns) - 1 + (1 if save_index else 0)  # Конечный столбец с учётом индекса
            ranges.append((start_row, end_row, start_col, end_col, save_index))  # Добавляем "save_index"

            # Запись датафрейма с учётом параметра индекса
            df.to_excel(writer, index=save_index, sheet_name=client_name, startrow=start_row - 1, startcol=start_col - 1)


    # Форматирование Excel
    logger.info(f"Formatting Sales svod for client {client_name}")

    wb = load_workbook(report_file_name)
    ws = wb[client_name]

    # Определяем стиль границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Применяем границы только к диапазону данных таблиц (без индекса)
    for start_row, end_row, start_col, end_col, save_index in ranges:
        for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col + (1 if save_index else 0),  # Сдвигаем начало диапазона вправо, если индекс записан
            max_col=end_col
        ):
            for cell in row:
                cell.border = thin_border

    # Применение числового форматирования с разделением разрядов (через пробелы)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):  # Проверяем, является ли значение числом
                cell.number_format = '#,##0'  # Формат с пробелами и двумя знаками после запятой

    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Номер текущего столбца
        column_letter = get_column_letter(column)  # Буква столбца
        for cell in col:
            try:
                if cell.value:  # Проверка на наличие данных
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # Сохранение обновленного файла
    wb.save(report_file_name)
    # GPT END ----

    # Сохраняем отчет
    # with pd.ExcelWriter(report_dir, engine='openpyxl') as writer:
    #     shop_name.to_excel(writer, sheet_name=client_name, startrow=1, startcol=1, index=False)
    #     df_plan.to_excel(writer, sheet_name=client_name, startrow=5, startcol=1)
    #     df_dates.to_excel(writer, sheet_name=client_name, startrow=11, startcol=1, index=False)


# Функция форматирования Excel
# def format_excel(date_report=str(date.today())):
#     logger.info(f"Formatting Sales svod for client {client_name}")
#     # Считываем отчет
#     report_file = f"Clients/{client_name}/SalesOrdersSvod/{date_report}_Свод_продажи_Ozon_{client_name}.xlsx"
#     wb = load_workbook(report_file)
#     ws = wb[client_name]

#     # Автоподбор ширины столбцов
#     for col in ws.columns:
#         max_length = 0
#         column = col[0].column  # Номер текущего столбца
#         column_letter = get_column_letter(column)  # Буква столбца
#         for cell in col:
#             try:
#                 if cell.value:  # Проверка на наличие данных
#                     max_length = max(max_length, len(str(cell.value)))
#             except:
#                 pass
#         ws.column_dimensions[column_letter].width = max_length + 2

#     # Применение числового форматирования с разделением разрядов (через пробелы)
#     for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#         for cell in row:
#             if isinstance(cell.value, (int, float)):  # Проверяем, является ли значение числом
#                 cell.number_format = '#,##0'  # Формат с пробелами и двумя знаками после запятой

#     # Добавление границ для всех ячеек с данными
#     thin_border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )

#     for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#         for cell in row:
#             if cell.value is not None:  # Проверка на содержание данных
#                 cell.border = thin_border

#     # Сохранение обновленного файла
#     wb.save(report_file)


# %% Вызов всех функций
if __name__ == '__main__':
    # Формируем диапазон дат
    # date_start = '2025-07-01T00:00:00.000Z'
    # date_end = '2025-07-10T23:59:59.000Z'
    date_start, date_end = generate_date_range(
        # start_date='2025-07-22',
        # end_date='2025-08-05'
    )
    # Планы продаж и заказов в штуках и рублях
    plan_orders_rub = 0
    plan_orders_amount = 0
    plan_sales_rub = 0
    plan_sales_amount = 0

    # Генерируем диапазон дат от начальной до конечной даты с интервалом в один день
    date_range_df = generate_date_range_df(date_start, date_end)

    logger.info(
        f"Calculating Sales svod for client {client_name} "
        f"for dates {date_start} - {date_end}"
    )
    # Выгружаем заказы FBO и FBS
    df_orders_all = get_orders(headers, date_start, date_end)
    # Выгружаем список транзакций
    df_transactions_list = getTransactionReport(headers, date_start, date_end)

    # Считаем заказы в штуках и рублях
    orders = calc_orders(df_orders_all, date_range_df)
    # Считаем продажи в штуках и рублях
    sales = calc_sales(df_transactions_list, date_range_df)
    # Объединяем продажи и заказы в один df
    df_orders_and_sales_wide = union_orders_and_sales(orders, sales)
    # Добавляем планы продаж и считаем отклонение плана от факта
    df_plan = add_plan(
        df_orders_and_sales_wide,
        client_name,
        plan_orders_rub,
        plan_orders_amount,
        plan_sales_rub,
        plan_sales_amount
    )
    # Создаем нужную директорию
    create_dirs()
    # Сохраняем Excel
    save_excel(df_plan, date_start, date_end)
    # Форматируем Excel
    # format_excel()
    logger.info(f"Finished calculating Sales svod for client {client_name}")
