
# %% Определение всех функций
import requests
import json
import time
from datetime import date,datetime,timedelta
import pandas as pd
import shutil
import os
import glob
from pathlib import Path
import csv
import zipfile
from zipfile import ZipFile
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from openpyxl.formatting.rule import ColorScaleRule
import re
from loguru import logger
import getopt
import sys
pd.set_option('future.no_silent_downcasting', True)

# Папка, где лежит текущий скрипт
BASE_DIR = Path(__file__).parent.parent


# Файл с настройками и номером клиента
# from options import settings, client_number, headers
# Некоторые константы
from ozon.scripts.constants import (
    headers,
    client_name,
    marketplace_dir_name,
    client_id_performance,
    client_secret_performance,
    ozon_performance_api_url,
    promotion_companies
)
# Функции выгрузки данных по API Seller
from ozon.scripts.uploadDataFromOzon import(
    get_ozon_product,
    getOrders,
    getStockReminders_fbo_v2,
    getStockReminders_fbs
)
# Функции выгрузки данных по API Seller (доп.)
from ozon.scripts.ozon_generic_functions import (
    get_orders,
    get_reminders_by_product
)
# Функции выгрузки данных по API Performane
from ozon.scripts.uploadDataFromOzonPerformance import (
    getAuthorizationToken,
    getCompanyList,
    getCompanyStatistics,
)
# Функция формирования статистики РК
from ozon.scripts.create_campaigns_report import(
    create_campaigns_report as upload_campaigns_report,
)
# Некоторые вспомогательные функции
from generic_functions import move_columns

# Функция создания директории, куда будет помещен итоговый отчет
def create_client_rk_svod_dir(client_name, date_report=str(date.today())):
    # Задаем путь к директории
    # upload_path_statistic = (
    #     f"{BASE_DIR}/"
    #     f"Clients/{client_name}/ClientSvod/"
    #     f"{date_report}/{date_report}_Кампании_API"
    # )
    client_rk_path = (
        f"{BASE_DIR}/"
        f"Clients/{client_name}/ClientRKSvod/"
    )
    if not os.path.exists(client_rk_path):
        os.makedirs(client_rk_path, exist_ok=True)
        logger.info(f"Creating Companies Upload directory:{client_rk_path}")

    return client_rk_path

# Функция создания директории, куда будет помещаться статистика кампаний
def create_upload_statistic_dir(client_name, date_report=str(date.today())):
    # Задаем путь к директории
    # upload_path_statistic = (
    #     f"{BASE_DIR}/"
    #     f"Clients/{client_name}/ClientSvod/"
    #     f"{date_report}/{date_report}_Кампании_API"
    # )
    upload_path_statistic = (
        f"{BASE_DIR}/"
        f"Clients/{client_name}/CampaignsReport/"
        f"{date_report}/{date_report}_Кампании_API"
    )
    if not os.path.exists(upload_path_statistic):
        os.makedirs(upload_path_statistic, exist_ok=True)
        logger.info(f"Creating Companies Upload directory:{upload_path_statistic}")

    return upload_path_statistic


# Функция создания диапазона дат
# GPT START ----
def generate_dates(reference_date: str = None, start_date: str = None, end_date: str = None) -> pd.DataFrame:
    # Проверка на несовместимость параметров
    if reference_date and (start_date or end_date):
        raise ValueError("Нельзя одновременно задавать reference_date и start_date/end_date.")

    # Если указан диапазон — используем его
    if start_date and end_date:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    elif start_date or end_date:
        raise ValueError("Нужно задать одновременно start_date и end_date.")
    else:
        # Старая логика по reference_date
        if reference_date is None:
            today = datetime.now()
        else:
            today = datetime.strptime(reference_date, "%Y-%m-%d")

        # Логика в зависимости от дня недели
        if today.weekday() == 0:
            # Если сегодня понедельник (weekday() == 0)
            # Понедельник предыдущей недели:
            start_dt = today - timedelta(days=7)
            start_dt = start_dt - timedelta(days=start_dt.weekday())
            # Воскресенье предыдущей недели:
            end_dt = start_dt + timedelta(days=6)
        else:
            # Если не понедельник:
            # Понедельник текущей недели:
            start_dt = today - timedelta(days=today.weekday())
            # Вчерашний день:
            end_dt = today - timedelta(days=1)

    start_date_str = start_dt.strftime('%Y-%m-%d')
    end_date_str = end_dt.strftime('%Y-%m-%d')
    start_date_iso = start_dt.strftime('%Y-%m-%dT00:00:00Z')
    end_date_iso = end_dt.strftime('%Y-%m-%dT23:59:59Z')

    df_dates = pd.DataFrame([{
        'date_start': start_date_str,
        'date_end': end_date_str,
        'datetime_start': start_date_iso,
        'datetime_end': end_date_iso
    }])

    return df_dates
# GPT END ----

# Функция получения начальной и конечной даты в формате строки
def get_start_end_date(df_dates, type_dates='companies'):
    # Если задан тип дат - даты для заказов, то берем даты со временм
    if type_dates == 'orders':
        date_start, date_end = df_dates.loc[0, ['datetime_start', 'datetime_end']]
    # В остальных случаях берем даты без времени
    else:
        date_start, date_end = df_dates.loc[0, ['date_start', 'date_end']]

    return date_start, date_end

# Функция создания списка РК с датами
def create_input_companies(promotion_companies, df_dates):
    # Создаем копию списка кампаний из констант
    input_companies = promotion_companies
    # Получаем даты для отчета РК
    date_start_performance, date_end_performance = get_start_end_date(df_dates, type_dates='companies')
    # Добавляем даты к списку РК
    for company in input_companies:
        company.extend([date_start_performance, date_end_performance])

    return input_companies


# Функция проверки наличия отчета по РК
def check_rk_report_exist(client_name, date_report=str(date.today())):
    # Директория, где хранятся отчеты по РК
    rk_report_dir = (
        f"{BASE_DIR}/"
        f"Clients/{client_name}/CampaignsReport/"
        f"{date_report}/"
    )
    # Имя файла с отчетом РК
    rk_file_name = (
        f"{date_report}_Отчет_РК_{client_name}_Ozon.xlsx"
    )
    # Полный путь к файлу
    filepath_rk_file = Path(f"{rk_report_dir}/{rk_file_name}")

    # Проверяем наличие файла в директории
    rk_report_exists = filepath_rk_file.exists()
    # Если файл существует, то возвращаем путь к файлу
    if rk_report_exists:
        return filepath_rk_file
    # Если файл не существует, то возвращаем False
    else:
        return False


# Функция выгрузки отчета по РК
def create_campaigns_report(
        df_dates,
        upload_path_statistic,
        client_name,
        headers,
        client_id_performance,
        client_secret_performance,
        to_save=True,
        delete_files=False,

    ):

    # Формируем список РК для отдельных клиентов
    input_companies = create_input_companies(
        promotion_companies,
        df_dates
    )
    # Формируем отчет РК
    df_campaigns_report = upload_campaigns_report(
        client_name,
        client_id_performance,
        client_secret_performance,
        headers,
        input_companies,
        upload_path_statistic,
        to_save=to_save,
        delete_files=delete_files
    )
    # Убираем сшитые товары
    df_campaigns_report = (
        df_campaigns_report
        .loc[df_campaigns_report['Тип товара'].isin(['Товар из РК']), :]
        .reset_index(drop=True)
    )
    # Переименовываем некоторые колонки
    df_campaigns_report = df_campaigns_report.rename(columns={
        'Расход с НДС, руб': 'РК',
        'ДРР, %': 'ДРР % по размеру'
    })

    return df_campaigns_report


# Функция обработка списка товаров АПИ
def process_product_list(df_products):
    # Создаем копию для избежания изменений в оригинальном df
    df_products_processed = df_products.copy()
    # Убираем ненужные колонки
    # df_products_processed = df_products_processed.loc[:, ~df_products_processed.columns.isin([
    #     'FBS OZON SKU ID', 'Контент-рейтинг', 'Рейтинг', 'Причины скрытия', 'Бренд', 'Размер', 'Цвет',
    #     'Статус товара', 'Видимость FBO', 'Причины скрытия FBO (при наличии)',
    #     'Видимость FBS', 'Причины скрытия FBS (при наличии)', 'Дата создания',
    #     'Категория комиссии', 'Объем товара, л', 'Объемный вес, кг',
    #     'Доступно к продаже по схеме FBO, шт.',
    #     'Вывезти и нанести КИЗ (кроме Твери), шт', 'Зарезервировано, шт',
    #     'Доступно к продаже по схеме FBS, шт.',
    #     'Доступно к продаже по схеме realFBS, шт.',
    #     'Зарезервировано на моих складах, шт',
    #     'Актуальная ссылка на рыночную цену', 'Размер НДС, %'
    #    ])]
    df_products_processed = df_products_processed.loc[:, df_products_processed.columns.isin([
        'Артикул',
        'Название товара',
        'SKU',
        # 'Отзывы',
        # 'Текущая цена с учетом скидки, ₽',
    ])]
    # Удаляем лишние символы из артикула
    df_products_processed['Артикул'] = df_products_processed['Артикул'].str.replace("'", "", regex=False)
    # Переводим баркод в число
    # df_products_processed['Barcode'] = pd.to_numeric(df_products_processed['Barcode'], errors='coerce')
    # df_products_processed['Barcode'] = df_products_processed['Barcode'].apply(lambda x: format(x, 'f') if pd.notnull(x) else x)
    # df_products_processed['Barcode'] = df_products_processed['Barcode'].apply(lambda x: str(x).split('.')[0] if pd.notnull(x) else x).astype('Int64')
    # Переименовываем колонки с ценами, которые приходят из отчета по товарам из апи
    # чтобы потом не было путаницы с ценами, которые мы получаем из отдельного метода
    df_products_processed = df_products_processed.rename(columns={
        'Текущая цена с учетом скидки, ₽': 'Текущая цена с учетом скидки (из отчета по товарам)',
        # 'Цена до скидки (перечеркнутая цена), ₽': 'Цена до скидки (перечеркнутая цена) (из отчета по товарам)',
        # 'Рыночная цена, ₽': 'Рыночная цена (из отчета по товарам)',
        # 'Цена Premium, ₽': 'Цена Premium (из отчета по товарам)',
    })
    # Переименовываем колонки с нужными ценами
    df_products_processed = df_products_processed.rename(columns={
        # 'Цена до скидки (перечеркнутая цена) (из отчета по товарам)': 'Цена до скидки',
        'Текущая цена с учетом скидки (из отчета по товарам)': 'Текущая цена с учетом скидки',
        # 'Минимальная цена после применения всех скидок': 'Мин. цена Ozon',
        # 'Цена до учета скидок (зачеркнутая)': 'Цена до скидки',
        # 'Цена с учетом скидок (на карточке товара)': 'Цена после скидки',
        # 'Минимальная цена после применения всех скидок': 'Мин. цена Ozon',
        # 'Цена с учетом акций продавца': 'Цена по акции',
        # 'Цена с учетом всех акций': 'Цена с баллами Озон',
    })
    # Переводим SKU в строку чтобы не было ошибок при merge
    df_products_processed['SKU'] = df_products_processed['SKU'].astype(str)
    # Переводим Артикул в строку
    df_products_processed['Артикул'] = df_products_processed['Артикул'].astype(str)
    # Выбираем нужные колонки

    return df_products_processed


# Функция выгрузки заказов
def get_orders_for_svod(df_dates, headers):
    # Получаем даты для отчета по заказам
    date_start_orders, date_end_orders = get_start_end_date(df_dates, type_dates='orders')
    # Выгружаем заказы
    df_orders_all = get_orders(
        headers,
        date_start_orders,
        date_end_orders
    )

    return df_orders_all


# Функция расчета заказов по артикулам
def calc_orders(df_orders):
    # Убираем отмененные отправления
    df_orders_filtered = (
        df_orders
        .loc[~df_orders['Статус'].isin(['Отменён', 'Отменен']), :]
        .copy()
    )
    # Считаем заказы по артикулам
    df_orders_by_products = (
        df_orders_filtered
        .groupby(['Артикул'], as_index=False)
        .agg(**{
            'Сумма заказов': ('Заказы руб', 'sum'),
            'Кол-во заказов': ('Заказы шт', 'sum')
        })
    )

    return df_orders_by_products

# Функция добавления данных к списку товаров
def add_data_to_product_list(
        df_products_processed,
        df_campaigns_report,
        df_reminders,
        df_orders_by_products,
    ):
    # Создаем список датафреймов, которые объединяем
    merge_params = [
        {"df": df_reminders, "columns": ["Остаток",], "on": "Артикул"},
        {"df": df_campaigns_report, "columns": ["РК", "ДРР % по размеру"], "on": "Артикул"},
        {"df": df_orders_by_products, "columns": ["Сумма заказов", "Кол-во заказов"], "on": "Артикул"}
    ]

    # Начинаем с основного датафрейма
    df_products_svod = df_products_processed.copy()

    for params in merge_params:
        # Ключ, по которому мерджим
        key = params["on"]
        # Берем указанные колонки из текущего df + колонку, по которой мерджим
        df_to_merge = params["df"][params["columns"] + [key]].copy()
        df_products_svod = df_products_svod.merge(df_to_merge, how='left', on=key, suffixes=('', f'_{key}_right'))

    return df_products_svod

# Функция расчета доп колонок
def calc_additional_columns(df_products_svod):
    # Создаем копию для избежания изменений в оригинальном df
    df_client_svod = df_products_svod.copy()
    # Заполняем пропуски
    columns_to_fillna = ['Остаток', 'Сумма заказов', 'Кол-во заказов']
    df_client_svod[columns_to_fillna] = (
        df_client_svod[columns_to_fillna]
        .fillna(0)
        .infer_objects(copy=False)
    )
    # Переводим некоторые колонки в integer
    columns_to_int = ['Остаток', 'Кол-во заказов']
    df_client_svod[columns_to_int] = (
        df_client_svod[columns_to_int]
        .astype(int)
    )
    # Средняя цена продажи
    df_client_svod['Средняя цена продажи'] = np.where(
        df_client_svod['Кол-во заказов'] > 0,
        df_client_svod['Сумма заказов'] / df_client_svod['Кол-во заказов'],
        0
    )
    # ДРР %
    df_client_svod['ДРР, %'] = df_client_svod['РК'] / df_client_svod['Сумма заказов'] * 100
    df_client_svod['ДРР, %'] = df_client_svod['ДРР, %'].replace([-np.inf, np.inf], np.nan)

    return df_client_svod

# Функция создания df для записи в excel
def create_client_svod_excel(df_client_svod, date_report):
    # Создаем копию для избежания изменений в оригинальном df
    df_client_excel_svod = df_client_svod.copy()
    # Добавляем дату в колонку заказов
    date_report_formatted = pd.to_datetime(date_report).strftime('%d.%m')
    df_client_excel_svod = df_client_excel_svod.rename(columns={
        'Остаток': f'Остаток {date_report_formatted}'
    })
    # Делаем порядок колонок
    column_order = [
        'Артикул', 'SKU', 'Название товара',
        f'Остаток {date_report_formatted}',
        # 'Тип кампании',
        'РК',
        'Сумма заказов', 'Кол-во заказов',
        'Средняя цена продажи',
        'ДРР, %',
        # 'Отзывы',
        # 'Текущая цена с учетом скидки (из отчета по товарам)',
        'ДРР % по размеру',

    ]
    df_client_excel_svod = df_client_excel_svod[column_order]

    # Сортировка по артикулу
    df_client_excel_svod = df_client_excel_svod.sort_values(by=['Артикул'], ignore_index=True)

    return df_client_excel_svod

# Функция сохранения и форматирования excel
def save_and_format_excel(
        df_client_excel_svod,
        df_campaigns_report,
        client_rk_svod_dir,
        date_report,
        df_dates
    ):
    # Формируем директорию для сохранения
    path_report = (
        f"{client_rk_svod_dir}/{date_report}/"
    )
    name_report = (
        f"{date_report}_Сводная_РК_{client_name}_Ozon.xlsx"
    )
    file_name_report = f"{path_report}/{name_report}"
    if not os.path.exists(path_report):
        os.makedirs(path_report, exist_ok=True)
    # Формируем даты для названия листа
    date_start_excel = (
        pd.to_datetime(df_dates['date_start'])
        .iloc[0]
        .strftime('%d.%m')
    )
    date_end_excel = (
        pd.to_datetime(df_dates['date_end'])
        .iloc[0]
        .strftime('%d.%m')
    )
    # Задаем колонки с процентным форматом
    percent_cols = ['ДРР, %', 'ДРР % по размеру']
    # Делим на 100, чтобы отображался процент
    # df_client_excel_svod[percent_cols] = df_client_excel_svod[percent_cols] / 100

    # Формируем название листа
    sheet_name_client_svod = f"{client_name} {date_start_excel}-{date_end_excel}"
    # Сохраняем книгу
    with pd.ExcelWriter(file_name_report) as w:
        df_client_excel_svod.to_excel(w, index=False, sheet_name=sheet_name_client_svod)
        df_campaigns_report.to_excel(w, index=False, sheet_name=f'Отчет РК {client_name}')

    # ---- Форматируем файл с отчетом ----

    # Открываем книгу excel
    wb = load_workbook(file_name_report)

    # Общее форматирование для всех листов
    for work_sheet in wb.sheetnames:
        # Выбираем нужный лист
        ws = wb[work_sheet]

        # Определяем стиль границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Создаем границы
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

        # Форматирование числовых ячеек
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    if float(cell.value).is_integer():
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'

        # Автоподбор ширины столбцов + фильтр на столбцы
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for i, cell in enumerate(col):
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Если это заголовок (первая строка), учитываем значок фильтра
            head_len = len(str(col[0].value)) if col[0].value else 0
            filter_icon_length = 3 if head_len == max_len else 2  # 3 для filter icon
            ws.column_dimensions[col_letter].width = max_len + filter_icon_length

        # Значок фильтра на столбцы
        ws.auto_filter.ref = ws.dimensions


    # Форматирование для листа со сводной таблицей по клиенту
    ws = wb[sheet_name_client_svod]
    # Применение числового форматирования с разделением разрядов (через пробелы)
    # Получаем типы данных из исходного датафрейма
    columns = list(df_client_excel_svod.columns)
    dtypes = df_client_excel_svod.dtypes

    # for idx, col in enumerate(columns, start=1):
    #     col_dtype = dtypes[col]
    #     for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
    #         for c in cell:
    #             val = c.value
    #             if val is None:
    #                 continue
    #             if pd.api.types.is_integer_dtype(col_dtype):
    #                 c.number_format = '# ##0'
    #             elif pd.api.types.is_float_dtype(col_dtype):
    #                 # Проверяем, целое ли число в float
    #                 if isinstance(val, float) and val.is_integer():
    #                     c.number_format = '# ##0'
    #                 else:
    #                     c.number_format = '# ##0.00'

    # ---  Форматирование процентных колонок ---
    for col_name in percent_cols:
        col_idx = list(df_client_excel_svod.columns).index(col_name) + 1  # Excel columns start from 1
        for row in range(2, ws.max_row + 1):  # пропустить заголовки
            cell = ws.cell(row=row, column=col_idx)
            if cell.value:
                cell.value = cell.value * 0.01 # Делим на 100 для корректного отображения %
            cell.number_format = '0.00%'  # формат: проценты с 2 знаками

    # ---  Условное форматирование градиентом для выбранных колонок ---
    gradient_cols = ['ДРР, %', 'ДРР % по размеру']
    for col_name in gradient_cols:
        col_idx = list(df_client_excel_svod.columns).index(col_name) + 1
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        # Диапазон от второй строки до конца в данном столбце (без заголовков)
        cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"

        # Устанавливаем цвет: min (зеленый), max (оранжевый), середина (по желанию, желтый)
        rule = ColorScaleRule(
            start_type='min', start_color='63BE7B',   # Зеленый
            mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Желтый
            end_type='max', end_color='E06967',       # Красный
        )
        ws.conditional_formatting.add(cell_range, rule)


    # Сохранение обновленного файла
    wb.save(file_name_report)



# %% Вызов всех функций
if __name__ == '__main__':
    # Дата, за который считался (или считается) отчет
    date_report=str(date.today())
    # Директория, куда будет сохраняться итоговый отчет
    client_rk_svod_dir = create_client_rk_svod_dir(client_name, date_report)
    # Директория, куда будет загружена статистика кампаний
    upload_path_statistic = create_upload_statistic_dir(client_name, date_report)
    # Генерируем диапазон дат
    df_dates= generate_dates()
    # Выбираем начальную и конечную дату
    date_start_print = df_dates['date_start'].iloc[0]
    date_end_print = df_dates['date_end'].iloc[0]
    logger.info(
        f"\nCreating client rk svod for client {client_name}\n"
        f"dates: {date_start_print} - {date_end_print}"
    )

    # Получаем список товаров АПИ
    df_products = get_ozon_product(headers, to_save=False)
    # Обрабатываем список товаров
    df_products_processed = process_product_list(df_products)

    # Получаем отчет РК
    df_campaigns_report = create_campaigns_report(
        df_dates,
        upload_path_statistic,
        client_name,
        headers,
        client_id_performance,
        client_secret_performance,
        to_save=True,
        delete_files=False
    )
    # Получаем заказы
    df_orders = get_orders_for_svod(df_dates, headers)
    # Получаем остатки
    df_reminders = get_reminders_by_product(df_products, headers)
    # Считаем заказы по артикулу
    df_orders_by_products = calc_orders(df_orders)
    # Добавляем данные к списку товаров
    df_products_svod = add_data_to_product_list(
        df_products_processed,
        df_campaigns_report,
        df_reminders,
        df_orders_by_products,
    )
    # Расчитываем доп. колонки
    df_client_svod = calc_additional_columns(df_products_svod)
    # Формируем таблицу для записи в excel
    df_client_excel_svod = create_client_svod_excel(
        df_client_svod,
        date_report
    )
    # Сохраняем excel
    save_and_format_excel(
        df_client_excel_svod,
        df_campaigns_report,
        client_rk_svod_dir,
        date_report,
        df_dates
    )
    logger.info(f"Done creating rk client svod for client {client_name}")
