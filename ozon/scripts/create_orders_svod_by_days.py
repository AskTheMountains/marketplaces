# %% Определение всех функций
import requests
import json
import time
from datetime import date,datetime,timedelta
import pandas as pd
import shutil
import os
import glob
from pathlib import Path
import csv
import zipfile
from zipfile import ZipFile
import numpy as np
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from openpyxl.formatting.rule import ColorScaleRule
import re
from loguru import logger
import getopt
import sys
pd.set_option('future.no_silent_downcasting', True)


# Словарь русских месяцев
months_ru = {
    1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель', 5: 'Май', 6: 'Июнь',
    7: 'Июль', 8: 'Август', 9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
}


# Папка, где лежит текущий скрипт
BASE_DIR = Path(__file__).parent.parent


# Файл с настройками и номером клиента
# from options import settings, client_number, headers
# Некоторые константы
from ozon.scripts.constants import (
    headers,
    client_name,
    marketplace_dir_name,
    client_id_performance,
    client_secret_performance,
    ozon_performance_api_url,
    promotion_companies
)
# Функции выгрузки данных по API Seller
from ozon.scripts.uploadDataFromOzon import(
    get_ozon_product,
    getOrders,
    getStockReminders_fbo_v2,
    getStockReminders_fbs
)
# Функции выгрузки данных по API Seller (доп.)
from ozon.scripts.ozon_generic_functions import (
    get_orders,
    get_reminders_by_product
)
# Функции выгрузки данных по API Performane
from ozon.scripts.uploadDataFromOzonPerformance import (
    getAuthorizationToken,
    getCompanyList,
    getCompanyStatistics,
)
# Функция формирования статистики РК
from ozon.scripts.create_campaigns_report import(
    create_campaigns_report as upload_campaigns_report,
)
# Некоторые вспомогательные функции
from generic_functions import move_columns

# Функция создания началной и конечной даты
# GPT START ----
def generate_dates(
        reference_date: str = None,
        start_date: str = None,
        end_date: str = None
    ) -> pd.DataFrame:
    """
    reference_date, start_date, end_date — обязательно str в формате 'YYYY-MM-DD' либо None!
    """
    # Проверка формата, если даты заданы
    def parse_date(date_str):
        return datetime.strptime(date_str, "%Y-%m-%d")

    # Сегодня и вчера
    today = datetime.now()
    yesterday = today - timedelta(days=1)

    # Конечная дата
    if end_date is not None:
        end_date_dt = parse_date(end_date)
    else:
        end_date_dt = yesterday

    # Начальная дата
    if start_date is not None:
        start_date_dt = parse_date(start_date)
    else:
        if reference_date is None:
            start_date_dt = today.replace(day=1)
        else:
            ref_dt = parse_date(reference_date)
            start_date_dt = ref_dt + timedelta(days=1)

    # Формирование результатов
    start_date_str = start_date_dt.strftime('%Y-%m-%d')
    end_date_str = end_date_dt.strftime('%Y-%m-%d')
    start_date_iso = start_date_dt.strftime('%Y-%m-%dT00:00:00Z')
    end_date_iso = end_date_dt.strftime('%Y-%m-%dT23:59:59Z')

    df_dates = pd.DataFrame([{
        'date_start': start_date_str,
        'date_end': end_date_str,
        'datetime_start': start_date_iso,
        'datetime_end': end_date_iso
    }])
    return df_dates
# GPT END ----


# Функция получения начальной и конечной даты в формате строки
def get_start_end_date(df_dates, type_dates='companies'):
    # Если задан тип дат - даты для заказов, то берем даты со временем
    if type_dates == 'orders':
        date_start, date_end = df_dates.loc[0, ['datetime_start', 'datetime_end']]
    # В остальных случаях берем даты без времени
    else:
        date_start, date_end = df_dates.loc[0, ['date_start', 'date_end']]

    return date_start, date_end


# Функция создания диапазона дат с частотой одного дня
def generate_date_range(df_dates):
    # Получаем начальную и конечную дату
    date_start_range, date_end_range = get_start_end_date(df_dates)
    # Генерируем диапазон дат от начальной до конечной даты с интервалом в один день
    df_date_range = pd.DataFrame({'datetime': pd.date_range(date_start_range, date_end_range, freq='d')})
    # Переводим в формат даты, чтобы убрать время
    df_date_range['date'] = df_date_range['datetime'].dt.date

    return df_date_range


# Функция создания директории, куда будет помещен итоговый отчет
def create_orders_svod_dir(client_name, date_report=str(date.today())):
    orders_svod_path = (
        f"{BASE_DIR}/"
        f"Clients/{client_name}/OrdersSvod/"
    )
    if not os.path.exists(orders_svod_path):
        os.makedirs(orders_svod_path, exist_ok=True)
        logger.info(f"Creating Companies Upload directory:{orders_svod_path}")

    return orders_svod_path

# Функция чтения предыдущих отчетов
def read_report_files(
        client_name,
        orders_svod_dir,
        # date_current_report=str(date.today())
    ):
    # Маска поиска файлов
    orders_reports_file_mask = f'*Свод_Заказы_{client_name}_Ozon.xlsx'
    search_pattern = os.path.join(orders_svod_dir, orders_reports_file_mask)

    # Фильтрация временных файлов excel
    files = [
        file_path
        for file_path in glob.glob(search_pattern)
        if not os.path.basename(file_path).startswith('~$')
    ]

    # Сбор информации о файлах
    data = []
    for file_path in files:
        file_name = os.path.basename(file_path)
        modified_time = os.path.getmtime(file_path)
        modified_datetime = datetime.fromtimestamp(modified_time)
        data.append({
            'Имя файла': file_name,
            'Дата изменения': modified_datetime.strftime('%Y-%m-%d %H:%M:%S'),
            'Полный путь': file_path
        })

    # Создание DataFrame
    columns = ['Имя файла', 'Дата изменения', 'Полный путь']
    df_orders_report_files = pd.DataFrame(data, columns=columns)

    # Сортировка по дате изменения
    df_orders_report_files = df_orders_report_files.sort_values(
        by=['Дата изменения'],
        ascending=False
    )

    return df_orders_report_files


# Функция получения последней даты заказов из предыдущего отчета
def get_last_date_from_previous_report(
        df_orders_report_files,
    ):
    # Получаем путь к файлу с последним отчетом
    last_report_path = df_orders_report_files['Полный путь'].iloc[0]
    # Считываем последний отчет
    df_last_report = pd.read_excel(last_report_path)
    # Ищем максимальную дату среди колонок
    date_candidates = []
    # Перебираем все колонки в датафрейме
    for col in df_last_report.columns:
        col_str = str(col)  # Приводим название столбца к строке
        dt = None           # Здесь будем хранить найденную дату

        # Пробуем разные форматы строковой даты
        for fmt in ("%Y-%m-%d %H:%M:%S", "%d.%m.%y", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(col_str, fmt)
                break  # Если успешно спарсили — выходим из цикла форматов
            except ValueError:
                pass   # Если формат не подошёл — пробуем следующий

        # Если не удалось конвертировать через strptime, пробуем через pandas
        if dt is None:
            try:
                dt = pd.to_datetime(col_str, dayfirst=True, errors='raise')
            except Exception:
                pass  # Если ошибка — просто идём дальше

        # Если удалось преобразовать к дате — добавляем в список
        if dt is not None:
            date_candidates.append(dt)

    # Находим максимальную дату, если нашли хотя бы одну
    if date_candidates:
        max_dt = max(date_candidates)
        max_dt_str = max_dt.strftime('%Y-%m-%d')
    else:
        max_dt_str = str(date.today())  # Если не нашли ни одной даты


    return max_dt_str

# Функция даты заказов для текущего отчета
def get_date_range_for_current_report(
        client_name,
        orders_svod_dir,
    ):
    # Считываем список файлов с отчетами по текущему клиенту
    df_orders_report_files = read_report_files(client_name, orders_svod_dir)
    # Если список файлов пустой, то считаем,
    # что отчет создается в первый раз
    if df_orders_report_files.empty:
        current_report_date = str(date.today())
        logger.info('Previous report has not been found')
    # Если список файлов не пустой,
    # начинаем процесс нахождения максимальной даты предыдущего отчета
    else:
        logger.info(f"Getting last date from previous report")
        # Находим максимальную дату предыдущего отчета
        current_report_date = get_last_date_from_previous_report(df_orders_report_files)
        print(f"Last report max date: {current_report_date}")
        # Генерируем диапазон дат для текущего отчета

    df_dates = generate_dates(
        reference_date=current_report_date
    )

    return df_dates

# Функция обработка списка товаров АПИ
def process_product_list(df_products):
    # Создаем копию для избежания изменений в оригинальном df
    df_products_processed = df_products.copy()
    # Убираем ненужные колонки
    # df_products_processed = df_products_processed.loc[:, ~df_products_processed.columns.isin([
    #     'FBS OZON SKU ID', 'Контент-рейтинг', 'Рейтинг', 'Причины скрытия', 'Бренд', 'Размер', 'Цвет',
    #     'Статус товара', 'Видимость FBO', 'Причины скрытия FBO (при наличии)',
    #     'Видимость FBS', 'Причины скрытия FBS (при наличии)', 'Дата создания',
    #     'Категория комиссии', 'Объем товара, л', 'Объемный вес, кг',
    #     'Доступно к продаже по схеме FBO, шт.',
    #     'Вывезти и нанести КИЗ (кроме Твери), шт', 'Зарезервировано, шт',
    #     'Доступно к продаже по схеме FBS, шт.',
    #     'Доступно к продаже по схеме realFBS, шт.',
    #     'Зарезервировано на моих складах, шт',
    #     'Актуальная ссылка на рыночную цену', 'Размер НДС, %'
    #    ])]
    df_products_processed = df_products_processed.loc[:, df_products_processed.columns.isin([
        'Артикул',
        'Название товара',
        'SKU',
        # 'Отзывы',
        # 'Текущая цена с учетом скидки, ₽',
    ])]
    # Удаляем лишние символы из артикула
    df_products_processed['Артикул'] = df_products_processed['Артикул'].str.replace("'", "", regex=False)
    # Переводим баркод в число
    # df_products_processed['Barcode'] = pd.to_numeric(df_products_processed['Barcode'], errors='coerce')
    # df_products_processed['Barcode'] = df_products_processed['Barcode'].apply(lambda x: format(x, 'f') if pd.notnull(x) else x)
    # df_products_processed['Barcode'] = df_products_processed['Barcode'].apply(lambda x: str(x).split('.')[0] if pd.notnull(x) else x).astype('Int64')
    # Переименовываем колонки с ценами, которые приходят из отчета по товарам из апи
    # чтобы потом не было путаницы с ценами, которые мы получаем из отдельного метода
    df_products_processed = df_products_processed.rename(columns={
        'Текущая цена с учетом скидки, ₽': 'Текущая цена с учетом скидки (из отчета по товарам)',
        # 'Цена до скидки (перечеркнутая цена), ₽': 'Цена до скидки (перечеркнутая цена) (из отчета по товарам)',
        # 'Рыночная цена, ₽': 'Рыночная цена (из отчета по товарам)',
        # 'Цена Premium, ₽': 'Цена Premium (из отчета по товарам)',
    })
    # Переименовываем колонки с нужными ценами
    df_products_processed = df_products_processed.rename(columns={
        # 'Цена до скидки (перечеркнутая цена) (из отчета по товарам)': 'Цена до скидки',
        'Текущая цена с учетом скидки (из отчета по товарам)': 'Текущая цена с учетом скидки',
        # 'Минимальная цена после применения всех скидок': 'Мин. цена Ozon',
        # 'Цена до учета скидок (зачеркнутая)': 'Цена до скидки',
        # 'Цена с учетом скидок (на карточке товара)': 'Цена после скидки',
        # 'Минимальная цена после применения всех скидок': 'Мин. цена Ozon',
        # 'Цена с учетом акций продавца': 'Цена по акции',
        # 'Цена с учетом всех акций': 'Цена с баллами Озон',
    })
    # Переводим SKU в строку чтобы не было ошибок при merge
    df_products_processed['SKU'] = df_products_processed['SKU'].astype(str)
    # Переводим Артикул в строку
    df_products_processed['Артикул'] = df_products_processed['Артикул'].astype(str)
    # Выбираем нужные колонки

    return df_products_processed


# Функция выгрузки заказов
def get_orders_for_svod(df_dates, headers):
    # Получаем даты для отчета по заказам
    date_start_orders, date_end_orders = get_start_end_date(df_dates, type_dates='orders')
    # Выгружаем заказы
    df_orders_all = get_orders(
        headers,
        date_start_orders,
        date_end_orders
    )

    return df_orders_all

# Функция расчета заказов
def calc_orders(
        df_orders,
        df_date_range
    ):
    # # Убираем отмененные отправления
    df_orders_filtered = (
        df_orders
        .loc[~df_orders['Статус'].isin(['Отменён', 'Отменен']), :]
        .copy()
    )
    # Колонка, по котрой берем дату для расчета продаж
    date_column = 'Принят в обработку'
    # Переводим колонку с датой в timestamp
    df_orders_filtered[date_column] = pd.to_datetime(df_orders_filtered[date_column])
    # Оставляем только дату
    df_orders_filtered[date_column] = df_orders_filtered[date_column].dt.date
    # Переименовываем колонку для мерджа
    df_orders_filtered = df_orders_filtered.rename(columns={
        date_column: 'date',
    })
    # Считаем заказы по дням
    df_orders_by_days = (
        df_orders_filtered
        .groupby(['Артикул', 'date'], as_index=False)
        .agg(**{
            # 'Сумма заказов': ('Заказы руб', 'sum'),
            'Заказы шт': ('Заказы шт', 'sum')
        })
    )
    # Мерджим с диапазоном дат, т.к. бывает, что не все дни присутствуют в заказах
    df_orders_by_days = df_orders_by_days.merge(
        df_date_range[['date']],
        on='date',
        how='outer'
    )
    # Переводим колонку с датой в нужный формат
    df_orders_by_days['date_formatted'] = (
        pd.to_datetime(df_orders_by_days['date'])
        .dt.strftime('%d.%m.%Y')
    )
    # Переводим даты в столбцы
    df_current_report = (
        pd.pivot_table(
            df_orders_by_days,
            index='Артикул',
            columns='date_formatted',
            values='Заказы шт',
            aggfunc='sum',
            fill_value=0
        )
        .reset_index()
    )
    # Убираем имя у индекса
    df_current_report.index.name = None
    df_current_report.columns.name = None  # убираем имя у колонок (если появилось)

    return df_current_report

# Функция объединения списка товаров с заказами
def add_orders_to_product_list(
        df_products_processed,
        df_orders_for_svod
    ):

    # Объединяем список товаров с текущими заказами
    df_products_with_orders = df_products_processed.merge(
        df_orders_for_svod,
        on='Артикул',
        how='left'
    )

    return df_products_with_orders

# Функция расчета доп. колонок
def calc_additional_columns(df_products_with_orders):
    # Создаем копию для избежания изменений в оригинальном df
    df_current_report = df_products_with_orders.copy()
    # Заполняем пропуски во всех колонках, не относящихся к товарам
    sku_columns = ['Артикул', 'SKU', 'Название товара']
    other_columns = (
        df_current_report
        .columns[~df_current_report.columns.isin(sku_columns)]
        .to_list()
    )
    df_current_report[other_columns] = df_current_report[other_columns].fillna(0)

    # Добавляем колонку с датой добавления товара
    df_current_report['Дата добавления товара'] = str(date.today())
    df_current_report['Дата добавления товара'] = (
        pd.to_datetime(df_current_report['Дата добавления товара'])
        .dt.date
    )

    # Перемещаем колонку с датой добавления товара в начало df
    df_current_report = move_columns(
        df_current_report,
        'Дата добавления товара',
        'Название товара',
        'after',
    )

    return df_current_report

# Функция чтения последнего отчета
def read_last_report_file(last_report_filepath):
    # Считываем файл с последним отчетом
    df_previous_report = pd.read_excel(last_report_filepath, sheet_name=0)
    # Переводим все имена колонок с датой в формат %d.%m.%Y
    new_columns = []
    for col in df_previous_report.columns:
        col_str = str(col)
        dt = pd.to_datetime(col_str, errors='coerce', format='mixed', dayfirst=True)
        if not pd.isnull(dt):
            new_columns.append(dt.strftime('%d.%m.%Y'))  # формат: день.месяц.две последние цифры года
        else:
            new_columns.append(col)
    df_previous_report.columns = new_columns
    # Переводим колонки с характеристиками товара в строку
    sku_columns = ['Артикул', 'Название товара', 'SKU']
    for col in sku_columns:
        df_previous_report[col] = df_previous_report[col].astype(str)

    return df_previous_report

# Функция добавления новых товаров в предыдущий отчет
def add_new_products(
        df_orders_svod_files,
        df_current_report,
    ):

    # Создаем копию для избежания изменений в оригинальном df
    df_current_report_ = df_current_report.copy()
    # Берем файл с датой последнего изменения
    last_report_filepath = df_orders_svod_files['Полный путь'].iloc[0]
    # Считываем файл с последним отчетом
    df_previous_report = read_last_report_file(last_report_filepath)

    # Разделяем объединение на два этапа:
    #  товары, которые отсутствовали в предыдущем отчете, мерджим через concat
    #  остальные товары через join

    # Создаем df с новыми артикулами
    df_new_products = (
        df_current_report_
        # .loc[:, ['Артикул', 'SKU', 'Название товара', 'Дата добавления товара']]
        .merge(
            df_previous_report[['Артикул']],
            on='Артикул',
            how='left',
            indicator='Наличие артикула'
        )
        .pipe(lambda df:
                df.loc[df['Наличие артикула'] == 'left_only', :]
        )
        .drop(columns=['Наличие артикула'])
    )
    if not df_new_products.empty:
        logger.info(f"Found {df_new_products.shape[0]} new products")
    # Формируем список с новыми артикулами
    new_products_list = df_new_products['Артикул'].to_list()

    # Создаем df с теми артикулами, которые уже были в предыдущих отчетах
    df_current_report_previous_products = (
        df_current_report_
        .loc[
            ~df_current_report['Артикул'].isin(new_products_list),
            ~df_current_report.columns.isin([
                'SKU',
                'Название товара',
                'Дата добавления товара'
            ])
        ]
        .copy()
    )

    # Старые товары объединяем через join
    df_current_and_previous_report = (
        df_previous_report
        .merge(
            df_current_report_previous_products,
            on='Артикул',
            how='left'
        )
    )
    # Новые товары объединяем через concat
    df_current_and_previous_report = pd.concat([
        df_current_and_previous_report,
        df_new_products
    ])

    # Сбрасываем index и делаем сортировку
    df_current_and_previous_report = (
        df_current_and_previous_report
        .reset_index(drop=True)
        .sort_values(by=['Артикул'], ignore_index=True)
    )
    # Заполняем пропуски в выбранных колонках
    columns_to_fill = (
        df_current_and_previous_report
        .columns[
            ~df_current_and_previous_report
            .columns.str.contains('Артикул|SKU|Название товара|Дата добавления товара', regex=True)
        ]
        .to_list()
    )
    df_current_and_previous_report[columns_to_fill] = df_current_and_previous_report[columns_to_fill].fillna(0)

    # Переводим колонку с датой добавления в datetime для удобства
    df_current_and_previous_report['Дата добавления товара'] = (
        pd.to_datetime(df_current_and_previous_report['Дата добавления товара'])
        .dt.date
    )
    # Перемещаем колонку с датой добавления товара в начало df
    df_current_and_previous_report = move_columns(
        df_current_and_previous_report,
        'Дата добавления товара',
        'Название товара',
        'after',
    )

    return df_current_and_previous_report

# Функция добавления новых товаров в предыдущий отчет (v2)
# НЕ РАБОТАЕТ, НУЖНА ДОРАБОТКА
def add_new_products_v2(
        df_orders_svod_files,
        df_current_report,
    ):

    # Создаем копию для избежания изменений в оригинальном df
    df_current_report_ = df_current_report.copy()
    # Берем файл с датой последнего изменения
    last_report_filepath = df_orders_svod_files['Полный путь'].iloc[0]
    # Считываем файл с последним отчетом
    df_previous_report = read_last_report_file(last_report_filepath)
    # Создаем df с новыми артикулами
    df_new_products = (
        df_current_report_
        # .loc[:, ['Артикул', 'SKU', 'Название товара', 'Дата добавления товара']]
        .merge(
            df_previous_report[['Артикул']],
            on='Артикул',
            how='left',
            indicator='Наличие артикула'
        )
        .pipe(lambda df:
                df.loc[df['Наличие артикула'] == 'left_only', :]
        )
        .drop(columns=['Наличие артикула'])
    )
    if not df_new_products.empty:
        logger.info(f"Found {df_new_products.shape[0]} new products")

    # Формируем список колонок, относящихся к товару
    sku_columns = ['Артикул', 'SKU', 'Название товара', 'Дата добавления товара']
    # Объединяем два отчета в один
    df_current_and_previous_report = (
        # Объединяем два отчета в один
        pd.concat([
            df_previous_report,
            df_current_report_
        ])
        # Считаем сумму после concat
        .groupby(sku_columns, as_index=False, dropna=False)
        .sum()
        # Сбрасываем index и делаем сортировку
        # .reset_index(drop=True)
        .sort_values(by=['Артикул'], ignore_index=True)
    )

    return df_current_and_previous_report

# Функция объединения текущего отчета с предыдущим
def add_data_from_previous_report(
        client_name,
        orders_svod_dir,
        df_current_report,
    ):
    # Считываем список файлов с отчетами по текущему клиенту
    df_orders_svod_files = read_report_files(client_name, orders_svod_dir)
    # Если список файлов пустой, то считаем,
    # что отчет создается в первый раз и возвращаем копию исходного df
    if df_orders_svod_files.empty:
        logger.info('Previous report has not been found, creating first report')
        df_current_and_previous_report = df_current_report.copy()
        return df_current_and_previous_report
    # Если список файлов не пустой, начинаем процесс объединения
    else:
        logger.info(f"Found previous report")
        df_current_and_previous_report = add_new_products(
            df_orders_svod_files,
            df_current_report
        )

    return df_current_and_previous_report


# Функция расчета итоговых значений за месяц
# GPT START ----
def calc_month_result_column(
        df_current_and_previous_report,
        calc_month_result=False
    ):
    months_ru = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
        7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
    }

    df_excel_orders_svod = df_current_and_previous_report.copy()
    if not calc_month_result:
        return df_excel_orders_svod

    # 1. Ищем все колонки-даты и формируем словарь (год, месяц): [колонки]
    date_col_to_dt = {}
    month_year_to_cols = defaultdict(list)
    for col in df_excel_orders_svod.columns:
        col_str = str(col)
        if re.fullmatch(r"\d{2}\.\d{2}\.\d{2,4}", col_str):
            dt = pd.to_datetime(col_str, errors='coerce', dayfirst=True)
            if not pd.isnull(dt):
                date_col_to_dt[col] = dt
                month_year_to_cols[(dt.year, dt.month)].append(col)

    # 2. Добавляем итоговые столбцы за месяц
    total_cols = {}
    for (year, month), cols in month_year_to_cols.items():
        cols_sorted = sorted(cols, key=lambda c: date_col_to_dt[c])
        total_col = f"Итого {months_ru[month]} {year}"
        df_excel_orders_svod[total_col] = df_excel_orders_svod[cols_sorted].sum(axis=1)
        total_cols[(year, month)] = total_col

    # 3. Характеристики товара — это не даты и не "Итого ..."
    date_cols = set(date_col_to_dt.keys())
    total_name_pattern = re.compile(r"Итого [А-Яа-я]+ \d{4}$")
    characteristics_cols = [c for c in df_excel_orders_svod.columns if (c not in date_cols) and (not total_name_pattern.match(str(c)))]

    # 4. Формируем порядок колонок: характеристики -> блоки месяц-даты + итог
    month_year_sorted = sorted(month_year_to_cols.keys())
    ordered_cols = []
    for year, month in month_year_sorted:
        month_cols = sorted(month_year_to_cols[(year, month)], key=lambda c: date_col_to_dt[c])
        ordered_cols.extend(month_cols)
        ordered_cols.append(total_cols[(year, month)])

    # 5. Собираем итоговый порядок
    full_col_order = characteristics_cols + ordered_cols
    # Если есть "лишние" столбцы (например, итоговые, уже имеющиеся), добавим их в конец
    missing_cols = [c for c in df_excel_orders_svod.columns if c not in full_col_order]
    result_cols = full_col_order + missing_cols

    return df_excel_orders_svod[result_cols]
# GPT END ----

# Функция сохранения и форматирования отчета excel
def save_and_format_excel(
        df_excel_orders_svod,
        orders_svod_dir,
        date_report,
    ):
    # Формируем директорию для сохранения
    path_report = (
        f"{orders_svod_dir}/"
    )
    name_report = (
        f"{date_report}_Свод_Заказы_{client_name}_Ozon.xlsx"
    )
    file_name_report = f"{path_report}/{name_report}"
    if not os.path.exists(path_report):
        os.makedirs(path_report, exist_ok=True)
    # Формируем даты для названия листа
    # date_start_excel = (
    #     pd.to_datetime(df_dates['date_start'])
    #     .iloc[0]
    #     .strftime('%d.%m')
    # )
    # date_end_excel = (
    #     pd.to_datetime(df_dates['date_end'])
    #     .iloc[0]
    #     .strftime('%d.%m')
    # )
    # Задаем колонки с процентным форматом
    # percent_cols = ['ДРР, %', 'ДРР % по размеру']
    # Делим на 100, чтобы отображался процент
    # df_client_excel_svod[percent_cols] = df_client_excel_svod[percent_cols] / 100

    # Формируем название листа
    sheet_name_client_cabinet_svod = f"{client_name}"
    # Сохраняем книгу
    with pd.ExcelWriter(file_name_report) as w:
        df_excel_orders_svod.to_excel(w, index=False, sheet_name=sheet_name_client_cabinet_svod)

    # ---- Форматируем файл с отчетом ----

    # Открываем книгу excel
    wb = load_workbook(file_name_report)

    # Общее форматирование для всех листов
    for work_sheet in wb.sheetnames:
        # Список колонок
        svod_columns = list(df_excel_orders_svod.columns)
        # Выбираем нужный лист
        ws = wb[work_sheet]

        # Определяем стиль границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Создаем границы
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

        # Форматирование числовых ячеек
        exclude_number_format_substrings = ['SKU']
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                # Определяем номер колонки с заголовком
                col_idx = cell.column - 1  # openpyxl: колонка начинается с 1
                col_name = svod_columns[col_idx]

                # Пропускаем форматирование некоторых столбцов
                if any(substr in str(col_name) for substr in exclude_number_format_substrings):
                    continue

                if isinstance(cell.value, (int, float)):
                    if float(cell.value).is_integer():
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'

        # --- Перенос текста во всех заголовках ---
        for idx in range(1, ws.max_column + 1):
            ws.cell(row=1, column=idx).alignment = Alignment(
                wrap_text=True,
                #horizontal=ws.cell(row=1, column=idx).alignment.horizontal or 'center',
                horizontal='center',
                vertical='center'
            )

        # --- Автоподбор ширины столбцов + фильтр на столбцы ---
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for i, cell in enumerate(col):
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Если это заголовок (первая строка), учитываем значок фильтра
            head_len = len(str(col[0].value)) if col[0].value else 0
            filter_icon_length = 3 if head_len == max_len else 2  # 3 для filter icon
            ws.column_dimensions[col_letter].width = max_len + filter_icon_length

        # --- Значок фильтра на столбцы ---
        ws.auto_filter.ref = ws.dimensions

        # --- Форматирование по подстрокам в заголовках ---
        columns_fix_center = ['Остаток Озон', 'РК', 'Кол-во заказов']
        columns_fix_left = ['Артикул', 'SKU', 'Название товара', 'Дата добавления товара']

        for idx, col_name in enumerate(svod_columns):
            col_letter = get_column_letter(idx + 1)

            # Проверяем на подстроки слева
            if any(sub in str(col_name) for sub in columns_fix_left):
                align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=idx+1).alignment = align
                ws.cell(row=1, column=idx+1).alignment = align  # заголовок тоже по левому краю

            # Остальные строки выравниваем по центру:
            else:
                ws.column_dimensions[col_letter].width = 18  # фикс ширина
                align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for row in range(2, ws.max_row + 1):  # пропустить заголовки
                    ws.cell(row=row, column=idx+1).alignment = align
                ws.cell(row=1, column=idx+1).alignment = align  # заголовок центрируем


        # ---  Форматирование процентных колонок ---
        # percent_substrings = ['%', 'ДРР']  # ищем колонки по наличию этих символов

        # # Получаем номера подходящих колонок по наличию подстроки
        # percent_col_indices = [
        #     idx + 1  # Excel columns start from 1
        #     for idx, name in enumerate(svod_columns)
        #     if any(substr in name for substr in percent_substrings)
        # ]

        # # Применяем форматирование к найденным колонкам
        # for col_idx in percent_col_indices:
        #     for row in range(2, ws.max_row + 1):  # пропустить заголовки
        #         cell = ws.cell(row=row, column=col_idx)
        #         if cell.value:
        #             cell.value = cell.value * 0.01  # для процентов
        #         cell.number_format = '0.00%'  # формат: проценты с 2 знаками

        # ---  Условное форматирование градиентом для выбранных колонок ---
        # gradient_cols = ['ДРР, %', 'ДРР % по размеру']
        # for col_name in gradient_cols:
        #     col_idx = list(df_excel_orders_svod.columns).index(col_name) + 1
        #     col_letter = ws.cell(row=1, column=col_idx).column_letter
        #     # Диапазон от второй строки до конца в данном столбце (без заголовков)
        #     cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"

        #     # Устанавливаем цвет: min (зеленый), max (оранжевый), середина (по желанию, желтый)
        #     rule = ColorScaleRule(
        #         start_type='min', start_color='63BE7B',   # Зеленый
        #         mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Желтый
        #         end_type='max', end_color='E06967',       # Красный
        #     )
        #     ws.conditional_formatting.add(cell_range, rule)

    # Сохранение обновленного файла
    wb.save(file_name_report)


# %% Вызов всех функций
if __name__ == '__main__':

    # Дата, за которую формируется (или формировался) отчет
    date_report = str(date.today())
    # Создаем директорию, куда будет помещен итоговый отчет
    orders_svod_dir = create_orders_svod_dir(client_name)
    # Формируем диапазон дат для текущего отчета
    df_dates = get_date_range_for_current_report(
        client_name,
        orders_svod_dir
    )
    date_start_current_report = df_dates['date_start'].iloc[0]
    date_end_current_report = df_dates['date_end'].iloc[0]
    logger.info(
        f"\nCreating orders svod for client {client_name}\n"
        f"for dates: {date_start_current_report} - {date_end_current_report}"
    )
    # Создаем диапазон дат с частотой одного дня
    df_date_range = generate_date_range(df_dates)
    # Выгружаем заказы
    df_orders = get_orders_for_svod(
        df_dates,
        headers,
    )
    # Получаем список товаров АПИ
    df_products = get_ozon_product(headers, to_save=False)
    # Обрабатываем список товаров
    df_products_processed = process_product_list(df_products)
    # Считаем заказы по дням
    df_orders_for_svod = calc_orders(
        df_orders,
        df_date_range,
    )
    # Объединяем заказы со списком товаров
    df_products_with_orders = add_orders_to_product_list(
        df_products_processed,
        df_orders_for_svod,
    )
    # Формируем текущую версию отчета
    df_current_report = calc_additional_columns(df_products_with_orders)

    # Объединяем текущий отчет с предыдущим
    df_current_and_previous_report = add_data_from_previous_report(
        client_name,
        orders_svod_dir,
        df_current_report
    )
    # Считаем итоговые суммы по месяцу
    df_excel_orders_svod = calc_month_result_column(
        df_current_and_previous_report,
        calc_month_result=True
    )
    # Сохраняем и форматируем excel
    save_and_format_excel(
        df_excel_orders_svod,
        orders_svod_dir,
        date_report,
    )
# %%
