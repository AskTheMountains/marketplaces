
# %% Определение всех функций
import pandas as pd
import numpy as np
import requests
import time
import os
import openpyxl
import json
from loguru import logger
from datetime import date,datetime,timedelta
from itertools import product
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_00

# Файл с некоторыми константами
from wb.scripts.constants import (
    headers,
    client_name,
    marketplace_dir_name
)

# Функции выгрузки данных по АПИ, которые уже готовы
from wb.scripts.uploadDataFromWB import getOrdersWB, getSalesWB
from wb.scripts.calc_svod_finance_report_WB_v2 import (
    upload_realization_report,
    rename_report_columns,
    add_new_columns,
    # filter_report_dates
)
# Доп. функции
from generic_functions import move_columns


# Создание нужных директорий
def create_dirs():
    # Папка с клиентом
    client_dir = f"{marketplace_dir_name}/Clients/{client_name}/"
    # Список директорий для создания
    dir_names = ['SaleSvod']
    for dir_name in dir_names:
        dir_path = os.path.join(client_dir, dir_name)
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)

# Функция создания диапазона дат
# GPT START ----
def generate_date_range(
        reference_date: str = None,
        start_date: str = None,
        end_date: str = None
    ):
    # Проверка на несовместимость параметров:
    # либо используем reference_date, либо вручную заданный диапазон
    if reference_date and (start_date or end_date):
        raise ValueError("Нельзя одновременно задавать reference_date и start_date/end_date.")

    # Если оба manual-диапазона заданы
    if start_date and end_date:
        # Парсим строки дат в объекты datetime
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    elif start_date or end_date:
        # Один из диапазонов не указан — ошибка
        raise ValueError("Нужно задать одновременно start_date и end_date.")
    else:
        # Ни диапазон, ни reference_date не заданы явно
        #  — значит работаем с reference_date (по умолчанию — сегодня)
        if reference_date is None:
            ref_dt = datetime.now()
        else:
            # Парсим reference_date из строки
            ref_dt = datetime.strptime(reference_date, "%Y-%m-%d")

        # Если дата — первое число месяца
        if ref_dt.day == 1:
            # Берём последний день прошлого месяца
            prev_month_last_day = ref_dt - timedelta(days=1)
            # Начало диапазона — первое число прошлого месяца
            start_dt = prev_month_last_day.replace(day=1)
            # Конец диапазона — последний день прошлого месяца
            end_dt = prev_month_last_day
        else:
            # Если не первое число месяца:
            # Начало диапазона — первое число текущего месяца
            start_dt = ref_dt.replace(day=1)
            # Конец диапазона — предыдущие сутки относительно reference_date
            end_dt = ref_dt - timedelta(days=1)

    # Форматы для строк и для ISO-даты по времени суток
    start_date_str = start_dt.strftime('%Y-%m-%d')
    end_date_str = end_dt.strftime('%Y-%m-%d')
    start_date_iso = start_dt.strftime('%Y-%m-%dT00:00:00Z')
    end_date_iso = end_dt.strftime('%Y-%m-%dT23:59:59Z')

    # Собираем результат в DataFrame
    df_dates = pd.DataFrame([{
        'date_start': start_date_str,       # начало диапазона (дд.мм.гггг)
        'date_end': end_date_str,           # конец диапазона (дд.мм.гггг)
        'datetime_start': start_date_iso,   # начало ISO-датой с 00:00:00
        'datetime_end': end_date_iso        # конец ISO-датой с 23:59:59
    }])

    return start_date_iso, end_date_iso
# GPT END ----


# Функция создания датафрейма с диапазоном дат
def generate_date_range_df(date_start, date_end):
    # Генерируем диапазон дат от начальной до конечной даты с интервалом в один день
    date_range_df = pd.DataFrame({'date': pd.date_range(date_start, date_end, freq='d')})
    # Переводим в формат даты, чтобы убрать время
    date_range_df['date'] = date_range_df['date'].dt.date

    return date_range_df

# Функция выгрузки отчета о заказах
def get_orders(headers, date_start, date_end):
    # Выгружаем отчет об отправлениях fbo и fbs
    df_orders = getOrdersWB(headers, date_start, date_end, to_save=False)

    return df_orders


# Функция выгрузки отчета о продажах
def get_sales(headers, date_start, date_end):
    # Выгружаем отчет об отправлениях fbo и fbs
    df_sales = getSalesWB(headers, date_start, date_end, to_save=False)

    return df_sales


# Функция фильтра отчета о реализации по датам
def filter_report_dates(date_start, date_end, df_realization_report_new_columns, filter_dates=False):
    # Создаем копию для избежания изменений в оригинальном df
    df_realization_report_date_filtered = df_realization_report_new_columns.copy()
    if filter_dates:
        # # Переводим даты выгрузки в Timestamp
        dt_start = pd.to_datetime(date_start, format='mixed')
        dt_end = pd.to_datetime(date_end, format='mixed')
        # Ищем id строк переходных недель
        tmp_df_transitional_weeks = df_realization_report_date_filtered.loc[
            (df_realization_report_date_filtered['Дата начала отчётного периода'] < dt_start) | (df_realization_report_date_filtered['Дата конца отчётного периода'] > dt_end),
            :]
        # Внутри переходных недель ищем id строк, у которых дата продажи не соответствует датам начала и окончания выгрузки
        transitional_weeks_ids = tmp_df_transitional_weeks.loc[(tmp_df_transitional_weeks['Дата продажи'] < dt_start) | (tmp_df_transitional_weeks['Дата продажи'] > dt_end), 'id'].to_list()
        # Убираем найденные строки из отчета
        df_realization_report_date_filtered = df_realization_report_date_filtered.loc[~df_realization_report_date_filtered['id'].isin(transitional_weeks_ids), :]
        # Делаем фильтр по дате продажи в пределах указанного диапазона
        # df_realization_report_date_filtered = df_realization_report_date_filtered.loc[(df_realization_report_date_filtered['Дата продажи'] >= dt_start) & (df_realization_report_date_filtered['Дата продажи'] <= dt_end), :]
        # Делаем reset_index после concat
        df_realization_report_date_filtered = df_realization_report_date_filtered.reset_index(drop=True)

    return df_realization_report_date_filtered


# Функция выгрузки отчета о реализации
def get_realization_report(headers, date_start, date_end):
    # Переводим даты в нужный формат
    date_start_ = datetime.fromisoformat(date_start).strftime('%Y-%m-%dT%H:%M:%S')
    date_end_ = datetime.fromisoformat(date_end).strftime('%Y-%m-%dT%H:%M:%S')
    # Получение отчета о реализации по АПИ
    df_realization_report = upload_realization_report(headers, date_start_, date_end_)
    # Переименование колонок отчета о реализации в русские названия
    df_realization_report_renamed = rename_report_columns(df_realization_report)
    # Расчет некоторых дополнительных колонок, которые будут участвовать в расчетах
    df_realization_report_new_columns = add_new_columns(df_realization_report_renamed)
    # Фильтр по переходным неделям
    df_realization_report_date_filtered = filter_report_dates(date_start_, date_end_,
                                                            df_realization_report_new_columns,
                                                            filter_dates=True)
    return df_realization_report_date_filtered


# Функция определения источника расчета продаж
def define_sales_source(df_realization_report, df_sales_api):
    # Переменные, куда будем помещать результат
    df_sales = pd.DataFrame()
    source = ''
    # Если отчет о реализации пустой, то источник - отчет апи,
    # иначе - отчет о реализации
    if df_realization_report.empty:
        df_sales = df_sales_api.copy()
        source = 'api_report'
    else:
        df_sales = df_realization_report.copy()
        source = 'realization_report'
    # Формируем словарь с результатами
    result_sales_source = {
        'df_sales': df_sales,
        'source': source
    }

    return result_sales_source


# Функция определения столбца расчета продаж в рублях
def define_sales_column(result_sales_source, client_name):
    # Определяем источник, по которому рассчитываем продажи
    source = result_sales_source['source']
    # Если источник - отчет о реализации
    if source == 'realization_report':
        if client_name in ['SENS', 'SENS_IP', 'KU_And_KU', 'Soyuz', 'TRIBE', 'Orsk_Combinat']:
            price_column = 'Цена розничная с учетом согласованной скидки'
        else:
            price_column = 'Сумма продаж (возвратов)'
    # Если источник - отчет АПИ
    elif source == 'api_report':
        if client_name in ['SENS', 'SENS_IP', 'KU_And_KU', 'Soyuz', 'TRIBE', 'Orsk_Combinat']:
            price_column = 'priceWithDisc'
        else:
            price_column = 'finishedPrice'

    return price_column


# Функция расчета заказов
def calc_orders(df_orders, date_range_df, price_column='priceWithDisc'):
    # Создаем копию для избежания изменений в оригинальном df
    df_orders_ = df_orders.copy()
    # Колонка, по котрой берем дату для расчета продаж
    date_column = 'date'
    # Колонка, по которой считаем итоговую сумму в рублях
    if client_name in ['SENS', 'SENS_IP', 'KU_And_KU', 'Soyuz', 'TRIBE', 'Orsk_Combinat']:
        price_column_ = 'priceWithDisc'
    elif client_name in ['new_client']:
        price_column_ = 'finishedPrice'
    # Колонка, по которой считаем количество в штуках
    amount_column = 'Заказы шт'
    # Переводим колонку с датой в timestamp
    df_orders_[date_column] = pd.to_datetime(df_orders_[date_column])
    # Оставляем только дату
    df_orders_[date_column] = df_orders_[date_column].dt.date
    # Создаем колонку с количеством заказов, 1 строка = 1 заказ
    df_orders_[amount_column] = 1
    # Переименовываем колонку для мерджа
    df_orders_ = df_orders_.rename(columns={
        date_column: 'date',
        price_column_: 'Заказы руб'
        })
    # Считаем заказы в штуках и рублях
    orders = (df_orders_
              .groupby(['date'])
              .agg(**{
                  'Заказы руб': ('Заказы руб', 'sum'),
                  'Заказы шт': ('Заказы шт', 'sum')
              })
              .reset_index()
    )

    # Мерджим с диапазоном дат, т.к. бывает, что не все дни присутствуют в заказах
    orders = orders.merge(
        date_range_df,
        on='date',
        how='outer'
    )
    # Заполняем пропуски после мерджа
    for col in ['Заказы руб', 'Заказы шт']:
        orders[col] = orders[col].fillna(0)

    return orders


# Функция расчета продаж по отчету о продажам
def calc_sales_from_api_report(df_sales_api, date_range_df, price_column='finishedPrice'):
    # Создаем копию для избежания изменений в оригинальном df
    df_sales_ = df_sales_api.copy()
    # Колонка, по котрой берем дату для расчета продаж
    date_column = 'date'
    # Колонка, по которой считаем сумму в рублях
    price_column_ = price_column
    # Колонка, по которой считаем количество в штуках
    # amount_column = 'Количество'
    # Переводим колонку с датой в timestamp
    df_sales_[date_column] = pd.to_datetime(df_sales_[date_column])
    # Оставляем только дату
    df_sales_[date_column] = df_sales_[date_column].dt.date

    # Создаем колонку с количеством продаж, 1 строка = 1 заказ
    # df_sales_[amount_column] = 1
    # Там, где тип заказа возврат, делаем отрицательное число штук товара,
    # чтобы потом просто посчитать сумму
    df_sales_['Продажи шт'] = np.where(df_sales_['saleID'].str.startswith('R'),
                                        -1,
                                         1)
    # Для продаж в рублях аналогично
    df_sales_['Продажи руб'] = df_sales_[price_column_]

    # Переименовываем колонку для мерджа
    df_sales_ = df_sales_.rename(columns={
        date_column: 'date',
        })

    # Считаем продажи в штуках и рублях
    sales = (
        df_sales_
        .groupby(['date'])
        .agg(**{
            'Продажи руб': ('Продажи руб', 'sum'),
            'Продажи шт': ('Продажи шт', 'sum')
            })
        .reset_index()
    )
    # Мерджим с диапазоном дат, т.к. бывает, что не все дни присутствуют в заказах
    sales = sales.merge(
        date_range_df,
        on='date',
        how='outer'
    )

    # Заполняем пропуски после мерджа
    for col in ['Продажи руб', 'Продажи шт']:
        sales[col] = sales[col].fillna(0)

    return sales


# Функция расчета продаж по отчету о реализации
def calc_sales_from_realization(df_realization_report, date_range_df, price_column='Сумма продаж (возвратов)'):
    # Создаем копию для избежания изменений в оригинальном df
    df_realization_report_ = df_realization_report.copy()
    # Колонка, по котрой берем дату для расчета продаж
    date_column = 'Дата продажи'
    # Колонка, по которой считаем сумму в рублях
    price_column_ = price_column
    # Колонка, по которой считаем количество в штуках
    amount_column = 'Количество'
    # Там, где тип начисления возврат, делаем отрицательное число штук товара,
    # чтобы потом просто посчитать сумму
    df_realization_report_['Продажи шт'] = np.where(df_realization_report_['Обоснование для оплаты'] == 'Возврат',
                                         df_realization_report_[amount_column] * (-1),
                                         df_realization_report_[amount_column])
    # Для продаж в рублях аналогично
    df_realization_report_['Продажи руб'] = np.where(df_realization_report_['Обоснование для оплаты'] == 'Возврат',
                                         df_realization_report_[price_column_] * (-1),
                                         df_realization_report_[price_column_])

    # Переводим колонку с датой в timestamp
    df_realization_report_[date_column] = pd.to_datetime(df_realization_report_[date_column])
    # Оставляем только дату
    df_realization_report_[date_column] = df_realization_report_[date_column].dt.date
    # Переименовываем колонку для мерджа
    df_realization_report_ = df_realization_report_.rename(columns={
        date_column: 'date',
        })

    # Делаем выборку по нужным типам начислений и считаем итоговые суммы
    sales = (
        df_realization_report_
        .loc[(df_realization_report_['Обоснование для оплаты'] \
              .isin(
                  ['Продажа', 'Возврат']
                  )
                ),
            :]
        .groupby(['date'])
        .agg(**{
            'Продажи руб': ('Продажи руб', 'sum'),
            'Продажи шт': ('Продажи шт', 'sum')
            })
        .reset_index()
    )
    # Мерджим с диапазоном дат, т.к. бывает, что не все дни присутствуют в заказах
    sales = sales.merge(
        date_range_df,
        on='date',
        how='outer'
    )


    # Заполняем пропуски после мерджа
    for col in ['Продажи руб', 'Продажи шт']:
        sales[col] = sales[col].fillna(0)

    return sales


# Функция расчета продаж (общая)
def calc_sales(df_realization_report, df_sales_api, date_range_df, client_name):
    # Определяем источник расчета продаж
    result_sales_source = define_sales_source(df_realization_report, df_sales_api)
    source = result_sales_source['source']
    # Определяем колонку, по которой будем считать продажи
    price_column = define_sales_column(result_sales_source, client_name)
    # Считаем продажи в зависимости от источника
    if source == 'realization_report':
        df_sales = calc_sales_from_realization(df_realization_report, date_range_df, price_column)
    elif source == 'api_report':
        df_sales = calc_sales_from_api_report(df_sales_api, date_range_df, price_column)

    return df_sales


# Объединяем заказы и продажи в одну таблицу
def union_orders_and_sales(orders, sales):
    # Объединяем в один df
    df_orders_and_sales = pd.concat([
        orders.set_index(orders['date']),
        sales.set_index(sales['date'])
    ], axis=1)
    # Удаляем ненужную колонку
    df_orders_and_sales = df_orders_and_sales.drop(columns=['date'])
    # Достаем дату из индекса
    df_orders_and_sales = df_orders_and_sales.reset_index()
    # Форматируем дату в нужный формат
    df_orders_and_sales['Дата'] = df_orders_and_sales['date'].apply(lambda x: x.strftime('%d.%m.%Y'))
    # Разворачиваем в широкий вид
    df_orders_and_sales_wide = pd.pivot_table(df_orders_and_sales,
                                              values=['Заказы руб', 'Заказы шт', 'Продажи руб', 'Продажи шт'],
                                              # index=['Заказы руб', 'Заказы шт', 'Продажи руб', 'Продажи шт'],
                                              columns=['date'],
                                              aggfunc='sum')
    # Делаем нужный формат даты у колонок
    df_orders_and_sales_wide.columns = pd.to_datetime(df_orders_and_sales_wide.columns).strftime('%d.%m.%Y')

    return df_orders_and_sales_wide


# Функция добавления плана и факта продаж
def add_plan(df_orders_and_sales_wide,
             client_name,
             plan_orders_rub = 0,
             plan_orders_amount = 0,
             plan_sales_rub = 0,
             plan_sales_amount = 0):
    # Создаем копию для избежания изменений в оригинальном df
    df_plan = df_orders_and_sales_wide.copy()
    # Считаем итоговые суммы
    df_plan['Факт'] = df_plan.sum(axis=1)
    # Добавляем План
    df_plan['План'] = np.nan
    df_plan.loc['Заказы руб', 'План'] = plan_orders_rub
    df_plan.loc['Заказы шт', 'План'] = plan_orders_amount
    df_plan.loc['Продажи руб', 'План'] = plan_sales_rub
    df_plan.loc['Продажи шт', 'План'] = plan_sales_amount
    # Считаем отклонение плана от факта
    df_plan['Факт/План, %'] = df_plan['Факт'] / df_plan['План'] * 100
    # Заменяем inf, если получилось деление на ноль
    df_plan['Факт/План, %'] = df_plan['Факт/План, %'].replace([-np.inf, np.inf], np.nan)
    # Добавляем имя магазина
    # df_plan['Магазин'] = client_name
    # # Добавление пустой строки в начало DataFrame
    # empty_row = pd.DataFrame([[np.nan] * df_plan.shape[1]], columns=df_plan.columns, index=['Магазин'])
    # df = pd.concat([empty_row, df_plan])
    # # Перемещаем колонки в начало df
    df_plan = move_columns(df_plan, ['План', 'Факт', 'Факт/План, %'], 0)
    # Убираем имя у индекса
    df_plan.index.name = None

    return df_plan


# Функция сохранения отчета в excel
def save_excel(df_plan, date_start, date_end, date_report=str(date.today())):
    # Создаем df с именем клиента
    shop_name = pd.DataFrame({'Магазин': [client_name]})
    # Создаем df с диапазоном формирования отчета
    df_dates = pd.DataFrame({'Начальная дата': [date_start],
                             'Конечная дата': [date_end]})
    # Даты для имени файла
    date_start_file = pd.to_datetime(date_start).strftime('%d-%m-%Y')
    date_end_file = pd.to_datetime(date_end).strftime('%d-%m-%Y')

    # Имя файла для сохранения отчета
    report_file_name = (f"{marketplace_dir_name}/Clients/{client_name}/SaleSvod/"
                        f"{date_start_file}_{date_end_file}"
                        f"_Свод_продажи_WB_{client_name}.xlsx"
    )

    # GPT START ----
    # Параметры для записи датафреймов
    dataframes = [
        (shop_name, 2, 2, False),  # Начало в строке 1, колонке 1, не записывать индекс датафрейма
        (df_plan, 6, 2, True),
        (df_dates, 12, 2, False),
    ]

    # Запись датафреймов на один лист Excel
    with pd.ExcelWriter(report_file_name, engine='openpyxl') as writer:
        ranges = []  # Здесь будем сохранять диапазоны (начало и конец) для границ
        for df, start_row, start_col, save_index in dataframes:
            # Если сохраняется индекс, добавляется 1 столбец
            end_row = start_row + len(df)  # Вычисляем конечную строку
            end_col = start_col + len(df.columns) - 1 + (1 if save_index else 0)  # Конечный столбец с учётом индекса
            ranges.append((start_row, end_row, start_col, end_col, save_index))  # Добавляем "save_index"

            # Запись датафрейма с учётом параметра индекса
            df.to_excel(writer, index=save_index, sheet_name=client_name, startrow=start_row - 1, startcol=start_col - 1)


    # Форматирование Excel
    logger.info(f"Formatting Sales svod for client {client_name}")

    wb = load_workbook(report_file_name)
    ws = wb[client_name]

    # Определяем стиль границ
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Применяем границы только к диапазону данных таблиц (без индекса)
    for start_row, end_row, start_col, end_col, save_index in ranges:
        for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col + (1 if save_index else 0),  # Сдвигаем начало диапазона вправо, если индекс записан
            max_col=end_col
        ):
            for cell in row:
                cell.border = thin_border

    # Применение числового форматирования с разделением разрядов (через пробелы)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):  # Проверяем, является ли значение числом
                cell.number_format = '#,##0'  # Формат с пробелами и двумя знаками после запятой

    # Автоподбор ширины столбцов
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Номер текущего столбца
        column_letter = get_column_letter(column)  # Буква столбца
        for cell in col:
            try:
                if cell.value:  # Проверка на наличие данных
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # Сохранение обновленного файла
    wb.save(report_file_name)
    # GPT END ----


# %% Вызов всех функций
if __name__ == '__main__':
    # Формируем диапазон дат
    # date_start = '2025-07-01T00:00:00.000Z'
    # date_end = '2025-07-08T23:59:59.000Z'
    # date_range = generate_date_range(date_start, date_end)

    date_start, date_end = generate_date_range()
    date_range_df = generate_date_range_df(date_start, date_end)


    logger.info(
        f"Calculating Sales svod for client {client_name} "
        f"for dates {date_start} - {date_end}"
    )
    # Планы продаж и заказов в штуках и рублях
    plan_orders_rub = 0
    plan_orders_amount = 0
    plan_sales_rub = 0
    plan_sales_amount = 0
    # Выгружаем заказы
    df_orders = get_orders(headers, date_start, date_end)
    # Выгружаем отчет о продажах
    df_sales_api = get_sales(headers, date_start, date_end)
    # Выгружаем отчет о реализации
    df_realization_report = get_realization_report(headers, date_start, date_end)

    # Рассчитываем заказы
    orders = calc_orders(df_orders, date_range_df)
    # Рассчитываем продажи
    sales = calc_sales(df_realization_report, df_sales_api, date_range_df, client_name)
    # Объединяем продажи и заказы в один df
    df_orders_and_sales_wide = union_orders_and_sales(orders, sales)
    # Добавляем планы продаж и считаем отклонение плана от факта
    df_plan = add_plan(
        df_orders_and_sales_wide,
        client_name,
        plan_orders_rub,
        plan_orders_amount,
        plan_sales_rub,
        plan_sales_amount
    )
    # Создаем нужную директорию
    create_dirs()
    # Сохраняем Excel
    save_excel(df_plan, date_start, date_end)
    # Форматируем Excel
    # format_excel()
    logger.info(f"Finished calculating Sales svod for client {client_name}")

# %%
