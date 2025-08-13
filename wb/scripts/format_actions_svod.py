import pandas as pd
import numpy as np
import os
import openpyxl
from loguru import logger
from datetime import date,datetime,timedelta
from itertools import product
from openpyxl import Workbook, formatting
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScale, ColorScaleRule, FormatObject
from openpyxl.utils import get_column_letter

from wb.scripts.constants import marketplace_dir_name, net_cost_koef

def format_excel_actions(
        client_name,
        svod_excel_actions,
        date_report_created
):
    # Путь к файлу с отчетом по акциям
    filepath_actions_svod = (
        f"{marketplace_dir_name}/Clients/{client_name}/Actions/"
        f"{date_report_created}_Таблица_по_акциям_{client_name}_WB.xlsx"
    )
    # Переводим дату создания свода из строки в дату
    date_report_created_ = datetime.strptime(date_report_created, '%Y-%m-%d')
    logger.info(f"Formatting action svod for client {client_name}")

    # Открываем файл с расчетами Excel
    wb = openpyxl.load_workbook(filepath_actions_svod)
    ws = wb['Акции']

    # Кол-во строк в df
    svod_len = len(svod_excel_actions)

    # Номер строки, откуда начинается запись
    row_start = 2

    # df с соответствием заголовков и названий столбцов
    excel_columns = pd.DataFrame({"column": svod_excel_actions.columns,
                                  "column_number": np.arange(1, len(svod_excel_actions.columns) + 1)})
    excel_columns['excel_column'] = excel_columns['column_number'].apply(lambda x: get_column_letter(x))

    # Вычисление даты предыдущего дня
    date_report_created_ = datetime.strptime(date_report_created, "%Y-%m-%d")
    date_next_day = date_report_created_ + timedelta(days=1)

    # Выборка различных групп колонок по их названиям
    # Мин. и макс. колонка
    min_col = excel_columns.loc[excel_columns['column_number'].idxmin(), 'excel_column']
    max_col = excel_columns.loc[excel_columns['column_number'].idxmax(), 'excel_column']
    # Заголовки
    header_cells = ws[f"{min_col}{row_start - 1}:{max_col}{row_start - 1}"]
    # Все колонки, кроме заголовков
    all_cells = ws[f"{min_col}{row_start}:{max_col}{svod_len + 1}"]
    # Штрихкод
    barcode = excel_columns.loc[excel_columns['column'] == 'Баркод', 'excel_column'].values[0]
    # РРЦ
    marketing_price_col = excel_columns.loc[excel_columns['column'] == 'РРЦ', 'excel_column'].values[0]
    # Себестоимость
    net_cost = excel_columns.loc[excel_columns['column'] == 'Себестоимость', 'excel_column'].values[0]
    # Скидка
    discount_col = excel_columns.loc[excel_columns['column'] == f"Скидка WB {date_report_created_.strftime('%d.%m')}", 'excel_column'].values[0]
    # Цена до скидки
    price_before_discount = excel_columns.loc[excel_columns['column'] == 'Цена до скидки', 'excel_column'].values[0]
    # Цена после скидки
    price_after_discount = excel_columns.loc[excel_columns['column'] == 'Цена после скидки', 'excel_column'].values[0]
    # Скидка до РРЦ
    discount_to_marketing_price = excel_columns.loc[excel_columns['column'] == 'Скидка до РРЦ', 'excel_column'].values[0]
    # Мин. цена
    min_price = excel_columns.loc[excel_columns['column'] == 'Минимальная цена маржинальная, руб.', 'excel_column'].values[0]
    # Max скидка, %
    # max_discount = excel_columns.loc[excel_columns['column'] == 'Max скидка, %', 'excel_column'].values[0]
    # Цена продажи без WB кошелек
    # price_no_wb_wallet = excel_columns.loc[excel_columns['column'] == 'Цена продажи без WB кошелек', 'excel_column'].values[0]
    # Ожидаемое количество на складе
    expected_on_warehouse = excel_columns.loc[excel_columns['column'] == 'Ожидаемое количество на складе', 'excel_column'].values[0]
    # Остатки
    remind_current_day = excel_columns.loc[excel_columns['column'] == f"Ост. {date_report_created_.strftime('%d.%m')}", 'excel_column'].values[0]
    # remind_next_day = excel_columns.loc[excel_columns['column'] == f"Ост. {date_next_day.strftime('%d.%m')}", 'excel_column'].values[0]
    # Итоговая цена и скидка
    # result_price = excel_columns.loc[excel_columns['column'] == "Итоговая цена", 'excel_column'].values[0]
    # result_discount = excel_columns.loc[excel_columns['column'] == "Итоговая скидка", 'excel_column'].values[0]

    # Колонки Участие в акции n
    action_participate_cols = excel_columns.loc[excel_columns['column'].str.contains('Участие в акции '), 'excel_column'].to_list()
    # Колонки Скидка по акции n
    action_discount_cols = excel_columns.loc[excel_columns['column'].str.contains('Скидка по акции '), 'excel_column'].to_list()
    # Колонки Цена по акции n
    action_price_cols = excel_columns.loc[excel_columns['column'].str.contains('Цена по акции '), 'excel_column'].to_list()
    # Колонки Скидка от РРЦ по акции n
    action_discount_from_marketing_price_cols = excel_columns.loc[excel_columns['column'].str.contains('Скидка от РРЦ по акции '), 'excel_column'].to_list()
    # Колонки Разница до мин. цены по акции n
    action_diff_to_min_price_cols = excel_columns.loc[excel_columns['column'].str.contains('Разница до мин. цены по акции '), 'excel_column'].to_list()

    # Формат заголовков
    thin_border = Side(border_style="thin", color="000000")
    for row in header_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11, bold=True)
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center',
                                       wrap_text=True)
            cell.border = Border(top = thin_border, bottom = thin_border,
                                 right = thin_border, left = thin_border)

    # Границы (сетка)
    for row in all_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11)
            cell.border = Border(top = thin_border, bottom = thin_border,
                                                        right = thin_border, left = thin_border)

    # Значок фильтра на столбцы
    ws.auto_filter.ref = ws.dimensions

    # Автоподбор ширины столбца
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # У столбца с баркодом делаем фиксированную ширину,
    # т.к. там бывает много значений
    ws.column_dimensions[barcode].width = 22

    # Формулы
    for i in range(row_start, svod_len + row_start):
        # Скидка до ррц
        ws[f"{discount_to_marketing_price}{i}"] = f"""= IFERROR(({price_before_discount}{i} - {marketing_price_col}{i}) / {marketing_price_col}{i}, "")"""

    # Формула разницы до мин. цены по акции
    for excel_column in zip(action_diff_to_min_price_cols, action_price_cols, action_participate_cols):
        for i in range(row_start, svod_len + row_start):
            # Если товар есть в списке на участие в акции и есть мин цена, считаем разницу до мин. цены
            # В остальных случаях оставляем разницу до мин. цены пустой
            if (ws[f"{excel_column[2]}{i}"].value is not None) & (ws[f"{min_price}{i}"].value is not None):
                ws[f"{excel_column[0]}{i}"] = f"= {excel_column[1]}{i} - {min_price}{i}"

    # Формула расчета разницы до РРЦ для каждой из акций (их может быть несколько)
    for action_column in zip(action_discount_from_marketing_price_cols, action_price_cols):
        for i in range(row_start, svod_len + row_start):
            if (ws[f"{action_column[1]}{i}"].value is not None) & (ws[f"{marketing_price_col}{i}"].value is not None):
                ws[f"{action_column[0]}{i}"] = f"= ({marketing_price_col}{i} - {action_column[1]}{i}) / {marketing_price_col}{i}"

    # Выравнивание по центру
    middle_cells = excel_columns.loc[~excel_columns['column'].str.contains('Баркод|Предмет|Артикул продавца|Наименование'), 'excel_column'].to_list()
    for i in range(len(middle_cells)):
        middle_cells[i] = f"{middle_cells[i]}{row_start}:{middle_cells[i]}{svod_len + 1}"
    for cell_range in middle_cells:
        cell_range = ws[cell_range]
        for row in cell_range:
            for cell in row:
                cell.alignment = Alignment(horizontal = "center",
                                                       vertical = "center")

    # Выравнивание слева
    left_cells = excel_columns.loc[excel_columns['column'].str.contains('Баркод|Предмет|Артикул продавца|Наименование'), 'excel_column'].to_list()
    for i in range(len(left_cells)):
        left_cells[i] = f"{left_cells[i]}{row_start}:{left_cells[i]}{svod_len + 1}"
    for cell_range in left_cells:
        cell_range = ws[cell_range]
        for row in cell_range:
            for cell in row:
                cell.alignment = Alignment(horizontal = "left",
                                                       vertical = "center")

    # Формат чисел
    # 0%
    percent_0_digit_cells = []
    percent_0_digit_cells.extend(action_discount_cols + [discount_col])
    for i in range(len(percent_0_digit_cells)):
        percent_0_digit_cells[i] = f"{percent_0_digit_cells[i]}{row_start}:{percent_0_digit_cells[i]}{svod_len + 1}"
    for cell_range in percent_0_digit_cells:
        cell_range = ws[cell_range]
        for row in cell_range:
            for cell in row:
                cell.number_format = '0%'

    # 0.00%
    percent_2_digit_cells = []
    percent_2_digit_cells.extend(action_discount_from_marketing_price_cols + [discount_to_marketing_price])
    for i in range(len(percent_2_digit_cells)):
        percent_2_digit_cells[i] = f"{percent_2_digit_cells[i]}{row_start}:{percent_2_digit_cells[i]}{svod_len + 1}"
    for cell_range in percent_2_digit_cells:
        cell_range = ws[cell_range]
        for row in cell_range:
            for cell in row:
                cell.number_format = '0.00%'

    # 0
    number_0_digit_cells = []
    number_0_digit_cells.extend([min_price, net_cost, price_before_discount, barcode])
    for i in range(len(number_0_digit_cells)):
        number_0_digit_cells[i] = f"{number_0_digit_cells[i]}{row_start}:{number_0_digit_cells[i]}{svod_len + 1}"

    for cell_range in number_0_digit_cells:
        cell_range = ws[cell_range]
        for row in cell_range:
            for cell in row:
                # Пробуем перевести текст в число
                if isinstance(cell.value, str):
                    try:
                        # Пробуем преобразовать строку в число
                        number = float(cell.value.replace(' ', '').replace(',', '.')) # если вдруг точка или запятая
                        cell.value = number
                        cell.number_format = '0'  # 0 знаков после запятой
                    except ValueError:
                        pass  # если не получилось, оставляем как есть


    # Условное форматирование
    red_font = Font(color='9C0006')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_font = Font(color='006100')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    # Колонки, которые нужно форматировать в зависимости от знака числа
    sign_format_cell_range = action_diff_to_min_price_cols + action_discount_from_marketing_price_cols
    for i in range(len(sign_format_cell_range)):
        sign_format_cell_range[i] = f"{sign_format_cell_range[i]}{row_start}:{sign_format_cell_range[i]}{svod_len + 1}"
    for cell_range in sign_format_cell_range:
        # cell_range_ws = ws[cell_range]
        ws.conditional_formatting.add(cell_range,
                                    formatting.rule.CellIsRule(
                                        operator='>',
                                        formula=[0],
                                        fill = green_fill, font = green_font,
                                    ))
        ws.conditional_formatting.add(cell_range,
                                    formatting.rule.CellIsRule(
                                        operator='<',
                                        formula=[0],
                                        fill = red_fill, font = red_font,
                                    ))

    # Колонки участия в акциях
    action_participate_cell_range = action_participate_cols.copy()
    for i in range(len(action_participate_cell_range)):
        action_participate_cell_range[i] = f"{action_participate_cell_range[i]}{row_start}:{action_participate_cell_range[i]}{svod_len + 1}"
    for cell_range in action_participate_cell_range:
        # cell_range_ws = ws[cell_range]
        ws.conditional_formatting.add(cell_range,
                                    formatting.rule.CellIsRule(
                                        operator='equal',
                                        formula=['\"Да\"'],
                                        fill = green_fill, font = green_font,
                                    ))
        ws.conditional_formatting.add(cell_range,
                                    formatting.rule.CellIsRule(
                                        operator='equal',
                                        formula=['\"Нет\"'],
                                        fill = red_fill, font = red_font,
                                    ))
    # Колонка РРЦ
    marketing_price_cell_range = [f"{marketing_price_col}{row_start}:{marketing_price_col}{svod_len + 1}"]
    for cell_range in marketing_price_cell_range:
        # cell_range_ws = ws[cell_range]
        ws.conditional_formatting.add(cell_range,
                                    formatting.rule.Rule(
                                        type='containsBlanks',
                                        # formula=[' '],
                                        dxf=DifferentialStyle(fill = red_fill, font = red_font)
                                    ))
    # Заливка
    gold_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF64', end_color='FFFF64', fill_type='solid')
    blue_fill = PatternFill(start_color='CFE2F3', end_color='CFE2F3', fill_type='solid')
    light_red_fill = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
    light_orange_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    beige_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    # Заливка для некоторых заголовков
    gold_fill_cells = excel_columns.loc[excel_columns['column'].str.contains('№|Артикул продавца|Предмет|Наименование|Баркод|Размер|Цвет|Статус|РРЦ|Себестоимость'), 'excel_column'].to_list()
    yellow_fill_cells = [discount_to_marketing_price] # + [result_price] + [result_discount]
    blue_fill_cells = action_discount_cols
    light_red_fill_cells = []
    # light_orange_fill_cells = [price_no_wb_wallet]
    beige_fill_cells = [expected_on_warehouse]
    # Заливка золотым цветом
    for col_name in gold_fill_cells:
        ws[f"{col_name}{row_start - 1}"].fill = gold_fill
    # Заливка желтым цветом
    for col_name in yellow_fill_cells:
        cell_range = ws[f"{col_name}{row_start - 1}:{col_name}{svod_len + 1}"]
        for row in cell_range:
            for cell in row:
                cell.fill = yellow_fill
    # Заливка синим цветом
    for col_name in blue_fill_cells:
        cell_range = ws[f"{col_name}{row_start - 1}:{col_name}{svod_len + 1}"]
        for row in cell_range:
            for cell in row:
                cell.fill = blue_fill
    # Заливка бардовым цветом
    for col_name in light_red_fill_cells:
        cell_range = ws[f"{col_name}{row_start - 1}:{col_name}{svod_len + 1}"]
        for row in cell_range:
            for cell in row:
                cell.fill = light_red_fill
    # Заливка оранжевым цветом
    # for col_name in light_orange_fill_cells:
    #     cell_range = ws[f"{col_name}{row_start - 1}:{col_name}{svod_len + 1}"]
    #     for row in cell_range:
    #         for cell in row:
    #             cell.fill = light_orange_fill
    # Заливка бежевым цветом
    for col_name in beige_fill_cells:
        cell_range = ws[f"{col_name}{row_start - 1}:{col_name}{svod_len + 1}"]
        for row in cell_range:
            for cell in row:
                cell.fill = beige_fill

    # Специфическое форматирование для отдельных клиентов
    if client_name in ['KU_And_KU', 'Soyuz', 'TRIBE', 'Orsk_Combinat']:
        # Маржинальность руб по акции n
        action_marginality_rub_cols = excel_columns.loc[excel_columns['column'].str.contains('Расчетная маржа, руб по акции '), 'excel_column'].to_list()
        # Маржинальность % по акции n
        action_marginality_percent_cols = excel_columns.loc[excel_columns['column'].str.contains('Расчетная маржа, % по акции '), 'excel_column'].to_list()

        # Формула расчета маржинальности по акции
        for excel_column in zip(action_price_cols, action_marginality_rub_cols, action_marginality_percent_cols):
            for i in range(row_start, svod_len + row_start):
                # Если товар есть в списке на участие в акции и есть себестоимость, считаем маржинальность
                # В остальных случаях оставляем маржинальность по акции пустой
                if (ws[f"{excel_column[0]}{i}"].value is not None) & (ws[f"{net_cost}{i}"].value is not None):
                    ws[f"{excel_column[1]}{i}"] = f"= {excel_column[0]}{i} - {net_cost_koef} * {excel_column[0]}{i} - {net_cost}{i}"
                    ws[f"{excel_column[2]}{i}"] = f"= {excel_column[1]}{i} / {excel_column[0]}{i}"

        # Формат числа 0% для колонки с маржинальностью
        action_marginality_percent_cell_range = action_marginality_percent_cols.copy()
        for i in range(len(action_marginality_percent_cell_range)):
            action_marginality_percent_cell_range[i] = f"{action_marginality_percent_cell_range[i]}{row_start}:{action_marginality_percent_cell_range[i]}{svod_len + 1}"

        for cell_range in action_marginality_percent_cell_range:
            cell_range = ws[cell_range]
            for row in cell_range:
                for cell in row:
                    cell.number_format = '0%'

        # Условное форматирование
        # Колонки с маржинальностью
        action_marginality_cell_range = action_marginality_percent_cols + action_marginality_rub_cols
        for i in range(len(action_marginality_cell_range)):
            action_marginality_cell_range[i] = f"{action_marginality_cell_range[i]}{row_start}:{action_marginality_cell_range[i]}{svod_len + 1}"
        for cell_range in action_marginality_cell_range:
            # cell_range_ws = ws[cell_range]
            ws.conditional_formatting.add(cell_range,
                                        formatting.rule.CellIsRule(
                                            operator='>',
                                            formula=[0],
                                            fill = green_fill, font = green_font,
                                        ))
            ws.conditional_formatting.add(cell_range,
                                        formatting.rule.CellIsRule(
                                            operator='<',
                                            formula=[0],
                                            fill = red_fill, font = red_font,
                                        ))


    # Сохранение
    wb.save(f"{filepath_actions_svod}_Formatted.xlsx")
    logger.info("Finished formatting actions file")



# Файл с настройками и номером клиента
# from options import settings, headers, client_number
# date_report_created = '2024-11-21'
# svod_excel_actions = pd.read_excel(f"Clients/{client_name}/Actions/{date_report_created}_Таблица_по_акциям_{client_name}_WB.xlsx")
# report_dates = pd.read_csv(f"Clients/{client_name}/UploadFiles/UploadFiles_{date_report_created}/{date_report_created}_dates_from_to.csv", sep=';')
# for col in report_dates:
#     report_dates[col] = pd.to_datetime(report_dates[col])
# date_start = report_dates['date_start_file'][0]
# date_end = report_dates['date_end_file'][0]

# format_excel_actions(settings, client_number, svod_excel_actions, date_report_created)
