
# %% Определение функций
import requests
import json
import time
from datetime import date,datetime,timedelta
import pandas as pd
import shutil
import os
import glob
import csv
import numpy as np
import re
from loguru import logger
import getopt
import sys
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color

pd.options.mode.chained_assignment = None  # default='warn'

# Некоторые константы
from wb.scripts.constants import (
    headers,
    client_name,
    marketplace_dir_name,
    catalog_finace_svod_columns

)

# Уже написанные ранее функции выгрузки данных из АПИ
from wb.scripts.uploadDataFromWB import getWBProduct, get_prices_WB, getOrdersWB

# Функция выгрузки для АПИ Продвижение
from wb.scripts.uploadDataPerformanceWB import get_costs_history

# Вспомогательные функции
from generic_functions import move_columns

# Функция выгрузки отчета о реализации по апи
def upload_realization_report(headers, date_start, date_end):
    logger.info(f"Uploading realization report for client {client_name} for dates {date_start} - {date_end}")

    # Параметры запроса
    params_realization_report = {
    "dateFrom": date_start,
    "dateTo": date_end,
    "rrdid": 0
}
    # Запрос к апи
    resp_data_realization_report = requests.get("https://statistics-api.wildberries.ru/api/v5/supplier/reportDetailByPeriod",
                                                headers=headers, params=params_realization_report).json()
    # Создаем df с отчетом
    df_realization_report = pd.DataFrame(resp_data_realization_report)
    # Добавляем столбцы года и месяца, за который был сформирован отчет
    df_realization_report['year'] = datetime.strptime(date_start, '%Y-%m-%dT%H:%M:%S').year
    df_realization_report['month'] = datetime.strptime(date_start, '%Y-%m-%dT%H:%M:%S').month
    # Добавляем столбец с id строки
    df_realization_report['id'] = np.arange(1, df_realization_report.shape[0] + 1)
    # Перемещаем столбец с id в начало отчета
    df_realization_report.insert(0, 'id', df_realization_report.pop('id'))

    logger.info("Finished uploading realization report")

    return df_realization_report


# Функция выгрузки отчета о реализации по апи
def upload_realization_report_v2(headers, date_start, date_end):

    # df, куда будем помещать итоговый результат
    df_realization_report = pd.DataFrame()
    # df, возвращаемый в методе API (делаем его не пустым для начала цикла)
    df_realization_report_api = pd.DataFrame({'A': [1]})
    rrd_id = 0

    # Если количество строк отчета превышает 100к,
    # то подключаем логику выгрузки отчета частями
    # if df_realization_report_api.shape[0] >= 100000:
    # Цикл до тех пор, пока выгружаемый отчет не будет пустым
    while not df_realization_report_api.empty:
        # Параметры запроса
        params_realization_report = {
            "dateFrom": date_start,
            "dateTo": date_end,
            "rrdid": rrd_id
        }
        logger.info(
            f"\nUploading realization report for client\n"
            f"{client_name}\n"
            f"for dates\n"
            f"{date_start} - {date_end}\n"
            f"rrd_id: {rrd_id}"
        )

        # Запрос к апи
        resp_data_realization_report = (
            requests
            .get(
                "https://statistics-api.wildberries.ru/api/v5/supplier/reportDetailByPeriod",
                headers=headers,
                params=params_realization_report
            )
            .json()
        )
        # Создаем df с отчетом
        df_realization_report_api = pd.DataFrame(resp_data_realization_report)
        # Добавляем столбцы года и месяца, за который был сформирован отчет
        df_realization_report_api['year'] = datetime.strptime(date_start, '%Y-%m-%dT%H:%M:%S').year
        df_realization_report_api['month'] = datetime.strptime(date_start, '%Y-%m-%dT%H:%M:%S').month
        # Добавляем столбец с id строки
        df_realization_report_api['id'] = np.arange(1, df_realization_report_api.shape[0] + 1)
        # Перемещаем столбец с id в начало отчета
        df_realization_report_api.insert(0, 'id', df_realization_report_api.pop('id'))
        # Объединяем с предыдущим проходом цикла
        df_realization_report = pd.concat([
            df_realization_report,
            df_realization_report_api
        ])

        # Число строк отчета, которое вернулось из API
        report_row_length = df_realization_report_api.shape[0]
        logger.info(f"Realization report rows: {report_row_length}")
        # Если количество строк отчета превышает 100к,
        # то подключаем логику выгрузки отчета частями
        if report_row_length >= 100000:
            logger.info(f"Returned 100 000 rows, uploading realization report by parts")
            # Получаем строку отчета, которую будем передавать в следующий запрос
            rrd_id = int(df_realization_report_api['rrd_id'].iloc[-1])
                # Ждем одну минуту перед следующим запросом
            logger.info(f"Waiting 1 minute before next request")
            time.sleep(60)
        # Если отчет пришел пустой, то выгружены все строки, выходим из цикла
        else:
            break

    # Сбрасываем index после concat
    df_realization_report = df_realization_report.reset_index(drop=True)

    logger.info("Finished uploading realization report")

    return df_realization_report


# Функция чтения минимальной и максимальной даты отчета
def read_report_dates(df_realization_report):
    # Выбираем мин. и макс. даты из отчета о реализации
    date_start_ = min(df_realization_report['date_from'])
    date_end_ = max(df_realization_report['date_to'])
    # Создаем нужный формат даты (нужна для некоторых методов апи)
    date_start_api = date_start_ + 'T00:00:00'
    date_end_api = date_end_ + 'T23:59:59'
    # Создаем df с датами
    df_report_dates = pd.DataFrame({'date_start': date_start_,
                                    'date_end': date_end_,
                                    'date_start_api': date_start_api,
                                    'date_end_api': date_end_api},
                                    index=[0])

    return df_report_dates


# Функция переименования столбцов отчета по реализации на русский
def rename_report_columns(df_realization_report):
    # Создаем копию для избежания изменений в оригинальном df
    df_realization_report_ = df_realization_report.copy()
    # Считывание файла с соответствием колонки из апи её описанию из документации
    api_report_columns = pd.read_csv(f'{marketplace_dir_name}/scripts/Столбцы отчета о реализации из апи.csv', sep=';')
    # Создаем df с колонками из отчета по апи
    df_realization_report_columns = pd.DataFrame({'api_name': df_realization_report_.columns})
    # Мерджим с колонками отчета, который выгрузили по апи
    all_columns = df_realization_report_columns.merge(api_report_columns,
                                                            how='outer',
                                                            on='api_name',
                                                            indicator=True)
    # Находим общие колонки
    columns_to_rename = all_columns.loc[all_columns['_merge'] == 'both', ['api_name', 'ru_name']]
    # Переименовываем колонки
    df_realization_report_ = df_realization_report_.rename(columns=columns_to_rename.set_index('api_name')['ru_name'])
    # Старый способ переименовывания колонок через цикл по каждой колонке
    # for i in range(api_report_columns.shape[0]):
    #     if api_report_columns['api_name'][i] in df_realization_report.columns_:
    #         df_realization_report_ = df_realization_report_.rename(columns={api_report_columns['api_name'][i]: api_report_columns['ru_name'][i]})

    df_realization_report_renamed = df_realization_report_.copy()

    return df_realization_report_renamed


# Функция создания доп. колонок для расчетов
def add_new_columns(df_realization_report_renamed):
    # Создаем копию для избежания изменений в оригинальном df
    df_realization_report_new_columns = df_realization_report_renamed.copy()

    # Создание колонок, которые не всегда приходят с апи
    variable_api_columns = ['Обоснование штрафов и доплат', 'Возмещение издержек по перевозке']
    for col in variable_api_columns:
        if col not in df_realization_report_new_columns.columns:
            df_realization_report_new_columns[col] = 0

    # Перевод в datetime некоторых колонок с датой
    for col in ['Дата начала отчётного периода', 'Дата конца отчётного периода', 'Дата заказа', 'Дата продажи', 'Дата операции']:
        # Переводим в datetime
        df_realization_report_new_columns[col] = pd.to_datetime(df_realization_report_new_columns[col], format='mixed')
        # Убираем миллисекунды
        df_realization_report_new_columns[col] = df_realization_report_new_columns[col].dt.floor('s')
        # Убираем тайм зоны
        df_realization_report_new_columns[col] = df_realization_report_new_columns[col].dt.tz_localize(None)
    # Год и месяц из даты продажи
    df_realization_report_new_columns['year_date_sold'] = df_realization_report_new_columns['Дата продажи'].dt.year
    df_realization_report_new_columns['month_date_sold'] = df_realization_report_new_columns['Дата продажи'].dt.month

    # Переносим строки с датами в начало df
    df_realization_report_new_columns = move_columns(df_realization_report_new_columns,
                                                     ['Дата начала отчётного периода', 'Дата конца отчётного периода',
                                                      'Дата заказа', 'Дата продажи','Дата операции',
                                                      'Дата формирования отчёта',
                                                      'Дата начала действия фиксации', 'Дата окончания действия фиксации',
                                                      ],
                                                      position='id',
                                                      insert_type='after')
    # Колонки, по которым считаем прочие расходы
    other_costs_columns = [
        'Стоимость логистики',
        'К перечислению продавцу за реализованный товар',
        'Возмещение за выдачу и возврат товаров на ПВЗ',
        'Возмещение издержек по эквайрингу',
        'Размер комиссии за эквайринг без НДС, %',
        'Вознаграждение WB без НДС',
        'НДС с вознаграждения WB',
        'Возмещение издержек по перевозке',
        'Стоимость хранения',
        'Прочие удержания/выплаты',
        'Стоимость платной приёмки',
        'Штрафы',
        'Доплаты'
    ]

    # Итоговая колонка с прочими расходами
    df_realization_report_new_columns['Итого руб'] = df_realization_report_new_columns.loc[:, other_costs_columns].sum(axis=1)

    return df_realization_report_new_columns

# Функция исключения переходных недель из отчета
def filter_report_dates(date_start, date_end, df_realization_report_new_columns, filter_dates=False):
    # Создаем копию для избежания изменений в оригинальном df
    df_realization_report_date_filtered = df_realization_report_new_columns.copy()
    # Если указан флаг фильтра по датам, то делаем фильтр по датам
    if filter_dates:
        # Переводим даты выгрузки в Timestamp
        dt_start = pd.to_datetime(date_start, format='mixed')
        dt_end = pd.to_datetime(date_end, format='mixed')
        # Ищем id строк переходных недель
        tmp_df_transitional_weeks = df_realization_report_date_filtered.loc[
            (df_realization_report_date_filtered['Дата начала отчётного периода'] < dt_start) | (df_realization_report_date_filtered['Дата конца отчётного периода'] > dt_end),
            :]
        # Внутри переходных недель ищем id строк, у которых дата продажи не соответствует датам начала и окончания выгрузки
        transitional_weeks_ids = tmp_df_transitional_weeks.loc[(tmp_df_transitional_weeks['Дата продажи'] < dt_start) | (tmp_df_transitional_weeks['Дата продажи'] > dt_end), 'id'].to_list()
        # Убираем найденные строки из отчета
        df_realization_report_date_filtered = df_realization_report_date_filtered.loc[~df_realization_report_date_filtered['id'].isin(transitional_weeks_ids), :]
        # Делаем фильтр по дате продажи в пределах указанного диапазона
        # df_realization_report_date_filtered = df_realization_report_date_filtered.loc[(df_realization_report_date_filtered['Дата продажи'] >= dt_start) & (df_realization_report_date_filtered['Дата продажи'] <= dt_end), :]
        # Делаем reset_index после concat
        df_realization_report_date_filtered = df_realization_report_date_filtered.reset_index(drop=True)

    return df_realization_report_date_filtered



# Функция получения дат отчета
def get_report_dates(df_realization_report_date_filtered):
    # Создание копии для избежания изменений в оригинальном df
    df_dates = df_realization_report_date_filtered.copy()
    # Убираем время в Дате продажи, иначе получается очень большой список
    df_dates['Дата продажи (без времени)'] = df_dates['Дата продажи'].dt.date
    # Формируем df со списком дат в отчете
    report_dates = (
        df_dates
        .groupby(['Дата начала отчётного периода', 'Дата конца отчётного периода', 'Дата продажи (без времени)'])
        .size()
        .reset_index()
        .drop(columns=0)
    )

    return report_dates


# Функция выгрузки заказов
def get_orders(headers, date_start, date_end, to_save=False):
    # Выгружаем отчет по заказам из АПИ
    df_orders_api = getOrdersWB(headers, date_start, date_end, to_save)
    # Создаем колонку, которую считаем за заказы
    df_orders_api['Заказы шт'] = 1
    # Выбираем нужные колонки
    df_orders = df_orders_api.loc[:, df_orders_api.columns.isin([
        'date',
        'Артикул WB',
        'Артикул продавца',
        'Размер',
        'Заказы шт'
        ])
    ]

    # Переводим артикул продавца с верхний регистр
    df_orders['Артикул продавца'] = df_orders['Артикул продавца'].str.upper()

    return df_orders


# Функция расчета заказов по Артикулу и Размеру
def calc_orders_by_size(df_orders):
    df_orders_stats = (
        df_orders
        .loc[:, ['Артикул WB', 'Артикул продавца', 'Размер', 'Заказы шт']]
        .groupby(['Артикул WB', 'Артикул продавца', 'Размер'])
        .agg(**{
            'Заказы шт': ('Заказы шт', 'sum')
        })
        .reset_index()
    )

    return df_orders_stats

# Функция расчета расходов, которые уже разделены по артикулам в отчете апи
def calc_sku_costs(df_realization_report_date_filtered, groupby_cols = ['Артикул WB', 'Артикул продавца', 'Размер']):
    # Выбираем расходы, где есть артикул
    df_sku_costs = df_realization_report_date_filtered.loc[df_realization_report_date_filtered['Артикул продавца'] != '', :]

    # Колонки, по которым будет производиться группировка
    # groupby_cols = ['Артикул WB', 'Артикул продавца', 'Размер']

    # df, в котором будут считаться расходы
    df_sku_sizes_costs_stats = pd.DataFrame()

    # Временная колонка с размерами групп
    df_sku_sizes_costs_stats['tmp_col'] = df_sku_costs.groupby(groupby_cols).size()

    # Колонка, по которой считаем продажи и возвраты в рублях
    if client_name in ['Orsk_Combinat']:
        sales_returns_col = 'Цена розничная с учетом согласованной скидки'
    else:
        sales_returns_col = 'Сумма продаж (возвратов)'

    # Расчет продаж
    sales_ids = df_sku_costs.loc[df_sku_costs['Обоснование для оплаты'] == 'Продажа', 'id'].to_list()
    df_sku_sizes_costs_stats['Продажи шт'] = df_sku_costs.loc[df_sku_costs['id'].isin(sales_ids), :] \
        .groupby(groupby_cols) \
        ['Количество'].sum()
    df_sku_sizes_costs_stats['Продажи руб'] = df_sku_costs.loc[df_sku_costs['id'].isin(sales_ids), :] \
        .groupby(groupby_cols) \
        [sales_returns_col].sum()

    # Расчет возвратов
    returns_ids = df_sku_costs.loc[df_sku_costs['Обоснование для оплаты'] == 'Возврат', 'id'].to_list()
    df_sku_sizes_costs_stats['Возвраты шт'] = df_sku_costs.loc[df_sku_costs['id'].isin(returns_ids), :] \
        .groupby(groupby_cols) \
        ['Количество'].sum()
    df_sku_sizes_costs_stats['Возвраты руб'] = df_sku_costs.loc[df_sku_costs['id'].isin(returns_ids), :] \
        .groupby(groupby_cols) \
        [sales_returns_col].sum()

    # Считаем итоговы суммы по Продажам и Возвратам (без разбивки по Артикулам)
    df_sku_sales_returns_by_operation = (
        df_sku_costs
        .loc[df_sku_costs['id'].isin(sales_ids + returns_ids), :]
        .groupby('Обоснование для оплаты')
        .agg(**{
            'Расходы руб': (sales_returns_col, 'sum')
        })
        .reset_index()
        .rename(columns={'Обоснование для оплаты': 'Группа расходов'})
    )

    # Заполняем пропуски
    df_sku_sizes_costs_stats = df_sku_sizes_costs_stats.fillna(0)

    # Итоговые колонки по продажам и возвратам
    df_sku_sizes_costs_stats['Продажи минус возвраты шт'] = df_sku_sizes_costs_stats['Продажи шт'] - df_sku_sizes_costs_stats['Возвраты шт']
    df_sku_sizes_costs_stats['Продажи минус возвраты руб'] = df_sku_sizes_costs_stats['Продажи руб'] - df_sku_sizes_costs_stats['Возвраты руб']

    # Считаем комиссию WB
    df_sku_sizes_costs_stats['Вознаграждение WB без НДС'] = df_sku_costs \
        .groupby(groupby_cols) \
        ['Вознаграждение WB без НДС'].sum()
    df_sku_sizes_costs_stats['НДС с вознаграждения WB'] = df_sku_costs \
        .groupby(groupby_cols) \
        ['НДС с вознаграждения WB'].sum()
    df_sku_sizes_costs_stats['Комиссия WB, %'] = (df_sku_sizes_costs_stats['Вознаграждение WB без НДС'] + df_sku_sizes_costs_stats['НДС с вознаграждения WB']) / df_sku_sizes_costs_stats['Продажи минус возвраты руб'] * 100


    # Заменяем деление на 0 на пропуски
    df_sku_sizes_costs_stats['Комиссия WB, %'] = df_sku_sizes_costs_stats['Комиссия WB, %'].replace(np.inf, np.nan).replace(-np.inf, np.nan)


    # Считаем суммы по Штрафам и Доплатам (по ним есть доп. детализация в столбце Обоснование штрафов и доплат)
    penalty_costs_ids = df_sku_costs.loc[df_sku_costs['Обоснование для оплаты'].isin(['Штрафы и доплаты', 'Штраф']), 'id'].to_list()
    df_sku_penalty_costs = df_sku_costs.loc[df_sku_costs['id'].isin(penalty_costs_ids), :]
    # Тут берем столбец 'Обоснование штрафов и доплат', а не 'Обоснование для оплаты', т.к. в нем есть доп. детализация
    for operation_type in df_sku_penalty_costs['Обоснование штрафов и доплат'].unique():
        df_sku_sizes_costs_stats[operation_type] = df_sku_costs.loc[df_sku_costs['Обоснование штрафов и доплат'] == operation_type, :] \
        .groupby(groupby_cols) \
        ['Итого руб'].sum()
    # Список с названиями и затратами по Штрафам и Доплатам
    df_sku_penalty_costs_by_operation = (
        df_sku_penalty_costs
        .groupby('Обоснование штрафов и доплат')
        .agg(**{
            'Расходы руб': ('Итого руб', 'sum')
        })
        .reset_index()
        .rename(columns={'Обоснование штрафов и доплат': 'Группа расходов'})
    )

    # Считаем суммы по оставшимся расходам
    other_costs_ids = df_sku_costs.loc[~df_sku_costs['id'].isin(sales_ids + returns_ids + penalty_costs_ids), 'id'].to_list()
    df_sku_other_costs = df_sku_costs.loc[df_sku_costs['id'].isin(other_costs_ids), :]
    for operation_type in df_sku_other_costs['Обоснование для оплаты'].unique():
        df_sku_sizes_costs_stats[operation_type] = df_sku_costs.loc[df_sku_costs['Обоснование для оплаты'] == operation_type, :] \
        .groupby(groupby_cols) \
        ['Итого руб'].sum()
    # Список с названиями и затратами по оставшимся расходам
    df_sku_other_costs_by_operation = (
        df_sku_other_costs
        .groupby('Обоснование для оплаты')
        .agg(**{
            'Расходы руб': ('Итого руб', 'sum')
        })
        .reset_index()
        .rename(columns={'Обоснование для оплаты': 'Группа расходов'})
    )

    # Удаляем временную колонку
    df_sku_sizes_costs_stats = df_sku_sizes_costs_stats.drop(columns=['tmp_col'])

    # Достаем доп. столбцы из index
    df_sku_sizes_costs_stats = df_sku_sizes_costs_stats.reset_index()

    # Переводим артикул продавца с верхний регистр
    df_sku_sizes_costs_stats['Артикул продавца'] = df_sku_sizes_costs_stats['Артикул продавца'].str.upper()


    # Перемещаем колонки в начало df
    # df_sku_sizes_costs_stats = move_columns(df_sku_sizes_costs_stats,
    #                                   ['Количество размеров в артикуле', 'Продажи в артикуле', 'Доля продаж в артикуле'],
    #                                    position='Продажи шт',
    #                                    insert_type='before'
    # )

    # Формируем list с id строк расходов по артикулам
    sku_costs_ids = sales_ids + returns_ids + penalty_costs_ids + other_costs_ids

    # Формируем df со списком расходов, которые привязаны к Артикулам
    df_sku_costs_by_operation = pd.concat([df_sku_sales_returns_by_operation, df_sku_penalty_costs_by_operation, df_sku_other_costs_by_operation]).reset_index(drop=True)

    # Формируем словарь с результатами
    result_sku_costs = {'df_sku_sizes_costs_stats': df_sku_sizes_costs_stats,
                        'df_sku_costs_by_operation': df_sku_costs_by_operation,
                        'sku_costs_ids': sku_costs_ids}

    return result_sku_costs



# Функция добавления заказов в отчет о реализации
def add_orders_to_sku_costs(result_sku_costs, df_orders):
    # Достаем df с расходами по SKU
    df_sku_costs_stats = result_sku_costs['df_sku_sizes_costs_stats'].copy()
    # Создаем копию для избежания изменений в оригинальном df
    df_orders_ = df_orders.copy()
    # Создаем колонку, по которой считаем заказы
    df_orders_['Заказы'] = 1
    # На всякий случай заполняем пропуски в размерах
    df_orders_['Размер'] = df_orders_['Размер'].fillna(0)
    # Считаем заказы на каждый артикул и размер
    df_orders_stats = (
        df_orders_
        .groupby(['Артикул WB', 'Артикул продавца', 'Размер'])
        .agg(**{
            'Заказы шт': ('Заказы', 'sum')
        })
        .reset_index()
    )
    # Объединяем заказы с расходами по SKU
    df_sku_and_orders = (
        pd.
        concat([df_sku_costs_stats, df_orders_stats])
        .groupby(['Артикул WB', 'Артикул продавца', 'Размер'])
        .sum()
        .reset_index()
    )
    # Перемещаем заказы в начало df для удобства
    df_sku_and_orders = move_columns(df_sku_and_orders, ['Заказы шт'], 'Продажи шт', 'before')
    # Считаем долю продаж в артикуле после добавления товаров
    df_sku_and_orders['Продажи в артикуле'] = df_sku_and_orders.groupby('Артикул WB')['Продажи шт'].transform('sum')
    df_sku_and_orders['Доля продаж в артикуле'] = df_sku_and_orders['Продажи шт'] / df_sku_and_orders['Продажи в артикуле']
    df_sku_and_orders['Количество размеров в артикуле'] = df_sku_and_orders.groupby('Артикул WB').transform('size')
    # Перемещаем колонки в начало df
    df_sku_and_orders = move_columns(
        df_sku_and_orders,
        ['Количество размеров в артикуле', 'Продажи в артикуле', 'Доля продаж в артикуле'],
        position='Продажи шт',
        insert_type='before'
    )

    return df_sku_and_orders



# Функция выгрузки и расчета расходов по платому хранению
def calc_paid_storage_costs(headers,
                            df_realization_report_date_filtered,
                            date_start,
                            date_end):

    # df, в который будем помещать итоговый результат
    df_paid_storage_stats = pd.DataFrame()

    # df, в который будем помещать результаты выгрузки отчетов апи
    df_paid_storage = pd.DataFrame()

    # Формируем недели, из которых был собран отчет о реализации и где были расходы на хранение
    paid_storage_dates = df_realization_report_date_filtered \
        .loc[df_realization_report_date_filtered['Обоснование для оплаты'].isin(['Хранение']), :] \
        .groupby(['Дата начала отчётного периода', 'Дата конца отчётного периода']) \
        .size() \
        .reset_index() \
        .rename(columns = {0: 'row_count',
                           'Дата начала отчётного периода': 'date_from',
                           'Дата конца отчётного периода': 'date_to'}
                           )

    paid_storage_dates = paid_storage_dates.loc[paid_storage_dates['row_count'] > 0, :]
    # Если были расходу по платному хранению, начинаем выгрузку отчетов
    if paid_storage_dates.shape[0] > 0:
        # df, в который будем помещать результат апи
        df_paid_storage = pd.DataFrame()

        # -- GPT START --
        # Преобразование строковых дат в тип данных datetime
        start_date = pd.to_datetime(date_start)
        end_date = pd.to_datetime(date_end)
        # Создаем список для хранения диапазона дат
        date_range_paid_storage = []
        # Генерируем даты с разницей не более 8 дней и без пересечений
        current_date = start_date
        while current_date <= end_date:
            # Устанавливаем конечную дату как текущую дату плюс 7 дней или дату окончания, если она раньше
            new_end_date = min(current_date + pd.Timedelta(days=7), end_date)
            date_range_paid_storage.append({'date_from': current_date, 'date_to': new_end_date})
            # Переходим к следующей начальной дате (добавляем 9 дней)
            current_date = new_end_date + pd.Timedelta(days=1)
        # Создаем DataFrame
        df_paid_storage_date_range = pd.DataFrame(date_range_paid_storage)
        # -- GPT END --

        # Делаем цикл по каждой из недель
        for i in range(df_paid_storage_date_range.shape[0]):
            time.sleep(5)
            # Выбираем одну неделю
            date_start_paid_storage = df_paid_storage_date_range['date_from'][i].isoformat()
            date_end_paid_storage = df_paid_storage_date_range['date_to'][i].isoformat()

            logger.info(f"Uploading paid storage for dates {date_start_paid_storage} - {date_end_paid_storage}")

            #  Параметры запроса на создание отчета
            params_report_generation = {
                'dateFrom': date_start_paid_storage,
                'dateTo': date_end_paid_storage
            }
            # Получаем id задания на создание отчета
            resp_data_task_id = requests.get("https://seller-analytics-api.wildberries.ru/api/v1/paid_storage", headers=headers, params=params_report_generation).json()
            task_id = resp_data_task_id['data']['taskId']

            # print(f"Task id: {task_id}")

            # Проверяем статус отчета
            report_status = 'new' # Начальное значение для цикла
            while report_status not in ['done', 'purged', 'canceled']:
                resp_data_report_status = requests.get(f"https://seller-analytics-api.wildberries.ru/api/v1/paid_storage/tasks/{task_id}/status", headers=headers).json()
                report_status = resp_data_report_status['data']['status']
                logger.info(f"Paid storage report status: {report_status}")
                time.sleep(10)

            # Загружаем отчет, если он готов
            if report_status == 'done':
                # Начальное значение для цикла while
                resp_data_paid_storage_report = {'data': []}
                # Делаем запрос на загрузку отчета, пока не скачаем его
                while not isinstance(resp_data_paid_storage_report, list):
                    resp_data_paid_storage_report = requests.get(f"https://seller-analytics-api.wildberries.ru/api/v1/paid_storage/tasks/{task_id}/download", headers=headers).json()
                    time.sleep(5)
                # Переводим отчет в df
                tmp_df = pd.DataFrame(resp_data_paid_storage_report)
                # Объединяем с предыдущим проходом цикла
                df_paid_storage = pd.concat([df_paid_storage, tmp_df])
                logger.info('Done uploading chunk')

            # Если ошибка, выводим текст ошибки
            else:
                print(f"Error downloading report, info: {resp_data_paid_storage_report}")

            # Ждем 1 минуту (ограничение на кол-во запросов),
            # если это не последняя неделя, за которую нужно выгрузить
            if i != (paid_storage_dates.shape[0] - 1):
                logger.info('Waiting 1 minute before downloading next week')
                time.sleep(60)

        logger.info("Done uploading paid storage reports")

        # Переименовываем некоторые столбцы для удобства
        df_paid_storage = df_paid_storage.rename(columns={
            'nmId': 'Артикул WB',
            'vendorCode': 'Артикул продавца',
            'size': 'Размер',
            'warehousePrice': 'Сумма хранения'
            })

        # Считаем стоимость хранения для каждого товара
        df_paid_storage_stats = df_paid_storage.groupby(['Артикул WB', 'Артикул продавца', 'Размер']).agg(
            Платное_хранение=('Сумма хранения', 'sum')
        ).reset_index()
        df_paid_storage_stats = df_paid_storage_stats.rename(columns={'Платное_хранение': 'Платное хранение (отдельный отчет)'})

        # Переводим артикул продавца с верхний регистр
        df_paid_storage_stats['Артикул продавца'] = df_paid_storage_stats['Артикул продавца'].str.upper()

    # Формируем словарь с результатами
    result_paid_storage = {
        'df_paid_storage': df_paid_storage,
        'df_paid_storage_stats': df_paid_storage_stats,
    }

    return result_paid_storage


# Функция объединения расходов по SKU, которые уже разбиты по размерам
def union_sku_sizes_costs(result_sku_costs, result_paid_storage):
    # Получаем df с результатами расходов по SKU
    df_sku_sizes_costs_stats = result_sku_costs['df_sku_sizes_costs_stats'].copy()
    # Получаем df с расходами на пл. хранение
    df_paid_storage_stats = result_paid_storage['df_paid_storage_stats'].copy()
    # Объединяем расходы по SKU из всех отчетов
    df_sku_sizes_costs_union = pd.concat([
        df_sku_sizes_costs_stats,
        #df_orders,
        df_paid_storage_stats
    ])
    # Колонки, по которым будет производиться группировка
    groupby_cols = ['Артикул WB', 'Артикул продавца', 'Размер']
    # Считаем суммы
    df_sku_sizes_costs_merged = (
        df_sku_sizes_costs_union
        .groupby(groupby_cols)
        .sum()
        .reset_index()
    )
    # Считаем долю продаж в каждом артикуле в каждом размере (изменится, если добавились новые артикулы)
    df_sku_sizes_costs_merged['Продажи в артикуле'] = df_sku_sizes_costs_merged.groupby('Артикул WB')['Продажи шт'].transform('sum')
    df_sku_sizes_costs_merged['Доля продаж в артикуле'] = df_sku_sizes_costs_merged['Продажи шт'] / df_sku_sizes_costs_merged['Продажи в артикуле']
    df_sku_sizes_costs_merged['Количество размеров в артикуле'] = df_sku_sizes_costs_merged.groupby('Артикул WB').transform('size')
    # Перемещаем колонки в начало df
    df_sku_sizes_costs_merged = move_columns(df_sku_sizes_costs_merged,
                                      ['Количество размеров в артикуле', 'Продажи в артикуле', 'Доля продаж в артикуле', 'Заказы шт'],
                                       position='Продажи шт',
                                       insert_type='before'
    )

    return df_sku_sizes_costs_merged


# Функция выгрузки и расчета расходов по платной приемке
def calc_paid_acceptance_costs(
        headers,
        df_sku_sizes_costs_merged,
        df_realization_report_date_filtered,
        date_start,
        date_end
):

     # df, в который будем помещать итоговый результат
    df_paid_acceptance_stats = pd.DataFrame()

    # df, в который будем помещать результаты выгрузки апи
    df_paid_acceptance = pd.DataFrame()

    # df, в который будем помещать результат разбивки платной приемки по артикулам
    df_paid_acceptance_sku_sizes = pd.DataFrame()

    # Создаем копию для избежания изменений в оригинальном df
    df_realization_report_ = df_realization_report_date_filtered.copy()

    # Проверяем, были ли расходы по платной приемке
    paid_acceptance_costs = df_realization_report_.loc[df_realization_report_['Обоснование для оплаты'].isin(['Платная приемка']), :]

    if paid_acceptance_costs.shape[0] > 0:
        # Переводим в datetime
        # for col in ['Дата начала отчётного периода', 'Дата конца отчётного периода']:
        #     paid_acceptance_costs[col] = pd.to_datetime(paid_acceptance_costs[col])

        # # Разбиваем df на равные интервалы по 30 дней
        # paid_acceptance_dates = paid_acceptance_costs \
        #     .groupby(['Дата начала отчётного периода', 'Дата конца отчётного периода']) \
        #     .size() \
        #     .reset_index() \
        #     .rename(columns = {0: 'row_count',
        #                     'Дата начала отчётного периода': 'date_from',
        #                     'Дата конца отчётного периода': 'date_to'}
        #                     )
        # # Переводим в datetime
        # for col in ['date_from', 'date_to']:
        #     paid_acceptance_dates[col] = pd.to_datetime(paid_acceptance_dates[col])

        # -- GPT START --
        # Преобразование строковых дат в тип данных datetime
        start_date = pd.to_datetime(date_start)
        end_date = pd.to_datetime(date_end)
        # Создаем список для хранения диапазона дат
        date_range_paid_storage = []
        # Генерируем даты с разницей не более 30 дней и без пересечений
        current_date = start_date
        while current_date <= end_date:
            # Устанавливаем конечную дату как текущую дату плюс 29 дней или дату окончания, если она раньше
            new_end_date = min(current_date + pd.Timedelta(days=30), end_date)
            date_range_paid_storage.append({'date_from': current_date, 'date_to': new_end_date})
            # Переходим к следующей начальной дате
            current_date = new_end_date + pd.Timedelta(days=1)
        # Создаем DataFrame
        df_paid_acceptance_date_range = pd.DataFrame(date_range_paid_storage)
        # -- GPT END --

        # # Находим минимальную дату
        # min_date = min(paid_acceptance_dates['date_from'])
        # # Считаем разницу с минимальной датой в днях
        # paid_acceptance_dates['days_diff'] = abs(paid_acceptance_dates['date_to'] - min_date).dt.days + 1
        # # Разбиваем на интервалы по 30 дней
        # paid_acceptance_dates['chunks'] = paid_acceptance_dates['days_diff'].apply(lambda x: int(x / 30) + 1)
        # # Сколько всего получилось интервалов
        # total_chunks = max(paid_acceptance_dates['chunks'])
        total_chunks =  df_paid_acceptance_date_range.shape[0]

        # Цикл по каждому интервалу
        for idx, row in df_paid_acceptance_date_range.iterrows():
            # Выбираем временной интервал не больше 30 дней
            # df_dates_tmp = paid_acceptance_dates.loc[paid_acceptance_dates['chunks'] == chunk, :]
            # Ищем мин. и макс дату
            min_date_chunk = str(row['date_from'].date())
            max_date_chunk = str(row['date_to'].date())

            logger.info(f"Uploading paid acceptance report for dates {min_date_chunk} - {max_date_chunk}")
            # Переводим даты в нужный формат
            # date_start_paid_acceptance = datetime.strptime(date_start, '%Y-%m-%dT%H:%M:%S').date().strftime('%Y-%m-%d')
            # date_end_paid_acceptance = datetime.strptime(date_end, '%Y-%m-%dT%H:%M:%S').date().strftime('%Y-%m-%d')

            # Параметры запроса
            params_paid_acceptance = {
                'dateFrom': min_date_chunk,
                'dateTo': max_date_chunk
            }
            # Формируем задание на генерацию отчета
            resp_data_paid_acceptance = (
                requests
                .get(
                    "https://seller-analytics-api.wildberries.ru/api/v1/acceptance_report",
                    headers=headers,
                    params=params_paid_acceptance
                )
                .json()
            )
            # Получаем ID задания на генерацию отчета
            task_id = resp_data_paid_acceptance['data']['taskId']
            # Проверяем статус отчета
            report_status = ''
            while report_status not in ['done', 'purged', 'canceled']:
                report_status = (
                    requests
                    .get(
                        f'https://seller-analytics-api.wildberries.ru/api/v1/acceptance_report/tasks/{task_id}/status',
                        headers=headers
                    )
                    .json()
                    ['data']['status']
                )
                logger.info(
                    f"\n"
                    f"Paid acceptance report status: {report_status}\n"
                    f"Dates: {min_date_chunk} - {max_date_chunk}\n"
                    f"taskID: {task_id}"
                )
                # Ограничение на запросы раз в 5 секунд
                time.sleep(6)


            # Если отчет выгружен успешно, добавляем его в df
            if report_status == 'done':
                # Получаем отчет
                resp_data_paid_acceptance_report = (
                    requests
                    .get(
                        f'https://seller-analytics-api.wildberries.ru/api/v1/acceptance_report/tasks/{task_id}/download',
                        headers=headers
                    )
                    .json()
                )
                # Переводим отчет в df
                tmp_df_paid_acceptance = pd.DataFrame(resp_data_paid_acceptance_report)
                # Объединяем с предыдущим проходом цикла
                df_paid_acceptance = pd.concat([df_paid_acceptance, tmp_df_paid_acceptance])
                logger.info("Done uploading chunk")
            else:
                print(f"Error uploading paid acceptance report, info: {resp_data_paid_acceptance}")

            # Ждем одну минуту до формирования отчета за следующие даты
            if (idx + 1) != total_chunks:
                logger.info("Waiting 1 minute before dowloading next report")
                time.sleep(60)

        # Если df с отчетами не пустой, начинаем обработку отчетов
        if not df_paid_acceptance.empty:
            # Переименовываем колонку с номером артикула
            df_paid_acceptance = df_paid_acceptance.rename(columns={
                'nmID': 'Артикул WB',
                'total': 'Суммарная стоимость приемки руб'
            })
            # Считаем итоговые расходы по платной приемке
            df_paid_acceptance_stats = (
                df_paid_acceptance
                .groupby('Артикул WB')
                .agg(
                    Платная_приемка=('Суммарная стоимость приемки руб', 'sum')
                )
                .reset_index()
            )
            df_paid_acceptance_stats = df_paid_acceptance_stats.rename(columns={'Платная_приемка': 'Платная приемка (отдельный отчет)'})

            # # Меняем знак на итоговых расходах на минус
            # df_paid_acceptance_stats['Платная_приемка'] = -df_paid_acceptance_stats['Платная_приемка']

            # # Считаем расходы на каждую поставку
            # df_paid_acceptance_by_incomes = df_paid_acceptance.groupby('incomeId').agg(
            #     Расходы_на_поставку=('total', 'sum')
            # ).reset_index()
            # df_paid_acceptance_by_incomes['Расходы_на_поставку'] = -df_paid_acceptance_by_incomes['Расходы_на_поставку']

            # Разбиваем список поставок на диапазоны по 1000 поставок
            # step = 1000
            # df_paid_acceptance_by_incomes['chunks'] = df_paid_acceptance_by_incomes

            # Переименовываем колонки в списке товаров для удобства
            # df_products_ = df_products.copy()
            # df_products_ = df_products_.rename(columns={'nmID': 'Артикул WB'})
            # df_products_ = df_products_.loc[:, ['Артикул WB', 'Артикул продавца', 'Размер']]

            # Получаем df с расходами по SKU
            df_sku_and_paid_storage_costs = df_sku_sizes_costs_merged.copy()
            # Мерджим с расходами на приемку, чтобы получить количество размеров в одном артикуле
            df_paid_acceptance_sku_sizes = df_paid_acceptance_stats.merge(
                df_sku_and_paid_storage_costs[[
                    'Артикул WB',
                    'Артикул продавца',
                    'Размер', 'Продажи шт',
                    'Количество размеров в артикуле',
                    'Продажи в артикуле',
                    'Доля продаж в артикуле'
                ]],
                on='Артикул WB',
                how='left',
                indicator='Наличие артикула'
            )
            # Если какого-то артикула не было в списке SKU по размерам, выводим предупреждение
            missing_articles = df_paid_acceptance_sku_sizes.loc[
                df_paid_acceptance_sku_sizes['Наличие артикула'] == 'left_only',
                ['Артикул WB']
            ]
            if not missing_articles.empty:
                logger.warning(
                    f"No size found for articles \n"
                    f"{missing_articles}"
                )

            # Заполняем пропуски в артикуле и размере
            fill_values = {
                'Артикул продавца': 'Неизвестный артикул',
                'Количество размеров в артикуле': 0,
                'Продажи в артикуле': 0,
                'Доля продаж в артикуле': 0,
                'Размер': 0
            }
            for col, fill_value in fill_values.items():
                df_paid_acceptance_sku_sizes[col] = df_paid_acceptance_sku_sizes[col].fillna(fill_value)

            # Разбиваем расходы по платной приемке на долю продаж в артикуле
            df_paid_acceptance_sku_sizes['Платная приемка (отдельный отчет)'] =(
                df_paid_acceptance_sku_sizes['Платная приемка (отдельный отчет)'] /
                df_paid_acceptance_sku_sizes['Количество размеров в артикуле']
            )

            # Выбираем нужные колонки
            df_paid_acceptance_sku_sizes = df_paid_acceptance_sku_sizes[['Артикул WB', 'Артикул продавца', 'Размер', 'Платная приемка (отдельный отчет)']]

    # Выгружаем список поставок
    #     logger.info('Uploading incomes list')
    #     df_incomes = pd.DataFrame()
    #     # Цикл по каждому интервалу
    #     for chunk in paid_acceptance_dates['chunks'].unique():
    #         # Выбираем временной интервал не больше 30 дней
    #         df_dates_tmp = paid_acceptance_dates.loc[paid_acceptance_dates['chunks'] == chunk, :]
    #         # Ищем мин. и макс дату, переводим в unix timestamp
    #         min_date_chunk = min(df_dates_tmp['date_from'])
    #         max_date_chunk = max(df_dates_tmp['date_to'])
    #         min_date_chunk_unix = min_date_chunk.timestamp()
    #         max_date_chunk_unix = max_date_chunk.timestamp()

    #         logger.info(f"Uploading paid incomes for dates {str(min_date_chunk.date())} - {str(max_date_chunk.date())}")
    #         # Переводим даты в нужный формат
    #         # date_start_paid_acceptance = datetime.strptime(date_start, '%Y-%m-%dT%H:%M:%S').date().strftime('%Y-%m-%d')
    #         # date_end_paid_acceptance = datetime.strptime(date_end, '%Y-%m-%dT%H:%M:%S').date().strftime('%Y-%m-%d')

    #         # Параметры запроса
    #         next = 1 # начальное значение для while
    #         params_incomes = {'limit': 1000,
    #             'next': 0,
    #             'dateFrom': int(min_date_chunk_unix),
    #             'dateTo': int(max_date_chunk_unix)
    #             }
    #         while next > 0:
    #             print(f"Next page incomes: {next not in [0, 1]}")
    #             # Делаем запрос, получаем отчет
    #             resp_data_incomes = requests.get("https://marketplace-api.wildberries.ru/api/v3/orders", headers=headers, params=params_incomes).json()
    #             # Переводим отчет в df
    #             tmp_df_incomes = pd.DataFrame(resp_data_incomes['orders'])
    #             # Объединяем с предыдущим проходом цикла
    #             df_incomes = pd.concat([df_incomes, tmp_df_incomes])
    #             # Пагинация
    #             next = resp_data_incomes['next']
    #             params_incomes = {'limit': 1000,
    #                             'next': next,
    #                             'dateFrom': int(min_date_chunk_unix),
    #                             'dateTo': int(max_date_chunk_unix)
    #                             }

    # Если расходов по платной приемке нет, то оставляем пустой df
    # else:
    #     df_paid_acceptance_sku_sizes = pd.DataFrame()

    # Формируем словарь с результатами
    result_paid_acceptance = {
        'df_paid_acceptance_sku_sizes': df_paid_acceptance_sku_sizes,
        'df_paid_acceptance_stats': df_paid_acceptance_stats,
        'df_paid_acceptance': df_paid_acceptance
    }

    return result_paid_acceptance


# Функция объединения расходов по SKU, пл.хранению и приемке
def union_sku_paid_storage_acceptance_costs(df_sku_and_paid_storage_stats, result_paid_acceptance):
    # Получаем df с результатами расходов по SKU пл. хранению и приемку
    df_sku_and_paid_storage_stats_ = df_sku_and_paid_storage_stats.copy()
    df_paid_acceptance_sku_ = result_paid_acceptance['df_paid_acceptance_sku'].copy()
    # Объединяем расходы по SKU пл. хранению и приемке
    df_sku_paid_storage_acceptance = pd.concat([df_sku_and_paid_storage_stats_, df_paid_acceptance_sku_])
    # Если есть расходы по пл. приемке, объединяем все расходы в один df
    if not df_paid_acceptance_sku_.empty:
        # Колонки, по которым будет производиться группировка
        groupby_cols = ['Артикул WB', 'Артикул продавца', 'Размер']
        # Считаем суммы
        df_sku_paid_storage_acceptance_stats = (
            df_sku_paid_storage_acceptance
            # Убираем колонки с количеством размеров в артикуле, т.к. их будем считать заново
            .loc[:, df_sku_and_paid_storage_stats_.columns.isin(['Продажи в артикуле', 'Доля продаж в артикуле', 'Количество размеров в артикуле'])]
            .groupby(groupby_cols)
            .sum()
            .reset_index()
        )
        # Считаем долю продаж в каждом артикуле в каждом размере (изменится, если добавились новые артикулы)
        df_sku_paid_storage_acceptance['Продажи в артикуле'] = df_sku_paid_storage_acceptance.groupby('Артикул WB')['Продажи шт'].transform('sum')
        df_sku_paid_storage_acceptance['Доля продаж в артикуле'] = df_sku_paid_storage_acceptance['Продажи шт'] / df_sku_paid_storage_acceptance['Продажи в артикуле']
        df_sku_paid_storage_acceptance['Количество размеров в артикуле'] = df_sku_paid_storage_acceptance.groupby('Артикул WB').transform('size')
        # Перемещаем колонки в начало df
        df_sku_paid_storage_acceptance_stats = move_columns(df_sku_paid_storage_acceptance_stats,
                                        ['Количество размеров в артикуле', 'Продажи в артикуле', 'Доля продаж в артикуле'],
                                        position='Продажи шт',
                                        insert_type='before'
        )
    # Если расходов по пл. приемке нет, то возвращаем исходный df с
    else:
        df_sku_paid_storage_acceptance_stats = df_sku_paid_storage_acceptance.copy()

    return df_sku_paid_storage_acceptance_stats


# Функция обработки файлов по рекламным кампаниям
def parse_companies_files(date_report):

    # df, куда будем помещать итоговый результат
    df_companies_stats = pd.DataFrame()

    # Директория, в которой лежат данные по рекламным кампаниям
    companies_stats_dir = f"{finance_reports_dir}/{date_report}_Кампании/"
    if not os.path.exists(companies_stats_dir):
        os.mkdir(companies_stats_dir)  # Создаем, если директория не существует
    # Получаем список файлов в директории
    companies_stats_files = os.listdir(companies_stats_dir)
    # Если папка пустая, выводим сообщение об ошибке и возвращаем пустой df
    if len(companies_stats_files) == 0:
        logger.warning("Нет данных по кампаниям из ЛК")
        return df_companies_stats

    # Если в папке есть файлы с кампаниями, начинаем их обработку
    else:
        logger.info(f"Parsing companies files")
        filenames_companies = {"path": []}
        path_companies = f"{companies_stats_dir}/статистика-*.xlsx"
        # Считывание файла с путем до него
        for file in glob.glob(path_companies):
            filenames_companies['path'].append(file)
        # Считывание только имени файла
        filenames_companies['file_name'] = [os.path.basename(x) for x in glob.glob(path_companies)]
        # Размер файла
        filenames_companies['file_size'] = [os.path.getsize(x) for x in glob.glob(path_companies)]
        # Перевод в df
        filenames_companies = pd.DataFrame(filenames_companies)
        # ID кампании
        filenames_companies['id'] = filenames_companies['file_name'].str.extract(r'-(\d+)\.xlsx$')

        # df, куда будем помещать результаты
        df_companies_parsed = pd.DataFrame()
        # Цикл по каждой кампании
        for i in range(filenames_companies.shape[0]):
            # Считываем файл со статистикой кампании
            tmp_df_company = pd.read_excel(filenames_companies['path'][i])
            # Добавляем id кампании
            tmp_df_company['id'] = filenames_companies['id'][i]
            # Удаляем последнюю строку из df
            tmp_df_company = tmp_df_company.iloc[:-1]
            # Объединяем с предыдущим проходом цикла
            df_companies_parsed = pd.concat([df_companies_parsed, tmp_df_company])

        # Сбрасываем index после concat
        df_companies_parsed = df_companies_parsed.reset_index(drop=True)

        # Считаем суммы по кампании
        df_companies_stats = (
            df_companies_parsed
            .rename(columns={
                'id': 'ID Кампании',
                'Номенклатура': 'Артикул WB',
                'Заказанные товары, шт.': 'Заказы из кампаний шт'
            })
            .groupby(['ID Кампании', 'Артикул WB'])
            .agg(**{
                'Заказы из кампаний шт': ('Заказы из кампаний шт', 'sum')
            })
            .reset_index()
        )
        # Добавляем столбец с суммарными заказами по каждой кампании
        df_companies_stats['Заказы внутри кампании шт'] = df_companies_stats.groupby(['ID Кампании'])['Заказы из кампаний шт'].transform('sum')
        # Добавляем столбец с количеством артикулов в кампании
        df_companies_stats['Количество артикулов в кампании'] = df_companies_stats.groupby(['ID Кампании']).transform('size')

        # Сохраняем обработанные данные по кампаниям в excel
        with pd.ExcelWriter(f"{companies_stats_dir}/{date_report}_Кампании_parsed.xlsx") as w:
            df_companies_parsed.to_excel(w, sheet_name='Кампании парсинг', index=False)
            df_companies_stats.to_excel(w, sheet_name='Кампании суммы', index=False)

        return df_companies_stats


# Функция распределения расходов рекламных кампаний по артикулам WB
def calc_companies_by_sku_and_size(df_companies_stats, df_costs_history, df_sku_sizes_costs_merged):
    # Создаем копию для избежания изменений в оригинальном df
    df_costs_history_ = df_costs_history.copy()
    df_companies_stats_ = df_companies_stats.copy()
    df_sku_sizes_costs_merged_ = df_sku_sizes_costs_merged.copy()
    # Переименовываем колонки для соответствия
    df_costs_history_ = df_costs_history_.rename(columns={
        'advertId': 'ID Кампании'
    })
    # Переводим колонку с ID кампании в int64
    df_companies_stats_['ID Кампании'] = df_companies_stats_['ID Кампании'].astype(np.int64)
    # Мерджим историю затрат (там, где "правильные расходы")
    # со статистикой кампаний (где содержатся артикулы кампаний)
    df_companies_stats_by_sku = df_companies_stats_.merge(
        df_costs_history_,
        on='ID Кампании',
        how='left',
        indicator='Наличие кампаний'
    )
    # Выводим предупреждение, если расходы по каким-то кампаниям не найдены
    missing_companies = df_companies_stats_by_sku.loc[
        df_companies_stats_by_sku['Наличие кампаний'] == 'left_only',
        ['ID Кампании', 'Заказы из кампаний шт']
    ]
    if not missing_companies.empty:
        logger.warning(f"No companies costs for companies\n"
                       f"{missing_companies}")
    # Сортировка
    df_companies_stats_by_sku = df_companies_stats_by_sku.sort_values(by=['ID Кампании', 'Артикул WB'], ignore_index=True)
    # Заполняем пропуски в расходах
    # (если кампании не было в Истории затрат, то считаем, что расходы по ней равны 0)
    df_companies_stats_by_sku['Расходы'] = df_companies_stats_by_sku['Расходы'].fillna(0)
    # Убираем кампании, которых не было в Истории затрат
    df_companies_stats_by_sku = df_companies_stats_by_sku.loc[df_companies_stats_by_sku['Наличие кампаний'] == 'both', :]
    # Распределяем расходы из Истории затрат на один артикул
    df_companies_stats_by_sku['Расходы по артикулу'] = (
        # df_companies_stats_by_sku['Заказы из кампаний шт']
        # / df_companies_stats_by_sku['Заказы внутри кампании шт']
        # * df_companies_stats_by_sku['Расходы']
        df_companies_stats_by_sku['Расходы'] /
        df_companies_stats_by_sku['Количество артикулов в кампании']
    )
    # Заполняем пропуски
    df_companies_stats_by_sku['Расходы по артикулу'] = df_companies_stats_by_sku['Расходы по артикулу'].fillna(0)
    # Распределяем расходы на один размер
    df_companies_stats_by_size = df_companies_stats_by_sku.merge(
        df_sku_sizes_costs_merged_[[
            'Артикул WB',
            'Артикул продавца',
            'Размер',
            'Количество размеров в артикуле',
            'Продажи в артикуле',
            'Доля продаж в артикуле',
            'Продажи шт'
        ]],
        how='left',
        on='Артикул WB',
        indicator='Наличие артикулов'
    )
    # Если какого-то артикула не было в списке SKU по размерам, выводим предупреждение
    missing_articles = df_companies_stats_by_size.loc[
        df_companies_stats_by_size['Наличие артикулов'] == 'left_only',
        ['Артикул WB']
    ]
    if not missing_articles.empty:
        logger.warning(f"No size found for articles\n"
                       f"{missing_articles}")

    # Заполняем пропуски
    fill_values = {
        'Артикул продавца': 'Неизвестный артикул',
        'Размер': 0,
        'Количество размеров в артикуле': 1,

    }
    df_companies_stats_by_size = df_companies_stats_by_size.fillna(fill_values)

    # Сортировка
    df_companies_stats_by_size = df_companies_stats_by_size.sort_values(by=['ID Кампании', 'Артикул WB', 'Артикул продавца', 'Размер'], ignore_index=True)
    # Считаем общее количество заказов
    # total_orders = df_companies_stats_by_size['Заказы шт'].sum()
    # Распределяем расходы на один размер
    df_companies_stats_by_size['Расходы по размеру'] = (
        df_companies_stats_by_size['Расходы по артикулу'] / df_companies_stats_by_size['Количество размеров в артикуле']
    )
    # Выбираем колонку, которая будет считаться итоговыми расходами
    df_companies_stats_by_size['Расходы по рекламным кампаниям'] = df_companies_stats_by_size['Расходы по размеру']
    # Выбираем нужные колонки
    df_companies_stats_by_size = df_companies_stats_by_size.loc[:, ['Артикул WB', 'Артикул продавца', 'Размер', 'Расходы по рекламным кампаниям']]

    return df_companies_stats_by_size


# Функция получения дат расходов по ВБ Продвижению
def get_promotion_dates(df_realization_report_date_filtered):
    # Ищем даты расходов по ВБ Продвижению
    promotion_dates = df_realization_report_date_filtered.loc[
        df_realization_report_date_filtered['Обоснование штрафов и доплат'] == 'Оказание услуг «ВБ.Продвижение»',
        ['Дата начала отчётного периода', 'Дата конца отчётного периода', 'Итого руб']
    ]
    if not promotion_dates.empty:
        # Переводим дату в нужный формат для выгрузки по апи рекламы
        date_start_promotion = (
            pd
            .to_datetime(min(promotion_dates['Дата начала отчётного периода']))
            .strftime('%Y-%m-%d')
        )
        date_end_promotion = (
            pd
            .to_datetime(max(promotion_dates['Дата конца отчётного периода']))
            .strftime('%Y-%m-%d')
        )
        # Переводим дату в нужный формат для выгрузки заказов
        date_start_orders = date_start_promotion + 'T00:00:00.000Z'
        date_end_orders = date_end_promotion + 'T23:59:59.000Z'
    else:
        date_start_promotion = None
        date_end_promotion = None
        date_start_orders = None
        date_end_orders = None

    # Формируем словарь с резульататми
    promotion_dates = {
        'date_start_promotion': date_start_promotion,
        'date_end_promotion': date_end_promotion,
        'date_start_orders': date_start_orders,
        'date_end_orders': date_end_orders,
    }

    return promotion_dates


# Функция расчета расходов по рекламным кампаниям
def calc_companies_costs(df_realization_report_date_filtered, df_sku_sizes_costs_merged, date_report):
    # Ищем даты, в которых были расходы на ВБ Продвижение
    promotion_dates = get_promotion_dates(df_realization_report_date_filtered)

    # Если расходов на ВБ Продвижение не было, возвращаем пустой df
    if any(date_promotion is None for date_promotion in promotion_dates.values()):
        logger.info("No promotion costs have been found")
        df_companies_stats_by_sku = pd.DataFrame()
    # Если расходы были, то начинаем расчет расходов по рекламным кампаниям
    else:
        # Обработка данных рекламных кампаний
        df_companies_stats = parse_companies_files(date_report)
        # # Выгружаем заказы
        # df_orders = get_orders(
        #     headers,
        #     promotion_dates['date_start_orders'],
        #     promotion_dates['date_end_orders'],
        #     to_save=False
        # )
        # # Считаем заказы по размерам
        # df_orders_stats = calc_orders_by_size(df_orders)
        # Получение истории затрат
        df_costs_history = get_costs_history(
            headers,
            promotion_dates['date_start_promotion'],
            promotion_dates['date_end_promotion']
        )
        # Распределяем расходы из истории затрат по артикулам в кампании
        df_companies_stats_by_sku = calc_companies_by_sku_and_size(
            df_companies_stats,
            df_costs_history,
            df_sku_sizes_costs_merged
        )

    return df_companies_stats_by_sku



# Функция расчета расходов, которые не привязаны к артикулам
def calc_other_costs_v2(result_sku_costs, df_sku_sizes_costs_merged):
    # Получаем list с id строк расходов по артикулам
    sku_costs_ids = result_sku_costs['sku_costs_ids']
    # df, в который будем помещать итоговый результат (df уже содержит расходы по sku)
    df_sku_and_other_costs = df_sku_sizes_costs_merged.copy()

    # Выбираем строки из отчета, которые не привязаны к артикулам
    df_other_costs = df_realization_report_date_filtered.loc[~df_realization_report_date_filtered['id'].isin(sku_costs_ids), :]
    other_costs_ids = df_realization_report_date_filtered.loc[~df_realization_report_date_filtered['id'].isin(sku_costs_ids), 'id'].to_list()

    # Выбираем расходы, у которых есть доп. детализация в столбце "Обоснование штрафов и доплат"
    df_other_costs_fees = df_other_costs.loc[
        (df_other_costs['Обоснование штрафов и доплат'] != '') & (~df_other_costs['Обоснование штрафов и доплат'].isna()),
    :]
    other_costs_fees_ids = df_other_costs_fees['id'].to_list()

    # Считаем итоговые суммы по Штрафам и доплатам
    df_other_costs_fees_by_operation = (
        df_other_costs_fees
        .groupby('Обоснование штрафов и доплат')
        .agg(
            Итого_прочие_расходы=('Итого руб', 'sum')
            )
            .reset_index()
            .rename(columns={'Обоснование штрафов и доплат': 'Тип расходов'})
    )

    # Выбираем расходы, у которых нет доп. детализации
    df_other_costs_leftover = df_other_costs.loc[~df_other_costs['id'].isin(other_costs_fees_ids), :]
    other_costs_leftover_ids = df_other_costs_leftover['id'].to_list()

    # Считаем итоговые суммы по прочим расходам
    df_other_costs_leftover_by_operation = (
        df_other_costs_leftover
        .groupby('Обоснование для оплаты')
        .agg(
            Итого_прочие_расходы=('Итого руб', 'sum')
            )
        .reset_index()
        .rename(columns={'Обоснование для оплаты': 'Тип расходов'})
    )

    # Объединяем группы расходов, полученные выше
    df_other_costs_by_operation = pd.concat([df_other_costs_fees_by_operation, df_other_costs_leftover_by_operation], axis = 0).reset_index(drop=True)

    # Добавляем прочие расходы в df с расходами по sku, разбивая их от количества продаж
    total_sales = df_sku_and_other_costs['Продажи шт'].sum()
    for operation_type in df_other_costs_by_operation['Тип расходов']:
        operation_costs = df_other_costs_by_operation.loc[df_other_costs_by_operation['Тип расходов'] == operation_type, 'Итого_прочие_расходы'].values[0]
        df_sku_and_other_costs[operation_type] = df_sku_and_other_costs['Продажи шт'] / total_sales * operation_costs

    # Удаляем некоторые группы расходов, т.к. по ним мы формируем отдельные отчеты
    df_sku_and_other_costs = df_sku_and_other_costs.loc[:, ~df_sku_and_other_costs.columns.isin([
        'Оказание услуг «ВБ.Продвижение»',
        # 'Платная приемка',
        # 'Корректировка приемки',
        'Хранение',
        'Корректировка хранения',
        'Удержания',
        ])]

    # Формируем словарь с результатами
    result_other_costs = {'df_sku_and_other_costs': df_sku_and_other_costs,
                          'df_other_costs_by_operation': df_other_costs_by_operation,
                          'other_costs_ids': other_costs_ids}

    return result_other_costs


# Функция расчета расходов, которые не привязаны к артикулам
def calc_other_costs(df_realization_report_date_filtered, result_sku_costs):
    # Получаем list с id строк расходов по артикулам
    sku_costs_ids = result_sku_costs['sku_costs_ids']
    # df, в который будем помещать итоговый результат (df уже содержит расходы по sku)
    df_sku_and_other_costs = result_sku_costs['df_sku_costs_stats'].copy()

    # Выбираем строки из отчета, которые не привязаны к артикулам
    df_other_costs = df_realization_report_date_filtered.loc[~df_realization_report_date_filtered['id'].isin(sku_costs_ids), :]
    other_costs_ids = df_realization_report_date_filtered.loc[~df_realization_report_date_filtered['id'].isin(sku_costs_ids), 'id'].to_list()

    # Выбираем расходы, у которых есть доп. детализация в столбце "Обоснование штрафов и доплат"
    df_other_costs_fees = df_other_costs.loc[
        (df_other_costs['Обоснование штрафов и доплат'] != '') & (~df_other_costs['Обоснование штрафов и доплат'].isna())
        , :]
    other_costs_fees_ids = df_other_costs_fees['id'].to_list()

    # Считаем итоговые суммы по Штрафам и доплатам
    df_other_costs_fees_by_operation = (
        df_other_costs_fees
        .groupby('Обоснование штрафов и доплат')
        .agg(
            Итого_прочие_расходы=('Итого руб', 'sum')
            )
            .reset_index()
            .rename(columns={'Обоснование штрафов и доплат': 'Тип расходов'})
    )

    # Выбираем расходы, у которых нет доп. детализации
    df_other_costs_leftover = df_other_costs.loc[~df_other_costs['id'].isin(other_costs_fees_ids), :]
    other_costs_leftover_ids = df_other_costs_leftover['id'].to_list()

    # Считаем итоговые суммы по прочим расходам
    df_other_costs_leftover_by_operation = (
        df_other_costs_leftover
        .groupby('Обоснование для оплаты')
        .agg(
            Итого_прочие_расходы=('Итого руб', 'sum')
            )
        .reset_index()
        .rename(columns={'Обоснование для оплаты': 'Тип расходов'})
    )

    # Объединяем группы расходов, полученные выше
    df_other_costs_by_operation = pd.concat([df_other_costs_fees_by_operation, df_other_costs_leftover_by_operation], axis = 0).reset_index(drop=True)

    # Добавляем прочие расходы в df с расходами по sku, разбивая их от количества продаж
    total_sales = df_sku_and_other_costs['Продажи шт'].sum()
    for operation_type in df_other_costs_by_operation['Тип расходов']:
        operation_costs = df_other_costs_by_operation.loc[df_other_costs_by_operation['Тип расходов'] == operation_type, 'Итого_прочие_расходы'].values[0]
        df_sku_and_other_costs[operation_type] = df_sku_and_other_costs['Продажи шт'] / total_sales * operation_costs

    # Удаляем столбцы с приемкой и хранением, т.к. расходы по ним мы получаем отдельным отчетом
    # for col in ['Корректировка приемки','Корректировка хранения', 'Платная приемка', 'Удержания', 'Хранение']:
    #     if col in df_sku_and_other_costs.columns:
    #         df_sku_and_other_costs = df_sku_and_other_costs.drop(columns=col)
    df_sku_and_other_costs = df_sku_and_other_costs.loc[:, df_sku_and_other_costs.columns.isin([
        'Платная приемка',
        'Корректировка приемки',
        # 'Хранение'
        # 'Корректировка хранения',
        'Удержания'
        ])]

    # Формируем словарь с результатами
    result_other_costs = {'df_sku_and_other_costs': df_sku_and_other_costs,
                          'df_other_costs_by_operation': df_other_costs_by_operation,
                          'other_costs_ids': other_costs_ids}

    return result_other_costs


# Функция объединения расходов
def merge_all_costs(df_sku_sizes_costs_merged,
                    result_paid_acceptance,
                    df_companies_costs):
    # Получаем df с расходами по платной приемке
    df_paid_acceptance_costs = result_paid_acceptance['df_paid_acceptance_sku_sizes']
    # Объединяем с расходами по платной приемке и хранению
    df_sku_sizes_costs_union = pd.concat([df_sku_sizes_costs_merged, df_paid_acceptance_costs, df_companies_costs])
    # Сводим все по артикулам после concat
    df_sku_sizes_costs_all = df_sku_sizes_costs_union.groupby(['Артикул WB', 'Артикул продавца', 'Размер']).sum().reset_index()

    return df_sku_sizes_costs_all


# Функция добавления данных из справочной таблицы
def add_data_from_catalog(result_other_costs):
    # Чтение справочной таблицы
    catalog = pd.read_excel(f"{marketplace_dir_name}/Clients/{client_name}/catalog/Справочная_таблица_{client_name}_WB.xlsx")
    # Создаем колонку размера, если её нет в справочной таблице
    if 'Размер' not in catalog.columns:
        catalog['Размер'] = 0

    # Получаем таблицу с расходами из результатов по всем расходам
    df_sku_and_other_costs = result_other_costs['df_sku_and_other_costs'].copy()

    # Создаем колонки для мерджа
    catalog['Артикул_Размер'] = catalog[['Артикул продавца', 'Размер']].apply(lambda row: '_size_'.join(row.values.astype(str)), axis=1)
    df_sku_and_other_costs['Артикул_Размер'] = df_sku_and_other_costs[['Артикул продавца', 'Размер']].apply(lambda row: '_size_'.join(row.values.astype(str)), axis=1)

    # Убираем артикул и размер
    catalog = catalog.loc[:, ~catalog.columns.isin(['Артикул продавца', 'Размер'])]
    df_sku_and_other_costs = df_sku_and_other_costs.loc[:, ~df_sku_and_other_costs.columns.isin(['Артикул продавца', 'Размер'])]

    # Мерджим расходы и справочную таблицу
    df_sku_and_other_costs_with_catalog = df_sku_and_other_costs.merge(catalog[['Артикул_Размер', 'Себестоимость']],
                                                                       how='left',
                                                                       on='Артикул_Размер')


    # Достаем колонки с Артикулом и Размером из общей колонки
    df_sku_and_other_costs_with_catalog = df_sku_and_other_costs_with_catalog.assign(**{
        'Артикул продавца': df_sku_and_other_costs_with_catalog['Артикул_Размер'].str.split('_size_', expand=True)[0],
        'Размер': df_sku_and_other_costs_with_catalog['Артикул_Размер'].str.split('_size_', expand=True)[1]
    })
    # Перемещаем колонку с Артикулом+размером в начало df
    df_sku_and_other_costs_with_catalog = move_columns(
        df_sku_and_other_costs_with_catalog,
        columns_to_move=['Артикул_Размер', 'Артикул продавца', 'Размер'],
        position='Артикул WB',
        insert_type='after'
    )
    # Считаем себестоимость
    df_sku_and_other_costs_with_catalog['Себестоимость * Факт продаж'] = abs(df_sku_and_other_costs_with_catalog['Себестоимость'] * df_sku_and_other_costs_with_catalog['Продажи минус возвраты шт'])

    return df_sku_and_other_costs_with_catalog


# Функция расчета финальных расходов
def calc_final_costs(df_all_costs_stats_with_catalog):
    # Создание копии для избежания изменений в оригинальном df
    df_all_costs_stats_ = df_all_costs_stats_with_catalog.copy()
    # Колонки, по которым нужно поменять знаки расходов с + на -
    change_sign_columns = [
        'Возмещение издержек по перевозке',
        'Добровольная компенсация при возврате',
        'Компенсация ущерба',
        'Логистика сторно'
    ]
    # Изменяем знак некоторых расходов с + на -
    for col in change_sign_columns:
        if col in df_all_costs_stats_.columns:
            df_all_costs_stats_[col] = df_all_costs_stats_[col] * (-1)

    # Колонки, которые нужно исключить из затрат
    columns_to_exclude = [
        'Артикул WB',
        'Артикул продавца',
        'Размер',
        'Артикул_Размер',
        'Статус',
        'Заказы шт',
        'Продажи шт',
        'Продажи руб',
        'Возвраты шт',
        'Количество размеров в артикуле',
        'Продажи в артикуле',
        'Доля продаж в артикуле',
        'Продажи руб',
        'Маржинальность руб',
        'Возвраты шт',
        'Продажи минус возвраты руб',
        'Продажи минус возвраты шт',
        'Себестоимость',
        # 'Вознаграждение WB без НДС',
        # 'НДС с вознаграждения WB',
        'Комиссия WB, %',
        # 'Платное хранение (отдельный отчет)',
        'Платная приемка (отдельный отчет)'
    ]

    # Расчет маржинальности в рублях
    df_all_costs_stats_['Маржинальность руб'] = df_all_costs_stats_['Продажи руб']
    for col in df_all_costs_stats_.columns[~df_all_costs_stats_.columns.isin(columns_to_exclude + ['Итого руб'])]:
        df_all_costs_stats_['Маржинальность руб'] = df_all_costs_stats_['Маржинальность руб'] - df_all_costs_stats_[col]

    # Расчет Маржинальности в процентах
    # Там, где нет себестоимости, маржинальность не считаем
    df_all_costs_stats_.loc[df_all_costs_stats_['Себестоимость'].isna(), 'Маржинальность руб'] = np.nan
    df_all_costs_stats_['Маржинальность %'] = df_all_costs_stats_['Маржинальность руб'] / df_all_costs_stats_['Продажи руб'] * 100
    # Меняем inf на nan
    df_all_costs_stats_['Маржинальность %'] = df_all_costs_stats_['Маржинальность %'].replace([-np.inf, np.inf], np.nan)

    return df_all_costs_stats_


# Функция расчета сводной таблицы по всем расходам
def create_svod_by_operations(df_final_costs):
    # Список колонок, по котором мы не считаем суммы
    columns_to_exclude = [
        'Артикул WB',
        'Артикул_Размер',
        'Артикул продавца',
        'Размер',
        'Количество размеров в артикуле',
        'Продажи в артикуле',
        'Доля продаж в артикуле',
        'Комиссия WB, %',
        'Себестоимость',
        # 'Себестоимость * Факт продаж',
        # 'Все затраты (руб)',
        # 'Маржинальность руб',
        'Маржинальность %',
    ]

    # Считаем суммы по остальным столбцам
    df_final_costs_by_operation = (
        df_final_costs
        # Убираем колонки, которые не должны быть включены в сумму
        .loc[
            :,
            ~df_final_costs.columns.isin(columns_to_exclude)
        ]
        # Считаем суммы по остальным
        .sum()
        # Достаем столбцы из index
        .reset_index()
        # Переименовываем колонки
        .rename(columns={
            'index': 'Тип операции',
            0: 'Итого руб'
        })

    )
    # Считаем сумму, от которой будем считать процент
    sum_columns = [
        'Продажи минус возвраты руб',
        # 'Баллы за скидки',
    ]
    # Считается сумма по тем колонкам, которые есть в df_final_costs
    sum_for_percent = (
        df_final_costs
        .loc[:, df_final_costs.columns.isin(sum_columns)]
        .sum().sum()
    )
    # Считаем процент
    df_final_costs_by_operation['Итого, %'] = (
        df_final_costs_by_operation['Итого руб'] / sum_for_percent * 100
    )
    # Делаем процент пропуском в некоторых расходах
    no_percent_operations = [
        'Продажи шт',
        'Возвраты шт',
        'Продажи минус возвраты шт',
    ]
    # Маска для loc
    mask_no_operations = df_final_costs_by_operation['Тип операции'].isin(no_percent_operations)
    # Создание пропусков
    df_final_costs_by_operation.loc[mask_no_operations, 'Итого, %'] = np.nan

    return df_final_costs_by_operation


# Функция сохранения результатов в excel
def save_excel(
        finance_reports_dir, date_report,
        df_realization_report_new_columns,
        report_dates,
        df_final_costs,
        df_final_costs_by_operation,
        result_sku_costs,
        result_other_costs,
        result_paid_storage,
        result_paid_acceptance
        ):
    # Получаем результаты по расходам, которые уже были разбиты на артикулы
    df_sku_costs_by_operation = result_sku_costs['df_sku_costs_by_operation']
    # Получаем результаты по платному хранению
    df_paid_storage = result_paid_storage['df_paid_storage']
    # Получаем результаты по платной приемке
    df_paid_acceptance = result_paid_acceptance['df_paid_acceptance']
    # Получаем результаты по прочим расходам
    df_other_costs_by_operation = result_other_costs['df_other_costs_by_operation']
    with pd.ExcelWriter(f"{finance_reports_dir}/{date_report}_Свод_реализации_WB_{client_name}.xlsx") as w:
        report_dates.to_excel(w, sheet_name='Список отчетных недель', index=False)
        df_realization_report_new_columns.to_excel(w, sheet_name='Исходный отчет АПИ', index=False)
        df_final_costs.to_excel(w, sheet_name='Итоговая таблица', index=False)
        df_final_costs_by_operation.to_excel(w, sheet_name='Сводная', index=False)
        df_sku_costs_by_operation.to_excel(w, sheet_name='Расходы по SKU', index=False)
        df_other_costs_by_operation.to_excel(w, sheet_name='Прочие расходы', index=False)
        df_paid_acceptance.to_excel(w, sheet_name='Отчет пл. приемка', index=False)
        df_paid_storage.to_excel(w, sheet_name='Отчет пл. хранение', index=False)


# Функция форматирования отчета Excel
def format_excel(finance_reports_dir, date_report, df_final_costs):
    # Открываем файл с расчетами Excel
    wb = openpyxl.load_workbook(f"{finance_reports_dir}/{date_report}_Свод_реализации_WB_{client_name}.xlsx")
    # Автоподбор ширины столбца
    for worksheet in wb.sheetnames:
        ws = wb[worksheet]
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        # Значок фильтра на столбцы
        ws.auto_filter.ref = ws.dimensions
    final_costs_worksheet = wb['Итоговая таблица']

    # df с соответствием заголовков и названий столбцов
    excel_columns = pd.DataFrame({"column": df_final_costs.columns,
                                  "column_number": np.arange(1, len(df_final_costs.columns) + 1)})
    excel_columns['excel_column'] = excel_columns['column_number'].apply(lambda x: get_column_letter(x))
    # Кол-во строк в df
    svod_len = df_final_costs.shape[0]
    # Номер строки, откуда начинается запись
    row_start = 2

    # Мин. и макс. колонка
    min_col = excel_columns.loc[excel_columns['column_number'].idxmin(), 'excel_column']
    max_col = excel_columns.loc[excel_columns['column_number'].idxmax(), 'excel_column']
    # Заголовки
    header_cells = final_costs_worksheet[f"{min_col}{row_start - 1}:{max_col}{row_start - 1}"]
    # Все колонки, кроме заголовков
    all_cells = final_costs_worksheet[f"{min_col}{row_start}:{max_col}{svod_len + 1}"]

    # Формат заголовков
    thin_border = Side(border_style="thin", color="000000")
    for row in header_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11, bold=True)
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center',
                                       wrap_text=True)
            cell.border = Border(top = thin_border, bottom = thin_border,
                                 right = thin_border, left = thin_border)
    # Границы (сетка)
    for row in all_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11)
            cell.border = Border(top = thin_border, bottom = thin_border,
                                 right = thin_border, left = thin_border)


    # Колонка с маржинальностью
    marginality_percent = excel_columns.loc[excel_columns['column'].str.contains('Маржинальность %'), 'excel_column'].values[0]

    # Формат числа 0% у маржинальности
    # percent_0_digit_cells = [marginality_percent]
    # for i in range(len(percent_0_digit_cells)):
    #     percent_0_digit_cells[i] = f"{percent_0_digit_cells[i]}{row_start}:{percent_0_digit_cells[i]}{svod_len + 1}"

    # for cell_range in percent_0_digit_cells:
    #     cell_range = final_costs_worksheet[cell_range]
    #     for row in cell_range:
    #         for cell in row:
    #             cell.number_format = '0%'

    wb.save(f"{finance_reports_dir}/{date_report}_Свод_реализации_WB_{client_name}.xlsx")


# %% Вызов всех функций
if __name__ == '__main__':
    # Имя папки, в которую будет сохранен отчет
    date_report = '2025_07'
    # Даты, за которые нужно сформировать отчет
    date_start = '2025-07-21T00:00:00'
    date_end = '2025-07-31T23:59:59'
    # Директория для сохранения результатов
    finance_reports_dir = f"{marketplace_dir_name}/Clients/{client_name}/FinanceReports/{date_report}"
    if not os.path.exists(finance_reports_dir):
        os.makedirs(finance_reports_dir)

    # Получение отчета о реализации по АПИ
    df_realization_report = upload_realization_report_v2(
        headers,
        date_start,
        date_end
    )
    # Переименование колонок отчета о реализации в русские названия
    df_realization_report_renamed = rename_report_columns(
        df_realization_report
    )
    # Расчет некоторых дополнительных колонок, которые будут участвовать в расчетах
    df_realization_report_new_columns = add_new_columns(
        df_realization_report_renamed
    )
    # Фильтр по переходным неделям
    df_realization_report_date_filtered = filter_report_dates(
        date_start,
        date_end,
        df_realization_report_new_columns,
        filter_dates=True
    )
    # Получаем список дат отчета
    report_dates = get_report_dates(
        df_realization_report_date_filtered
    )

    # Расчет затрат, которые уже разбиты по артикулам
    result_sku_costs = calc_sku_costs(
        df_realization_report_date_filtered
    )
    # # Выгрузка заказов
    # df_orders = get_orders(
    #     headers,
    #     date_start,
    #     date_end,
    #     to_save=False
    # )
    # Добавляем заказы в отчет
    # df_sku_costs_and_orders = add_orders_to_sku_costs(result_sku_costs, df_orders)
    # Расходы на платное хранение (отдельный отчет)
    result_paid_storage = calc_paid_storage_costs(
        headers,
        df_realization_report_date_filtered,
        date_start,
        date_end
    )
    # Объединяем расходы, которые уже разбиты по размерам
    df_sku_sizes_costs_merged = union_sku_sizes_costs(
        result_sku_costs,
        result_paid_storage
    )
    # Расходы на платную приемку (отдельный отчет)
    result_paid_acceptance = calc_paid_acceptance_costs(
        headers,
        df_sku_sizes_costs_merged,
        df_realization_report_date_filtered,
        date_start,
        date_end
    )
    # Считаем расходы на рекламные кампании
    df_companies_costs = calc_companies_costs(
        df_realization_report_date_filtered,
        df_sku_sizes_costs_merged,
        date_report
    )
    # Объединяем расходы на пл. приемку с расходами по SKU и хранению
    # df_sku_paid_storage_acceptance_stats = union_sku_paid_storage_acceptance_costs(df_sku_and_paid_storage_stats, result_paid_acceptance)
    # Объединяем все расходы из отдельных отчетов в один df
    df_sku_sizes_costs_all = merge_all_costs(
        df_sku_sizes_costs_merged,
        result_paid_acceptance,
        df_companies_costs
    )
    # Расчет затрат, которые не привязаны к артикулам
    # result_other_costs = calc_other_costs(df_realization_report_date_filtered, result_sku_costs)
    result_other_costs = calc_other_costs_v2(
        result_sku_costs,
        df_sku_sizes_costs_all
    )
    # Список текущих товаров по апи
    # df_products = getWBProduct(headers, to_save=False)
    # Объединение расходов
    # df_all_costs_stats = merge_all_costs(result_other_costs,
    #                                      result_paid_storage,
    #                                      result_paid_acceptance)
    # Достаем получившиеся расходы после расчета прочих расходов
    df_all_costs_stats = result_other_costs['df_sku_and_other_costs'].copy()
    # Добавляем данные из справочной таблицы
    df_all_costs_stats_with_catalog = add_data_from_catalog(
        result_other_costs
    )
    # Расчет итоговой колонки по затратам
    df_final_costs = calc_final_costs(
        df_all_costs_stats_with_catalog
    )
    # Считаем сводную по затратам
    df_final_costs_by_operation = create_svod_by_operations(df_final_costs)
    # Сохранение результатов в excel
    save_excel(
        finance_reports_dir,
        date_report,
        df_realization_report_new_columns,
        report_dates,
        df_final_costs,
        df_final_costs_by_operation,
        result_sku_costs,
        result_other_costs,
        result_paid_storage,
        result_paid_acceptance
    )
    # Форматирование файла excel
    format_excel(
        finance_reports_dir,
        date_report,
        df_final_costs
    )
    print(f"\033[32m\033[47mDone calculating Finance Report for Client {client_name}")


# %%
