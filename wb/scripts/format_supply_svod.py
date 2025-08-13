import pandas as pd
import numpy as np
import os
import shutil
from loguru import logger
import openpyxl
from datetime import date,datetime,timedelta
from itertools import product
from openpyxl import Workbook, formatting
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.formatting.rule import ColorScale, ColorScaleRule, FormatObject
from openpyxl.utils import get_column_letter

# Некоторые константы
# from wb.scripts.constants import client_name

# date_report_created = '2025-02-03'
# svod_excel = pd.read_excel(f"Clients/{client_name}/SupplySvod/{date_report_created}_Расчет_Поставок_{client_name}_WB.xlsx", sheet_name='Всего')
# svod_excel_clusters = pd.read_excel(f"Clients/{client_name}/SupplySvod/{date_report_created}_Расчет_Поставок_{client_name}_WB.xlsx", sheet_name='По кластерам')
# clusters_mapping_df = pd.read_excel('wb_warehouses_mapping.xlsx', sheet_name='Монопалеты')
# cluster_column = 'Группировка'

# Создание копии оригинального файла, в котором будет производиться форматирование
def copy_supply_svod_file(path_supply_svod, client_name, date_report_created):
    src_file = f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_WB.xlsx"
    dst_file = f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_WB_formatted.xlsx"
    shutil.copy(src_file, dst_file)

# Форматирование листа "Всего"
def format_sheet_total(
        path_supply_svod,
        client_name,
        date_report_created,
        svod_excel,
        svod_excel_clusters,
        clusters_mapping_df,
        cluster_column,
        sheet_name='Всего'
    ):

    logger.info(f"Formatting sheet \"{sheet_name}\"")

    # Открываем файл с расчетами Excel
    wb = openpyxl.load_workbook(f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_WB_formatted.xlsx")
    ws = wb[sheet_name]

    # Кол-во строк в df
    svod_len = svod_excel.shape[0]

    # Номер строки, откуда начинается запись
    row_start = 2

    # df с соответствием заголовков и названий столбцов
    excel_columns = pd.DataFrame({"column": svod_excel.columns,
                                  "column_number": np.arange(1, len(svod_excel.columns) + 1)})
    excel_columns['excel_column'] = excel_columns['column_number'].apply(lambda x: get_column_letter(x))

    # Выборка различных групп колонок по их названиям
    # Мин. и макс. колонка
    min_col = excel_columns.loc[excel_columns['column_number'].idxmin(), 'excel_column']
    max_col = excel_columns.loc[excel_columns['column_number'].idxmax(), 'excel_column']
    # Заголовки
    header_cells = ws[f"{min_col}{row_start - 1}:{max_col}{row_start - 1}"]
    # Все колонки, кроме заголовков
    all_cells = ws[f"{min_col}{row_start}:{max_col}{svod_len + 1}"]
    # Штрихкод
    barcode = excel_columns.loc[excel_columns['column'] == 'Штрихкод', 'excel_column'].values[0]
    # Артикул
    article = excel_columns.loc[excel_columns['column'] == 'Артикул продавца', 'excel_column'].values[0]
    # Размер
    product_size = excel_columns.loc[excel_columns['column'] == 'Размер', 'excel_column'].values[0]
    # Артикул+Размер
    article_size = excel_columns.loc[excel_columns['column'] == 'Артикул_Размер', 'excel_column'].values[0]
    # Оборачиваемость
    turnover = excel_columns.loc[excel_columns['column'] == 'Оборачиваемость', 'excel_column'].values[0]
    # Заказы
    orders = excel_columns.loc[excel_columns['column'].str.contains('ЗАКАЗЫ'), 'excel_column'].values[0]
    # Продажи
    sales = excel_columns.loc[excel_columns['column'].str.contains('ПРОДАЖИ'), 'excel_column'].values[0]
    # Остатки
    reminders = excel_columns.loc[excel_columns['column'].str.contains('ОСТАТОК') & ~(excel_columns['column'].str.contains('FBS')), 'excel_column'].values[0]
    # Остатки FBS
    reminders_fbs = excel_columns.loc[excel_columns['column'].str.contains('ОСТАТОК') & excel_columns['column'].str.contains('FBS'), 'excel_column'].values[0]
    # Потребность на 40 дней
    demand_40_days = excel_columns.loc[excel_columns['column'] == 'Потребность на 40 дней', 'excel_column'].values[0]
    # Потребность на 60 дней
    demand_60_days = excel_columns.loc[excel_columns['column'] == 'Потребность на 60 дней', 'excel_column'].values[0]
    # Дефицит\Избыток 40 дней
    deficit_40_days = excel_columns.loc[excel_columns['column'] == 'Дефицит/Избыток 40 дней', 'excel_column'].values[0]
    # Дефицит\Избыток 40 дней с учетом FBS
    deficit_40_days_fbs = excel_columns.loc[excel_columns['column'] == 'Дефицит/Избыток 40 дней с учетом FBS', 'excel_column'].values[0]
    # Дефицит\Избыток 60 дней
    deficit_60_days = excel_columns.loc[excel_columns['column'] == 'Дефицит/Избыток 60 дней', 'excel_column'].values[0]
    # Потребность на 40 дней (округл.)
    demand_40_days_rounded = excel_columns.loc[excel_columns['column'] == 'Потребность на 40 дней (округл.)', 'excel_column'].values[0]
    # Потребность на 40 дней с учетом FBS (округл.)
    demand_40_days_rounded_fbs = excel_columns.loc[excel_columns['column'] == 'Потребность на 40 дней с учетом FBS (округл.)', 'excel_column'].values[0]
    # Потребность на 60 дней (округл.)
    demand_60_days_rounded = excel_columns.loc[excel_columns['column'] == 'Потребность на 60 дней (округл.)', 'excel_column'].values[0]
    # Список кластеров
    cluster_list = excel_columns.loc[excel_columns['column'].str.contains('Потребность(?! на).*'), 'column'].str.replace('Потребность ', '').to_list()

    # Нужные колонки с листа По кластерам
    # Кол-во строк в df
    svod_len_clusters = svod_excel_clusters.shape[0]

    # Номер строки, откуда начинается запись
    row_start_clusters = 2

    # df с соответствием заголовков и названий столбцов
    excel_columns_clusters = pd.DataFrame({"column": svod_excel_clusters.columns,
                                           "column_number": np.arange(1, len(svod_excel_clusters.columns) + 1)})
    excel_columns_clusters['excel_column'] = excel_columns_clusters['column_number'].apply(lambda x: get_column_letter(x))

    # Артикул (Кластеры)
    article_from_clusters = excel_columns_clusters.loc[excel_columns_clusters['column'] == 'Артикул продавца', 'excel_column'].values[0]
    # Размер (Кластеры)
    product_size_from_clusters = excel_columns_clusters.loc[excel_columns_clusters['column'] == 'Размер', 'excel_column'].values[0]
    # Артикул+Размер (Кластеры)
    article_size_from_clusters = excel_columns.loc[excel_columns['column'] == 'Артикул_Размер', 'excel_column'].values[0]
    # Кластер (Кластеры)
    cluster_from_clusters = excel_columns_clusters.loc[excel_columns_clusters['column'] == 'Склад', 'excel_column'].values[0]
    # Корректировка с учетом нулевых остатков (Кластеры)
    correction_null_reminders_from_clusters = excel_columns_clusters.loc[excel_columns_clusters['column'] == 'Корректировка с учетом нулевых остатков', 'excel_column'].values[0]


    # Форматирование
    # Стили границ
    thin = Side(border_style="thin", color="000000")

    # Форматирование заголовков таблицы
    for row in header_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11,
                                       bold=True)
            cell.alignment = Alignment(horizontal='center',
                                                 wrap_text=True)
            cell.border = Border(top = thin, bottom = thin,
                                                        right = thin, left = thin)
    # Значок фильтра на столбцы
    ws.auto_filter.ref = ws.dimensions

    # Границы (сетка)
    for row in all_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11)
            cell.border = Border(top = thin, bottom = thin,
                                                        right = thin, left = thin)
    # Автоподбор ширины столбца
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # У столбца с баркодом делаем фиксированную ширину,
    # т.к. там бывает много значений
    ws.column_dimensions[barcode].width = 22

    # Суммирование данных с листа по кластерам
    clusters_mapping_df_ = clusters_mapping_df.copy()
    # Выбираем нужные колонки из таблицы соответствия кластеров
    clusters_mapping_df_ = clusters_mapping_df_.loc[:, ['Склад', cluster_column]]
    # Объединяем список субкластеров в одну строку
    clusters_mapping_df_ = clusters_mapping_df_.groupby([cluster_column]).agg(list).reset_index()
    # Выбираем нужные кластеры (только те, которые встретились в выгрузках)
    clusters_mapping_df_ = clusters_mapping_df_.loc[clusters_mapping_df_[cluster_column].isin(cluster_list), :]
    clusters_mapping_df_ = clusters_mapping_df_.reset_index(drop=True)
    # Переименовываем колонки со складами для удобства
    clusters_mapping_df_ = clusters_mapping_df_.rename(columns={
        'Склад': 'subcluster',
        cluster_column: 'cluster'
    })
    # clusters_mapping_df_['cluster_and_subcluster'] = clusters_mapping_df_['cluster'].map(list) + clusters_mapping_df_['subcluster']
    for i in range(clusters_mapping_df_.shape[0]):
        # Выбираем список кластеров из таблицы соответствия
        cluster = clusters_mapping_df_['cluster'][i]
        subcluster_list = clusters_mapping_df_['subcluster'][i].copy()
        # Добавляем кластер в список субкластеров
        subcluster_list.insert(0, cluster)
        # Удаляем дубликаты из субкластеров
        subcluster_list = list(dict.fromkeys(subcluster_list))

        # Выбираем колонку с потребностью для кластера на листе всего
        demand_cluster_column = excel_columns.loc[excel_columns['column'] == f"Потребность {cluster}", 'excel_column'].values[0]
        for k in range(row_start, svod_len + row_start):
            # Строка, куда будет помещаться итоговая формула
            subcluster_formula = ''
            for subcluster in subcluster_list:
                # Прописываем формулу, включающую в себя сумму по субкластерам
                subcluster_formula_tmp = (
                     f"SUMIFS('По кластерам'!${correction_null_reminders_from_clusters}${row_start_clusters}:${correction_null_reminders_from_clusters}${svod_len_clusters + 1},"
                     f"'По кластерам'!${article_size_from_clusters}${row_start_clusters}:${article_size_from_clusters}${svod_len_clusters + 1}, ${article_size}{k},"
                     f"'По кластерам'!${cluster_from_clusters}${row_start_clusters}:${cluster_from_clusters}${svod_len_clusters + 1}, \"{subcluster}\""
                     f")\n"
                )
                # Суммируем с предыдущим кластером
                subcluster_formula = f"{subcluster_formula_tmp} + {subcluster_formula}"
                # Удаляем лишние символы в конце строки, если это последний субкластер
                subcluster_formula = subcluster_formula.rstrip(" + ''")
            # Добавляем = вначале формулы, чтобы эксель воспринимал это как формулу
            subcluster_formula = f"= {subcluster_formula}"
            ws[f'{demand_cluster_column}{k}'] = subcluster_formula

    # Формулы
    for i in range(row_start, svod_len + row_start):
        # Оборачиваемость
        ws[f'{turnover}{i}'] = f'= ({orders}{i} + {sales}{i}) / 2 / 30'
        # Потребность на 40 дней
        ws[f'{demand_40_days}{i}'] = f'= {turnover}{i} * 40'
        # Потребность на 60 дней
        ws[f'{demand_60_days}{i}'] = f'= {turnover}{i} * 60'
        # Дефицит/избыток 40 дней
        ws[f'{deficit_40_days}{i}'] = f'= {reminders}{i} - {demand_40_days}{i}'
        # Дефицит/избыток 40 дней с учетом FBS
        ws[f'{deficit_40_days_fbs}{i}'] = f'= {reminders}{i} + {reminders_fbs}{i} - {demand_40_days}{i}'
        # Дефицит/избыток 60 дней
        ws[f'{deficit_60_days}{i}'] = f'= {reminders}{i} - {demand_60_days}{i}'
        # Потребность на 40 дней (округл.)
        ws[f'{demand_40_days_rounded}{i}'] = f'= ROUNDUP(IF({deficit_40_days}{i} > 0, 0, -{deficit_40_days}{i}),0)'
        # Потребность на 40 дней с учетом FBS (округл.)
        ws[f'{demand_40_days_rounded_fbs}{i}'] = f'= ROUNDUP(IF({deficit_40_days_fbs}{i} > 0, 0, -{deficit_40_days_fbs}{i}),0)'
        # Потребность на 60 дней (округл.)
        ws[f'{demand_60_days_rounded}{i}'] = f'= ROUNDUP(IF({deficit_60_days}{i} > 0, 0, -{deficit_60_days}{i}),0)'

    # Условное форматирование для колонок Дефицит\Избыток на 40 и 60 дней
    red_font = Font(color='9C0006')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_font = Font(color='006100')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    deficit_cols = [deficit_40_days, deficit_40_days_fbs, deficit_60_days]
    for i in range(len(deficit_cols)):
        deficit_cols[i] = f"{deficit_cols[i]}{row_start}:{deficit_cols[i]}{svod_len + 1}"
    for cell_range in deficit_cols:
        ws.conditional_formatting.add(cell_range,
                                formatting.rule.CellIsRule(
                                    operator='lessThan',
                                    formula=['0'],
                                    fill=red_fill, font=red_font))
        ws.conditional_formatting.add(cell_range,
                                formatting.rule.CellIsRule(
                                    operator='greaterThan',
                                    formula=['0'],
                                    fill=green_fill, font=green_font))
    # 2 знака после запятой у чисел
    svod_2num_cells = [turnover, demand_40_days, demand_60_days]
    for i in range(len(svod_2num_cells)):
        svod_2num_cells[i] = f"{svod_2num_cells[i]}{row_start}:{svod_2num_cells[i]}{svod_len + 1}"
    for cell_range in svod_2num_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.number_format = "0.00"

    # 1 знак после запятой у чисел
    svod_1num_cells = [deficit_40_days, deficit_40_days_fbs, deficit_60_days]
    for i in range(len(svod_1num_cells)):
        svod_1num_cells[i] = f"{svod_1num_cells[i]}{row_start}:{svod_1num_cells[i]}{svod_len + 1}"
    for cell_range in svod_1num_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.number_format = "0.0"

    # 0 знаков после запятой у чисел
    svod_0num_cells = [barcode]
    for i in range(len(svod_0num_cells)):
        svod_0num_cells[i] = f"{svod_0num_cells[i]}{row_start}:{svod_0num_cells[i]}{svod_len + 1}"
    # Цикл по набору ячеек
    for cell_range in svod_0num_cells:
        # Выбираем набор ячеек (напр, A1:A100)
        cell_range_ws = ws[cell_range]
        # Цикл по строке в наборе ячеек
        for row in cell_range_ws:
            # Цикл по ячейке в строке
            for cell in row:
                # Если в ячейке число, пробуем перевести его в int
                if isinstance(cell.value, (float,int)):
                    cell.number_format = '0'
                # Если в ячейке строка, содержащая E+, переводим во float
                # и задаем 0 знаков после запятой
                elif isinstance(cell.value, str) and "E+" in cell.value:
                    # Если число в строковом виде
                    cell.value = int(float(cell.value))
                    cell.number_format = '0'

    # Выравнивание по центру
    svod_middle_cells = excel_columns.loc[~excel_columns['column'].str.contains('Артикул_Размер|Штрихкод|Предмет|Артикул продавца|Размер|Наименование'), 'excel_column'].to_list()
    for i in range(len(svod_middle_cells)):
        svod_middle_cells[i] = f"{svod_middle_cells[i]}{row_start}:{svod_middle_cells[i]}{svod_len + 1}"
    for cell_range in svod_middle_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.alignment = Alignment(horizontal = "center",
                                                            vertical = "center")
    # Выравнивание слева
    svod_left_cells = excel_columns.loc[excel_columns['column'].str.contains('Артикул_Размер|Штрихкод|Предмет|Артикул продавца|Размер|Наименование'), 'excel_column'].to_list()
    for i in range(len(svod_left_cells)):
        svod_left_cells[i] = f"{svod_left_cells[i]}{row_start}:{svod_left_cells[i]}{svod_len + 1}"
    for cell_range in svod_left_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.alignment = Alignment(horizontal = "left",
                                                            vertical = "center")

    #return wb
    # Сохранение файла Excel
    wb.save(f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_WB_formatted.xlsx")
    logger.info(f"Finished formatting sheet {sheet_name}")

# Форматирование листа "По кластерам"
def format_sheet_clusters(
        path_supply_svod,
        client_name,
        date_report_created,
        svod_excel,
        sheet_name='По кластерам'
    ):

    logger.info(f"Formatting sheet \"{sheet_name}\"")

    # Открываем файл с расчетами Excel
    wb = openpyxl.load_workbook(f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_WB_formatted.xlsx")
    ws = wb[sheet_name]

    # Кол-во строк в df
    svod_len = svod_excel.shape[0]

    # Номер строки, откуда начинается запись
    row_start = 2

    # df с соответствием заголовков и названий столбцов
    excel_columns = pd.DataFrame({"column": svod_excel.columns,
                                  "column_number": np.arange(1, len(svod_excel.columns) + 1)})
    excel_columns['excel_column'] = excel_columns['column_number'].apply(lambda x: get_column_letter(x))

    # Выборка различных групп колонок по их названиям
    # Мин. и макс. колонка
    min_col = excel_columns.loc[excel_columns['column_number'].idxmin(), 'excel_column']
    max_col = excel_columns.loc[excel_columns['column_number'].idxmax(), 'excel_column']
    # Заголовки
    header_cells = ws[f"{min_col}{row_start - 1}:{max_col}{row_start - 1}"]
    # Все колонки, кроме заголовков
    all_cells = ws[f"{min_col}{row_start}:{max_col}{svod_len + 1}"]
    # Штрихкод
    barcode = excel_columns.loc[excel_columns['column'] == 'Штрихкод', 'excel_column'].values[0]
    # Оборачиваемость
    turnover = excel_columns.loc[excel_columns['column'] == 'Оборачиваемость', 'excel_column'].values[0]
    # Заказы
    orders = excel_columns.loc[excel_columns['column'].str.contains('ЗАКАЗЫ'), 'excel_column'].values[0]
    # Продажи
    sales = excel_columns.loc[excel_columns['column'].str.contains('ПРОДАЖИ'), 'excel_column'].values[0]
    # Остатки
    reminders = excel_columns.loc[excel_columns['column'].str.contains('ОСТАТОК'), 'excel_column'].values[0]
    # Потребность на 40 дней
    demand_40_days = excel_columns.loc[excel_columns['column'] == 'Потребность на 40 дней', 'excel_column'].values[0]
    # Потребность на 60 дней
    demand_60_days = excel_columns.loc[excel_columns['column'] == 'Потребность на 60 дней', 'excel_column'].values[0]
    # Дефицит\Избыток 40 дней
    deficit_40_days = excel_columns.loc[excel_columns['column'] == 'Дефицит/Избыток 40 дней', 'excel_column'].values[0]
    # Дефицит\Избыток 60 дней
    deficit_60_days = excel_columns.loc[excel_columns['column'] == 'Дефицит/Избыток 60 дней', 'excel_column'].values[0]
    # Потребность на 40 дней (округл.)
    demand_40_days_rounded = excel_columns.loc[excel_columns['column'] == 'Потребность на 40 дней (округл.)', 'excel_column'].values[0]
    # Потребность на 60 дней (округл.)
    demand_60_days_rounded = excel_columns.loc[excel_columns['column'] == 'Потребность на 60 дней (округл.)', 'excel_column'].values[0]
    # Корректировка с учетом нулевых остатков
    correction_null_reminders = excel_columns.loc[excel_columns['column'] == 'Корректировка с учетом нулевых остатков', 'excel_column'].values[0]

    # Форматирование
    # Стили границ
    thin = Side(border_style="thin", color="000000")

    # Форматирование заголовков таблицы
    for row in header_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11,
                                       bold=True)
            cell.alignment = Alignment(horizontal='center',
                                                 wrap_text=True)
            cell.border = Border(top = thin, bottom = thin,
                                                        right = thin, left = thin)
    # Значок фильтра на столбцы
    ws.auto_filter.ref = ws.dimensions

    # Границы (сетка)
    for row in all_cells:
        for cell in row:
            cell.font = Font(name = "Calibri", size = 11)
            cell.border = Border(top = thin, bottom = thin,
                                                        right = thin, left = thin)
    # Автоподбор ширины столбца
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # У столбца с баркодом делаем фиксированную ширину,
    # т.к. там бывает много значений
    ws.column_dimensions[barcode].width = 22


    # Формулы
    for i in range(row_start, svod_len + row_start):
        # Оборачиваемость
        ws[f'{turnover}{i}'] = f'= ({orders}{i} + {sales}{i}) / 2 / 30'
        # Потребность на 40 дней
        ws[f'{demand_40_days}{i}'] = f'= {turnover}{i} * 40'
        # Потребность на 60 дней
        ws[f'{demand_60_days}{i}'] = f'= {turnover}{i} * 60'
        # Дефицит/избыток 40 дней
        ws[f'{deficit_40_days}{i}'] = f'= {reminders}{i} - {demand_40_days}{i}'
        # Дефицит/избыток 60 дней
        ws[f'{deficit_60_days}{i}'] = f'= {reminders}{i} - {demand_60_days}{i}'
        # Потребность на 40 дней (округл.)
        ws[f'{demand_40_days_rounded}{i}'] = f'= ROUNDUP(IF({deficit_40_days}{i} > 0, 0, -{deficit_40_days}{i}),0)'
        # Потребность на 60 дней (округл.)
        ws[f'{demand_60_days_rounded}{i}'] = f'= ROUNDUP(IF({deficit_60_days}{i} > 0, 0, -{deficit_60_days}{i}),0)'
        # Корректировка с учетом нулевых остатков
        ws[f'{correction_null_reminders}{i}'] = f'= {demand_60_days_rounded}{i}'

    # Условное форматирование для колонок Дефицит\Избыток на 40 и 60 дней
    red_font = Font(color='9C0006')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    green_font = Font(color='006100')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    deficit_cols = [deficit_40_days, deficit_60_days]
    for i in range(len(deficit_cols)):
        deficit_cols[i] = f"{deficit_cols[i]}{row_start}:{deficit_cols[i]}{svod_len + 1}"
    for cell_range in deficit_cols:
        ws.conditional_formatting.add(cell_range,
                                formatting.rule.CellIsRule(
                                    operator='lessThan',
                                    formula=['0'],
                                    fill=red_fill, font=red_font))
        ws.conditional_formatting.add(cell_range,
                                formatting.rule.CellIsRule(
                                    operator='greaterThan',
                                    formula=['0'],
                                    fill=green_fill, font=green_font))
    # 2 знака после запятой у чисел
    svod_2num_cells = [turnover, demand_40_days, demand_60_days]
    for i in range(len(svod_2num_cells)):
        svod_2num_cells[i] = f"{svod_2num_cells[i]}{row_start}:{svod_2num_cells[i]}{svod_len + 1}"
    for cell_range in svod_2num_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.number_format = "0.00"

    # 1 знак после запятой у чисел
    svod_1num_cells = [deficit_40_days, deficit_60_days]
    for i in range(len(svod_1num_cells)):
        svod_1num_cells[i] = f"{svod_1num_cells[i]}{row_start}:{svod_1num_cells[i]}{svod_len + 1}"
    for cell_range in svod_1num_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.number_format = "0.0"

    # 0 знаков после запятой у чисел
    svod_0num_cells = [barcode]
    for i in range(len(svod_0num_cells)):
        svod_0num_cells[i] = f"{svod_0num_cells[i]}{row_start}:{svod_0num_cells[i]}{svod_len + 1}"
    # Цикл по набору ячеек
    for cell_range in svod_0num_cells:
        # Выбираем набор ячеек (напр, A1:A100)
        cell_range_ws = ws[cell_range]
        # Цикл по строке в наборе ячеек
        for row in cell_range_ws:
            # Цикл по ячейке в строке
            for cell in row:
                # Если в ячейке число, пробуем перевести его в int
                if isinstance(cell.value, (float,int)):
                    cell.number_format = '0'
                # Если в ячейке строка, содержащая E+, переводим во float
                # и задаем 0 знаков после запятой
                elif isinstance(cell.value, str) and "E+" in cell.value:
                    # Если число в строковом виде
                    cell.value = int(float(cell.value))
                    cell.number_format = '0'


    # Выравнивание по центру
    svod_middle_cells = excel_columns.loc[~excel_columns['column'].str.contains('Штрихкод|Предмет|Артикул продавца|Размер|Наименование|Склад'), 'excel_column'].to_list()
    for i in range(len(svod_middle_cells)):
        svod_middle_cells[i] = f"{svod_middle_cells[i]}{row_start}:{svod_middle_cells[i]}{svod_len + 1}"
    for cell_range in svod_middle_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.alignment = Alignment(horizontal = "center",
                                                            vertical = "center")
    # Выравнивание слева
    svod_left_cells = excel_columns.loc[excel_columns['column'].str.contains('Штрихкод|Предмет|Артикул продавца|Размер|Наименование|Склад'), 'excel_column'].to_list()
    for i in range(len(svod_left_cells)):
        svod_left_cells[i] = f"{svod_left_cells[i]}{row_start}:{svod_left_cells[i]}{svod_len + 1}"
    for cell_range in svod_left_cells:
        cell_range_ws = ws[cell_range]
        for row in cell_range_ws:
            for cell in row:
                cell.alignment = Alignment(horizontal = "left",
                                                            vertical = "center")

    #return wb
    # Сохранение файла Excel
    wb.save(f"{path_supply_svod}/{date_report_created}_Расчет_Поставок_{client_name}_WB_formatted.xlsx")
    logger.info(f"Finished formatting sheet {sheet_name}")
