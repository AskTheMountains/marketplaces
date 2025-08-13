# %% Определение функций
import requests
import json
import time
from datetime import date,datetime,timedelta
import pandas as pd
import shutil
import os
from pathlib import Path
import csv
import numpy as np
import openpyxl
from openpyxl.styles import Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
import re
from loguru import logger
import getopt
import sys
pd.options.mode.chained_assignment = None  # default='warn'

# Файл с настройками и номером клиента
from wb.scripts.constants import (
    headers,
    client_name,
    marketplace_dir_name
)
# Функции выгрузки данных
from wb.scripts.uploadDataFromWB import(
    getWBProduct,
    getOrdersWB
)
# Некоторые вспомогательные функции
from generic_functions import move_columns
# Папка, где лежит текущий скрипт
BASE_DIR = Path(__file__).parent.parent


# Создание директории для выгрузки статистики РК за конкретные даты
def create_report_path(date_campaign_report, client_name):
    # Задаем путь к директории
    # upload_path = (
    #     f"Clients/{client_name}/CampaignsReport/{str(date.today())}/"
    #     f"{dt_start_dir}-{dt_end_dir}_API/"
    # )
    report_path = (
        f"{BASE_DIR}/"
        f"Clients/{client_name}/CampaignsReport/"
        f"{date_campaign_report}"
    )
    if not os.path.exists(report_path):
        os.makedirs(report_path, exist_ok=True)
        logger.info(f"Creating Companies Upload directory:{report_path}")

    return report_path


# Функция создания датафрейма со списком кампаний и датами их выгрузки
def create_companies_upload_df(
        input_companies,
):
    # Создаем df из словаря
    df_input_companies = (
        pd.DataFrame(
            input_companies,
            columns=['Номер кампании', 'date_start', 'date_end']
        )
        # .reset_index()
        # .rename(columns={
        #     'index': 'Номер кампании'
        # })
    )
    # Удаляем дубликаты
    df_input_companies = (
        df_input_companies
        .loc[~df_input_companies['Номер кампании'].isin(['']), :]
        .drop_duplicates(subset=['Номер кампании', 'date_start', 'date_end'])
    )
    # Присваиваем каждой кампании уникальный ID по номеру кампании и дате
    df_input_companies['unique_company_id'] = (
        df_input_companies
        .groupby(['Номер кампании', 'date_start', 'date_end'], as_index=False)
        .ngroup() + 1
    )
    # Сортировка
    df_input_companies = df_input_companies.sort_values(by=['unique_company_id'], ignore_index=True)
    # Перемещаем колонку в начало df
    df_input_companies = move_columns(
        df_input_companies,
        columns_to_move=['unique_company_id'],
        position=0
    )

    return df_input_companies

# Функция обработки списка товаров
def process_product_list(df_products):
    # Создаем копию для избежания изменений в оригинальном df
    df_products_processed = df_products.copy()

    return df_products_processed


# Функция получения списка кампаний
def get_company_list(
     headers
):
    # Делаем запрос для получения списка кампаний
    resp_data_company_list = requests.get("https://advert-api.wildberries.ru/adv/v1/promotion/count", headers=headers).json()
    # Переводим в df
    df_company_list_stats = pd.DataFrame(resp_data_company_list['adverts'])
    # Распаковываем список кампаний
    df_company_list_unpacked = df_company_list_stats.explode(column='advert_list').reset_index()
    # Получаем id кампаний
    df_company_ids = pd.json_normalize(df_company_list_unpacked['advert_list'])
    # Объединяем с исходным df
    df_company_list = pd.concat([df_company_ids, df_company_list_unpacked], axis=1)
    # Переводим колонку с датой в tiemstamp
    df_company_list['changeTime'] = pd.to_datetime(df_company_list['changeTime'], format='mixed').dt.tz_localize(None)

    return df_company_list


# Функция получения информации о кампаниях
def get_companies_info(headers, df_input_companies):
    # Создаем копию для избежания изменений в оригинальном df
    df_company_list_ = df_input_companies.copy()
    # Разбиваем df с кампаниями на интервалы по 50 кампаний
    step = 50
    df_company_list_['id'] = np.arange(0, len(df_company_list_))
    df_company_list_['chunks'] = df_company_list_['id'].apply(lambda x: int(x/step) + 1)
    # df для каждого из типов кампаний
    df_companies_6 = pd.DataFrame()
    df_companies_8 = pd.DataFrame()
    df_companies_9 = pd.DataFrame()
    logger.info("Uploading companies info")
    # Цикл по каждому интервалу
    for chunk in df_company_list_['chunks'].unique():
        # Получаем список кампаний
        df_company_list_chunk = df_company_list_.loc[df_company_list_['chunks'] == chunk, :]
        company_list = df_company_list_chunk['Номер кампании'].to_list()
        # Переводим номера кампаний в int для запроса
        # company_list_int = [int(elem) for elem in company_list]
        # Параметры запроса
        # params_companies_info = {company_list}
        resp_data_companies_info = requests.post("https://advert-api.wildberries.ru/adv/v1/promotion/adverts", headers=headers, json=company_list).json()
        # Переводим в df
        tmp_df_companies_info = pd.DataFrame(resp_data_companies_info)
        # Переводим колонки с датой в timestamp
        for col in ['endTime', 'createTime', 'changeTime', 'startTime']:
            tmp_df_companies_info[col] = pd.to_datetime(tmp_df_companies_info[col], format='mixed')
            # Убираем нахуй таймзону
            tmp_df_companies_info[col] = tmp_df_companies_info[col].dt.tz_localize(None)
        # Делаем выборку на каждый из типов кампаний
        tmp_df_companies_6 = tmp_df_companies_info.loc[~tmp_df_companies_info['type'].isin([8, 9]), :]
        tmp_df_companies_8 = tmp_df_companies_info.loc[tmp_df_companies_info['type'] == 8, :]
        tmp_df_companies_9 = tmp_df_companies_info.loc[tmp_df_companies_info['type'] == 9, :]
        # Объединяем с предыдущей итерацией цикла
        df_companies_6 = pd.concat([df_companies_6, tmp_df_companies_6])
        df_companies_8 = pd.concat([df_companies_8, tmp_df_companies_8])
        df_companies_9 = pd.concat([df_companies_9, tmp_df_companies_9])

    # Обработка данных кампаний со статусами, отличными от 8 и 9
    if not df_companies_6.empty:
        # Делаем reset_index после loc
        df_companies_6 = df_companies_6.reset_index(drop=True)
        # Убираем лишние колонки
        df_companies_6 = df_companies_6.loc[:, ~df_companies_6.columns.isin(['nmCPM', 'unitedParams', 'autoParams'])]
        # Переводим колонки с датой в timestamp
        # for col in ['endTime', 'createTime', 'changeTime', 'startTime']:
        #     df_companies_6[col] = pd.to_datetime(df_companies_6[col], format='mixed')
        # Получаем параметры кампании из списка
        df_companies_6_params = (
            df_companies_6
            .loc[:, ['endTime', 'createTime', 'changeTime', 'startTime', 'advertId', 'params']]
            .explode(column='params')
            .reset_index(drop=True)
        )
        # Получаем параметры кампании в виде столбцов df
        df_companies_6_params = pd.concat([
            df_companies_6_params,
            pd.json_normalize(df_companies_6_params['params'])
        ], axis=1)
        # Получаем список номенклатур кампании
        df_companies_6_nms = df_companies_6_params.explode(column='nms').reset_index(drop=True)
        # Получаем список номенклатур кампании в виде столбцов df
        df_companies_6_nms = pd.concat([df_companies_6_nms, pd.json_normalize(df_companies_6_nms['nms'])], axis=1)
        # Убираем лишние колонки
        df_companies_6_nms = df_companies_6_nms.loc[:, ['endTime', 'createTime', 'changeTime', 'startTime', 'advertId', 'nm']]
        # Переименовываем колонку чтобы дальше использовать concat
        df_companies_6_nms = df_companies_6_nms.rename(columns={'nm': 'nms'})
    else:
        df_companies_6_nms = pd.DataFrame()

    # Обработка данных кампаний со статусом 8
    if not df_companies_8.empty:
        # Делаем reset_index после loc
        df_companies_8 = df_companies_8.reset_index(drop=True)
        # Убираем лишние колонки
        df_companies_8 = df_companies_8.loc[:, ~df_companies_8.columns.isin(['nmCPM', 'unitedParams'])]
        # Переводим колонки с датой в timestamp
        # for col in ['endTime', 'createTime', 'changeTime', 'startTime']:
        #     df_companies_8[col] = pd.to_datetime(df_companies_8[col], format='mixed')
        # Получаем параметры кампании из списка
        df_companies_8_params = (
            df_companies_8
            .loc[:, ['endTime', 'createTime', 'changeTime', 'startTime', 'advertId', 'autoParams']]
        )
        # Получаем параметры кампании в виде столбцов df
        df_companies_8_params = pd.concat([df_companies_8_params, pd.json_normalize(df_companies_8_params['autoParams'])], axis=1)
        # Получаем список номенклатур
        df_companies_8_nms = df_companies_8_params.explode(column='nms')
        # Убираем лишние колонки
        df_companies_8_nms = df_companies_8_nms.loc[:, ['endTime', 'createTime', 'changeTime', 'startTime', 'advertId', 'nms']]
    else:
        df_companies_8_nms = pd.DataFrame()

    # Обработка данных кампаний со статусами, отличными от 8 и 9
    if not df_companies_9.empty:
        # Делаем reset_index после loc
        df_companies_9 = df_companies_9.reset_index(drop=True)
        # Убираем лишние колонки
        df_companies_9 = df_companies_9.loc[:, ~df_companies_9.columns.isin(['nmCPM', 'autoParams'])]
        # Переводим колонки с датой в timestamp
        # for col in ['endTime', 'createTime', 'changeTime', 'startTime']:
        #     df_companies_9[col] = pd.to_datetime(df_companies_9[col], format='mixed')
        # Получаем параметры кампании из списка
        df_companies_9_params = (
            df_companies_9
            .loc[:, ['endTime', 'createTime', 'changeTime', 'startTime', 'advertId', 'unitedParams']]
            .explode(column='unitedParams')
            .reset_index(drop=True)
        )
        # Получаем параметры кампании в виде столбцов df
        df_companies_9_params = pd.concat([
            df_companies_9_params,
            pd.json_normalize(df_companies_9_params['unitedParams'])
        ], axis=1)
        # Получаем список номенклатур кампании
        df_companies_9_nms = df_companies_9_params.explode(column='nms').reset_index(drop=True)
        # Получаем список номенклатур кампании в виде столбцов df
        # df_companies_9_nms = pd.concat([df_companies_9_nms, pd.json_normalize(df_companies_9_nms['nms'])], axis=1)
        # Убираем лишние колонки
        df_companies_9_nms = df_companies_9_nms.loc[:, ['endTime', 'createTime', 'changeTime', 'startTime', 'advertId', 'nms']]
    else:
        df_companies_9_nms = pd.DataFrame()

    # Формируем итоговый список номенклатур по всем кампаниям
    df_companies_nms_all = pd.concat([df_companies_6_nms, df_companies_8_nms, df_companies_9_nms])
    # Переименовываем некоторые колонки для удобства
    df_companies_nms_all = df_companies_nms_all.rename(columns={
        'nms': 'Артикул WB',
        'advertId': 'Номер кампании',
    })
    # Получаем количество артикулов в кампании
    df_companies_nms_all['Количество артикулов в кампании'] = df_companies_nms_all.groupby('Номер кампании').transform('size')
    # Перемещаем колонки в начало для удобства
    df_companies_nms_all = move_columns(
        df_companies_nms_all,
        ['Номер кампании', 'Артикул WB', 'Количество артикулов в кампании'],
        position=0
    )
    # Формируем словарь с результатами
    result_companies_info = {
        'df_companies_list': df_company_list,
        'df_companies_nms_all': df_companies_nms_all
    }

    return df_companies_nms_all


# Функция получения статистики кампаний
def get_company_stats(
        headers,
        df_input_companies,
        request_type='interval'
):
    # Создаем копию для избежания изменений в оригинальном df
    df_companies_list = df_input_companies.copy()
    # df, в который будем помещать результаты выгрузки по апи
    df_companies_stats = pd.DataFrame()

    # Для запроса с интервалом
    # Для каждой кампании формируем интервал запроса данных статистики
    df_companies_list['interval'] = df_companies_list.apply(
        lambda row:  {'begin': row['date_start'], 'end': row['date_end']},
        axis=1
    )

    # Для запроса по дням
    # Переводим в timestamp
    # df_companies_list['datetime_start'] = pd.to_datetime(df_companies_list['date_start'])
    # df_companies_list['datetime_end'] = pd.to_datetime(df_companies_list['date_end'])
    # Создаем список дат с частотой одного дня
    df_companies_list['dates'] = df_companies_list.apply(
        lambda row: pd.date_range(row['date_start'], row['date_end'], freq='D').strftime('%Y-%m-%d').tolist(),
        axis=1
    )
    # Удаляем вспомогательные колонки
    df_companies_list = df_companies_list.loc[:, ~df_companies_list.columns.isin([
        'date_start', 'date_end', 'datetime_start', 'datetime_end'
    ])]

    # Разбиваем df на диапазоны по 100 измерений
    step = 100
    df_companies_list['id'] = np.arange(0, len(df_companies_list))
    df_companies_list['chunks'] = df_companies_list['id'].apply(lambda x: int(x/step) + 1)
    max_chunks = max(df_companies_list['chunks'])

    logger.info("Uploading companies stats")
    # Цикл по каждому диапазону
    for chunk in df_companies_list['chunks'].unique():
        # Выбираем нужный диапазон и колонки
        tmp_df_companies_list = df_companies_list.loc[
            df_companies_list['chunks'] == chunk,
            ['Номер кампании'] + [request_type]
        ]
        # Переименовываем колонку для соответствия параметрам запроса
        tmp_df_companies_list = tmp_df_companies_list.rename(columns={'Номер кампании': 'id'})
        # Переводим в словарь (это и будут параметры запроса)
        params_companies_stats = tmp_df_companies_list.to_dict(orient='records')
        logger.info(f"Uploading {chunk} of {max_chunks} chunks")
        # Делаем запрос
        resp_data_companies_stats = requests.post("https://advert-api.wildberries.ru/adv/v2/fullstats", headers=headers, json=params_companies_stats)
        # Если пришли данные по кампаниям
        if resp_data_companies_stats.status_code == 200:
            # Переводим в словарь
            resp_data_companies_stats = resp_data_companies_stats.json()
            # Переводим в df
            tmp_df_companies_stats = pd.DataFrame(resp_data_companies_stats)
            # Объединяем с предыдущим проходом цикла
            df_companies_stats = pd.concat([df_companies_stats, tmp_df_companies_stats])
            # Ждем 1 минуту до выгрузки следующего диапазона кампаний (если это не последний диапазон)
            logger.info("Waiting 1 minute before uploading next chunk")
            if chunk != max_chunks:
                time.sleep(60)
        # Если не пришли данные по кампаниям, выводим ошибку
        else:
            print(f"{resp_data_companies_stats.json()}")
            if chunk != max_chunks:
                time.sleep(60)
                # Ждем 1 минуту до выгрузки следующего диапазона кампаний (если это не последний диапазон)
                logger.info("Waiting 1 minute before uploading next chunk")

    # Убираем лишние колонки
    # df_companies_stats.loc[:, ~df_companies_stats.columns.isin(['dates', 'days', 'boosterStats'])]

    # Получаем df с номенклатурами кампаний
    # df_companies_nms = result_companies_info['df_companies_nms_all']

    # Добавляем в df статистики кампаний исходные интервалы запроса
    # поскольку интервал, возвращаемый в ответе, не соответствует интервалу запроса
    df_request_interval = (
        pd.concat([
            df_companies_list[['Номер кампании']],
            pd.json_normalize(df_companies_list['interval'])
        ], axis=1)
        .rename(columns={
            'Номер кампании': 'advertId',
            'begin': 'Начальная дата запроса статистики РК',
            'end': 'Конечная дата запроса статистики РК'
        })
    )
    df_companies_stats_with_request_interval = (
        df_companies_stats
        .merge(
            df_request_interval,
            on='advertId',
            how='left'
        )
    )

    # Достаем interval из словаря
    df_companies_stats_with_response_interval = (
        pd.concat([
            df_companies_stats_with_request_interval,
            pd.json_normalize(df_companies_stats_with_request_interval['interval'])
        ], axis=1)
        # .rename(columns={
        #     'begin': 'date_start',
        #     'end': 'date_end',
        # })
    )
    # Переводим даты в timestamp
    # Создаем колонки с датой
    df_companies_stats_with_response_interval['Начальная дата ответа статистики РК'] = pd.to_datetime(df_companies_stats_with_response_interval['begin']).dt.date
    df_companies_stats_with_response_interval['Конечная дата ответа статистики РК'] = pd.to_datetime(df_companies_stats_with_response_interval['end']).dt.date
    # У конечной даты добавляем 23:59:59
    # time_to_add = timedelta(hours=23, minutes=29, seconds=59)
    # df_companies_stats_with_response_interval['Конечная дата ответа статистики РК'] = (
    #      df_companies_stats_with_response_interval['Конечная дата ответа статистики РК'] + time_to_add
    # )

    # Перемещаем колонки с ID кампании и датами в начало df
    companies_columns = [
        'advertId',
        'Начальная дата запроса статистики РК', 'Конечная дата запроса статистики РК',
        'Начальная дата ответа статистики РК', 'Конечная дата ответа статистики РК'
    ]
    df_companies_stats_with_response_interval = move_columns(
        df_companies_stats_with_response_interval,
        companies_columns,
        position=0
    )

    # Распаковываем статистику по дням
    df_companies_stats_unpacked = df_companies_stats_with_response_interval.explode(column='days', ignore_index=True)
    df_companies_stats_by_days = pd.concat([
        df_companies_stats_unpacked[companies_columns],
        pd.json_normalize(df_companies_stats_unpacked['days']),
        ], axis=1)

    # Распаковываем статистику по платформам
    df_by_days_unpacked = df_companies_stats_by_days.explode(column='apps', ignore_index=True)
    df_by_apps = pd.concat([
        df_by_days_unpacked[companies_columns],
        pd.json_normalize(df_by_days_unpacked['apps']),
        ], axis=1)

    # Распаковываем статистику по артикулам
    df_by_apps_unpacked = df_by_apps.explode(column='nm', ignore_index=True)
    df_by_nms = pd.concat([
        df_by_apps_unpacked[companies_columns],
        pd.json_normalize(df_by_apps_unpacked['nm']),
        ], axis=1)

    # Колонки, по которым производится группировка
    groupby_cols = companies_columns + ['nmId', 'name']
    # Считаем суммы по кампании и артикулу
    df_companies_stats_nms = (
        df_by_nms
        .loc[:,
             groupby_cols +
             ['views', 'clicks', 'sum', 'atbs', 'orders', 'shks', 'sum_price']]
        .groupby(groupby_cols)
        .sum()
        # .agg(**{
        #     'views': ('views', 'sum'),
        #     'clicks': ('clicks', 'sum')
        # })
        .reset_index()
    )
    # Считаем оставшиеся метрики
    df_companies_stats_nms = df_companies_stats_nms.assign(
        ctr=lambda df: df['clicks'] / df['views'] * 100,
        cr=lambda df: df['shks'] / df['clicks'] * 100,
        cpc=lambda df: df['sum'] / df['clicks'],
        cpm=lambda df: df['sum'] / df['views'] * 1000,
    )
    # Переименовываем колонки
    df_companies_stats_nms = df_companies_stats_nms.rename(columns={
        'advertId': 'Номер кампании',
        'nmId': 'Артикул WB',
        'name': 'Наименование товара',
        'views': 'Показы',
        'clicks': 'Клики',
        'sum': 'Затраты руб',
        'atbs': 'Добавлений в корзину',
        'orders': 'Заказы по рекламе шт (из отчета РК)',
        'sum_price': 'Заказы по рекламе руб (из отчета РК)',
        'shks': 'Заказанные товары шт (из отчета РК)',
        'ctr': 'CTR, %',
        'cr': 'CR, %',
        'cpc': 'CPC',
        'cpm': 'CPM',
    })

    # Создаем столбец с уникальным ID кампании по дате
    df_companies_stats_nms['unique_company_id'] = (
        df_companies_stats_nms
        .groupby(['Номер кампании', 'Начальная дата запроса статистики РК', 'Конечная дата запроса статистики РК'], as_index=False)
        .ngroup() + 1
    )
    # Порядок колонок
    df_companies_stats_nms = df_companies_stats_nms.loc[:,
        [
            'unique_company_id',
            'Номер кампании',
            'Начальная дата запроса статистики РК', 'Конечная дата запроса статистики РК',
            'Начальная дата ответа статистики РК', 'Конечная дата ответа статистики РК',
            'Артикул WB', 'Наименование товара',
            'Показы', 'Клики',
            'CTR, %', 'CR, %', 'CPC','CPM',
            'Затраты руб',
            'Добавлений в корзину',
            'Заказы по рекламе шт (из отчета РК)',
            'Заказанные товары шт (из отчета РК)',
            'Заказы по рекламе руб (из отчета РК)',
       ]
    ]
    # Перемещаем колонку в начало df
    # df_companies_stats_nms = move_columns(
    #     df_companies_stats_nms,
    #     columns_to_move=['unique_company_id'],
    #     position=0
    # )

    return df_companies_stats_nms


# Функция определения минимальной и максимальной даты в кампаниях
def get_min_max_dates(df_input_companies):
    # Создаем копию для избежания изменений в оригинальном df
    df_input_companies_ = df_input_companies.copy()
    # Переводим колонки в datetime
    df_input_companies_['datetime_start'] = pd.to_datetime(df_input_companies_['date_start'])
    df_input_companies_['datetime_end'] = pd.to_datetime(df_input_companies_['date_end'])
    # Находим мин. и макс. дату
    # GPT START----
    # Создаем столбец с минимумом по строке
    df_input_companies_['row_min'] = df_input_companies_[['datetime_start', 'datetime_end']].min(axis=1)
    df_input_companies_['row_max'] = df_input_companies_[['datetime_start', 'datetime_end']].max(axis=1)

    # Индекс общего минимума среди этих двух столбцов (по всем строкам)
    idx_min = df_input_companies_['row_min'].idxmin()
    idx_max = df_input_companies_['row_max'].idxmax()
    # Получаем label для строки с общим минимумом
    min_date = df_input_companies_.loc[idx_min, 'date_start']
    max_date = df_input_companies_.loc[idx_max, 'date_end']
    # Переводим в нужный формат
    min_date = min_date + 'T00:00:00'
    max_date = max_date + 'T23:59:59'
    # GPT END----

    return min_date, max_date


# Функция выгрузки заказов
def get_orders(headers, date_start, date_end):
    # Выгружаем заказа
    df_orders = getOrdersWB(headers, date_start, date_end, to_save=False)
    # Переводим колонку со временем заказа в timestamp
    df_orders['datetime_orders'] = pd.to_datetime(df_orders['date'])
    return df_orders


# Функция расчета заказов для отдельной рекламной кампании с фильтром по дате
def calc_orders_for_company(
        datetime_start,
        datetime_end,
        # begin_date_company,
        # end_date_company,
        # unique_company_id,
        # company_number,
        df_orders,
        tmp_df_company,
):
    # Создаем копию для избежания изменений в оригинальном df
    # df_orders_for_company = df_orders.copy()
    # Фильтруем заказы по датам кампании
    df_orders_date_filtered = (
        df_orders
        .loc[
            df_orders['datetime_orders'].between(datetime_start, datetime_end),
              :
            ]
            .copy()
    )
    # Считаем заказы
    df_orders_for_company = (
        df_orders_date_filtered
        .groupby(['Артикул WB'], as_index=False)
        .agg(**{
            'Заказы артикула шт (из отчета по заказам)' : ('Артикул WB', 'count'),
            'Заказы артикула руб (из отчета по заказам)': ('priceWithDisc', 'sum'),
        })
    )

    # Добавляем номер кампании к заказам
    # df_orders_for_company = df_orders_for_company.assign(**{
    #     'unique_company_id': unique_company_id,
    #     'Номер кампании': company_number,
    #     'begin': begin_date_company,
    #     'end': end_date_company,

    # })
    # Колонки для заполнения пропусков
    columns_to_fillna = ['Заказы артикула шт (из отчета по заказам)', 'Заказы артикула руб (из отчета по заказам)']
    # Мерджим заказы с компанией
    tmp_df_company_with_orders = (
        tmp_df_company
        .loc[:, ~tmp_df_company.columns.isin(['unique_company_id'])]
        .merge(
            df_orders_for_company,
            how='left',
            on='Артикул WB'
        )
        # Заполняем пропуски в колонках заказов
        .fillna({col: 0 for col in columns_to_fillna})
    )

    return tmp_df_company_with_orders

# Функция получения ассоциированных артикулов БЕЗ артикулов кампании
def get_associated_products(
        company_sku_list,
        df_products_processed
):
    # Удаляем дубликаты списка товаров, поскольку он распакован по размерам
    df_products_no_duplicates = (
        df_products_processed
        .loc[:, ['nmID', 'imtID', 'Артикул продавца', 'Наименование товара']]
        .drop_duplicates(ignore_index=True)
    )
    # Получаем список связанных артикулов артикула РК из списка товаров
    associated_skus = (
        df_products_processed
        .loc[df_products_processed['nmID'].isin(company_sku_list), 'imtID']
        .drop_duplicates()
        .to_list()
    )
    # Получаем df соответствия артикулов РК и их связанных артикулов
    df_associated_sku_all = (
        df_products_processed
        # Делаем выборку по ассоциированным артикулам
        .loc[
            df_products_processed['imtID'].isin(associated_skus),
            ['nmID', 'imtID', 'Артикул продавца', 'Наименование товара']
        ]
        # Удаляем дубликаты по артикулу
        .drop_duplicates(ignore_index=True)
        # .pipe(lambda df:
        #       # Получаем
        #       df.loc[df_associated_skus['nmID'].isin(company_sku_list), :]
        # )
    )
    # Получаем df артикула РК и ID общего артикула
    df_company_sku = (
        df_associated_sku_all
        .loc[df_associated_sku_all['nmID'].isin(company_sku_list), ['nmID', 'imtID']]
        # .groupby(['imtID'], as_index=False)
        # .agg({
        #     'nmID': lambda x: ';'.join(map(str, x))
        # })
        .rename(columns={
            'nmID': 'Артикул РК'
        })
    )
    # Получаем df связанных артикулов НЕ ВКЛЮЧАЯ артикулы РК
    df_associated_sku_company = (
        # Добавляем в df ассоциированных артикулов артикул РК
        # через merge по ID связанного артикула
        df_associated_sku_all
        .merge(
            df_company_sku,
            how='inner',
            on='imtID'
        )
        # Удаляем артикулы из РК
        .pipe(lambda df:
              df.loc[
                  ~df['nmID'].isin(company_sku_list),
                  ~df.columns.isin(['imtID'])
                ]
        )
        .rename(columns={
            'nmID': 'Ассоциированный артикул'
        })
    )

    # Один пайплайн (нужно тестить)
    # df_associated_sku_company = (
    #     df_products_no_duplicates
    #     # Делаем выборку по артикулам РК
    #     .loc[df_products_no_duplicates['nmID'].isin(company_sku_list), ['nmID', 'imtID']]
    #     # Удаляем дубликаты по артикулу
    #     # .drop_duplicates(ignore_index=True)
    #     # Переименовываем колонку артикулом РК
    #     .rename(columns={
    #         'nmID': 'Артикул РК'
    #     })
    #     .merge(
    #         df_products_no_duplicates[['nmID', 'imtID']],
    #         how='inner',
    #         on='imtID'
    #     )
    #     # Удаляем дубликаты по артикулу (поскольку список товаров распакован по размерам)
    #     # .drop_duplicates(ignore_index=True)
    #     # Удаляем артикулы из РК
    #     .pipe(lambda df:
    #           df.loc[
    #               ~df['nmID'].isin(company_sku_list),
    #               ~df.columns.isin(['imtID'])
    #             ]
    #     )
    #     # Переименовываем колонку с артикулами, которые пришли из списка товаров
    #     .rename(columns={
    #         'nmID': 'Ассоциированный артикул'
    #     })
    # )

    return df_associated_sku_company

# Функция расчета заказов по ассоциированным артикулам
def calc_associated_orders(
        datetime_start,
        datetime_end,
        company_sku_list,
        # begin_date_company,
        # end_date_company,
        # unique_company_id,
        company_number,
        df_products_processed,
        df_orders,
        tmp_df_company,
):
    # Получаем список ассоциированных артикулов кампании
    # БЕЗ самого артикула РК
    df_associated_sku_company = get_associated_products(
        company_sku_list,
        df_products_processed
    )
    # Получаем список для фильтрации заказов
    associated_sku_list = df_associated_sku_company['Ассоциированный артикул'].unique().tolist()
    # Создаем копию для избежания изменений в оригинальном df
    # df_orders_for_company = df_orders.copy()
    # Фильтруем заказы по датам кампании
    df_orders_date_filtered = (
        df_orders
        .loc[
            df_orders['datetime_orders'].between(datetime_start, datetime_end),
              :
            ]
            .copy()
    )
    # Фильтруем заказы по ассоциированным артикулам
    df_orders_sku_date_filtered = (
        df_orders_date_filtered
        .loc[
            df_orders_date_filtered['Артикул WB'].isin(associated_sku_list),
            :
        ]
    )
    # Считаем заказы
    df_orders_for_company = (
        df_orders_sku_date_filtered
        .groupby(['Артикул WB'], as_index=False)
        .agg(**{
            'Заказы артикула шт (из отчета по заказам)' : ('Артикул WB', 'count'),
            'Заказы артикула руб (из отчета по заказам)': ('priceWithDisc', 'sum'),
        })
        .rename(columns={
            'Артикул WB': 'Ассоциированный артикул'
        })
    )
    # Добавляем заказы к списку ассоицированнных артикулов
    df_associated_sku_with_orders = df_associated_sku_company.merge(
        df_orders_for_company,
        on='Ассоциированный артикул',
        how='left'
    )
    # Заполняем пропуски
    df_associated_sku_with_orders = df_associated_sku_with_orders.fillna(0)

    # Добавляем номер кампании к заказам
    df_associated_sku_with_orders = df_associated_sku_with_orders.assign(**{
        # 'unique_company_id': unique_company_id,
        'Номер кампании': company_number,
        'Начальная дата ответа статистики РК': datetime_start,
        'Конечная дата ответа статистики РК': datetime_end,
    })
    # Делаем порядок колонок
    df_associated_sku_with_orders = df_associated_sku_with_orders.loc[
        :,
        [
            'Номер кампании',
            'Начальная дата ответа статистики РК', 'Конечная дата ответа статистики РК',
            'Ассоциированный артикул', 'Артикул РК',
            'Артикул продавца', 'Наименование товара',
            'Заказы артикула шт (из отчета по заказам)',
            'Заказы артикула руб (из отчета по заказам)',
        ]
    ]

    return df_associated_sku_with_orders

# Функция расчета заказов для каждой из рекламных кампаний
def calc_orders_for_companies(
        df_input_companies,
        df_orders,
        df_companies_stats
):
    # df, куда будем помещать итоговый результат
    df_companies_stats_with_orders = pd.DataFrame()
    df_associated_orders = pd.DataFrame()
    # Цикл по каждой кампании
    for unique_company_id in df_input_companies['unique_company_id'].unique():
        # Временный df с данными текущей кампании
        tmp_df_company = (
            df_companies_stats
            .loc[df_companies_stats['unique_company_id'] == unique_company_id, :]
            # .drop_duplicates(subset=['Номер кампании'])
        )
        # Если вернулась статистика по данной РК, то начинаем расчет статистики
        if not tmp_df_company.empty:
            # Получаем номер кампании
            company_number = tmp_df_company['Номер кампании'].iloc[0]
            # Получаем даты начала и окончания кампании (для фильтра заказов)
            datetime_start_orders = (
                pd.to_datetime(tmp_df_company['Начальная дата ответа статистики РК'])
                .dt.tz_localize(None)
                .iloc[0]
            )
            datetime_end_orders = (
                pd.to_datetime(tmp_df_company['Конечная дата ответа статистики РК'])
                .dt.tz_localize(None)
                # У конечной даты заменяем время на 23:59:59
                .apply(lambda x: x.replace(hour=23, minute=59, second=59))
                .iloc[0]
            )
            # Эти даты нужны для мерджа
            begin_date_company = tmp_df_company['Начальная дата ответа статистики РК'].iloc[0]
            end_date_company = tmp_df_company['Конечная дата ответа статистики РК'].iloc[0]
            # Получаем список исходных SKU кампании
            company_sku_list = tmp_df_company['Артикул WB'].to_list()
            # Считаем заказы с фильтром по данным датам
            # и добавляем данные по заказам к кампании (расчет + merge в одной функции)
            tmp_df_company_with_orders = calc_orders_for_company(
                datetime_start_orders,
                datetime_end_orders,
                # begin_date_company,
                # end_date_company,
                # unique_company_id,
                # company_number,
                df_orders,
                tmp_df_company
            )
            # Считаем заказы по ассоциированным артикулам
            tmp_df_associated_orders = calc_associated_orders(
                datetime_start_orders,
                datetime_end_orders,
                company_sku_list,
                company_number,
                df_products_processed,
                df_orders,
                tmp_df_company
            )
        # Если статистика по РК не вернулась, то заказы по ней не считаем
        else:
            company_with_no_stats = df_input_companies.loc[
                df_input_companies['unique_company_id'] == unique_company_id,
                ['Номер кампании', 'date_start', 'date_end']
            ]
            logger.warning(f"No company stats for company\n {company_with_no_stats}")
            tmp_df_company_with_orders = pd.DataFrame()
            tmp_df_associated_orders = pd.DataFrame()
        # Объединяем с предыдущей кампанией
        df_companies_stats_with_orders = pd.concat([
                df_companies_stats_with_orders,
                tmp_df_company_with_orders
                ],
            ignore_index=True
        )
        df_associated_orders = pd.concat([
                df_associated_orders,
                tmp_df_associated_orders
                ],
            ignore_index=True
        )

    # Формируем словарь с результатами
    result_companies_orders = {
        'Статистика РК': df_companies_stats_with_orders,
        'Ассоциированные заказы': df_associated_orders
    }

    return df_companies_stats_with_orders, df_associated_orders


# Функция расчета итоговых колонок в статистике РК
def calc_final_orders(df_companies_stats_with_orders):
    # Достаем df со статистикой РК из словаря
    df_companies_stats_final_orders = df_companies_stats_with_orders.copy()
    # Считаем ДРР
    df_companies_stats_final_orders['ДРР, %'] = np.where(
        df_companies_stats_final_orders['Заказы артикула руб (из отчета по заказам)'] > 0,
        (
            df_companies_stats_final_orders['Затраты руб']
            / df_companies_stats_final_orders['Заказы артикула руб (из отчета по заказам)']
            * 100
        ),
        np.nan
    )

    return df_companies_stats_final_orders


# Функция выгрузки статистики по ключевым словам
def get_keywords_stats(headers, df_input_companies):


    # df, куда будем помещать итоговый результат
    df_keyword_company_stats_api = pd.DataFrame()

    # Делаем цикл по каждой кампании
    for idx, row in df_input_companies.iterrows():
        # Получаем номер кампании
        company_number = row['Номер кампании']
        # Получаем даты, за которые нужно получить статистику
        date_start_company = pd.to_datetime(row['date_start'])
        date_end_company = pd.to_datetime(row['date_end'])
        # Разбиваем диапазон дат кампании на интервалы по 7 дней
        date_range = []
        # Генерируем даты с разницей не более 8 дней и без пересечений
        current_date = date_start_company
        while current_date <= date_end_company:
            # Устанавливаем конечную дату как текущую дату плюс 7 дней или дату окончания, если она раньше
            new_end_date = min(current_date + pd.Timedelta(days=6), date_end_company)
            date_range.append({
                'advertId': company_number,
                'date_from': current_date,
                'date_to': new_end_date
            })
            # Переходим к следующей начальной дате (добавляем 8 дней)
            current_date = new_end_date + pd.Timedelta(days=1)
        # Создаем DataFrame
        df_date_range = pd.DataFrame(date_range)
        # Создаем даты для запроса
        df_date_range = df_date_range.assign(
            date_from_request=lambda df: df['date_from'].dt.strftime('%Y-%m-%d'),
            date_to_request=lambda df: df['date_to'].dt.strftime('%Y-%m-%d'),

        )

        # Считаем, сколько получилось интервалов
        total_intervals = df_date_range.shape[0]
        # Сколько выгружено интервалов
        uploaded_intervals = 0

        # Цикл по каждому интервалу
        for idx, row in df_date_range.iterrows():
            logger.info(
                f"\n"
                f"Uploading keyword statistics: \n"
                f"advertID: {company_number}: \n"
                f"dates: {row['date_from_request']} - {row['date_to_request']}"
            )
            # Формируем параметры запроса
            params_keyword_companies = {
                'advert_id': company_number,
                'from': row['date_from_request'],
                'to': row['date_to_request']
            }
            # Делаем запрос
            resp_data_keyword_companies = requests.get(
                'https://advert-api.wildberries.ru/adv/v0/stats/keywords',
                headers=headers,
                params=params_keyword_companies
            )
            # Если запрос был успешным, создаем df из данных апи
            if resp_data_keyword_companies.status_code == 200:
                resp_data_keyword_companies = resp_data_keyword_companies.json()
                # Переводим в df
                tmp_df_keyword_companies = pd.DataFrame(resp_data_keyword_companies)
                # Добавляем номер кампании
                tmp_df_keyword_companies['Номер кампании'] = company_number

                # Объединяем с предыдущим проходом цикла\
                df_keyword_company_stats_api = pd.concat([
                    tmp_df_keyword_companies,
                    df_keyword_company_stats_api,
                ], ignore_index=True)
            # Если была ошибка, выводим текст ошибки
            else:
                print(resp_data_keyword_companies.text)

    # Обработка данных кампаний поиска по ключевым фразам
    # Достаем данные из словарей
    df_companies_keywords_normalized = (
        pd.concat([
            df_keyword_company_stats_api,
            pd.json_normalize(df_keyword_company_stats_api['keywords'])
        ], axis=1)
    )
    # Достаем данные из списков
    df_companies_keywords_unpacked = df_companies_keywords_normalized.explode(['stats'], ignore_index=True)
    # Достаем данные из словарей списков
    df_companies_keywords_stats = (
        pd.concat([
            df_companies_keywords_unpacked,
            pd.json_normalize(df_companies_keywords_unpacked['stats']),
        ], axis=1)
        # Убираем колонку со словарями
        .pipe(
            lambda df:
                df.loc[:, ~df.columns.isin(['stats'])]
        )
        # Переименовываем колонки
        .rename(columns={
            'date': 'Дата',
            'keyword': 'Ключевая фраза',
            'views': 'Просмотры',
            'clicks': 'Клики',
            'ctr': 'CTR',
            'sum': 'Затраты'
        })
    )
    # Переводим колонку с датой в datetime
    df_companies_keywords_stats['Дата'] = pd.to_datetime(df_companies_keywords_stats['Дата'])
    # Сортируем по затратам
    df_companies_keywords_stats = (
        df_companies_keywords_stats
        .sort_values(
            by=['Номер кампании', 'Затраты'],
            ascending=False,
            ignore_index=True
        )
    )
    # Делаем удобный порядок колонок
    df_companies_keywords_stats = df_companies_keywords_stats.loc[
        :,
        [
            'Номер кампании',
            'Дата',
            'Ключевая фраза',
            'Затраты',
            'Клики',
            'Просмотры',
            'CTR'
        ]
    ]

    return df_companies_keywords_stats


# Функция формирования словаря с итоговыми результатами
def create_result_dictionary(
        df_final_orders,
        df_associated_orders,
        df_companies_keywords_stats
):
    # Добавляем статистику по ключевым словам в словарь с кампаниями
    result_companies_report = {
        'Статистика РК': df_final_orders,
        'Ассоциированные заказы': df_associated_orders,
        'Ключевые слова': df_companies_keywords_stats
    }

    return result_companies_report


# Функция сохранения excel
def save_and_format_excel(
        report_path,
        date_campaign_report,
        client_name,
        result_companies_report
):
    logger.info(f"Formatting campaigns report for client {client_name}")

    # Имя файла для сохранения
    file_name = f"{report_path}/{date_campaign_report}_Отчет_РК_{client_name}_WB.xlsx"

    # Создаем книгу excel и сохраняем в неё df из словаря с результатами
    with pd.ExcelWriter(file_name) as w:
        for sheet_name, df in result_companies_report.items():
            df.to_excel(w, index=False, sheet_name=sheet_name)

    # Форматирование файла с отчетом
    # GPT START----
    wb = openpyxl.load_workbook(file_name)

    # Создаем границы для ячеек
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)


    # 1. Задаем список названий столбцов, которые НЕ надо форматировать как числа
    exclude_format_headers = [
    'Артикул', 'Ассоциированный артикул', 'Номер кампании',
    ]
    # Цикл по каждому листу
    for ws in wb.worksheets:
        # 2. Сохраняем названия колонок для дальнейшей фильтрации
        headers = [cell.value for cell in ws[1]]

        # 3. Определяем индексы столбцов, которые НЕ нужно форматировать как числа
        exclude_format_idx = [
            idx + 1  # openpyxl считает колонки с 1
            for idx, header in enumerate(headers)
            if header in exclude_format_headers
        ]

        # 4. Обходим все ячейки данных (начиная со 2-й строки)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for col_idx, cell in enumerate(row, start=1):
                # 5. Пропускаем столбцы, для которых не нужно форматирование чисел
                if col_idx in exclude_format_idx:
                    continue
                # 6. Форматируем целые числа: разделитель тысяч, без дробной части
                if isinstance(cell.value, int):
                    cell.number_format = '#,##0'
                # 7. Форматируем дробные числа:
                #    - Если значение "целое" по факту (например, 10000.0), показываем без знаков после запятой
                #    - Иначе — 2 знака после запятой, разделитель тысяч
                elif isinstance(cell.value, float):
                    if cell.value % 1 == 0:
                        cell.number_format = '#,##0'
                    else:
                        cell.number_format = '#,##0.00'

        # Границы и автоподбор ширины
        column_widths = {}
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell_value_len = len(str(cell.value)) if cell.value is not None else 0
                column_widths[cell.column] = max(column_widths.get(cell.column, 0), cell_value_len)

        for col_idx, width in column_widths.items():
            # Для строк заголовков добавляем запас под фильтр
            header_len = len(str(ws.cell(row=1, column=col_idx).value)) if ws.cell(row=1, column=col_idx).value else 0
            # Даём больше места, если ширина определилась по заголовку
            reserve = 5 if width == header_len else 2
            ws.column_dimensions[get_column_letter(col_idx)].width = width + reserve

        # Высота загловков
        ws.row_dimensions[1].height = 25
        # Фильтр на первый ряд (заголовки)
        ws.auto_filter.ref = ws.dimensions

        # Перенос текста в заголовках (первая строка)
        for cell in ws[1]:
            cell.alignment = Alignment(
                vertical='center',
                horizontal='left',
                wrap_text=True
            )

    wb.save(file_name)

    # GPT END----


# Функция получения истории затрат
# def get_costs_history(headers, date_start, date_end):
#     # Разбиваем диапазон дат на периоды по 1 месяцу каждый

#     # GPT START ----
#     # Преобразование строковых дат в datetime
#     dt_start = pd.to_datetime(date_start).tz_localize(None)
#     dt_end = pd.to_datetime(date_end).tz_localize(None)

#     # Создание списка интервалов
#     intervals = []

#     # Первый интервал от начальной даты до конца месяца
#     first_month_end = (dt_start + pd.DateOffset(months=1)).replace(day=1) - pd.Timedelta(seconds=1)
#     if first_month_end > dt_end:
#         first_month_end = dt_end

#     intervals.append({
#         'date_start': dt_start.strftime('%Y-%m-%d'),
#         'date_end': first_month_end.strftime('%Y-%m-%d'),
#         'dt_start': dt_start,
#         'dt_end': first_month_end
#     })

#     # Следующий интервал от начала следующего месяца до конца месяца
#     current_start = first_month_end + pd.Timedelta(seconds=1)
#     while current_start <= dt_end:
#         monthly_start = current_start.replace(day=1)
#         monthly_end = (monthly_start + pd.DateOffset(months=1)).replace(day=1) - pd.Timedelta(seconds=1)

#         if monthly_end > dt_end:
#             monthly_end = dt_end

#         intervals.append({
#             'date_start': monthly_start.strftime('%Y-%m-%d'),
#             'date_end': monthly_end.strftime('%Y-%m-%d'),
#             'dt_start': monthly_start,
#             'dt_end': monthly_end
#         })

#         # Переход к следующему месяцу
#         current_start = monthly_start + pd.DateOffset(months=1)

#     # Создание датафрейма
#     date_range_df = pd.DataFrame(intervals)

#     # df, куда будем помещать результаты истории затрат
#     df_history_costs = pd.DataFrame()

#     # Сколько всего нужно выгрузить интервалов
#     total_intervals = date_range_df.shape[0]

#     # Делаем запрос для каждого интервала
#     for i in range(date_range_df.shape[0]):
#         logger.info(f"Uploading history costs for dates "
#                     f"{date_range_df['date_start'][i]} - {date_range_df['date_end'][i]}"
#                     )
#         # Параметры запроса
#         params_history_costs = {
#             'from': date_range_df['date_start'][i],
#             'to': date_range_df['date_end'][i]
#         }
#         # Делаем запрос
#         result_history_costs = requests.get('https://advert-api.wildberries.ru/adv/v1/upd',
#                                             headers=headers,
#                                             params=params_history_costs).json()
#         # Получаем результаты
#         tmp_df_history_costs = pd.DataFrame(result_history_costs)
#         # Добавляем в резульаты даты выгрузки
#         tmp_df_history_costs = tmp_df_history_costs.assign(
#             date_start = date_range_df['date_start'][i],
#             date_end = date_range_df['date_end'][i]
#         )

#         # Объединяем с предыдущим проходом цикла
#         df_history_costs = pd.concat([df_history_costs, tmp_df_history_costs])

#         # Ждем 1 минуту перед следующим запросом, если это не последний интервал
#         if (i + 1) != total_intervals:
#             logger.info("Wating 1 minute before uploading next interval")
#             time.sleep(60)
#         else:
#             logger.info("Done uploading history costs")

#     # Сбрасываем index после concat
#     df_history_costs = df_history_costs.reset_index(drop=True)

#     # Считаем расходы на каждую из кампаний
#     df_history_costs_stats = (
#         df_history_costs
#         .groupby(['advertId'])
#         #.groupby(['updNum', 'advertId'])
#         .agg(
#             Расходы=('updSum', 'sum')
#         )
#         .reset_index()
#         .rename(columns={
#             'updNum': 'Номер документа',
#             'advertId': 'ID Кампании'
#         })
#     )

#     return df_history_costs_stats

# # Функция разбивки рекламных расходов по артикулам
# def calc_costs_by_sku(result_companies_info, df_costs_history_stats):
#     # Получаем список номенклатур в кампаниях
#     df_companies_nms_all = result_companies_info['df_companies_nms_all'].copy()
#     # Убираем колонку с количеством артикулов в кампании, чтобы посчитать её заново
#     df_companies_nms_all = df_companies_nms_all.loc[:, ~df_companies_nms_all.columns.isin(['Количество артикулов в кампании'])]
#     # Получаем количество артикулов в кампании
#     df_companies_nms_all['Количество артикулов в кампании'] = df_companies_nms_all.groupby('advertId').transform('size')

#     # Мерджим список номенклатур с корректными расходами на кампанию
#     df_companies_sku = df_costs_history_stats.merge(df_companies_nms_all,
#                                                     on='advertId',
#                                                     how='left')

#     # Разбиваем расходы по артикулам, если их было несколько в одной кампании
#     df_companies_sku['Разбивка расходов на кампанию'] = df_companies_sku['Расходы'] / df_companies_sku['Количество артикулов в кампании']

#     # Считаем расходов по каждому артикулу
#     df_companies_sku_stats = (
#         df_companies_sku
#         .groupby(['Артикул WB'])
#         .agg(
#             Расходы_на_продвижение=('Разбивка расходов на кампанию', 'sum')
#             )
#         .reset_index()
#     )

#     return df_companies_sku_stats


# %% Вызов всех функций
if __name__ == '__main__':
    # Список кампаний и даты для каждой кампании,
    # за которые нужно сформировать отчет
    input_companies = [
        # ID Кампании: ['Начальная дата', 'Конечная дата']
        # 24246401: ['2025-05-19', '2025-05-25'],
        # 26111907: ['2025-05-29', '2025-06-04'],
        [27163894, '2025-07-21', '2025-07-28'],  # Орский Комбинат
        [27098220, '2025-07-21', '2025-07-28'],  # Орский Комбинат
        [27218787, '2025-07-21', '2025-07-28'],  # Орский Комбинат
        [27361337, '2025-07-21', '2025-07-28'],  # Орский Комбинат
        [27361332, '2025-07-21', '2025-07-28'],  # Орский Комбинат
    ]
    # Дата формирования отчета
    date_campaign_report = str(date.today())
    # Создаем df с нужным списком кампаний и датами
    df_input_companies = create_companies_upload_df(input_companies)
    # Получаем список кампаний
    df_company_list = get_company_list(headers)
    # Получаем список номенклатур кампании
    df_companies_nms = get_companies_info(headers, df_input_companies)
    # Получаем статистику кампаний
    df_companies_stats = get_company_stats(headers, df_input_companies, request_type='interval')
    # Получаем список товаров
    df_products = getWBProduct(headers, to_save=False)
    # Обрабатываем список товаров
    df_products_processed = process_product_list(df_products)
    # Находим мин. и макс. дату для получения заказов
    date_start_orders, date_end_orders = get_min_max_dates(df_input_companies)
    # Выгружаем заказы
    df_orders = get_orders(headers, date_start_orders, date_end_orders)
    # Считаем заказы для каждой кампании
    df_companies_stats_with_orders, df_associated_orders = calc_orders_for_companies(
        df_input_companies,
        df_orders,
        df_companies_stats
    )
    # Получаем статистику по ключевым фразам в поиске
    df_companies_keywords_stats = get_keywords_stats(
        headers,
        df_input_companies
    )
    # Получаем историю затрат (для корректных расходов на кампании)
    # df_costs_history_stats = get_costs_history(headers, date_start, date_end)
    # Считаем расходы на номенклатуры в кампании
    # df_companies_sku_stats = calc_costs_by_sku(result_companies_info, df_costs_history_stats)

    # Считаем ДРР и итоговые заказы
    df_final_orders = calc_final_orders(df_companies_stats_with_orders)
    # Создаем итоговый словарь с результатами
    result_companies_report = create_result_dictionary(
        df_final_orders,
        df_associated_orders,
        df_companies_keywords_stats,
    )
    # Создаем директорию, куда будет сохранен отчет
    report_path = create_report_path(date_campaign_report, client_name)
    # Сохраняем и форматируем файл с отчетом
    save_and_format_excel(
        report_path,
        date_campaign_report,
        client_name,
        result_companies_report
    )
    logger.info(f"Done creating campaigns report for client {client_name}")
# %%
